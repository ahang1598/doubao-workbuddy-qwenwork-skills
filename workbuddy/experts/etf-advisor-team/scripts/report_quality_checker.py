#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
"""股票交易报告质量校验器

目的：把报告规范中的结构、三段式、关键要素覆盖要求落成可执行门禁，
避免把“快速回答”误交付成“正式报告”，并把质量标准从“篇幅优先”升级为“内容完整性优先”。

用法示例：
    python report_quality_checker.py OutputReport/交易决策报告_300394_天孚通信.md --style swing --require-supply-demand
    python report_quality_checker.py OutputReport/交易决策报告_000776_广发证券.md --style long
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple
import sys

# Windows PowerShell / CMD 下强制 stdout/stderr 使用 UTF-8 编码，
# 避免特殊字符被 GBK/CP936 吃掉导致整段输出丢失。
if sys.platform == "win32":
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

STYLE_LABELS = {
    "ultra_short": "超短线",
    "short": "短线",
    "swing": "波段",
    "long": "中长线",
    "full": "全周期覆盖",
}

COMMON_REQUIRED = [
    "### 一、核心结论与操作指令",
    "### 二、风险红线排查",
    "### 三、大盘与板块环境",
    "### 四、多维度分析",
    "### 五、",
    "### 补充说明",
    "**数据来源**",
]

# 兼容两种章节结构：
#  旧结构（6 章）：五、精准交易计划 → 六、风险预案
#  新结构（7 章）：五、综合研判 → 六、精准交易计划 → 七、风险预案
# 至少存在其中一种"风险预案"章节即可。
RISK_PLAN_ALTERNATIVES = [
    "### 六、风险预案",  # 旧结构
    "### 七、风险预案",  # 新结构（新增"五、综合研判"章后编号顺延）
]

# ═══════════════════════════════════════════════════════════════════════
# 六个面框架（v2）报告类型识别
# ───────────────────────────────────────────────────────────────────────
# 本团队报告按"六个面框架"组织，门禁按文件名前缀分流（详见
# references/six_dimension_framework.md §七 术语映射规范）：
#   交易决策报告_  → 汇总决策报告（Intent-1）：report_type="trade" 完整门禁
#   基本面_ → 基本面单面报告：report_type="fundamental"
#        v10：一律视为深度研究，无条件触发三件套 GATE0 + 推导链审计（不再区分深度/普通基本面）
#   政策面_/资金面_/筹码面_/技术面_/消息面_ → 其余五个面单面报告：
#       v22 起**与 Intent-1 分面深稿共用同一 face_contract + 同一 _eval_single_face_draft**，
#       深度（字数/表格/小节/脚标/信源URL表/门禁自检/required_elements 分析纵深）严格相等；
#       仅【摘要卡】为 Intent-1 阶段A 交接专用、单面报告不强制。
# 说明：report_type 的字面值 "trade"/"fundamental" 是程序内部标识（非用户可见的
#       意图分类文案），保留；所有用户可见提示一律用业务名（汇总决策报告/基本面报告）。
# ═══════════════════════════════════════════════════════════════════════
DIMENSION_REPORT_PREFIXES = ("政策面_", "资金面_", "筹码面_", "技术面_", "消息面_")
# 深度基本面标记：v10 起所有基本面报告一律走深度门禁，本清单不再作为"是否深度"的门槛开关，
# 仅保留供推导链审计/其它启发式参考。
DEEP_FUNDAMENTAL_MARKERS = (
    "三表预测", "4.1.4B", "盈利预测表", "forecast", "EPS 预测", "EPS预测",
    "DCF", "自由现金流折现", "三表勾稽",
)


def is_dimension_single_report(name: str) -> bool:
    """是否为政策/资金/筹码/技术/消息 五个面之一的单面报告。
    v22：这五个面单面报告与 Intent-1 分面深稿共用同一 face_contract 契约校验，深度相等。"""
    return any(name.startswith(p) for p in DIMENSION_REPORT_PREFIXES)


def is_fundamental_report_name(name: str) -> bool:
    """是否为基本面单面报告（文件名前缀 基本面_）。"""
    return name.startswith("基本面_")


def needs_deep_fundamental_gate(report_path: Path) -> bool:
    """是否触发深度基本面三件套 GATE0 + 推导链审计。

    v10 取消「深度/普通基本面」二分——**所有 `基本面_` 前缀报告一律视为深度研究**，
    无条件触发三件套 GATE0 + 推导链审计（不再依赖 DEEP_FUNDAMENTAL_MARKERS 标记）。
    即：每一份基本面报告都必须具备 预测三件套 + 卖方级三表 + 可比/激励/ESG 等深度要素，
    不存在"轻量基本面单面报告"这一低配档。
    """
    return report_path.name.startswith("基本面_")


def needs_trade_fundamental_gate(report_text: str) -> bool:
    """Intent-1 汇总决策报告（交易决策报告_）是否一并强制三件套 GATE0。

    v1.23 修复（路由层漏洞 A）：旧版三件套 GATE0 仅 `基本面_` 前缀触发，
    导致 `交易决策报告_` 前缀的 Intent-1 报告（含完整 4.1.4 盈利预测五段式）
    完全绕过三件套硬拦截——直接违反团队铁律2「**含 Intent-1 的基本面章节**，
    必须先跑 historical→assumptions→forecast 三件套」。

    触发判据：报告含 4.1.4 盈利预测子节（即做了卖方级盈利预测/三情景 EPS）。
    短线/超短线仅做排雷级基本面、无 4.1.4，则不触发（避免误伤）。
    """
    if not report_text:
        return False
    has_414 = "4.1.4" in report_text
    has_forecast_intent = any(
        kw in report_text for kw in ("盈利预测", "情景", "EPS", "归母净利", "净利预测")
    )
    return has_414 and has_forecast_intent


def _forecast_scenarios_complete(forecast_path: Path):
    """校验 forecast_engine 产物 forecast.json 是否真含 bull/base/bear 三档非空 EPS/净利。

    这是「assumptions 真被填 + 引擎真算出结果」的充分证据：若假设留空（待填），
    forecast_engine 会输出 eps=None / net_profit_parent_yi=None。
    返回 (ok: bool, reason: str)。
    """
    try:
        fc = json.loads(forecast_path.read_text(encoding="utf-8"))
    except Exception as e:  # noqa: BLE001
        return False, f"解析失败：{e}"
    l4 = fc.get("L4")
    if not isinstance(l4, dict):
        return False, "缺少 L4 利润表预测节（forecast_engine 未正常产出）"
    for sc in ("bull", "base", "bear"):
        sc_obj = l4.get(sc)
        if not isinstance(sc_obj, dict) or not sc_obj:
            return False, f"L4 缺少情景档 {sc}"
        y1 = sc_obj.get("year_1") or {}
        if y1.get("net_profit_parent_yi") is None or y1.get("eps") is None:
            return False, (
                f"情景 {sc} 的 year_1 净利/EPS 为空——assumptions.yaml 命门假设"
                "（营收增速/毛利率）仍为'待填'，或 forecast_engine 未真正算出"
            )
    return True, "ok"


def _assumptions_critical_unfilled(yaml_path: Path):
    """检查 assumptions.yaml 利润表预测命门字段（L4 营收增速/毛利率三档）是否仍为'待填'。

    纯文本扫描，避免引入 yaml 运行时依赖。返回 (unfilled: bool, fields: List[str])。
    这两个字段直接决定 EPS 预测，是「三件套是否真填了」的最小命门集。
    """
    try:
        txt = yaml_path.read_text(encoding="utf-8")
    except Exception:  # noqa: BLE001
        return True, ["<无法读取 assumptions.yaml>"]
    unfilled: List[str] = []
    for field in ("revenue_growth_pct", "gross_margin_pct"):
        m = re.search(rf"{field}:\s*\n((?:[ \t]+\S.*\n?){{1,8}})", txt)
        if m and "待填" in m.group(1):
            unfilled.append(field)
    return (len(unfilled) > 0), unfilled


# ════════════════════════════════════════════════════════════════════════
# v1.32 问题1：根治"为过线而填数字"的数字游戏（财务预测假设必须源自调研，非凑门禁）
# ────────────────────────────────────────────────────────────────────────
# 用户投诉根因：历史上 base 增速/毛利率假设被反复回调，且 comment 直接写成
#   『…EPS 预计约14元，在一致预期27.01元-48%偏离(门禁50%以内)』
# —— 用"门禁阈值"而非"年报/季报/一致预期"论证数字，这是把财务预测当数字游戏的铁证。
# 两道硬校验（命中即 GATE0 FAIL）：
#   ①【命门假设须带依据+信源】每个命门假设块必须有非空 comment（推导依据）+ source（信源），
#      不得裸填一个数字。这把"信源/推导依据"从'软建议'升级为'硬门禁'，强制研究背书。
#   ②【数字游戏红旗】comment/source 出现"门禁/过线/凑/卡线"等表述 → 证明数字是为通过门禁
#      而设、非源自调研 → FAIL，强制改回研究口径。
# ════════════════════════════════════════════════════════════════════════
# 命门假设字段（直接驱动 EPS/估值）：营收增速 + 毛利率 + 四费率 + 等效税率 + WACC
_CRITICAL_ASSUMPTION_FIELDS = (
    "revenue_growth_pct", "gross_margin_pct",
    "sales_expense_ratio_pct", "management_expense_ratio_pct",
    "rd_expense_ratio_pct", "financial_expense_ratio_pct",
    "effective_tax_rate_pct", "wacc_pct",
)
# 【v1.31 bugfix · 字段命名漂移】assumptions_yaml_generator.py 的文档注释写的是不带
# _pct 后缀的字段名，但实际生成代码（yb(..., "xxx_expense_ratio_pct", ...)）一直是带
# _pct 后缀命名；而 historical_summary.margins 下恰好存在同名（不带 _pct）的只读参考字段
# （由脚本自动汇总、按设计本就没有 comment/source）。旧的 _CRITICAL_ASSUMPTION_FIELDS 用
# 不带 _pct 的名字去匹配，会先命中 historical_summary 里的参考字段（因 _iter_assumption_blocks
# 按 seen 去重、取文件里第一个同名块），对其报"缺 comment/source"——这是把只读历史数据误判为
# 裸填假设的假阳性，而真正的 L4_income_statement/valuation_inputs 里带 _pct 后缀的假设字段反而
# 从未被这条规则实际检查到。现改为与生成器实际产出的字段名一致（带 _pct），从根源修复误报。
# "为过线而填"的红旗词：依据里出现这些＝用门禁阈值倒推数字，而非用调研推导
_NUMBER_GAME_FLAGS = (
    "门禁", "过线", "硬凑", "凑过", "凑数", "凑到", "卡线", "压线", "踩线",
    "卡在", "卡到", "为了通过", "为通过", "刚好低于", "刚好<", "刚好在", "阈值以内", "门槛内",
)
_PLACEHOLDER_VALUES = {"", "待填", "tbd", "todo", "-", "—", "null", "none", "n/a", "na", "？", "?", "xxx"}


def _kv_value(blob: str, key: str) -> Optional[str]:
    """从块体里取 `key: value` 的标量值（去引号/空白）；无该键返回 None。"""
    m = re.search(rf"(?m)^\s*{key}:\s*(.*)$", blob)
    if not m:
        return None
    return m.group(1).strip().strip("\"'").strip()


def _kv_filled(blob: str, key: str) -> bool:
    v = _kv_value(blob, key)
    return v is not None and v.lower() not in _PLACEHOLDER_VALUES


def _iter_assumption_blocks(txt: str):
    """按缩进切出每个 `字段名:` 块体（字段行 + 其下更深缩进的连续行）。
    生成 (field_name, start_line_no(1基), block_text)。纯文本扫描，无 yaml 运行时依赖。"""
    lines = txt.splitlines()
    for i, line in enumerate(lines):
        m = re.match(r"^([ \t]*)([A-Za-z_]\w*):\s*$", line)
        if not m:
            continue
        indent = len(m.group(1))
        field = m.group(2)
        body: List[str] = []
        for j in range(i + 1, len(lines)):
            lj = lines[j]
            if lj.strip() == "":
                body.append(lj)
                continue
            ind_j = len(lj) - len(lj.lstrip())
            if ind_j <= indent:
                break
            body.append(lj)
        yield field, i + 1, "\n".join(body)


def _assumptions_sources_missing(yaml_path: Path) -> List[str]:
    """① 命门假设必须带 comment（推导依据）+ source（信源），否则=裸填数字 → 问题清单。"""
    try:
        txt = yaml_path.read_text(encoding="utf-8")
    except Exception:  # noqa: BLE001
        return ["<无法读取 assumptions.yaml>"]
    problems: List[str] = []
    seen: set = set()
    for field, ln, blob in _iter_assumption_blocks(txt):
        if field not in _CRITICAL_ASSUMPTION_FIELDS or field in seen:
            continue
        seen.add(field)
        miss = []
        if not _kv_filled(blob, "comment"):
            miss.append("comment 推导依据")
        if not _kv_filled(blob, "source"):
            miss.append("source 信源")
        if miss:
            problems.append(f"L{ln} `{field}` 缺{'、'.join(miss)}（命门假设不得裸填数字，须摊明依据与信源）")
    return problems


def _assumptions_number_game_redflags(yaml_path: Path) -> List[str]:
    """② 扫描全 yaml 的 comment/source，识别"为通过门禁而设数字"的红旗表述 → 问题清单。"""
    try:
        txt = yaml_path.read_text(encoding="utf-8")
    except Exception:  # noqa: BLE001
        return []
    hits: List[str] = []
    for i, line in enumerate(txt.splitlines(), start=1):
        s = line.strip()
        if not (s.startswith("comment:") or s.startswith("source:")):
            continue
        for flag in _NUMBER_GAME_FLAGS:
            if flag in s:
                snippet = s if len(s) <= 80 else s[:77] + "…"
                hits.append(f"L{i} 出现『{flag}』：{snippet}")
                break
    return hits


# ── v1.28 base 档信源校准合理性阈值（base 预测 vs 市场一致预期的相对偏离 %）──
_BASE_VS_CONSENSUS_WARN_PCT = 30.0   # 30%~50% → WARN（建议在 §5.x 逆向估值显式说明分歧）
_BASE_VS_CONSENSUS_FAIL_PCT = 50.0   # >50%    → FAIL（base 已偏到准 bull/bear 区，须重校准）


def _base_scenario_vs_consensus_divergence(forecast_path: Path):
    """v1.28 信源校准合理性硬校验：base（中性）档预测与市场一致预期偏离过大拦截。

    【设计动机 · 历史事故】assumptions.yaml 的 base 档量价/毛利率假设由 LLM 手填，门禁此前
    只查「三档非空 + 与正文一致 + 分部加总」，**不查 base 假设本身是否合理**。曾发生 LLM 把
    "保守判断"误塞进 base 档（base EPS 10.94 vs 一致预期 26.92，偏离 -59%），等于把准 bear
    情景当 base，导致 DCF/PE-Band 目标价与现价背离约 6 倍仍能 PASS。本校验把"base 不得无理由
    严重背离一致预期"做成机器兜底——base 档定位是『最可能情形』，保守应放 bear 档、乐观放 bull 档。

    【取数】全部来自 forecast.json（纯 JSON，无 yaml 运行时依赖）：
      year_1: base_eps=L4.base.year_1.eps  vs  market_consensus.eps_2026e.median
      year_2: base_eps=L4.base.year_2.eps  vs  market_consensus.eps_2027e.median（v1.30 新增）
      base_rev = L4.base.year_1.revenue_yi vs market_consensus.revenue_2026e.median
      偏离% = (base - consensus) / consensus × 100

    【v1.30 远期一致性补强（远期 EPS 一致性校验）】历史漏洞：本校验
      此前**只查 year_1**（次年 EPS，通常较温和），完全不查 year_2（300308 实测 base year_1
      EPS≈21.6 vs 一致预期 26.9 仅偏 -20% 顺利 PASS，但 year_2 EPS 被失控外推到 80.77
      vs 一致预期 46.36 偏 +74% 却无人拦截）。v1.30 把 year_2 一并纳入：取 year_1/year_2
      EPS 偏离的**较差者**驱动分级；year_2 一致预期缺失时仅跳过该年、不误伤。

    【分级】呼应团队规范「一致预期仅作参照系、分析师可有据偏离」，只在"极端背离"才 FAIL：
      |EPS 偏离| ≤ 30%             → ok（合理范围）
      30% < |EPS 偏离| ≤ 50%       → warn（建议 §5.x 逆向估值/「我vs市场」显式说明分歧）
      |EPS 偏离| > 50%             → fail（base 偏到准 bull/bear 区，须重校准或改用一致预期锚）
      （营收偏离 > 50% 但 EPS 在阈内 → 至少 warn：提示毛利率假设可能在对冲，值得复核）

    【不误伤】year_1 一致预期 EPS 缺失/为 0（冷门无覆盖标的）→ status='skip'，不参与拦截。

    返回 dict：{status, detail, eps_div(year_1), eps_div_y2, rev_div}
    """
    try:
        fc = json.loads(forecast_path.read_text(encoding="utf-8"))
    except Exception as e:  # noqa: BLE001
        return {"status": "skip", "detail": f"forecast.json 解析失败：{e}", "eps_div": None, "rev_div": None}

    base_node = (fc.get("L4", {}) or {}).get("base", {}) or {}
    b = base_node.get("year_1", {}) or {}
    b2 = base_node.get("year_2", {}) or {}
    mc = fc.get("market_consensus", {}) or {}

    def _safe(x):
        try:
            v = float(x)
            return v
        except (TypeError, ValueError):
            return None

    base_eps = _safe(b.get("eps"))
    base_rev = _safe(b.get("revenue_yi"))
    cons_eps = _safe((mc.get("eps_2026e", {}) or {}).get("median"))
    cons_rev = _safe((mc.get("revenue_2026e", {}) or {}).get("median"))

    if not cons_eps or cons_eps <= 0:
        return {"status": "skip",
                "detail": "无一致预期 EPS（market_consensus.eps_2026e.median 缺失/为0），跳过 base 校准合理性校验",
                "eps_div": None, "eps_div_y2": None, "rev_div": None}
    if base_eps is None:
        return {"status": "skip",
                "detail": "forecast.json base EPS 为空（应已被三档完整性校验拦截）",
                "eps_div": None, "eps_div_y2": None, "rev_div": None}

    eps_div = (base_eps - cons_eps) / cons_eps * 100.0
    rev_div = None
    if cons_rev and cons_rev > 0 and base_rev is not None:
        rev_div = (base_rev - cons_rev) / cons_rev * 100.0

    y1_label = b.get("year_label") or "次年"
    parts = [f"base {y1_label}E EPS={round(base_eps, 2)} vs 一致预期中位数={round(cons_eps, 2)} → 偏离 {eps_div:+.1f}%"]
    if rev_div is not None:
        parts.append(f"base 营收={round(base_rev, 1)}亿 vs 一致预期={round(cons_rev, 1)}亿 → 偏离 {rev_div:+.1f}%")

    # ── v1.30 远期一致性：year_2 vs eps_2027e（缺失则跳过该年，不误伤）──
    base_eps_y2 = _safe(b2.get("eps"))
    cons_eps_y2 = _safe((mc.get("eps_2027e", {}) or {}).get("median"))
    eps_div_y2 = None
    if base_eps_y2 is not None and cons_eps_y2 and cons_eps_y2 > 0:
        eps_div_y2 = (base_eps_y2 - cons_eps_y2) / cons_eps_y2 * 100.0
        y2_label = b2.get("year_label") or "后年"
        parts.append(f"base {y2_label}E EPS={round(base_eps_y2, 2)} vs 一致预期={round(cons_eps_y2, 2)} → 偏离 {eps_div_y2:+.1f}%")
    detail = "；".join(parts)

    # 取 year_1 / year_2 EPS 偏离的「较差者」驱动分级
    abs_eps = max(abs(eps_div), abs(eps_div_y2) if eps_div_y2 is not None else 0.0)
    abs_rev = abs(rev_div) if rev_div is not None else 0.0
    if abs_eps > _BASE_VS_CONSENSUS_FAIL_PCT:
        status = "fail"
    elif abs_eps > _BASE_VS_CONSENSUS_WARN_PCT or abs_rev > _BASE_VS_CONSENSUS_FAIL_PCT:
        status = "warn"
    else:
        status = "ok"
    return {"status": status, "detail": detail, "eps_div": eps_div, "eps_div_y2": eps_div_y2, "rev_div": rev_div}


def _render_forecast_diagnostic(forecast_path: Path) -> str:
    """v1.31 GATE0 FAIL 根因诊断渲染器（让 LLM 看见三件套 FAIL 的可计算根因）。

    【设计动机 · 请求 G/H】门禁此前只告诉 LLM「base 偏离一致预期 > 50% / 营收失控外推 → 重校准」，
    但**不告诉它哪里不合理**——LLM 看到 FAIL 往往只会盲目调一个总增速数字硬凑过线，而非反思
    「分部是否覆盖全部业务线？每个分部量价假设是否高估/低估？是否漏了营收来源？」。本函数从
    forecast.json 的 base 档 revenue_segments_detail + revenue_sanity + market_consensus 抽取**纯可
    计算事实**并结构化呈现，引导 LLM 做"算无遗策"式根因排查。

    【铁律 · 只做算术/一致性/定位，不替 LLM 做业务判断】本函数严禁输出"应该调到 X%""1.6T 太低"这类
    业务结论；只回答三类客观问题：
      ① 算术：各分部加总 = 多少？覆盖度（分部基数合计 ÷ 基准年实际营收）= 多少？是否 < 95%（疑漏业务线）？
      ② 一致性：分部加总反算的隐含总增速 vs 一致预期营收增速差多少？base EPS vs 一致预期差多少？
      ③ 定位：哪个分部贡献最大 / 哪个最拖累（负增长）？失控外推哨兵命中哪些档/年？
    最终"该不该这么填、调哪个假设"由 LLM 对照年报/季报自行判断。

    返回 markdown 字符串（无数据/解析失败则返回空串，调用方据此决定是否拼接）。
    """
    try:
        fc = json.loads(forecast_path.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        return ""

    def _f(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None

    base = (fc.get("L4", {}) or {}).get("base", {}) or {}
    y1 = base.get("year_1", {}) or {}
    rsd = y1.get("revenue_segments_detail") or {}
    mc = fc.get("market_consensus", {}) or {}
    lines: List[str] = []

    # ── A. 分部覆盖度（算术）：分部基数合计 vs 基准年实际营收 ──
    base_total = _f(rsd.get("base_total_yi"))
    base_actual = _f(rsd.get("base_year_actual_revenue_yi"))
    segments = rsd.get("segments") or []
    implied_g = _f(rsd.get("implied_year1_growth_pct"))
    method = rsd.get("method") or y1.get("revenue_method") or "未知"

    if segments:
        lines.append("### 🔬 GATE0 根因诊断：自下而上营收分部加总解剖（base 档 year_1）")
        lines.append("")
        lines.append(f"> 取数：`forecast.json :: L4.base.year_1.revenue_segments_detail`（method={method}）。"
                     f"以下全为**可计算事实**，不含业务结论；请据此对照年报/季报逐条排查。")
        lines.append("")
        # A1 覆盖度
        if base_total is not None and base_actual is not None and base_actual > 0:
            cover = base_total / base_actual * 100.0
            flag = "✅ 充分" if cover >= 95.0 else ("⚠️ **疑漏业务线**" if cover < 95.0 else "")
            lines.append(f"**① 分部覆盖度（是否漏了营收来源？）**")
            lines.append("")
            lines.append(f"- 各分部基数合计 = **{round(base_total, 2)} 亿** ÷ 基准年实际总营收 "
                         f"**{round(base_actual, 2)} 亿** = **覆盖度 {cover:.1f}%** {flag}")
            if cover < 95.0:
                lines.append(f"- ⚠️ 缺口 **{round(base_actual - base_total, 2)} 亿（{100.0 - cover:.1f}%）** 未被任何分部覆盖"
                             f"——请核对年报『分行业/分产品营业收入』，确认是否漏列业务线（如其他业务/海外/系统集成等），"
                             f"漏列会系统性低估总营收与 EPS。")
            lines.append("")

        # A2 分部明细 + 贡献 + 拖累定位
        seg_rows = []
        max_drag = None   # 最拖累（增速最低）
        max_drive = None  # 最大贡献（year_1 营收最高）
        sum_y1 = 0.0
        for s in segments:
            nm = s.get("name", "?")
            br = _f(s.get("base_revenue_yi"))
            vg = _f(s.get("volume_growth_pct"))
            pc = _f(s.get("price_change_pct"))
            gp = _f(s.get("growth_pct"))
            yr = _f(s.get("year_1_revenue_yi"))
            if yr is not None:
                sum_y1 += yr
            seg_rows.append((nm, br, vg, pc, gp, yr))
            if gp is not None and (max_drag is None or gp < max_drag[1]):
                max_drag = (nm, gp)
            if yr is not None and (max_drive is None or yr > max_drive[1]):
                max_drive = (nm, yr)
        lines.append("**② 各分部量价假设与贡献（每个分部的假设是否合理？）**")
        lines.append("")
        lines.append("| 分部 | 基数(亿) | 量增% | 价变% | 反算增速% | year_1(亿) | 占比% |")
        lines.append("|---|---|---|---|---|---|---|")
        for (nm, br, vg, pc, gp, yr) in seg_rows:
            pctshare = (yr / sum_y1 * 100.0) if (yr is not None and sum_y1 > 0) else None
            lines.append(
                f"| {nm} | {('%.1f' % br) if br is not None else '—'} "
                f"| {('%+.1f' % vg) if vg is not None else '—'} "
                f"| {('%+.1f' % pc) if pc is not None else '—'} "
                f"| {('%+.1f' % gp) if gp is not None else '—'} "
                f"| {('%.2f' % yr) if yr is not None else '—'} "
                f"| {('%.1f' % pctshare) if pctshare is not None else '—'} |"
            )
        lines.append(f"| **合计** | **{round(base_total, 1) if base_total is not None else '—'}** | — | — "
                     f"| — | **{round(sum_y1, 2)}** | 100.0 |")
        lines.append("")
        if max_drive:
            lines.append(f"- 🟢 最大贡献分部：**{max_drive[0]}**（year_1 ≈ {round(max_drive[1], 1)} 亿）"
                         f"——其量价假设直接决定总盘，请优先核对。")
        if max_drag:
            drag_flag = "（负增长，拖累总盘）" if max_drag[1] < 0 else ""
            lines.append(f"- 🔻 增速最低分部：**{max_drag[0]}**（反算增速 {max_drag[1]:+.1f}%）{drag_flag}"
                         f"——若年报显示该业务并未萎缩，则量增/价变假设可能过保守。")
        lines.append("")

        # A3 隐含总增速 vs 一致预期（一致性）
        cons_rev = _f((mc.get("revenue_2026e", {}) or {}).get("median"))
        lines.append("**③ 隐含总增速 vs 市场一致预期（你的自下而上 vs 卖方共识差多少？）**")
        lines.append("")
        if implied_g is not None:
            lines.append(f"- 分部加总反算隐含总增速 = **{implied_g:+.1f}%**（= Σ各分部 year_1 ÷ 基数合计 − 1）")
        l4_g = _f(y1.get("revenue_growth_pct"))
        if l4_g is not None and implied_g is not None and abs(l4_g - implied_g) > 0.5:
            lines.append(f"- ⚠️ 注意：L4 汇总行 revenue_growth_pct={l4_g:+.1f}% 与分部反算 {implied_g:+.1f}% **不一致**"
                         f"——汇总行口头增速被分部加总覆盖；以分部明细为准，请检查 assumptions.yaml 是否两处填了矛盾数字。")
        if cons_rev is not None and base_actual is not None and base_actual > 0:
            cons_g = (cons_rev - base_actual) / base_actual * 100.0
            if sum_y1 > 0:
                gap = sum_y1 - cons_rev
                lines.append(f"- 一致预期 year_1 营收 = **{round(cons_rev, 1)} 亿**（隐含同比 {cons_g:+.1f}%）"
                             f"；你的自下而上 = **{round(sum_y1, 2)} 亿** → 差 **{gap:+.1f} 亿**")
                if implied_g is not None and abs(implied_g - cons_g) > 30.0:
                    direction = "低于" if implied_g < cons_g else "高于"
                    lines.append(f"- ⚠️ 你的隐含增速 {implied_g:+.1f}% 显著{direction}一致预期 {cons_g:+.1f}%（差 "
                                 f"{abs(implied_g - cons_g):.0f} 个百分点）——若非有据逆向判断，"
                                 f"请回到分部层面核对：是哪个分部的量增/价变假设造成了这个缺口？是否漏了高增长新品（如更高速率产品/新客户放量）？")
        lines.append("")

    # ── B. 失控外推哨兵命中明细（定位）──
    sanity = fc.get("revenue_sanity", []) or []
    if sanity:
        lines.append("**④ 失控外推 / 天花板哨兵命中（远期是否脱离行业天花板？）**")
        lines.append("")
        lines.append("| 档 | 年 | 类型 | 说明 |")
        lines.append("|---|---|---|---|")
        for w in sanity:
            lines.append(f"| {w.get('scenario', '?')} | {w.get('year_label', '?')} | {w.get('type', '?')} "
                         f"| {w.get('msg', '')} |")
        lines.append("")

    if not lines:
        return ""

    lines.append("> 📌 以上仅为可计算事实定位。**请勿盲目调一个总增速数字硬凑过线**——"
                 "对照三件套 FAIL 根因排查 SOP（见「基本面深度三件套铁律」v1.31 八步排查 SOP），逐分部回到年报/季报核对量价假设、"
                 "确认分部覆盖全部业务线、锁定缺口来源后再重跑 forecast_engine.py。")
    lines.append("")
    return "\n".join(lines)


# 从基本面报告文件名中提取 6 位股票代码：基本面_{code}_
FUNDAMENTAL_CODE_RE = re.compile(r"基本面_(\d{6})_")

COMMON_SUBSECTIONS = [
    "#### 3.1 大盘环境",
    "#### 3.2 所属板块",
]

LONG_FORM_SUBSECTIONS = [
    "#### 4.1 基本面",
    "##### 4.1.1",
    "##### 4.1.2",
    "##### 4.1.3",
    "##### 4.1.4",
    "##### 4.1.5",
    "#### 4.2",
    "#### 4.3",
    "#### 4.4",
    "#### 4.5",
    "#### 4.6",
    "#### 4.7",
]

SECTION_MIN_TABLES = {
    "4.1.1": 1,
    "4.1.2": 3,
    "4.1.3": 3,
    "4.1.4": 4,
    "4.1.5": 3,
    # v11：五个非基本面面（4.2~4.6）的章节深度需对齐"单独输出该面报告"的深度，
    # 表格数随之从 1~2 提升到 2（单面报告普遍含 数据表 + 推导/价位/明细表 ≥2 张）。
    "4.2": 2,
    "4.3": 2,
    "4.4": 2,
    "4.5": 2,
    "4.6": 2,
    "4.7": 1,
}

# ═════════════════════════════════════════════════════════════════════
# 内容深度硬下限（基于历史优秀报告的分位数基准，v7 新增；v11 对齐"单面深度"重订）
#
# 基准参考（中际旭创/立讯精密/凯格精机/中国稀土 4 份优秀报告的效率字数均值）：
#   4.1.1 ~575   4.1.2 ~1773   4.1.3 ~1714   4.1.4 ~782   4.1.5 ~754
#   4.2 ~373     4.3 ~457      4.4 ~389      4.5 ~363    4.6 ~271    4.7 ~322
#   五、综合研判 ~100（部分报告五章写在其它地方）
#
# 五段式基本面（4.1.1~4.1.5）下限取均值的 60~70%，对"勉强合格但不够深"BLOCK。
#
# ── v11 关键变更（思路2：Intent-1 不再拆 6 份独立面文件，改出单份深稿）──
# 既然不再产出"政策面_/资金面_/筹码面_/技术面_/消息面_"独立报告，那么汇总决策报告里
# 这五个面的章节（4.2~4.6）就是该面的"唯一承载"，其深度必须**对齐单独输出该面报告
# 的深度**。实测单面报告效率字数：政策面~1370 / 技术面~1615 / 资金面~1364 /
# 筹码面~1386 / 消息面~1815。扣除标题/元信息/附录信源等报告级开销（约 15~25%），
# 章节核心分析正文 ≈ 1100~1500。故 4.2~4.6 下限由原 200~350 大幅提升到 850~1000，
# 强制每个面的章节"写到单面报告的深度"，而非汇总时的薄片化收束。
# ═════════════════════════════════════════════════════════════════════
SECTION_MIN_EFFECTIVE_LEN = {
    # ══════════════════════════════════════════════════════════════════════
    # v23 合稿内聚地板重锚定（根治"合稿章节可以只有深稿 1/7"的结构性漏洞）
    # ──────────────────────────────────────────────────────────────────────
    # 病根复盘（本次整改根因）：旧版 4.2~4.6 仅 800~950，而对应深稿地板是 5200~5800
    # （见各面 face_contract.min_eff_len），即门禁**明文允许合稿章节只有深稿的 ~1/7**。
    # 于是 LLM 把深稿"轻飘飘几句引用/请见"带过即可过门禁——这正是用户投诉
    # "合稿没把深稿全部纳入、只写了几句详见"的制度性根源。
    # v23 修复：合稿每个面章节是该面深度的**唯一最终承载**，深度须 = Intent-2 单面报告。
    # 故把 4.1~4.6 字数地板统一重锚定到"合格深稿同面**核心正文**（去顶部结论卡）效率字数下限 × 0.90"，
    # 并全部硬化为 BLOCK（见文末降级逻辑）。下列数值由一次性历史校准固化（effective_length 口径、
    # 去顶部结论卡），是**绝对底线、非理想目标**，不依赖任何具体个股报告或外部输出物：
    #   基本面 10443 / 政策 7371 / 技术 5927 / 资金 6111 / 筹码 5215 / 消息 5990。
    # 与 check_merge_cohesion() 的"合稿章节 ≥ 本次深稿核心 × 0.90"相对校验互为双保险：
    #   绝对地板锚定校准下限，相对校验确保"本次深稿被逐字内聚"。
    # ── §4.1 基本面：5 个子节合计 ≈ 10443 × 0.90 ≈ 9400（仍 BLOCK，v16 内聚铁律延续）──
    "4.1.1": 750,   # 宏观层：宏观驱动表 + 国际权威信源(Reuters/Bloomberg/FT) + 传导链 + 小结
    "4.1.2": 2500,  # 产业层：行业规模/竞争格局CR3/代际/ASP/供应链 + 商业模式 + 3-5年竞争演化
    "4.1.3": 3000,  # 企业层：财务时序 + 护城河五维 + ROIC/WACC + 资本配置 + 客户集中度 + M-Score
    "4.1.4": 1600,  # 盈利预测：供给/需求/成本/费用/三情景 + 分部加总量价 + 一致预期比较 + 灵敏度
    "4.1.5": 1550,  # 估值：DCF/PE-Band/可比/IRR 四法 + 隐含假设反推 + 敏感性矩阵
    # ── 4.2~4.6：= 同面深稿核心正文下限 × 0.90（v23 全部硬 BLOCK，不再是软建议）──
    "4.2": 6600,    # 政策面（核心下限 7371 × 0.90）：政策清单/传导链/受益弹性 + 百分制评分 + 减仓硬条件
    "4.3": 5300,    # 技术面（核心下限 5927 × 0.90）：均线/支撑压力/量价/形态/多周期 + 关键价位 + 评分
    "4.4": 5500,    # 资金面（核心下限 6111 × 0.90）：主力四阶段/北向/融资/G1-G6/逐日资金 + 评分
    "4.5": 4700,    # 筹码面（核心下限 5215 × 0.90）：股东/集中度/质押/解禁/CYQ成本分布 + 评分
    "4.6": 5400,    # 消息面（核心下限 5990 × 0.90）：事件定性/传导链/量化影响/情绪 + 研报辩证 + 评分
    "4.7": 250,
}

# 整份报告效率字数下限（参见篇幅预算规范）
# 效率字数 = 汉字数 + 英文单词数；略小于"肉眼字符数"的 40~50%
#   波段 建议篇幅 5000-9000 字符 → 效率字数 ≈ 6500+（取历史均值 11000 的 60%）
#   中长线/全周期 → 效率字数 ≈ 9000+
# v11：思路2 下 Intent-1 出单份深稿（五段式基本面 + 五个面各对齐单面深度），
#   整份字数自然抬高，下限随之上调，防止"各面薄片化拼装"绕过单面深度要求。
MIN_EFFECTIVE_LENGTH = {
    # v16：基本面 §4.1 深度对齐独立报告后（4.1 合计下限 ≈6850），整份下限相应上调，
    # 防止"基本面写够了、其余面薄片化拼装"或反之。基本面 ≈6850 + 五面 ≈4300 + 决策/结论 ≈2000。
    "swing": 11000,
    "long": 13000,
    "full": 14000,
    "short": 1500,
    "ultra_short": 800,
}

HEADING_RE = re.compile(r"^(#{2,5})\s+(.+)$")


def effective_length(text: str) -> int:
    """近似字数：中文字符 + 英文单词数。仅作信息参考，不直接作为拦截条件。"""
    chinese_chars = len(re.findall(r"[\u4e00-\u9fff]", text))
    english_words = len(re.findall(r"[A-Za-z0-9_]+", text))
    return chinese_chars + english_words


def detect_style(text: str) -> Optional[str]:
    for key, label in STYLE_LABELS.items():
        if f"**交易风格**: {label}" in text:
            return key
    if "全周期覆盖" in text:
        return "full"
    if "中长线" in text:
        return "long"
    if "波段" in text:
        return "swing"
    if "短线" in text:
        return "short"
    if "超短线" in text:
        return "ultra_short"
    return None


def collect_headings(lines: List[str]) -> List[Tuple[int, int, str]]:
    headings: List[Tuple[int, int, str]] = []
    for idx, line in enumerate(lines):
        match = HEADING_RE.match(line.strip())
        if match:
            level = len(match.group(1))
            headings.append((idx, level, line.strip()))
    return headings


def find_heading(headings: List[Tuple[int, int, str]], prefix: str) -> Optional[Tuple[int, int, str]]:
    for item in headings:
        if item[2].startswith(prefix):
            return item
    return None


# ════════════════════════════════════════════════════════════════════════
# v1.32 章节定位去脆弱化（根治"老因章节编号/标题层级 # 数量不符而误判 FAIL"）
# ────────────────────────────────────────────────────────────────────────
# 病根：旧版章节"存在性"判定用 `"##### 4.1.1" in text` 精确字符串匹配、章节"区间
# 切分"用 `next_level <= level`（即按 # 字面数量截断）。一旦深稿内聚进合稿后标题层级
# 与门禁基准（4.1.x=#####、4.x=####）的 # 数量不一致（例：深稿 ### 被整体下沉成 ####，
# 比 ##### 层级更高），就会同时触发"缺章节"误判 + 区间被提前截断导致字数/小节覆盖 FAIL。
# 修复：章节身份以**点分编号语义**（4.1 / 4.1.1）为准，容忍任意 # 数量；区间切分按
# **编号层级深度**而非 # 字面数量——4.1 一直延伸到 4.2（兄弟编号）或上级中文章节，
# 其内部的 4.1.x（更深编号）与散文小标题（更深 # 层级）不再误截。
# ════════════════════════════════════════════════════════════════════════
_NUM_HEADING_LINE_RE = re.compile(r"^#{1,6}\s+(\d+(?:\.\d+)*)")


def _secnum_depth(num: str) -> int:
    """点分编号深度：'4'→1，'4.1'→2，'4.1.1'→3。"""
    return num.count(".") + 1


def _prefix_secnum(prefix: str) -> Optional[str]:
    """从调用方 prefix（如 '#### 4.1' / '##### 4.1.4'）提取点分编号；中文章节前缀返回 None。"""
    m = re.search(r"(\d+(?:\.\d+)*)", prefix or "")
    return m.group(1) if m else None


def scan_numbered_headings(lines: Sequence[str]) -> List[Tuple[int, int, str, int]]:
    """扫描全部点分编号标题（容忍 1~6 个 #），返回 [(行号, # 数量, 编号, 编号深度)]。
    与 collect_headings（上限 5 个 #、不限编号）正交：本扫描专供编号语义切分使用，
    覆盖深稿内聚后可能出现的 ###### 级标题。"""
    out: List[Tuple[int, int, str, int]] = []
    for idx, line in enumerate(lines):
        s = line.strip()
        m = _NUM_HEADING_LINE_RE.match(s)
        if m:
            hashes = len(s) - len(s.lstrip("#"))
            num = m.group(1)
            out.append((idx, hashes, num, _secnum_depth(num)))
    return out


def has_numbered_heading(lines: Sequence[str], secnum: str) -> bool:
    """是否存在编号恰为 secnum 的标题（容忍任意 # 数量、容忍标题文字差异）。"""
    return any(num == secnum for (_i, _h, num, _d) in scan_numbered_headings(lines))


def _subsection_present(lines: Sequence[str], text: str, entry: str) -> bool:
    """章节"存在性"判定（v1.32 去脆弱化）：
      · entry 含点分编号（如 '#### 4.1 基本面' / '##### 4.1.1'）→ 只要存在同编号标题即视为存在，
        **容忍 # 数量与标题文字差异**（不再要求 `"##### 4.1.1" in text` 逐字符精确出现）；
      · entry 无点分编号（中文章节 '### 一、' / '**数据来源**'）→ 回退原全文包含判定。"""
    num = _prefix_secnum(entry)
    if num is not None:
        return has_numbered_heading(lines, num)
    return entry in text


def get_section_text(lines: List[str], headings: List[Tuple[int, int, str]], prefix: str) -> Optional[str]:
    """切出某章节正文。

    双模式：
      ① **编号语义模式**（prefix 含点分编号，如 '#### 4.1' / '##### 4.1.4'）——
         起点 = 编号恰为该编号的标题（容忍 # 数量）；终点 = 此后第一个满足以下任一的标题：
           · **同一编号家族**的兄弟/上级编号节（如 4.1 止于 4.2、4.1.4 止于 4.1.5）——
             合稿内联的**异家族**编号（深稿原始 2.x/3.x）一律不算边界；
           · 处于**全文最浅 # 层级**的真·顶层中文章节（如 4.6 止于 '## 五、综合研判'）。
         本节内部更深编号（4.1.x）、异家族编号（2.x）、更深散文/中文小标题（### 一、/######）
         一律不截断，彻底消除"合稿双编号体系 + 内联浅层中文章节 → 区间被提前截断"的误判。
      ② **回退模式**（prefix 无编号，如中文章节 '### 五、'）——沿用原 # 层级截断逻辑。
    """
    target_num = _prefix_secnum(prefix)
    if target_num is not None:
        sec_parts = target_num.split(".")
        nums = scan_numbered_headings(lines)
        start = next(((i, h, n, d) for (i, h, n, d) in nums if n == target_num), None)
        if start is None:
            return None
        start_idx, start_hashes, _n, target_depth = start
        end_idx = len(lines)
        seen = False

        def _num_is_boundary(cand_num: str) -> bool:
            """编号标题 cand_num 是否构成本节（sec_parts）的截断边界。
            关键：合稿会把深稿原始编号（如 2.4 / 2.4.1）内联进汇总编号（4.1.4）之下，
            形成**双编号体系**。深稿的 2.x 与本节 4.x **不同家族**，绝不能当成"更浅的兄弟节"
            而把本节正文提前截断（历史回归 bug：4.1.4 一进门就遇到内联的 #### 2.4 被切到只剩标题）。
            规则：① 本节后代（以 sec 为前缀且更深）→ 不截断（属节内）；② 顶层章节号(深度1)且
            与本节顶层不同 → 截断（新章节）；③ 同一家族（共享顶层）且深度≤本节深度 → 截断
            （兄弟/上级节，如 4.1.5 / 4.2 截断 4.1.4）；④ 其它家族（内联深稿 2.x 等）→ 不截断。"""
            cp = cand_num.split(".")
            if len(cp) > len(sec_parts) and cp[: len(sec_parts)] == sec_parts:
                return False  # 后代，节内
            if len(cp) == 1:
                return cp != sec_parts[:1]  # 不同顶层章节才算边界
            return cp[0] == sec_parts[0] and len(cp) <= len(sec_parts)

        # 合并编号标题与（用于上级中文章节边界判断的）collect_headings 标题，按行号排序
        boundary: List[Tuple[int, Optional[str], Optional[int]]] = []  # (行号, 编号串|None, # 层级)
        num_str_idx = {i: n for (i, _h, n, _d) in nums}
        for idx, lv, raw in headings:
            boundary.append((idx, num_str_idx.get(idx), lv))
        # collect_headings 上限 5 个 #，补入编号扫描里 collect 漏掉的标题（如 ###### / 编号行）
        seen_idx = {idx for (idx, _s, _l) in boundary}
        for (i, h, n, d) in nums:
            if i not in seen_idx:
                boundary.append((i, n, h))
        boundary.sort(key=lambda x: x[0])
        # 全文最浅 # 层级 = 顶层"章节"级（合稿里通常是 ## 一、…六、）。合稿会把深稿原样内联，
        # 深稿自带的 ### 一、/### 二、等中文小章节**比顶层更深**，绝不能当成汇总级章节边界
        # （历史回归 bug：#### 4.1 一进门就撞上内联的 ### 一、核心结论 被切到只剩标题）。
        # 故无编号标题只有处于**全文最浅层级**（真·顶层章节，如 ## 五、综合研判）才算边界。
        _doc_levels = [lv for (_i, lv, _r) in headings if lv is not None]
        min_doc_level = min(_doc_levels) if _doc_levels else start_hashes
        for idx, cand_num, lv in boundary:
            if idx == start_idx:
                seen = True
                continue
            if not seen:
                continue
            if cand_num is not None:
                if _num_is_boundary(cand_num):
                    end_idx = idx
                    break
            else:
                # 无编号标题：仅当其 # 层级 = 全文最浅层级（真顶层章节）才截断；
                # 内联深稿的更深中文章节（### 一、等）一律视为节内，不截断。
                if lv is not None and lv <= min_doc_level and lv < start_hashes:
                    end_idx = idx
                    break
        return "\n".join(lines[start_idx:end_idx]).strip()

    # ── 回退模式：prefix 无点分编号（中文章节等），沿用原 # 层级截断 ──
    current = find_heading(headings, prefix)
    if not current:
        return None
    start_idx, level, _ = current
    end_idx = len(lines)
    found_current = False
    for idx, next_level, _ in headings:
        if idx == start_idx:
            found_current = True
            continue
        if found_current and next_level <= level:
            end_idx = idx
            break
    return "\n".join(lines[start_idx:end_idx]).strip()


def count_tables(lines: Sequence[str]) -> int:
    count = 0
    in_table = False
    for line in lines:
        stripped = line.strip()
        is_table_line = stripped.startswith("|") and stripped.endswith("|")
        if is_table_line and not in_table:
            count += 1
            in_table = True
        elif not is_table_line:
            in_table = False
    return count


def check_markers(section_text: str) -> Dict[str, bool]:
    return {
        "has_data": "📊" in section_text,
        "has_analysis": "🔍" in section_text,
        "has_conclusion": "📌" in section_text,
    }


def contains_any(text: str, keywords: Sequence[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def missing_semantic_keywords(text: str, keyword_groups: Sequence[Tuple[str, Sequence[str]]]) -> List[str]:
    missing: List[str] = []
    for canonical, aliases in keyword_groups:
        if not contains_any(text, aliases):
            missing.append(canonical)
    return missing


def validate_marker_and_table_rules(section_id: str, section_text: str) -> List[str]:
    issues: List[str] = []

    # ── 三段式标记：按配对检查（不只是存在某个 emoji）
    marker_state = check_markers(section_text)
    if not all(marker_state.values()):
        missing = [key for key, value in marker_state.items() if not value]
        label_map = {"has_data": "📊 数据段", "has_analysis": "🔍 推导段", "has_conclusion": "📌 结论段"}
        issues.append(
            f"章节 {section_id} 缺少三段式标记：{', '.join(label_map[k] for k in missing)}（每个分析章节至少各 1 个）"
        )

    # ── 表格数量
    table_count = count_tables(section_text.splitlines())
    required_tables = SECTION_MIN_TABLES.get(section_id, 1)
    if table_count < required_tables:
        issues.append(f"章节 {section_id} 表格数量不足：当前 {table_count}，至少需要 {required_tables}")

    # ── 章节字数（v7 新增：防止章节空壳）
    # P1-1：纯字数不足降级为软建议（带 [字数建议] 前缀，主流程会从 issues 分流到 warnings，
    #        不再阻断 PASS）。真正决定深度的"要素/表格/三段式"仍是硬 FAIL。
    section_len = effective_length(section_text)
    required_len = SECTION_MIN_EFFECTIVE_LEN.get(section_id)
    if required_len and section_len < required_len:
        gap = required_len - section_len
        issues.append(
            f"[字数建议] 章节 {section_id} 篇幅偏短：当前效率字数 {section_len}，建议 ≥{required_len}（还差约 {gap} 字）。"
            f"字数不足不阻断 PASS（要素/表格/三段式齐全即视为达标），但建议补充数据表/推导段以提升深度。"
        )

    return issues


def check_duplicate_paragraphs(text: str, min_effective: int = 60,
                               sim_threshold: float = 0.90) -> List[str]:
    """v29 反复制粘贴填充闸（硬 FAIL）。

    背景：历史反复出现"同一段『万能分析』被逐字/近似复制进多个主题不同的小节"来凑字数的
    质量顽疾（如把讲财务应收账款的段落原样搬进讲估值的小节），造成"文字堆砌"。其**制度性
    诱因**是决策稿 §4.x 残留了 v23 合稿时代的字数目标，撰写者为凑字数而复制粘贴。字数目标
    在 v29 已对 faces-split 决策稿移除（见 section_snapshot / 字数建议分流），本闸作为**结果侧
    的最后一道拦截**：无论动机如何，只要整段近似重复即判不合格，强制"要么写原创、要么就精简"。

    规则：按空行切段，剔除标题/表格/列表/引用/图片等结构行；对"有效字数 ≥ min_effective"的
    正文段做规范化（仅保留中文字/字母数字）后两两比较，出现完全相同或相似度 ≥ sim_threshold
    的段落对，即判定为复制粘贴填充 → 硬 FAIL。阈值取高（0.90）以只抓真复制、避免误伤模板句式。
    """
    import difflib

    issues: List[str] = []
    # 1) 切段：空行分隔，聚合连续非空行为一个块
    blocks: List[str] = []
    buf: List[str] = []
    for raw in text.splitlines():
        if raw.strip() == "":
            if buf:
                blocks.append("\n".join(buf))
                buf = []
        else:
            buf.append(raw)
    if buf:
        blocks.append("\n".join(buf))

    def _is_structural(block: str) -> bool:
        # 表格（任一行含 |）、标题、列表、引用、图片、水平线等结构块不参与查重
        first = block.lstrip()
        if first.startswith(("#", ">", "-", "*", "!", "|", "```")):
            return True
        if "|" in block:  # 表格
            return True
        return False

    # 2) 归一化 + 过滤长度
    candidates = []  # (index, normalized, eff_len, preview)
    for idx, block in enumerate(blocks):
        if _is_structural(block):
            continue
        norm = re.sub(r"[^\u4e00-\u9fffA-Za-z0-9]", "", block)
        eff = effective_length(block)
        if eff >= min_effective:
            preview = re.sub(r"\s+", "", block)[:32]
            candidates.append((idx, norm, eff, preview))

    # 3) 两两比较（段落数不多，O(n^2) 可接受）
    reported = set()
    for i in range(len(candidates)):
        for j in range(i + 1, len(candidates)):
            a, b = candidates[i], candidates[j]
            if not a[1] or not b[1]:
                continue
            if a[1] == b[1]:
                ratio = 1.0
            else:
                ratio = difflib.SequenceMatcher(None, a[1], b[1]).ratio()
            if ratio >= sim_threshold:
                key = (a[3], b[3])
                if key in reported:
                    continue
                reported.add(key)
                issues.append(
                    f"[反重复·BLOCK] 发现整段近似重复（相似度 {ratio:.0%}，疑似复制粘贴凑字数）："
                    f"『{a[3]}…』与『{b[3]}…』内容雷同。同一段落不得套用到多个小节，"
                    f"请针对各小节主题分别改写为原创内容，或直接删除冗余段（决策稿 §4.x 只需结论速览，不追求字数）。"
                )
    return issues


def validate_section_411(section_text: str) -> List[str]:
    issues = validate_marker_and_table_rules("4.1.1", section_text)
    if not contains_any(section_text, ["CNBC", "Reuters", "路透", "FT", "Financial Times", "Bloomberg"]):
        issues.append("章节 4.1.1 缺少国际权威信源引用")
    if not contains_any(section_text, ["传导", "影响路径", "资本开支", "需求", "订单", "利润"]):
        issues.append("章节 4.1.1 缺少宏观到企业的传导链条")
    return issues


def validate_section_412(section_text: str) -> List[str]:
    issues = validate_marker_and_table_rules("4.1.2", section_text)
    keyword_groups = [
        ("需求", ["需求", "市场规模", "出货量", "渗透率"]),
        ("ASP", ["ASP", "价格趋势", "单价"]),
        ("竞争格局", ["竞争格局", "市占率", "CR3", "CR5", "对手"]),
        ("技术路线", ["技术路线", "代际", "CPO", "NPO", "1.6T", "800G"]),
        ("供应链", ["供应链", "供给", "瓶颈", "上游"]),
    ]
    missing = missing_semantic_keywords(section_text, keyword_groups)
    if missing:
        issues.append(f"章节 4.1.2 缺少产业关键要素：{', '.join(missing)}")
    # 商业模式画布分析检测
    if not contains_any(section_text, ["商业模式", "收入模式", "价值主张", "增长飞轮"]):
        issues.append("章节 4.1.2 缺少商业模式画布分析（七要素框架）")
    return issues


def validate_section_413(section_text: str) -> List[str]:
    issues = validate_marker_and_table_rules("4.1.3", section_text)
    if not contains_any(section_text, ["分季度", "季度", "Q1", "Q2", "Q3", "Q4"]):
        issues.append("章节 4.1.3 缺少分季度趋势分析")
    if "护城河" not in section_text:
        issues.append("章节 4.1.3 缺少护城河评估")
    if not contains_any(section_text, ["风险", "隐忧"]):
        issues.append("章节 4.1.3 缺少关键风险与隐忧")
    if not contains_any(section_text, ["经营现金流", "FCF", "自由现金流"]):
        issues.append("章节 4.1.3 缺少现金流质量分析")
    if not contains_any(section_text, ["ROE", "毛利率", "净利润", "营收"]):
        issues.append("章节 4.1.3 缺少核心财务指标支撑")
    # ROIC vs WACC 增长质量验证
    if not contains_any(section_text, ["ROIC", "投入资本回报率"]):
        issues.append("章节 4.1.3 缺少ROIC增长质量验证（ROIC vs WACC）")
    # 财务造假预警排查
    if not contains_any(section_text, ["M-Score", "Beneish", "财务造假", "造假预警", "造假排查"]):
        issues.append("章节 4.1.3 缺少财务造假预警排查（Beneish M-Score或定性排查）")
    return issues


def validate_section_414(section_text: str, require_supply_demand: bool) -> List[str]:
    issues = validate_marker_and_table_rules("4.1.4", section_text)
    keyword_groups = [
        ("需求", ["需求", "收入侧"]),
        ("收入", ["收入"]),
        ("成本", ["成本", "成本侧"]),
        ("费用", ["费用", "费用率"]),
        ("净利润", ["净利润"]),
    ]
    if require_supply_demand:
        keyword_groups.insert(0, ("供给", ["供给", "供需"]))
    missing = missing_semantic_keywords(section_text, keyword_groups)
    if missing:
        issues.append(f"章节 4.1.4 缺少盈利预测关键要素：{', '.join(missing)}")
    if "三情景" not in section_text:
        issues.append("章节 4.1.4 缺少三情景利润预测")
    if not contains_any(section_text, ["一致预期", "卖方一致预期"]):
        issues.append("章节 4.1.4 缺少与卖方一致预期比较")
    if "灵敏度" not in section_text:
        issues.append("章节 4.1.4 缺少灵敏度分析")
    return issues


def validate_section_415(section_text: str) -> List[str]:
    issues = validate_marker_and_table_rules("4.1.5", section_text)
    if not contains_any(section_text, ["估值", "PE", "PB", "IRR", "EV/EBITDA", "目标价"]):
        issues.append("章节 4.1.5 缺少估值指标或估值推导")
    if not contains_any(section_text, ["安全边际", "目标价", "合理市值", "定价"]):
        issues.append("章节 4.1.5 缺少定价/安全边际结论")
    if not contains_any(section_text, ["中性情景", "4.1.4"]):
        issues.append("章节 4.1.5 未显式引用 4.1.4 盈利预测输入")
    if not contains_any(section_text, ["隐含假设", "市场定价"]):
        issues.append("章节 4.1.5 缺少市场隐含假设分析")
    # DCF敏感性矩阵检测（FCFF可预测的标的强制执行）
    has_dcf = contains_any(section_text, ["DCF敏感性矩阵", "DCF敏感性", "WACC vs", "WACC=", "永续增长率"])
    has_dcf_skip = contains_any(section_text, ["不适用DCF", "不适用 DCF", "周期股", "亏损", "跳过DCF", "跳过 DCF"])
    if not has_dcf and not has_dcf_skip:
        issues.append("章节 4.1.5 缺少DCF敏感性矩阵（FCFF可预测标的强制执行；不适用时请标注原因）")
    return issues


def validate_generic_section(section_id: str, section_text: str) -> List[str]:
    issues = validate_marker_and_table_rules(section_id, section_text)
    if section_id == "4.3" and not contains_any(section_text, ["支撑位", "压力位", "均线", "量价"]):
        issues.append("章节 4.3 缺少支撑/压力/均线/量价四要素之一")
    if section_id == "4.4" and not contains_any(section_text, ["主力净流入", "北向", "融资", "龙虎榜"]):
        issues.append("章节 4.4 缺少资金流向核心数据")
    if section_id == "4.4" and not contains_any(section_text, ["大宗交易", "机构调研", "调研次数", "近3月调研"]):
        issues.append("章节 4.4 缺少大宗交易或机构调研说明（有数据写数据，无数据也需写明“近期无”及其影响）")

    if section_id == "4.5" and not contains_any(section_text, ["股东", "筹码", "质押", "解禁"]):
        issues.append("章节 4.5 缺少筹码稳定性核心数据")
    if section_id == "4.6" and not contains_any(section_text, ["情绪", "涨停", "炸板", "上涨", "下跌"]):
        issues.append("章节 4.6 缺少市场情绪核心数据")
    if section_id == "4.7" and not contains_any(section_text, ["研报", "机构", "一致预期", "外部观点"]):
        issues.append("章节 4.7 缺少研报或外部观点审视")
    return issues


# ═══════════════════════════════════════════════════════════════════
# v8 行业信源硬门禁（v11 重构：取消"6 份独立面报告"连带检查）
# ───────────────────────────────────────────────────────────────────
# 背景：report_quality_checker 历来只校验"单个文件"，曾放过一类系统性事故——
#   trade_advisor 行业接口（Push2）失效 → 行业数据缺失，正文用估算/编造填充，
#   而门禁只校验 4.1.2 含"竞争格局/市占率"等关键词 + 表格数，编造也能命中。
# 对策：校验"交易决策报告_"汇总报告时，连带强制校验——
#   FinancialData/{code}_industry.md 行业信源文件存在、非空、含关键维度 + 信源标注。
#   不满足即 FAIL，从根上杜绝"行业数据无信源裸填"。
#
# ── v11 关键变更（思路2，已被 v15 部分推翻，见下方 v17 说明）──
# 取消原 "A. 同目录 6 份独立面报告齐全" 连带检查：Intent-1 报告模式不再产出
# 『6 份独立面报告 + 1 份汇总』，改为直接产出 1 份深度汇总决策报告。
#
# ── ⚠️ v15 / v17 关键变更（教义-执法对齐，重新引入"分面深稿"硬门禁）──
# v15 铁律#1 把 Intent-1 综合买卖决策**全周期统一**改回「分面深写三阶段流水线」：
#   阶段A·分面深写 → 六个面各自独立深写成 6 份单面深稿（OutputReport/分面深稿_…），
#   阶段B·交叉勾稽 → 六稿事实对齐 + 六面方向矩阵，
#   阶段C·汇总裁决 → 把六稿核心内容内聚进最终汇总报告。
# 然而 v15 之后**门禁代码长期未跟随更新**——团队规范写了"6 份分面深稿必须落盘"，
# 但本脚本（决定 PASS/FAIL 的硬门禁）从不校验它，仅 cross_face_reconciliation
# 软门禁会瞄一眼且"0 份静默通过"。结果：阶段A 可被零成本完全跳过、汇总报告
# 各面被压缩成薄片（历史事故：基本面 §4.1 仅 ~4-5K vs 独立报告 ~26K），门禁却放行。
# v17 修复：新增 check_stage_a_face_drafts() 硬门禁——Intent-1 汇总报告交付前，
#   必须存在 6 份「分面深稿_{面}_{stem尾}.md」、每份达最小深度、且含【摘要卡】，
#   缺一即 FAIL。这把 v15 的流水线铁律从"文档纪律"升级为"机器可验证的交付前提"。
# 可用 --no-companion-check 关闭连带门禁（仅限确有特殊场景，默认开启）。
# ═══════════════════════════════════════════════════════════════════
FACE_REPORT_PREFIXES = ("基本面_", "政策面_", "资金面_", "筹码面_", "技术面_", "消息面_")
TRADE_CODE_RE = re.compile(r"交易决策报告_(\d{6})_")

# ═══════════════════════════════════════════════════════════════════
# v24 新架构开关：Intent-1 决策稿「不合稿」模式（faces-split）
# ───────────────────────────────────────────────────────────────────
# 背景：v23 强制把 6 份分面深稿【逐字内聚】进决策稿 §4.1-4.6（合稿内聚铁律），
#   导致决策稿膨胀到 ~95K 字、深稿与合稿两份正文易漂移、合稿环节高失败率且无新增价值。
# v24 新交付架构：六个面【不再合稿】，各自作为独立深稿 .md 保留落盘；决策稿只承载
#   「核心决策 + 风险红线 + §五综合研判 + §四 六面结论速览(指向深稿)」。HTML 由
#   md2html_report.py 把 7 份生成为单文件 Tab 多页面 HTML(7 个标签页 + 顶部导航栏切换)。
# v25 更新：faces-split 已从"可选架构"升级为**强制项、无回落路径**——
#   validate_report() 在识别 report_type=="trade" 后，若正文缺该标记，
#   直接记一条 [GATE0·架构标记缺失] 硬 FAIL 并把 _faces_split 强制置 True，
#   不再静默沿用 v23 合稿内聚口径（check_merge_cohesion 已无调用点）。
# ═══════════════════════════════════════════════════════════════════
INTENT1_FACES_SPLIT_RE = re.compile(r"<!--\s*INTENT1_ARCH\s*:\s*faces-split", re.I)
INTENT1_FACE_NAMES = ("基本面", "政策面", "技术面", "资金面", "筹码面", "消息面")


def is_faces_split_report(text: str) -> bool:
    """决策稿是否声明启用 v24 faces-split（不合稿）架构。"""
    return bool(INTENT1_FACES_SPLIT_RE.search(text or ""))


def _read_fundamental_draft_text(report_path: Path) -> Optional[str]:
    """faces-split 架构下，§4.1.4 深度盈利预测承载在【基本面深稿】里。
    定位 分面深稿_基本面_{tail}.md 并返回其正文，供 forecast 类门禁重定向校验。"""
    m = re.match(r"交易决策报告_(.+)$", report_path.stem)
    if not m:
        return None
    draft = report_path.parent / f"分面深稿_基本面_{m.group(1)}.md"
    if not draft.exists():
        return None
    try:
        return draft.read_text(encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        return None

# v17 阶段A·分面深稿硬门禁参数 —————————————————————————————————————
# 6 个面（与流水线铁律#1 / cross_face_reconciliation_validator 完全一致）
STAGE_A_FACES = ("基本面", "政策面", "资金面", "筹码面", "技术面", "消息面")

# ═══════════════════════════════════════════════════════════════════════
# v22 架构级根治：每个面的「深度契约」统一为唯一可信源（SSOT）
# ───────────────────────────────────────────────────────────────────────
# 病根（屡修屡犯的真正架构原因）：
#   ① 两个 Intent 走两条不同强度的校验路径——单独出某面报告(Intent-2)走轻量
#      check_dimension_report()，交易决策报告里的分面深稿(Intent-1)走严格
#      _eval_single_face_draft()，二者"合格线"不相等 → 不变式从未被保证；
#   ② "深度要求"分散在 4 处（faces/<面>.md 正文 + 本文件多个 STAGE_A_* 常量 +
#      _analytical_depth_defects 正则 + 团队规范表格）且互相漂移 → 改 faces 规范
#      门禁根本不知道，改进无法传导到最终报告。
# v22 修复：每个面的深度契约**只在 references/faces/<面>.md 的 ```json face_contract
#   块**里定义一次（字数/表格/小节/脚标下限 + 门禁前缀 + 必填明细表 + required_elements
#   分析要素正则）。本文件所有 per-face 阈值与 _analytical_depth_defects 一律由
#   _load_face_contract() 从该块加载；且 Intent-1 分面深稿与 Intent-2 单面报告**共用
#   同一个 _eval_single_face_draft()**。于是：用户改进某面 = 编辑该面 face_contract 一处
#   → 两个 Intent 的门禁同步收紧 → 改进必然落到最终交易决策报告。
# 内置 _FACE_CONTRACT_DEFAULTS 与各面文档内嵌契约保持一致，仅作"文档缺失/解析失败"
# 时的回退，保证脚本永不因契约缺失而崩溃或行为突变。
# ═══════════════════════════════════════════════════════════════════════
_FACES_DIR = Path(__file__).resolve().parent.parent / "references" / "faces"
_FACE_CONTRACT_DEFAULTS: Dict[str, Dict] = {
    "基本面": {"gate_prefix": "R-FU", "min_eff_len": 9700, "min_tables": 18,
              "min_subsections": 20, "min_footnotes": 10,
              "min_core_len": 3800, "max_table_ratio": 0.65,
              "detail_table": {"min_rows": 8, "kind": "逐季/逐年 P&L"},
              "required_elements": [
                  {"id": "devils_advocate", "desc": "对手方论证/Devil's Advocate（空头核心论点+我方反驳+反驳的可证伪条件+置信度调整声明）", "any": ["对手方论证", "空头核心论点", "空头论点", "看空逻辑"]},
                  {"id": "consensus_compare", "desc": "与卖方一致预期对比（本报告预测 vs 卖方一致预期 + 偏离根因）", "any": ["卖方一致预期", "一致预期对比", "一致预期对照", "consensus"]},
                  {"id": "falsification", "desc": "盈利/估值/评级的量化证伪条件", "any": ["证伪条件", "证伪边界"]},
              ]},
    "政策面": {"gate_prefix": "R-PL", "min_eff_len": 7100, "min_tables": 8,
              "min_subsections": 9, "min_footnotes": 8,
              "min_core_len": 2500, "max_table_ratio": 0.66,
              "required_elements": [
                  {"id": "pct_score", "desc": "百分制综合评分（须给 XX/100 分项加权评分表，不能只用 ★/5 主观星级）", "any": ["\\d{1,3}\\s*/\\s*100", "综合评分[^\\n]{0,40}100"]},
                  {"id": "reduce_trigger", "desc": "『触发减仓/卖出的硬条件』（须列可执行的政策/资金触发线，见本面 §4.4 减仓硬条件规范）", "any": ["触发减仓", "减仓的三条", "减仓硬条件", "清仓.{0,6}条件", "卖出.{0,6}条件"]},
              ]},
    "资金面": {"gate_prefix": "R-CF", "min_eff_len": 5900, "min_tables": 11,
              "min_subsections": 9, "min_footnotes": 8,
              "min_core_len": 3800, "max_table_ratio": 0.50,
              "detail_table": {"min_rows": 10, "kind": "逐日资金流向"},
              "required_elements": [
                  {"id": "pct_score", "desc": "百分制综合评分（须给 XX/100 分项加权评分表，不能只用 ★/5 主观星级）", "any": ["\\d{1,3}\\s*/\\s*100", "综合评分[^\\n]{0,40}100"]},
                  {"id": "reduce_trigger", "desc": "『触发减仓/卖出的硬条件』（须列可执行的破位/资金触发线，见本面 §4.4 减仓硬条件规范）", "any": ["触发减仓", "减仓的三条", "减仓硬条件", "清仓.{0,6}条件", "卖出.{0,6}条件"]},
              ]},
    "筹码面": {"gate_prefix": "R-CH", "min_eff_len": 5050, "min_tables": 12,
              "min_subsections": 10, "min_footnotes": 8,
              "min_core_len": 2800, "max_table_ratio": 0.45,
              "required_elements": [
                  {"id": "pct_score", "desc": "百分制综合评分（须给 XX/100 分项加权评分表，不能只用 ★/5 主观星级）", "any": ["\\d{1,3}\\s*/\\s*100", "综合评分[^\\n]{0,40}100"]},
                  {"id": "reduce_trigger", "desc": "『触发减仓/卖出的硬条件』（须列可执行的破位/筹码触发线，见本面 §4.4 减仓硬条件规范）", "any": ["触发减仓", "减仓的三条", "减仓硬条件", "清仓.{0,6}条件", "卖出.{0,6}条件"]},
                  {"id": "cyq_table", "desc": "CYQ 筹码分布表（须按价格区间逐行列筹码占比/性质/距现价距离，支撑成本锚定结论）", "any": ["CYQ", "筹码分布", "成本分布", "价格区间.{0,8}筹码占比"]},
              ]},
    "技术面": {"gate_prefix": "R-TC", "min_eff_len": 5650, "min_tables": 12,
              "min_subsections": 9, "min_footnotes": 6,
              "min_core_len": 3000, "max_table_ratio": 0.47,
              "detail_table": {"min_rows": 10, "kind": "逐节点量价"},
              "required_elements": [
                  {"id": "pct_score", "desc": "百分制综合评分（须给 XX/100 分项加权评分表，不能只用 ★/5 主观星级）", "any": ["\\d{1,3}\\s*/\\s*100", "综合评分[^\\n]{0,40}100"]},
                  {"id": "reduce_trigger", "desc": "『触发减仓/卖出的硬条件』（须列可执行的破位/量价触发线，见本面 §4.4 减仓硬条件规范）", "any": ["触发减仓", "减仓的三条", "减仓硬条件", "清仓.{0,6}条件", "卖出.{0,6}条件"]},
              ]},
    "消息面": {"gate_prefix": "R-NW", "min_eff_len": 5700, "min_tables": 11,
              "min_subsections": 11, "min_footnotes": 8,
              "min_core_len": 3300, "max_table_ratio": 0.53,
              "required_elements": [
                  {"id": "pct_score", "desc": "百分制综合评分（须给 XX/100 分项加权评分表，不能只用 ★/5 主观星级）", "any": ["\\d{1,3}\\s*/\\s*100", "综合评分[^\\n]{0,40}100"]},
                  {"id": "reduce_trigger", "desc": "『触发减仓/卖出的硬条件』（须列可执行的事件/舆情触发线，见本面 §4.4 减仓硬条件规范）", "any": ["触发减仓", "减仓的三条", "减仓硬条件", "清仓.{0,6}条件", "卖出.{0,6}条件"]},
              ]},
}
_FACE_CONTRACT_CACHE: Dict[str, Dict] = {}
# v28 fail-closed：记录每个面契约块的解析错误（""=无错）。当 <面>.md 存在 ```json
# face_contract 块但解析失败/结构非法时，绝不静默回退默认放行，而是留痕供门禁强制不通过。
_FACE_CONTRACT_ERROR: Dict[str, str] = {}


def _load_face_contract(face: str) -> Dict:
    """v22 SSOT：读取 references/faces/<面>.md 内嵌的 ```json face_contract 块作为该面
    深度契约的唯一可信源；缺文件则回退 _FACE_CONTRACT_DEFAULTS（保证永不崩溃）。
    v28：块存在但解析失败/非 dict 时记录错误（fail-closed，由门禁暴露为不通过），
    不再静默回退默认。结果缓存。两个 Intent 的门禁与文档据此同源执法。"""
    if face in _FACE_CONTRACT_CACHE:
        return _FACE_CONTRACT_CACHE[face]
    contract: Dict = dict(_FACE_CONTRACT_DEFAULTS.get(face, {}))
    err = ""
    try:
        md = (_FACES_DIR / f"{face}.md").read_text(encoding="utf-8", errors="replace")
        m = re.search(r"```json\s+face_contract\s*\n(.*?)\n```", md, re.S)
        if m:
            # 块存在即视为"契约声明"：解析失败/非 dict → 记录错误供 fail-closed 暴露，
            # 而非静默回退默认（防"改坏契约块却门禁照旧放行"的隐患）。
            try:
                parsed = json.loads(m.group(1))
            except Exception as e:  # noqa: BLE001
                parsed = None
                err = f"face_contract JSON 解析失败：{e}"
            if isinstance(parsed, dict):
                contract.update(parsed)  # 文档内嵌契约优先于内置默认
            elif not err:
                err = "face_contract 块内容非 JSON 对象（应为 {...}）"
    except FileNotFoundError:
        pass  # 附属单面文档仅带自身面文档，其余面缺文件属正常回退，不算契约错误
    except Exception as e:  # noqa: BLE001 — 读取异常回退默认，不影响主门禁
        err = f"契约文档读取失败：{e}"
    _FACE_CONTRACT_ERROR[face] = err
    _FACE_CONTRACT_CACHE[face] = contract
    return contract
# 每份单面深稿的最小效率字数下限（v18 抬高·对齐"单面专项报告"应有纵深）：
# 设计动机——历史事故复盘发现旧阈值（其余面统一 2500）远低于各面 §⭐报告输出规范
# 的"总字数下限"（资金 3500 / 筹码 3800 / 技术 3200 / 政策 3500 / 消息 3500，基本面 9000），
# 导致即便达标也只有单面报告 ~60-70% 的体量、且结构（必填表格/评级/信源表）完全不被校验，
# 分面深稿被压成薄片仍能放行（实测薄片：资金面 753 字 / 基本面 2595 字 vs 合格单面 ~8000-10000）。
# v18 把字数阈值抬到各面单面规范"总字数下限"附近（effective_length 含表格内字数、口径更宽，
# 故取规范文字下限的 ~85-95% 作安全线，既逼出真实纵深又不误伤合规深稿），并叠加结构硬校验。
# ⚠️ v21/v23 基准锚定（消除"门禁线 = 天花板"病根）：
#   病根复盘——旧阈值仅为一份合格单面深稿实测字数的 40-61%（基本60%/政策40%/资金49%/
#   筹码61%/技术47%/消息50%），即"写到合格线一半即 PASS"。LLM 天然向门禁线对齐而非
#   合格线，导致"合规却只有半个深度"的浅稿长期被放行（这正是"分面深稿深度差远了"
#   屡修屡犯的数学根源）。v21 起把字数/表格/小节三类阈值统一抬到历史校准下限附近，
#   使"过门禁"≈"达合格深度"。下列校准下限由一次性历史校准固化（绝对底线、非理想目标，
#   不依赖任何具体个股报告/外部输出物）：
#     字数  基本10832 政策7950 资金6568 筹码5619 技术6322 消息6372
#     表格  基本27   政策9   资金14   筹码15   技术15   消息14
#     小节  基本30   政策12  资金11   筹码13   技术12   消息15
#   字数下限 v23 已取校准值×0.90（见各面 face_contract.min_eff_len），表格/小节取 ≈×0.8，
#   既逼出真实纵深、又对达标深稿零误伤。
STAGE_A_MIN_EFF_LEN = {f: _load_face_contract(f).get("min_eff_len", 2500) for f in STAGE_A_FACES}
# v18/v21 结构硬校验：分面深稿 = 完整单面专项报告，必须具备其骨架的结构纵深，
# 不能只靠字数堆砌。表格数 = markdown 表格分隔行计数；推导小节 = ### 级标题数。
# v21 抬到历史校准表格/小节数下限的 ~80%（对达标深稿零误伤）。
STAGE_A_MIN_TABLES = {f: _load_face_contract(f).get("min_tables", 5) for f in STAGE_A_FACES}
STAGE_A_MIN_SUBSECTIONS = {f: _load_face_contract(f).get("min_subsections", 5) for f in STAGE_A_FACES}


def _count_md_tables(text: str) -> int:
    """统计 markdown 表格数：以"分隔行"（形如 |---|---| 的整行）为一张表的标志。"""
    return len(re.findall(r"(?m)^\s*\|?[\s:]*-{3,}[\s:|\-]*\|", text))


def _max_table_rows(text: str) -> int:
    """返回全文所有 markdown 表格中"数据行数"的最大值（颗粒度代理指标）。
    数据行 = 以 | 开头的整行，剔除表头行与 |---| 分隔行。用于校验是否存在
    一张承载逐日/逐季/逐区间明细的"长表"，反"散文+约数代替明细表"。"""
    best = 0
    cur = 0  # 当前连续表格块内"非分隔行"的行数（含表头）
    for ln in text.splitlines():
        s = ln.strip()
        is_row = s.startswith("|") and s.count("|") >= 2
        is_sep = bool(re.match(r"^\|?[\s:]*-{3,}[\s:|\-]*\|", s))
        if is_row and not is_sep:
            cur += 1
        elif is_sep:
            pass  # 分隔行不计入、也不打断表格块
        else:
            # 表格块结束：数据行 = 块内行数 - 1（表头），取最大
            best = max(best, cur - 1)
            cur = 0
    best = max(best, cur - 1)
    return max(best, 0)


def _count_subsections(text: str) -> int:
    """统计 ### 及更深级别小节标题数（推导分节的代理指标）。"""
    return len(re.findall(r"(?m)^\s*#{3,}\s+\S", text))


def _has_source_table(text: str) -> bool:
    """是否含信源汇总表/足量信源引用：标题含"信源汇总"，或 [^srcN]/[srcN] 引用 ≥5 处。"""
    if re.search(r"信源汇总|数据信源|信源[一二三四五六]?[、:：]", text):
        return True
    return len(re.findall(r"\[\^?src\d", text)) >= 5


# ── v19 GATE0·分面深写 结构纵深增强（对应 Intent-2 单面专项报告的"区分性要素"）──
# 设计动机：v18 已校验字数/表格/小节/★/信源表，但与合格单面报告对比仍存三处盲区——
#   ① 信源汇总表只判"有标题 OR ≥5 脚标"二选一，导致深稿可写一张无 URL/无时效的简易表就免脚标；
#   ② 完全不校验"本面硬门禁条款自检表"（合格深稿每份均有 R-CF1~6 / R-TC1~8 / R-PL1~7 等逐条 ✅/❌ 自检）；
#   ③ 脚标密度无下限（合格深稿正文密集 [^srcN] / <sup>N</sup>，薄稿几乎与信源脱钩）。
# v19 把这三项升格为 GATE0 硬校验，使分面深稿真正达到"信源可溯 + 门禁自证 + 推导带引用"的下限规格。
FACE_GATE_PREFIX = {f: _load_face_contract(f).get("gate_prefix") for f in STAGE_A_FACES}
# 各面脚标引用密度下限（历史校准：基本面/政策面/资金面/筹码面/消息面 12-20 处；
# 技术面以脚本行情源为主、外部公开信源偏少，故放宽）。同时识别 [^srcN] 与 <sup>N</sup> 两种形态。
STAGE_A_MIN_FOOTNOTES = {f: _load_face_contract(f).get("min_footnotes", 8) for f in STAGE_A_FACES}


def _count_footnote_citations(text: str) -> int:
    """统计正文脚标引用数：兼容 [^srcN] / [^N] 与 <sup>N</sup> 两种形态（两类都识别）。"""
    n = len(re.findall(r"\[\^[A-Za-z]*\d+\]", text))
    n += len(re.findall(r"<sup>\s*\d+\s*</sup>", text))
    return n


def _source_table_has_url_and_time(text: str) -> bool:
    """信源汇总表是否达可溯源规格：含 ≥3 个公开 http(s) URL + 含"抓取/数据时效/数据截止/时效"时间列关键词。
    付费终端域名（wind/bloomberg 等）不计入有效公开 URL（与信源四铁律 D 类黑名单一致）。"""
    urls = re.findall(r"https?://[^\s)|]+", text)
    bad = ("wind.com.cn", "bloomberg.com", "ihs.com", "omdia.com",
           "lightcounting.com", "gartner.com", "idc.com")
    good_urls = [u for u in urls if not any(b in u.lower() for b in bad)]
    has_time_col = re.search(r"抓取时间|数据时效|数据截止|时效|获取时间|更新时间", text) is not None
    return len(good_urls) >= 3 and has_time_col


# ── v27 GATE0·脚标定义完整性硬校验（根治 HTML 终检 C3 裸脚标泄漏的"事后才发现"问题）──
# 设计动机（本次真实事故复盘）：000063 中兴通讯任务里，六面深稿全部通过了 v18-v20 的
# 所有结构校验（字数/表格/★/信源表/脚标密度/门禁自检），却在最后一步 html_gate.py 终检
# 才发现 100+ 处裸脚标泄漏——根因是正文引用了 [^src9]~[^src16] 等编号，但信源表/已有
# 脚注定义块里根本没有对应定义（或因信源表标题措辞漂移导致 inject_footnote_definitions()
# 未能解析出定义）。旧校验只看"有没有信源表/脚标够不够密"，从不校验"每个引用编号是否
# 真的有定义"——这类"引用-定义不闭合"的缺陷只能靠 md2html_report.py 转换后肉眼/终检发现，
# 回改成本极高（要跨 6 份深稿 + 决策报告逐条排查差集）。
# v27 把这一校验前移到阶段1单面检查点：直接在 markdown 源文本层面统计"正文引用了哪些
# [^srcN] 编号"与"信源表/定义块里实际定义了哪些编号"的差集，缺失即 FAIL，把"能否被
# HTML 正确渲染为脚注"这件事从"HTML 产物级事后检测"降级为"markdown 源码级前置检测"。
_INLINE_CODE_RE = re.compile(r"`[^`\n]*`")
_FENCED_CODE_RE = re.compile(r"```.*?```", re.S)


def _strip_code_spans(text: str) -> bool:
    """剔除行内代码 `...` 与围栏代码块 ```...```，避免自检表格里的字面量示例
    （如 `` `[^srcN]` `` 中 N 非具体数字）被误判为"引用但未定义"的假阳性。"""
    t = _FENCED_CODE_RE.sub(" ", text)
    t = _INLINE_CODE_RE.sub(" ", t)
    return t


def _undefined_footnote_refs(text: str) -> List[str]:
    """返回"正文引用过、但全文找不到对应 [^id]: 定义"的脚注编号列表（已排序去重）。

    定义来源两类（与 md2html_report.py::inject_footnote_definitions 的解析逻辑对齐，
    确保"门禁判定"与"实际渲染"同一套真相，不漂移）：
      ① 显式脚注定义行 `[^id]: 说明`（正文任何位置，含手写的）；
      ② 信源汇总表/数据信源汇总表 表格首列的数字编号（如 `9 | 中兴通讯官网 | ... | url | 时效`），
         该表格会被 md2html_report.py 自动转成 `[^src{N}]:` 定义，故表中出现的编号也算已定义。
    """
    scan_text = _strip_code_spans(text)
    used = set(re.findall(r"\[\^(src[\w\-]+)\]", scan_text))
    if not used:
        return []

    defined = set(re.findall(r"(?m)^\s*\[\^(src[\w\-]+)\]:", text))

    m_anchor = re.search(r"(?:数据)?信源汇总表", text)
    if m_anchor:
        tail = text[m_anchor.end():]
        end_m = re.search(r"\n#{1,6}\s|\n---\s*\n|\n【自用声明】", tail)
        block = tail[: end_m.start()] if end_m else tail
        for line in block.split("\n"):
            s = line.strip()
            if not s.startswith("|") or re.match(r"^\|\s*[-:|\s]+\|", s):
                continue
            cells = [c.strip() for c in s.strip("|").split("|")]
            if not cells:
                continue
            m_num = re.search(r"\b(\d+)\b", cells[0])
            if m_num:
                defined.add(f"src{m_num.group(1)}")

    missing = sorted(used - defined, key=lambda x: (len(x), x))
    return missing


def _has_face_gate_selfcheck(text: str, face: str) -> bool:
    """是否含"本面硬门禁条款自检表"：正文出现对应面的 R-XX 门禁编号 ≥3 处（如 R-CF1/R-CF2…）。
    合格单面报告均以一张表逐条自检 R-XX 门禁（✅通过/未通过+说明）。"""
    prefix = FACE_GATE_PREFIX.get(face)
    if not prefix:
        return True
    return len(re.findall(re.escape(prefix) + r"\d", text)) >= 3


# ── v20 GATE0·分面深写 分析纵深增强（对齐"可量化 + 可执行 + 可证伪"的决策要素）──
# 设计动机：v18/v19 已把"字数/表格/小节/信源表 URL/脚标密度/门禁自检"
# 升格为硬校验，能拦住"结构薄片/不可溯源"。但历史校准发现，仍有三类"分析纵深/可执行性"
# 差异要素被薄稿系统性省略，而机构级深稿每份必备——这正是"合规薄稿"与"机构级深稿"的分水岭：
#   ① 百分制综合评分：合格深稿 5 个非基本面面均给「XX/100 分项加权评分表」；薄稿只给 ★/5 主观
#      星级（且常自相矛盾，如技术面正文 3.05/5 vs 评级 3.5/5）——不可比、不可审计。
#   ② 触发减仓/卖出的硬条件：合格深稿每份 §4.4「触发减仓的三条硬条件」逐条给可执行破位/资金/
#      基本面触发线；薄稿多数面缺失——结论不可落地。
#   ③ 基本面专属三件（机构级标配）：对手方论证(空头论点+反驳+反驳的可证伪条件+置信度调整)
#      + 与卖方一致预期对比(含偏离根因) + 盈利/估值/评级证伪条件；薄稿全缺。
# 上述判据均经历史校准验证：对达标深稿不误伤、对薄稿缺陷可精准拦截。
# v22：上述判据已从硬编码下沉为各面 face_contract（SSOT）的 required_elements / detail_table，
# 由 _analytical_depth_defects 统一据契约执法。STAGE_A_NEED_* / STAGE_A_MIN_DETAIL_ROWS 旧常量
# 已废弃（不再决定行为），删除以根除"门禁要求在多处重复定义并漂移"的病根。


def _analytical_depth_defects(text: str, face: str) -> List[str]:
    """v22：分析纵深/可执行性缺陷检测，**全部由该面 face_contract（唯一可信源）驱动**——
    遍历 required_elements（每条 `any` 正则任一命中即视为具备该要素）+ detail_table
    （全文最长明细表行数下限）。用户在 references/faces/<面>.md 的 ```json face_contract
    块里增删 required_elements / 调整 detail_table，本函数即自动随之收紧或放松；且因
    Intent-1 分面深稿与 Intent-2 单面报告共用 _eval_single_face_draft → 改进同源生效、
    必然传导到最终交易决策报告。返回缺陷描述列表（空 = 通过）。"""
    d: List[str] = []
    c = _load_face_contract(face)
    for el in c.get("required_elements", []):
        pats = el.get("any") or []
        if pats and not any(re.search(p, text, re.I) for p in pats):
            d.append("缺" + str(el.get("desc") or el.get("id") or "必备分析要素"))
    dt = c.get("detail_table")
    if isinstance(dt, dict) and dt.get("min_rows"):
        min_rows = int(dt["min_rows"])
        max_rows = _max_table_rows(text)
        if max_rows < min_rows:
            kind = dt.get("kind", "明细")
            d.append(
                f"数据颗粒度不足（最长表仅 {max_rows} 行 < {min_rows} 行）：须含一张{kind}明细表"
                f"（≥{min_rows} 行逐行），禁止用散文+约数代替逐笔明细"
            )
    return d

INDUSTRY_DIMENSION_KEYWORDS = [
    ("市场规模/景气度", ["市场规模", "行业规模", "景气", "CAGR", "出货量", "渗透率"]),
    ("竞争格局/市占", ["竞争格局", "市占", "市场份额", "份额", "CR3", "CR5", "排名", "龙头"]),
    ("同业估值对比", ["同业", "可比", "估值对比", "PE", "PB", "PEG"]),
    ("产业链/技术路线", ["产业链", "技术路线", "上游", "下游", "代际", "供应链"]),
    ("行业风险", ["风险", "降价", "产能", "出清", "周期"]),
]


def check_companion_deliverables(report_path: Path) -> List[str]:
    """汇总决策报告（交易决策报告_）专用：校验行业信源文件充分。

    v11：取消"6 份独立面报告齐全"连带检查（思路2——Intent-1 直接出单份深稿，
    不再拆 6 份独立面文件）；仅保留行业信源文件 FinancialData/{code}_industry.md
    的存在/非空/关键维度/信源标注校验。
    返回 issues 列表（空 = 通过）。非标准命名的报告直接返回空（不误伤）。
    """
    issues: List[str] = []
    m = TRADE_CODE_RE.search(report_path.name)
    if not m:
        return issues
    code = m.group(1)
    outdir = report_path.parent

    # ── 行业信源文件存在 + 非空 + 关键维度 + 信源标注 ──────
    industry_file = outdir.parent / "FinancialData" / f"{code}_industry.md"
    if not industry_file.exists():
        issues.append(
            f"[行业信源] 缺少行业数据信源文件 FinancialData/{code}_industry.md。"
            "在写产业层/同业对比前，必须先判定所属行业并用脚本/ web_search 充分采集"
            "『行业规模/竞争格局/同业估值/产业链/行业风险』并落盘该文件；"
            "trade_advisor 行业接口（Push2）失效时尤须 web_search 补采，"
            "严禁用估算或编造数字填充正文（信源诚信铁律）。"
        )
    else:
        try:
            ind_text = industry_file.read_text(encoding="utf-8", errors="replace")
        except Exception:
            ind_text = ""
        ind_len = effective_length(ind_text)
        if ind_len < 600:
            # P1-1：纯字数过短降级为软建议；行业"维度/信源数量"缺失仍为硬 FAIL（见下）
            issues.append(
                f"[字数建议][行业信源] 行业数据文件偏短（效率字数 {ind_len} < 600，还差约 {600 - ind_len} 字）："
                f"建议 FinancialData/{code}_industry.md 覆盖行业规模/竞争格局/同业估值/产业链/风险并逐条标注信源。"
            )
        hit_dims = [n for n, kws in INDUSTRY_DIMENSION_KEYWORDS if contains_any(ind_text, kws)]
        if len(hit_dims) < 3:
            missing_dims = [n for n, kws in INDUSTRY_DIMENSION_KEYWORDS if not contains_any(ind_text, kws)]
            issues.append(
                f"[行业信源] 行业研究维度不足：仅覆盖 {len(hit_dims)}/5 个关键维度，"
                f"缺少『{('、'.join(missing_dims))}』（至少覆盖 3 个）。"
            )
        src_hits = (
            len(re.findall(r"来源", ind_text))
            + len(re.findall(r"https?://", ind_text))
            + len(re.findall(r"<sup>\d+</sup>", ind_text))
            + len(re.findall(r"20\d{2}[-/年]", ind_text))
        )
        if src_hits < 4:
            issues.append(
                f"[行业信源] 行业数据信源标注不足（来源/URL/日期合计仅 {src_hits} 处，要求 ≥4）："
                "每条行业数据须可溯源，禁止无信源裸数字。"
            )
    return issues


# ── v28 GATE0·核心分析纵深 + 反灌水 + 分层字数（根治"表格/脚注堆字数、议论说理薄"）──
# 设计动机（本次整改）：effective_length 含表格内字数与脚注定义行字数，历史校准发现薄稿
# 可用一张大表 + 密集脚注定义把 effective_length 顶过门禁线，但真正承载分析的"表外散文
# 议论"极薄（实测薄片：资金面表格占比 66%、核心分析字数仅 539）。v28 增设三道与既有字数
# 门禁正交、且全部经健康样例校准（对达标深稿零误伤）的判据：
#   ① 核心分析字数下限 min_core_len：核心字数 = 效率字数 − 表格内字数 − 脚注定义行字数，
#      逼出"表格之外的推导议论"真实纵深（健康样例各面核心字数 min 2881~4886）；
#   ② 表格贡献占比上限 max_table_ratio：表格内字数 / 效率字数，反"整篇几乎全是表格"
#      （健康样例各面占比 max 0.31~0.59，阈值取 max+~0.10 松弛，薄片实测 0.61~0.74 必被拦）；
#   ③ 段落重复率上限 max_dup_ratio：规范化后完全相同的实质段落占比，反复制粘贴灌水（宽松兜底）；
#   ④ 超长审查 target_hi：效率字数远超合格深稿区间时给"提示"（软、非拦截），供人工复核注水。
# 上述阈值均可被各面 face_contract 覆盖（SSOT），未声明则由 min_eff_len 派生安全默认。
def _table_effective_length(text: str) -> int:
    """markdown 表格行（以 | 开头、含 ≥2 个 |）贡献的效率字数。"""
    total = 0
    for ln in text.splitlines():
        s = ln.strip()
        if s.startswith("|") and s.count("|") >= 2:
            total += effective_length(ln)
    return total


def _footnote_def_effective_length(text: str) -> int:
    """脚注定义行 `[^id]: 说明` 贡献的效率字数（信源脚注定义堆砌不应计入核心分析）。"""
    total = 0
    for ln in text.splitlines():
        if re.match(r"^\s*\[\^[\w\-]+\]:", ln):
            total += effective_length(ln)
    return total


def _core_prose_length(text: str) -> int:
    """核心分析字数 = 效率字数 − 表格内字数 − 脚注定义行字数（表格/信源之外的推导议论体量）。"""
    return max(0, effective_length(text) - _table_effective_length(text)
               - _footnote_def_effective_length(text))


_DUP_NORM_RE = re.compile(r"[\s、，。：:；;·\-—~（）()【】\[\]『』「」/\\|*#>！!？?\"'`]+")


def _duplicate_paragraph_ratio(text: str) -> Tuple[float, int]:
    """反灌水：统计正文实质段落中"规范化后完全相同"的重复占比。
    返回 (重复占比, 参与统计的实质段落数)。排除表格行、标题行与过短段落。"""
    seen: Dict[str, int] = {}
    dup = total = 0
    for para in re.split(r"\n\s*\n", text):
        p = para.strip()
        if not p or p.lstrip().startswith("|") or p.lstrip().startswith("#"):
            continue
        key = _DUP_NORM_RE.sub("", p)
        if len(key) < 30:
            continue
        total += 1
        if key in seen:
            dup += 1
        else:
            seen[key] = 1
    return ((dup / total) if total else 0.0), total


def _core_len_floor(contract: Dict) -> int:
    """核心分析字数下限：契约 min_core_len 优先，否则由 min_eff_len 派生（×0.55 安全线）。"""
    v = contract.get("min_core_len")
    if isinstance(v, (int, float)) and v > 0:
        return int(v)
    return int(round(contract.get("min_eff_len", 2500) * 0.55))


def _table_ratio_cap(contract: Dict) -> float:
    """表格贡献占比上限：契约 max_table_ratio 优先，否则默认 0.62。"""
    v = contract.get("max_table_ratio")
    if isinstance(v, (int, float)) and 0 < v <= 1:
        return float(v)
    return 0.62


def _dup_ratio_cap(contract: Dict) -> float:
    """段落重复率上限：契约 max_dup_ratio 优先，否则宽松兜底 0.40（仅拦极端复制粘贴）。"""
    v = contract.get("max_dup_ratio")
    if isinstance(v, (int, float)) and 0 < v <= 1:
        return float(v)
    return 0.40


def _overlong_threshold(contract: Dict) -> int:
    """超长审查阈值：契约 target_hi 优先，否则由 min_eff_len 派生（×2.8）。"""
    v = contract.get("target_hi")
    if isinstance(v, (int, float)) and v > 0:
        return int(v)
    return int(round(contract.get("min_eff_len", 2500) * 2.8))


def _eval_single_face_draft(dtext: str, face: str) -> Dict[str, object]:
    """v21：对单份分面深稿正文做"字数 + 摘要卡 + v18/19/20 结构纵深"全套校验。
    v28 增设核心分析字数/表格占比/反灌水硬校验 + 超长软提示 + 契约 fail-closed。
    返回 {'too_thin': str|None, 'no_card': bool, 'defects': List[str], 'warnings': List[str], ...}。
    供批量门禁 check_stage_a_face_drafts 与单面检查点 check_single_face_draft 复用——
    单一可信源，避免两条校验路径漂移。"""
    eff = effective_length(dtext)
    floor = STAGE_A_MIN_EFF_LEN.get(face, 2500)
    too_thin = f"{face}(效率字数 {eff} < {floor})" if eff < floor else None
    no_card = "摘要卡" not in dtext
    # ── v18 结构硬校验：分面深稿须具备单面专项报告骨架的结构纵深 ──
    n_tables = _count_md_tables(dtext)
    n_subs = _count_subsections(dtext)
    min_tbl = STAGE_A_MIN_TABLES.get(face, 5)
    min_sub = STAGE_A_MIN_SUBSECTIONS.get(face, 5)
    defects: List[str] = []
    if n_tables < min_tbl:
        defects.append(f"表格 {n_tables}<{min_tbl} 张")
    if n_subs < min_sub:
        defects.append(f"推导小节 {n_subs}<{min_sub} 个")
    if "★" not in dtext:
        defects.append("缺 ★ 量化评级")
    if not _has_source_table(dtext):
        defects.append("缺信源汇总表/足量信源引用")
    # ── v19 区分性结构硬校验：对齐合格单面报告的三大可溯源要素 ──
    if not _source_table_has_url_and_time(dtext):
        defects.append("信源表缺公开URL+抓取时间列（需≥3条http(s)公开URL且含抓取时间/数据时效列）")
    n_fn = _count_footnote_citations(dtext)
    min_fn = STAGE_A_MIN_FOOTNOTES.get(face, 8)
    if n_fn < min_fn:
        defects.append(f"正文脚标引用不足（[^srcN]/<sup>N</sup> 仅 {n_fn}<{min_fn} 处）")
    if not _has_face_gate_selfcheck(dtext, face):
        defects.append(f"缺本面硬门禁条款自检表（{FACE_GATE_PREFIX.get(face)}1~ 逐条✅/❌）")
    # ── v20 分析纵深硬校验：对齐"可量化评分/可执行触发线/可证伪边界" ──
    defects.extend(_analytical_depth_defects(dtext, face))
    # ── v27 脚标定义完整性硬校验：正文引用的 [^srcN] 必须都能在信源表/定义块里找到对应定义，
    # 否则该编号在 HTML 终检时必然裸露成 C3 泄漏（此前只能靠事后临时脚本才能发现）。
    missing_refs = _undefined_footnote_refs(dtext)
    if missing_refs:
        defects.append(
            f"脚标引用缺定义（正文引用 {('、'.join('[^' + r + ']' for r in missing_refs))} "
            "但信源汇总表/脚注定义块未收录对应编号，HTML 转换后必然裸露泄漏；"
            "须在信源表补行或补写 `[^srcN]: 说明` 定义行）"
        )
    # ── v28 核心分析纵深 + 反灌水 + 分层字数（与既有 effective_length 门禁正交）──
    contract = _load_face_contract(face)
    core_len = _core_prose_length(dtext)
    core_floor = _core_len_floor(contract)
    if core_len < core_floor:
        defects.append(
            f"核心分析字数不足（核心字数 {core_len}<{core_floor}；核心字数=效率字数−表格内字数−脚注定义行字数）："
            "须补足表格/信源之外的推导议论纵深，禁止用大表/密集脚注堆字数而说理单薄"
        )
    tbl_ratio = (_table_effective_length(dtext) / eff) if eff else 0.0
    ratio_cap = _table_ratio_cap(contract)
    if tbl_ratio > ratio_cap:
        defects.append(
            f"表格占比过高（表格内字数占效率字数 {tbl_ratio:.0%}>{ratio_cap:.0%} 上限）："
            "整篇过度依赖表格罗列，须增加表外的推导论证与解读"
        )
    dup_ratio, dup_n = _duplicate_paragraph_ratio(dtext)
    dup_cap = _dup_ratio_cap(contract)
    if dup_n >= 8 and dup_ratio > dup_cap:
        defects.append(
            f"疑似重复灌水（{dup_ratio:.0%} 的实质段落为规范化后完全相同的重复内容 >{dup_cap:.0%} 上限）："
            "须删除复制粘贴段落、代之以增量分析"
        )
    # 契约 fail-closed：本面 face_contract 块存在但解析失败时，绝不静默放行
    c_err = _FACE_CONTRACT_ERROR.get(face)
    if c_err:
        defects.append(
            f"契约解析失败（references/faces/{face}.md 的 {c_err}）：门禁已回退内置默认并强制不通过，"
            "请修复该面 ```json face_contract 块后复跑"
        )
    # 超长审查（软提示，不拦截）：远超合格深稿区间时提请人工复核是否注水
    warnings: List[str] = []
    overlong = _overlong_threshold(contract)
    if eff > overlong:
        warnings.append(
            f"效率字数 {eff} 远超合格深稿区间（>{overlong}）：请人工复核是否存在冗余/注水，"
            "确认无冗余后可忽略本提示（非拦截项）"
        )
    return {"too_thin": too_thin, "no_card": no_card, "defects": defects,
            "warnings": warnings, "eff": eff, "floor": floor,
            "n_tables": n_tables, "n_subs": n_subs, "core_len": core_len,
            "core_floor": core_floor, "table_ratio": tbl_ratio, "dup_ratio": dup_ratio}


def check_single_face_draft(draft_path: Path, face: str) -> Tuple[bool, List[str]]:
    """v21 单面检查点（per-face checkpoint）：阶段A 每写完一个面的分面深稿，立即就地
    校验该单份是否已达"= 完整单面专项报告"的纵深，未过即返回待补清单。

    设计动机（屡修屡犯根因之三）：旧机制只有"终点批量门禁"——6 份全写完、汇总报告也写完
    后才一次性校验。长任务里该门禁经常根本不被运行，或 FAIL 也已"先交付"；且批量校验
    发生在 LLM 已切换到汇总收束、对单面上下文已遗忘之时，回改成本高、容易草草。
    本检查点把校验前移到"每面写完即卡"，趁该面上下文还在、未赶收束，必须 PASS 才进入下一面。
    正确 SOP：写完一个面的『分面深稿_{面}_…md』→ 立即跑本检查点 → FAIL 则按清单当场补深 →
    复跑直到 PASS → 再写下一个面。

    返回 (passed, issues)。passed=True 且 issues=[] 表示该面已达合格纵深下限。
    """
    issues: List[str] = []
    if face not in STAGE_A_FACES:
        return False, [f"未知面名『{face}』，须为：{('、'.join(STAGE_A_FACES))}"]
    if not draft_path.exists():
        return False, [f"分面深稿不存在：{draft_path}"]
    try:
        dtext = draft_path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        return False, [f"读取失败：{e}"]
    ev = _eval_single_face_draft(dtext, face)
    if ev["too_thin"]:
        issues.append(
            f"[单面深度] {ev['too_thin']}。分面深稿 = 完整单面专项报告，"
            f"效率字数须达本面历史校准下限（本面下限 {ev['floor']}，此为绝对底线、非理想目标）。"
        )
    if ev["no_card"]:
        issues.append("[摘要卡] 缺末尾【摘要卡】（方向▲/►/▼ + ★评级 + 3-5 关键数据 + 核心依据一句话 + 关键信源），供阶段C 二次综合。")
    if ev["defects"]:
        issues.append(f"[结构纵深] {('、'.join(ev['defects']))}。须严格套用 references/faces/{face}.md §⭐报告输出规范的章节骨架/必填表格/门禁自检表。")
    # 硬门禁判定仅取上述拦截项；超长审查为软提示（前缀 [超长审查·提示]），附于清单但不影响 passed。
    passed = (len(issues) == 0)
    if ev.get("warnings"):
        issues.append(f"[超长审查·提示] {('；'.join(ev['warnings']))}")
    return passed, issues


def check_stage_a_face_drafts(report_path: Path) -> List[str]:
    """v17 GATE0·分面深写前置：Intent-1 汇总决策报告（交易决策报告_）交付前，
    必须存在阶段A 产出的 6 份单面深稿，且每份达最小深度、含【摘要卡】。

    设计动机（本次整改根因）：团队铁律#1 要求"全周期统一走分面深写
    三阶段流水线"——阶段A 先把六个面各自独立深写成 6 份「分面深稿_{面}_…」中间稿
    落盘，阶段C 再把它们内聚进汇总报告。但门禁代码长期停留在 v11 哲学（"不再拆 6 份"），
    从不校验分面深稿是否真的存在；唯一相关的 cross_face_reconciliation 是软门禁，
    且"一份都没有时静默通过"（反向激励：彻底跳过阶段A 反而 0 WARN）。两者叠加导致
    阶段A 可被零成本完全跳过、汇总报告各面被压缩成薄片仍能放行。

    本门禁把"阶段A 是否真的执行"变成机器可验证的硬交付前提：
      ① 齐备性：6 个面的「分面深稿_{面}_{stem尾}.md」必须全部存在（缺任一 → FAIL）；
      ② 深度：每份效率字数 ≥ STAGE_A_MIN_EFF_LEN[面]（防桩文件/薄片 → FAIL）；
      ③ 摘要卡：每份须含【摘要卡】结构块（阶段C 二次综合的结构化输入 → 缺则 FAIL）。
    其中 stem尾 = 报告名去掉「交易决策报告_」前缀（= {code}_{简称}_{时间戳}），
    与最终报告同时间戳，便于成组追溯与人工抽查。

    非 Intent-1 命名（无"交易决策报告_"）直接返回空，不误伤单面/基本面报告。
    """
    issues: List[str] = []
    m = re.match(r"交易决策报告_(.+)$", report_path.stem)
    if not m:
        return issues
    tail = m.group(1)  # {code}_{简称}_{时间戳}
    outdir = report_path.parent

    missing: List[str] = []
    too_thin: List[str] = []
    no_card: List[str] = []
    weak_struct: List[str] = []  # v18：结构纵深不足（表格/小节/信源表）
    for face in STAGE_A_FACES:
        draft = outdir / f"分面深稿_{face}_{tail}.md"
        if not draft.exists():
            missing.append(face)
            continue
        try:
            dtext = draft.read_text(encoding="utf-8", errors="replace")
        except Exception:
            dtext = ""
        ev = _eval_single_face_draft(dtext, face)
        if ev["too_thin"]:
            too_thin.append(ev["too_thin"])
        if ev["no_card"]:
            no_card.append(face)
        if ev["defects"]:
            weak_struct.append(f"{face}（{'、'.join(ev['defects'])}）")

    if missing:
        issues.append(
            f"[GATE0·分面深写] 缺少阶段A 单面深稿：{('、'.join(missing))}（共缺 {len(missing)}/6 面）。"
            "团队铁律#1（v15·全周期统一）要求 Intent-1 综合决策必须先把六个面各自独立深写成 "
            "6 份『分面深稿_{面}_" + tail + ".md』并落盘，再做阶段B 交叉勾稽、阶段C 汇总裁决；"
            "严禁跳过阶段A 直接在汇总报告里压缩式写六面（历史事故：基本面被压成薄片 ~4-5K vs 独立报告 ~26K）。"
            "请逐面加载对应 references/faces/<面>.md 方法论独立深写后再交付汇总报告。"
        )
    if too_thin:
        issues.append(
            f"[GATE0·分面深写] 以下单面深稿深度不足（疑似桩文件/薄片）：{('；'.join(too_thin))}。"
            "分面深稿深度须对齐 Intent-2 单面专项报告（基本面五段式 4.1.1~4.1.5、其余面三段式+表格+★评级），"
            "不得用占位/摘要充数——这是汇总报告各面纵深的来源。"
        )
    if no_card:
        issues.append(
            f"[GATE0·分面深写] 以下单面深稿缺末尾【摘要卡】结构块：{('、'.join(no_card))}。"
            "每份深稿须产出摘要卡（方向 ▲/►/▼ + ★评级 + 3-5 关键数据 + 核心依据一句话 + 关键信源），"
            "供阶段C 汇总二次综合（§5.1.5 六面方向矩阵）使用，缺失会使汇总退化为重读全文、丢失结构化方向。"
        )
    if weak_struct:
        issues.append(
            f"[GATE0·分面深写] 以下单面深稿结构纵深不足（疑似薄片/赶收束/不可溯源/不可审计）：{('；'.join(weak_struct))}。"
            "分面深稿 = 一份完整的单面专项报告，必须严格套用 references/faces/<面>.md §⭐报告输出规范 的"
            "章节骨架，且需具备合格单面报告的三大『可溯源/可审计』结构要素：\n"
            "  ① 信源汇总表（≥3 条公开 http(s) URL + 含『抓取时间/数据时效』列，付费终端域名不计入公开URL）；\n"
            "  ② 正文密集脚标引用（[^srcN] 或 <sup>N</sup>，基本面≥10 / 政策·资金·筹码·消息≥8 / 技术≥6 处），关键数字逐一挂源；\n"
            "  ③ 本面硬门禁条款自检表（逐条列 R-XX 门禁编号 + ✅通过/❌未通过 + 说明，基本面 R-FU / 政策 R-PL / "
            "资金 R-CF / 筹码 R-CH / 技术 R-TC / 消息 R-NW）。\n"
            "并须达到合格单面报告的三类『分析纵深/可执行/可证伪』要素（v20）：\n"
            "  ④ 百分制综合评分：政策/资金/筹码/技术/消息面须给「XX/100 分项加权评分表」（不能只用 ★/5 主观星级，"
            "更不得正文星级与评级结论自相矛盾）；\n"
            "  ⑤ 触发减仓/卖出的硬条件：政策/资金/筹码/技术/消息面须列可执行的破位/资金/基本面触发线（见本面 §4.4『触发减仓的三条硬条件』规范），使结论可落地；\n"
            "  ⑥ 基本面机构级三件：对手方论证/Devil's Advocate（空头核心论点 + 我方反驳 + 反驳的可证伪条件 + 置信度调整声明）"
            "+ 与卖方一致预期对比（本报告预测 vs 卖方一致预期 + 偏离根因）+ 盈利/估值/评级的量化证伪条件。\n"
            "同时按对应面规范补齐表格（资金面主力四阶段表+三维联动矩阵+综合评分表、技术面均线+量价+指标+多周期+支撑压力多表、"
            "筹码面解禁减持质押多表等）后再交付。"
        )
    return issues


# ═══════════════════════════════════════════════════════════════════
# v23 新增：GATE0·合稿内聚门禁（根治"合稿把深稿轻飘飘几句引用带过"）
# ───────────────────────────────────────────────────────────────────
# 病根（本次整改根因）：流水线分"阶段A 写 6 份深稿 → 阶段C 合稿"两步。
#   check_stage_a_face_drafts 只校验【深稿】够不够深，却从不校验【合稿】是否真把深稿
#   全文内聚进来。于是合稿可以只写一句"> 详见 OutputReport/分面深稿_基本面_….md"
#   即过门禁——用户投诉"合稿没把深稿全部纳入、只写几句引用/请见"正源于此。
#   叠加旧版 §4.2~4.6 字数地板只有深稿的 1/7（软建议），合稿薄片化零成本。
# v23 修复：把"合稿确实内聚了深稿"变成机器可验证的硬交付前提。对每个面，定位其在
#   合稿报告中的章节（基本面§4.1 / 政策§4.2 / 技术§4.3 / 资金§4.4 / 筹码§4.5 / 消息§4.6），强制：
#     闸A·禁引用跳转：章节内出现"详见/请见 …OutputReport/*.md / 分面深稿"即 FAIL；
#     闸B·核心字数：章节效率字数 ≥ 对应深稿【核心正文】× 0.90（核心 = 去顶部结论卡 +
#        去尾部 摘要卡/信源汇总表/R-XX自检表 这三类合稿时合理移出的 QA 专用件）；
#     闸C·小节全覆盖：深稿核心正文每个小节标题（去编号后）须在合稿章节出现，杜绝整节丢弃。
#   与 SECTION_MIN_EFFECTIVE_LEN 绝对地板互为双保险：绝对地板锚定"同面历史校准下限"，
#   本函数相对校验确保"本次深稿被逐字内聚到合稿"。
# ═══════════════════════════════════════════════════════════════════
# face → 合稿章节号映射（与 references/templates/intent1_full_report.md 骨架严格一致）
FACE_TO_SECTION = {
    "基本面": "4.1", "政策面": "4.2", "技术面": "4.3",
    "资金面": "4.4", "筹码面": "4.5", "消息面": "4.6",
}
MERGE_COHESION_RATIO = 0.90  # 合稿面章节效率字数 ≥ 对应深稿核心正文 × 该比例（用户拍板 0.90）

# 合稿严禁引用深稿/外部 md 文件的跳转式写法（读者只看合稿即应获得该面全部信息）
_MERGE_REF_JUMP_RE = re.compile(
    r"(详见|请见|参见|另见|详细见|具体见|详情见|见)\s*[`《\"'（(]?\s*"
    r"(?:OutputReport[/\\][^\s`》\"'）)]*?\.md|分面深稿)"
)
# 深稿正文小节标题前缀（数字/中文序号 + 顿号/点），用于"去编号"后做标题命中比对
_HEADING_NUM_PREFIX_RE = re.compile(r"^[#>\s]*(?:[0-9零一二三四五六七八九十百]+[、\.．:：\s]+)+")
# 规范化用：剔除空白与常见标点，便于子串命中（容忍编号/标点差异）
_PUNCT_STRIP_RE = re.compile(r"[\s、，。：:·\-—~（）()【】\[\]『』「」/\\|*#>]+")


def _draft_core_for_merge(dtext: str) -> str:
    """提取深稿中应被合稿【逐字内聚】的核心正文——剥离三类合稿时合理移出的件：
      ① 顶部"一、核心结论/投资评级量化卡"（合稿后并入 §1 综合决策摘要，约 6% 去重）；
      ② 尾部 QA 专用件：摘要卡（→§5 方向矩阵）/ 信源汇总表（→附录）/ R-XX 硬门禁自检表（QA 件）。
    返回核心正文文本，用于与合稿章节做公平的 ×0.90 字数 / 小节覆盖对比。"""
    body = dtext
    parts = re.split(r"\n#{1,4}\s*[二2][、\.．\s]", dtext, maxsplit=1)
    if len(parts) > 1:
        body = "## 二、" + parts[1]
    cut = len(body)
    for marker in ("摘要卡", "信源汇总表", "信源表", "硬门禁自检表", "门禁自检表", "门禁条款自检"):
        m = re.search(r"\n#{1,6}[^\n]*" + re.escape(marker), body)
        if m:
            cut = min(cut, m.start())
    return body[:cut]


def _norm_text(s: str) -> str:
    return _PUNCT_STRIP_RE.sub("", s.strip())


def check_merge_cohesion(report_path: Path) -> List[str]:
    """v23 GATE0·合稿内聚（仅 交易决策报告_）：对每个面强制 闸A 禁引用跳转 + 闸B 核心字数
    ≥ 深稿核心 × 0.90 + 闸C 小节全覆盖。非 交易决策报告_ 直接返回空；深稿缺失由
    check_stage_a_face_drafts 负责，本函数对缺失面跳过（不重复报缺失）。"""
    issues: List[str] = []
    m = re.match(r"交易决策报告_(.+)$", report_path.stem)
    if not m:
        return issues
    tail = m.group(1)  # {code}_{简称}_{时间戳}
    outdir = report_path.parent
    try:
        text = report_path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return issues
    lines = text.splitlines()
    headings = collect_headings(lines)

    for face, sec in FACE_TO_SECTION.items():
        draft = outdir / f"分面深稿_{face}_{tail}.md"
        if not draft.exists():
            continue  # 缺失由 GATE0·分面深写处理
        try:
            dtext = draft.read_text(encoding="utf-8", errors="replace")
        except Exception:
            continue
        sec_text = get_section_text(lines, headings, f"#### {sec}")
        if sec_text is None:
            issues.append(
                f"[GATE0·合稿内聚] 合稿报告缺『#### {sec}』{face}章节——六个面深稿必须各自内聚为对应章节"
                "（基本面§4.1 / 政策§4.2 / 技术§4.3 / 资金§4.4 / 筹码§4.5 / 消息§4.6），不得整面缺章。"
            )
            continue

        # ── 闸A·禁引用跳转 ──
        ref = _MERGE_REF_JUMP_RE.search(sec_text)
        if ref or "分面深稿" in sec_text:
            hit = ref.group(0) if ref else "分面深稿"
            issues.append(
                f"[GATE0·合稿内聚·禁引用] {face}章节(§{sec})出现对深稿/外部 md 的引用跳转（命中『{hit}』）。"
                "合稿必须把深稿核心正文【逐字内聚】，读者只看合稿即应获得该面全部信息，"
                "严禁写『详见/请见 分面深稿 / OutputReport/*.md』等转引。"
            )

        # ── 闸B·核心字数 ≥ 深稿核心 × 0.90 ──
        core = _draft_core_for_merge(dtext)
        core_eff = effective_length(core)
        sec_eff = effective_length(sec_text)
        floor = int(core_eff * MERGE_COHESION_RATIO)
        if core_eff > 0 and sec_eff < floor:
            issues.append(
                f"[GATE0·合稿内聚·字数] {face}章节(§{sec})效率字数 {sec_eff} < 对应深稿核心正文 {core_eff} × "
                f"{MERGE_COHESION_RATIO:.2f} = {floor}（缺口约 {floor - sec_eff} 字）。"
                "合稿是『只增不减的逐字内聚』，深稿核心正文（已扣除顶部结论卡 + 尾部摘要卡/信源表/自检表）"
                "须基本全量进入该章节，禁止概括/压缩/薄片化。"
            )

        # ── 闸C·小节全覆盖（深稿每个分析小节标题须在合稿章节出现，仅可重新编号）──
        sec_blob = _norm_text(sec_text)
        missing_subs: List[str] = []
        for hm in re.finditer(r"(?m)^#{2,4}\s+(.+)$", core):
            raw = hm.group(1).strip()
            title = _norm_text(_HEADING_NUM_PREFIX_RE.sub("", raw))
            if len(title) < 3:  # 跳过过短/泛化标题，降低误伤
                continue
            if title not in sec_blob:
                missing_subs.append(raw)
        if missing_subs:
            shown = "、".join(missing_subs[:8]) + ("…" if len(missing_subs) > 8 else "")
            issues.append(
                f"[GATE0·合稿内聚·小节覆盖] {face}章节(§{sec})缺失深稿以下分析小节（标题去编号后未在合稿章节出现）：{shown}。"
                "合稿须逐字保留深稿每个分析小节（仅可重新编号/层级下沉），不得整节丢弃或概括为一句话。"
            )
    return issues


# ═══════════════════════════════════════════════════════════════════
# v24 GATE0·六面导览（faces-split 不合稿架构专用，取代 check_merge_cohesion）
# ───────────────────────────────────────────────────────────────────
# 决策稿不再逐字内聚六面，但必须对每个面给出【结论速览】并以引用记号
#   [详见：{面名}] 指向对应独立深稿（md 里是纯文本，Tab 多页面 HTML 下由 md2html_report.py
#   convert_face_refs 转成"切换到对应标签页"，读者点击即切到该面标签页）。
# 深稿是否够深由 check_stage_a_face_drafts 负责（与本函数互补，不重复）。
# ═══════════════════════════════════════════════════════════════════
def check_faces_split_navigation(report_path: Path, text: str) -> List[str]:
    """v24 不合稿架构：决策稿须对每个【已落盘】的面出现 [详见：{面名}] 深稿引用记号。
    非 交易决策报告_ 直接返回空；某面深稿缺失由 check_stage_a_face_drafts 报缺失。"""
    issues: List[str] = []
    m = re.match(r"交易决策报告_(.+)$", report_path.stem)
    if not m:
        return issues
    tail = m.group(1)
    outdir = report_path.parent
    # ① 决策稿须含「六面结论速览」承载章节（§四 多维度分析仍是其锚）
    if "六面结论速览" not in text and "多维度分析" not in text:
        issues.append(
            "[GATE0·六面导览] 决策稿(faces-split)缺『§四 六面结论速览』承载章节——"
            "不合稿架构下，决策稿须在 §四 对六个面各写一段结论速览并以 [详见：{面名}] 指向独立深稿。"
        )
    # ② 对每个已落盘的面，决策稿须出现 [详见：{面名}] 引用记号（容忍半/全角冒号）
    for face in INTENT1_FACE_NAMES:
        draft = outdir / f"分面深稿_{face}_{tail}.md"
        if not draft.exists():
            continue  # 缺失由 check_stage_a_face_drafts 报
        token_full = f"[详见：{face}]"
        token_half = f"[详见:{face}]"
        if token_full not in text and token_half not in text:
            issues.append(
                f"[GATE0·六面导览] 决策稿缺对『{face}』的深稿引用记号 `{token_full}`。"
                "不合稿架构下，决策稿须对每个面写结论速览并以 [详见：{面名}] 记号指向独立深稿"
                "（md 中为纯文本，Tab 多页面 HTML 下自动转为切换到对应标签页）。"
            )
    return issues


# ═══════════════════════════════════════════════════════════════════
# v9 新增（B 档）：场景化数据信源连带门禁（内容驱动，复用 v8 industry.md 模式）
# ───────────────────────────────────────────────────────────────────
# 背景：trade_advisor 仅自动并行 13 个核心脚本；风险红线 P0 + 筹码/股权激励/
#   海外可比/期权情绪等"按需脚本"全靠 LLM 自觉调用，无门禁兜底 → 漏采或裸编
#   数据仍能 PASS（v8 门禁只校验 industry.md，对其余脚本产出零校验）。
#   本门禁沿用 v8 industry.md 的成熟逻辑——"写了什么必须采了什么"：
#     ① A 档（风险红线 P0）：汇总决策第 2 步『全维度风险排查』为必跑项，要求
#        FinancialData/{code}_regulatory.json 或 _asset_quality.json 至少 1 个落盘。
#        （脚本即使 degraded 也会落盘，故本校验只证明 P0 脚本被调用，
#         不阻断脚本失败后用 web_search 兜底——因为兜底前脚本一定先跑过且落了盘。）
#     ② B 档（内容触发）：正文一旦出现特定维度主张（股权激励/海外可比/期权情绪/
#        机构持股明细），反查对应脚本产出文件存在。不声称就不查，最大限度避免
#        误伤轻量报告与纯 web_fetch 路径。
#   可经 --no-companion-check 与 v8 连带门禁一并关闭。
# ═══════════════════════════════════════════════════════════════════
def check_scenario_data_sources(report_path: Path, text: str) -> List[str]:
    """汇总决策报告（交易决策报告_）专用：风险红线 P0 必跑 + 内容触发的场景信源连带校验。

    返回 issues 列表（空 = 通过）。非标准命名的报告直接返回空（不误伤）。
    """
    issues: List[str] = []
    m = TRADE_CODE_RE.search(report_path.name)
    if not m:
        return issues
    code = m.group(1)
    fdata = report_path.parent.parent / "FinancialData"

    def _exists(*names: str) -> bool:
        return any((fdata / n).exists() for n in names)

    def _glob_exists(pattern: str) -> bool:
        try:
            return len(list(fdata.glob(pattern))) > 0
        except Exception:
            return False

    # ── A 档：风险红线 P0 必跑（regulatory / asset_quality 至少 1 个落盘）──
    if not _exists(f"{code}_regulatory.json", f"{code}_asset_quality.json"):
        issues.append(
            "[场景信源·风险红线P0] 汇总决策第 2 步『全维度风险排查』为必跑项，"
            f"但 FinancialData/ 下既无 {code}_regulatory.json 也无 {code}_asset_quality.json。"
            "请至少运行其一（脚本 degraded 也会落盘，本校验只验脚本是否被调用过）：\n"
            f"  python scripts/regulatory_action_scraper.py {code}\n"
            f"  python scripts/asset_quality_scraper.py {code}\n"
            "严禁跳过风险红线脚本直接在正文裸写『无重大风险/无监管处罚』。"
        )

    # ── A 档（v1.29 新增）：大股东/董监高/实控人增减持 无条件必采 ──
    # 病根复盘（历史校准根因）：筹码面深稿漏采大股东减持，正文裸写"管理层零减持传递信心"，
    #   而合格深稿应识别出实控人减持 650 万股 + 触发 CR1 红线——方向性误判（看多 vs 看空）。
    #   根因正是增减持数据从来不是"必采项"：筹码面 CR1-CR5 与 团队铁律#10 强依赖
    #   增减持方向，但门禁从不校验该脚本是否被调用过，漏采时仍能 PASS。
    #   本门禁把增减持采集升级为 A 档无条件硬交付前提（脚本 degraded 也会落盘，仅验是否跑过）。
    if not _exists(f"{code}_insider_trading.json"):
        issues.append(
            "[场景信源·增减持必采] 大股东/董监高/实控人增减持是筹码面 CR1-CR5 红线与团队铁律#10 "
            f"的强依赖项，但 FinancialData/ 下缺少 {code}_insider_trading.json。"
            "漏采增减持会直接导致『方向性误判』（历史事故：标的实际有大股东减持却被裸写为"
            "『管理层零减持传递信心』，看空错判成看多）。请运行（degraded 也会落盘，本校验只验是否跑过）：\n"
            f"  python scripts/insider_trading_scraper.py {code}\n"
            "严禁未采增减持就在正文裸写『无减持/零减持/管理层增持传递信心』等方向性结论。"
        )

    # ── B 档：内容驱动触发（正文声称该维度 → 必须有对应脚本产出文件）──
    # 每条规则：(触发关键词, 精确文件名列表[任一存在即通过], glob 模式[可空], 维度名, 补救命令)
    scenario_rules = [
        (
            ["股权激励", "限制性股票", "股票期权激励", "业绩对赌", "解锁批次", "行权价"],
            [f"{code}_equity_incentive.json"], None,
            "股权激励",
            f"python scripts/equity_incentive_scraper.py {code}",
        ),
        (
            ["海外可比", "全球龙头对标", "美股可比", "港股可比", "国际对标", "海外对标"],
            [f"{code}_overseas_peers.json"], None,
            "海外可比公司",
            f"python scripts/overseas_comparable_scraper.py {code}",
        ),
        (
            ["十大流通股东", "机构持股比例", "机构持仓占比", "QFII", "陆股通持股明细", "北向持股明细"],
            [f"{code}_institution_holding.json", f"{code}_northbound.json"], None,
            "机构/北向持股明细",
            f"python scripts/institution_holding_scraper.py {code}",
        ),
        (
            ["隐含波动率", "期权 IV", "期权IV", "PCR", "Put/Call", "Skew", "期权情绪"],
            [], "option_iv_*.json",
            "期权情绪 IV/PCR",
            "python scripts/option_iv_scraper.py 510300",
        ),
    ]
    for kws, exact_files, glob_pat, dim, cmd in scenario_rules:
        if not contains_any(text, kws):
            continue
        ok = (_exists(*exact_files) if exact_files else False) or (
            _glob_exists(glob_pat) if glob_pat else False
        )
        if not ok:
            target = "/".join(exact_files) if exact_files else glob_pat
            issues.append(
                f"[场景信源·{dim}] 正文出现『{dim}』相关主张，但 FinancialData/ 下缺少对应脚本产出"
                f"（{target}）。凡正文写到的维度必须先用脚本采集落盘，严禁裸编数字：\n"
                f"  {cmd}"
            )
    return issues


def check_decision_artifact(report_path: Path, text: str) -> List[str]:
    """P0-1：汇总决策报告（交易决策报告_）必须有 quant_scorer all 产物
    OutputReport/{stem}_decision.json，且核心数字（综合胜率）与正文一致。

    设计目的：堵死"六维加权 LLM 手算 / 边写边推"的口子——结论赖以成立的
    确定性计算（胜率/赔率/收益风险比/决策矩阵）必须先行落盘为结构化产物，
    报告正文只能引用该产物，不能在写作时手改/重算（对抗事后合理化）。
    非标准命名报告直接返回空（不误伤）。
    """
    issues: List[str] = []
    m = TRADE_CODE_RE.search(report_path.name)
    if not m:
        return issues
    code = m.group(1)
    stem = report_path.stem
    decision_path = report_path.parent / f"{stem}_decision.json"
    if not decision_path.exists():
        issues.append(
            f"[GATE0·决策固化] 缺少量化决策产物 OutputReport/{stem}_decision.json。"
            f"汇总决策的六维加权胜率/收益风险比/拥挤度/决策矩阵必须由 quant_scorer.py 确定性计算并**先行落盘**"
            f"（禁止 LLM 手算，确保结论先于正文固化）。修复：准备 FinancialData/{code}_quant_input.json"
            f"（含 win_rate / risk_reward / crowding / decision_matrix 节），再运行 "
            f"`python scripts/quant_scorer.py all "
            f"--json-input FinancialData/{code}_quant_input.json --output OutputReport/{stem}_decision.json`。"
        )
        return issues
    try:
        dobj = json.loads(decision_path.read_text(encoding="utf-8"))
    except Exception as e:
        issues.append(
            f"[GATE0·决策固化] {decision_path.name} 解析失败（{e}）：请用 quant_scorer all 重新生成合法 JSON。"
        )
        return issues
    if not isinstance(dobj, dict) or "win_rate" not in dobj:
        issues.append(
            f"[GATE0·决策固化] {decision_path.name} 缺少 win_rate 节（六维加权胜率）。"
            "请确保 quant_input.json 含 win_rate 节并用 quant_scorer all 全量计算。"
        )
        return issues
    # 数字一致性：决策产物综合胜率 vs 正文"综合胜率 XX%"，偏差 >5pct 即 FAIL
    try:
        wr_pct = dobj["win_rate"]["results"]["win_rate_pct"]
    except Exception:
        wr_pct = None
    if isinstance(wr_pct, (int, float)):
        m2 = re.search(r"综合胜率[^0-9]{0,8}(\d+(?:\.\d+)?)\s*%", text)
        if m2:
            text_wr = float(m2.group(1))
            if abs(text_wr - float(wr_pct)) > 5:
                issues.append(
                    f"[GATE0·决策固化] 正文综合胜率 {text_wr}% 与决策产物 {decision_path.name} "
                    f"的 {float(wr_pct):.1f}% 偏差 >5pct。结论数字必须锚定确定性产物，"
                    "禁止写报告时手改/重算（防事后合理化）。请核对二者口径或重跑 quant_scorer。"
                )
    return issues


def check_forecast_scenario_consistency(report_path: Path, text: str) -> List[str]:
    """P0-1c（v1.23 修复漏洞 C）：汇总决策报告正文三情景 EPS/净利，必须与
    forecast_engine 产物 {stem}_forecast.json 的 L4 三档 year_1 一致。

    设计目的：堵死"正文三情景预测系 LLM 手工锚定、与脚本产物脱钩"的口子——
    既然 GATE0 已要求 forecast.json 三档非空，正文引用就必须等于它（防伪「自下而上预测」）。
    forecast.json 不存在时返回空（缺失由 GATE0 负责拦截，此处不重复报）。
    """
    issues: List[str] = []
    if not TRADE_CODE_RE.search(report_path.name):
        return issues
    fc_path = report_path.parent / f"{report_path.stem}_forecast.json"
    if not fc_path.exists():
        return issues
    try:
        fc = json.loads(fc_path.read_text(encoding="utf-8"))
        l4 = fc.get("L4") or {}
    except Exception:  # noqa: BLE001
        return issues
    scn_cn = {"bull": "乐观", "base": "中性", "bear": "悲观"}
    # 匹配三情景表行：| 🟢 乐观 | <EPS> | <净利> | ...
    scn_re = {
        "bull": re.compile(r"乐观[^\n|]*\|\s*([\d.]+)\s*\|\s*([\d.]+)"),
        "base": re.compile(r"中性[^\n|]*\|\s*([\d.]+)\s*\|\s*([\d.]+)"),
        "bear": re.compile(r"悲观[^\n|]*\|\s*([\d.]+)\s*\|\s*([\d.]+)"),
    }
    EPS_TOL = 0.08          # EPS 绝对容差（元）
    NP_REL_TOL = 0.08       # 净利相对容差（8%）
    for sc in ("bull", "base", "bear"):
        y1 = ((l4.get(sc) or {}).get("year_1")) or {}
        fc_eps = y1.get("eps")
        fc_np = y1.get("net_profit_parent_yi")
        m = scn_re[sc].search(text)
        if not m:
            continue
        try:
            txt_eps = float(m.group(1))
            txt_np = float(m.group(2))
        except ValueError:
            continue
        if isinstance(fc_eps, (int, float)) and abs(txt_eps - fc_eps) > EPS_TOL:
            issues.append(
                f"[GATE0·三情景一致性] {scn_cn[sc]}情景正文 EPS {txt_eps} 与 forecast.json 的 "
                f"{fc_eps} 偏差 >{EPS_TOL}。正文三情景预测必须等于 forecast_engine 产物，"
                "禁止手工锚定/改写（防伪『自下而上预测』）。请重跑引擎或据产物回写正文。"
            )
        if isinstance(fc_np, (int, float)) and abs(txt_np - fc_np) / max(abs(fc_np), 1.0) > NP_REL_TOL:
            issues.append(
                f"[GATE0·三情景一致性] {scn_cn[sc]}情景正文净利 {txt_np}亿 与 forecast.json 的 "
                f"{round(fc_np, 1)}亿 相对偏差 >{int(NP_REL_TOL * 100)}%。请以引擎产物为准回写正文。"
            )
    return issues


def check_conclusion_derivability(report_path: Path, text: str) -> List[str]:
    """P0-1d（v1.24 ②结论可推导性硬门禁）：汇总决策报告的"总体结论"目标价
    必须可由 forecast_engine 三档目标价区间推导，杜绝"叶子数字被门禁强制更新、
    但首屏评级 / §5 目标价等总体结论纹丝不动"的事后合理化脱钩。

    以 {stem}_forecast.json 的 synthesis.factor_4_scenario_target_prices 三档目标价为锚：
      ① 包络：§5 概率加权矩阵每档目标价须落在 [bear_low, bull_high]（含 ±25% 战术容差）内；
      ② 中性锚：§5 中性档目标价中值与 forecast base 目标价偏离 ≤ 30%。
    forecast.json 不存在时返回空（缺失由 GATE0 拦截，此处不重复报）。
    """
    issues: List[str] = []
    if not TRADE_CODE_RE.search(report_path.name):
        return issues
    fc_path = report_path.parent / f"{report_path.stem}_forecast.json"
    if not fc_path.exists():
        return issues
    try:
        fc = json.loads(fc_path.read_text(encoding="utf-8"))
        f4 = (fc.get("synthesis") or {}).get("factor_4_scenario_target_prices") or {}
    except Exception:  # noqa: BLE001
        return issues

    def _g(sc: str, key: str) -> Optional[float]:
        v = (f4.get(sc) or {}).get(key)
        return float(v) if isinstance(v, (int, float)) else None

    bear_low = _g("bear", "target_price_low")
    base_mid = _g("base", "target_price_mid")
    bull_high = _g("bull", "target_price_high")
    if bear_low is None or base_mid is None or bull_high is None:
        return issues  # 目标价不全，交由其它门禁

    ENV_TOL = 0.25          # 战术容差（6 个月目标 vs 公允价值）
    NEUTRAL_REL_TOL = 0.30  # 中性档与 forecast base 公允价值的最大允许偏离
    env_low = bear_low * (1 - ENV_TOL)
    env_high = bull_high * (1 + ENV_TOL)

    # 提取 §5 概率加权矩阵三档目标价区间（形如 14.50-15.50，要求两端均含小数以避开"概率%"误抓）
    range_re = {
        "bull": re.compile(r"乐观[^\n]*?(\d+\.\d+)\s*[-~–]\s*(\d+\.\d+)"),
        "base": re.compile(r"中性[^\n]*?(\d+\.\d+)\s*[-~–]\s*(\d+\.\d+)"),
        "bear": re.compile(r"悲观[^\n]*?(\d+\.\d+)\s*[-~–]\s*(\d+\.\d+)"),
    }
    scn_cn = {"bull": "乐观", "base": "中性", "bear": "悲观"}
    for sc in ("bull", "base", "bear"):
        m = range_re[sc].search(text)
        if not m:
            continue
        try:
            lo, hi = float(m.group(1)), float(m.group(2))
        except ValueError:
            continue
        mid = (lo + hi) / 2
        if hi > env_high or lo < env_low:
            issues.append(
                f"[GATE0·结论可推导性] §5 {scn_cn[sc]}档目标价 {lo}-{hi} 超出 forecast 三档目标价"
                f"可推导包络 [{round(env_low, 2)}, {round(env_high, 2)}]"
                f"（锚 bear_low={bear_low}/bull_high={bull_high}，±{int(ENV_TOL * 100)}% 战术容差）。"
                "总体结论必须可由盈利预测推导——请重算结论或复核 forecast 产物。"
            )
        if sc == "base" and abs(mid - base_mid) / max(abs(base_mid), 1e-6) > NEUTRAL_REL_TOL:
            issues.append(
                f"[GATE0·结论可推导性] §5 中性档目标价中值 {round(mid, 2)} 与 forecast base 目标价 "
                f"{round(base_mid, 2)} 偏离 >{int(NEUTRAL_REL_TOL * 100)}%。疑似盈利预测已更新而总体结论未联动，"
                "请基于 forecast 产物重算中性目标价（或在正文显式说明 6 个月战术目标与公允价值的差异依据）。"
            )
    return issues


def check_assumptions_disclosure(report_path: Path, text: str) -> List[str]:
    """P0-1e（v1.24 ③假设条件显式披露）：含 §4.1.4 盈利预测的汇总决策报告，
    必须在正文显式披露驱动 EPS / 估值的关键假设——**数值（三档）+ 推导依据 + 信源**，
    源自 assumptions.yaml，便于用户独立核查"预测的灵魂（假设）"是否合理。

    设计动机：forecast_engine 只是确定性计算器，"预测对不对"取决于 assumptions.yaml
    里 LLM 填的假设。把这些假设连同依据/信源摊在报告里，用户才有可能发现"假设拍脑袋"。
    短线 / 超短线（无 §4.1.4）跳过，避免误伤。forecast.json 不存在时由 GATE0 拦截。
    """
    issues: List[str] = []
    if not TRADE_CODE_RE.search(report_path.name):
        return issues
    # 仅对真正做了盈利预测（§4.1.4）的报告生效
    if "4.1.4" not in text or not any(k in text for k in ("盈利预测", "EPS", "归母净利")):
        return issues

    # ① 透明度：必须让用户能核查"依据 + 信源"
    has_basis = any(k in text for k in ("推导依据", "驱动假设", "核心假设", "假设依据", "假设依据"))
    has_source = any(k in text for k in ("信源", "来源", "assumptions.yaml", "一致预期"))
    # ② 命门假设项必须逐项显式出现（驱动 EPS 的最小集 + 估值锚）
    key_items = {
        "营收增速": ["营收增速", "收入增速", "营业收入增速"],
        "毛利率": ["毛利率"],
        "费用率/四费": ["费用率", "四费", "期间费用"],
        "等效税率/所得税": ["等效税率", "税率", "所得税"],
        "WACC/折现率": ["WACC", "折现率", "加权平均资本成本"],
    }
    missing = [name for name, kws in key_items.items() if not any(k in text for k in kws)]

    if not has_basis:
        issues.append(
            "[GATE0·假设披露] §4.1.4 盈利预测缺少『推导依据/驱动假设』说明：每条关键假设必须写明"
            "推导依据（源自 assumptions.yaml 的 comment），否则用户无法核查『预测的灵魂』是否合理。"
        )
    if not has_source:
        issues.append(
            "[GATE0·假设披露] §4.1.4 盈利预测的关键假设缺少『信源』标注：须注明来源"
            "（年报/卖方一致预期/Mysteel 等，源自 assumptions.yaml 的 source），便于用户溯源核查。"
        )
    if missing:
        issues.append(
            f"[GATE0·假设披露] §4.1.4 未显式披露驱动 EPS/估值的命门假设：{', '.join(missing)}。"
            "请在『关键假设披露表』中逐项列出三档数值（乐观/中性/悲观）+ 推导依据 + 信源，便于用户检查。"
        )

    # ③ 反伪造（轻校验）：若正文披露了中性毛利率，须与 forecast.json L4 base 一致
    fc_path = report_path.parent / f"{report_path.stem}_forecast.json"
    if fc_path.exists():
        try:
            fc = json.loads(fc_path.read_text(encoding="utf-8"))
            b = ((fc.get("L4") or {}).get("base") or {}).get("year_1") or {}
            fc_gm = b.get("gross_margin_pct")
        except Exception:  # noqa: BLE001
            fc_gm = None
        if isinstance(fc_gm, (int, float)):
            # 抓"中性 ... 毛利率 ... XX%" 或 "毛利率 ... 中性 ... XX%"形态的中性档毛利率
            gm_re = re.compile(r"中性[^\n]{0,40}?毛利率[^\n%]{0,8}?(\d+(?:\.\d+)?)\s*%")
            gm_re2 = re.compile(r"毛利率[^\n%]{0,8}?(\d+(?:\.\d+)?)\s*%[^\n]{0,30}?中性")
            m = gm_re.search(text) or gm_re2.search(text)
            if m:
                try:
                    txt_gm = float(m.group(1))
                    if abs(txt_gm - fc_gm) > 0.5:
                        issues.append(
                            f"[GATE0·假设披露] 正文披露的中性毛利率 {txt_gm}% 与 forecast.json L4 base "
                            f"的 {fc_gm}% 偏差 >0.5pct。披露假设必须等于喂给引擎的实际假设，"
                            "禁止披露一套、计算另一套（防『假设橱窗』）。"
                        )
                except ValueError:
                    pass
    return issues


# ═══════════════════════════════════════════════════════════════════════
# P0-1f（v1.25 ④盈利预测「逐行推导计算链」硬门禁）
# ───────────────────────────────────────────────────────────────────────
# 痛点：P0-1e 只保证"假设值 + 依据 + 信源"被披露，但不保证报告展示了
#       "假设 → EPS 的逐行计算过程"。用户实测发现报告直接给出三情景 EPS/净利/
#       营收终值，却没写明这些数字是怎么从假设一行行算出来的（营收→毛利→减四费
#       →营业利润→减税→减少数股东→归母净利→÷股本=EPS），导致读者无法复算、
#       无法判断数字可信度。本门禁强制 §4.1.4 必须摊开这条利润表计算瀑布。
# ───────────────────────────────────────────────────────────────────────
# 判定（三者须同时具备，缺一即 FAIL）：
#   ① 推导指引词：出现「逐行推导/逐行计算/推导计算/计算式/计算过程/推导链/Step」之一；
#   ② 算式证据：出现 ≥3 处乘除运算符（× ÷ 乘 除），证明确有逐行运算而非罗列终值；
#   ③ 利润表链路：营收 / 毛利 / 营业利润 / 归母净利(或净利润) / EPS 五个节点至少命中 4 个。
# 仅对做了 §4.1.4 盈利预测的汇总决策报告生效；短线/超短线无 §4.1.4 自动跳过。
# ═══════════════════════════════════════════════════════════════════════
_DERIV_HINT_WORDS = (
    "逐行推导", "逐行计算", "推导计算", "计算式", "计算过程",
    "推导链", "逐步推导", "Step", "step",
)
_DERIV_ARITH_RE = re.compile(r"[×÷]|乘以|除以|乘|除")
_DERIV_CHAIN_NODES = (
    ("营收/收入", ("营收", "营业收入", "总收入")),
    ("毛利", ("毛利",)),
    ("营业利润", ("营业利润", "经营利润")),
    ("归母净利/净利润", ("归母净利", "归母净利润", "净利润")),
    ("EPS", ("EPS", "每股收益")),
)


def _extract_section_414(text: str) -> str:
    """切出 §4.1.4 章节正文。v1.32 改用编号语义切分（容忍标题 # 数量），
    取不到时退化为全文，保证不漏判。"""
    lines = text.splitlines()
    headings = collect_headings(lines)
    sec = get_section_text(lines, headings, "##### 4.1.4")
    return sec if sec else text


def check_forecast_derivation_chain(report_path: Path, text: str) -> List[str]:
    """P0-1f：§4.1.4 盈利预测必须展示「假设 → EPS」的逐行计算推导链，
    而不能只摆三情景终值。缺失则 GATE0 FAIL。"""
    issues: List[str] = []
    if not TRADE_CODE_RE.search(report_path.name):
        return issues
    if "4.1.4" not in text or not any(k in text for k in ("盈利预测", "EPS", "归母净利")):
        return issues

    sec = _extract_section_414(text)
    has_hint = any(w in sec for w in _DERIV_HINT_WORDS)
    arith_hits = len(_DERIV_ARITH_RE.findall(sec))
    chain_hits = sum(
        1 for _name, kws in _DERIV_CHAIN_NODES if any(k in sec for k in kws)
    )

    lacks: List[str] = []
    if not has_hint:
        lacks.append(
            "缺『逐行推导/计算式/计算过程』指引（须有一段把假设代入逐行计算的文字或推导表）"
        )
    if arith_hits < 3:
        lacks.append(
            f"算式运算符不足（× ÷ 乘 除 仅 {arith_hits} 处，需 ≥3，证明确有逐行运算而非罗列终值）"
        )
    if chain_hits < 4:
        missing_nodes = [n for n, kws in _DERIV_CHAIN_NODES if not any(k in sec for k in kws)]
        lacks.append(
            f"利润表链路节点仅命中 {chain_hits}/5（缺：{', '.join(missing_nodes)}；"
            "需覆盖 营收→毛利→营业利润→归母净利→EPS 中至少 4 个）"
        )

    if lacks:
        issues.append(
            "[GATE0·逐行推导] §4.1.4 盈利预测只给了三情景终值，未展示『假设 → EPS』的逐行计算过程："
            + "；".join(lacks)
            + "。请补一张『逐行推导计算表』，按 营收=Σ各分部(量×价)（见分部加总表，禁止直接拍整体增速）→ 毛利=营收×毛利率 → "
            "减销售/管理/研发/财务四费 → 营业利润 → ×(1-税率)=净利总额 → ×(1-少数股东占比)=归母净利 → "
            "÷总股本=EPS 逐行写出计算式与结果，让读者能独立复算（不得只摆终值）。"
        )
    return issues


# ═══════════════════════════════════════════════════════════════════════
# P0-1g（v1.25 ⑤营收「分部加总法（自下而上）」硬门禁）
# ───────────────────────────────────────────────────────────────────────
# 痛点：盈利预测最致命的偷懒是——营收直接拍一个「整体增速 +XX%」（自上而下），
#       而不把营收拆到不可再拆的产品/业务条线、用「量×价×份额」自下而上加总。
#       用户质疑："营收凭什么用整体增速假设？不应该分部加总、拆到没法拆吗？"
#       —— 完全成立。本门禁强制 §4.1.4 必须含『分部加总营收表』：≥2 条业务线
#       （单一业务公司须显式声明不可再拆才豁免），每条给量/价驱动，末行 Σ 加总，
#       整体增速只能是加总后的「反算校验值」，禁止作为拍脑袋输入。
# ───────────────────────────────────────────────────────────────────────
# 判定（缺分部加总证据且无单一业务豁免声明 → FAIL）：
#   分部关键词（分部/分业务/分产品/分条线/业务条线/分部加总/自下而上）
#   + 量价驱动（销量/出货量/出货量 + ASP/单价/均价/单价）
#   + 加总行（合计/加总/Σ/汇总/总营收）
# 豁免：正文显式写「单一业务/不可再拆/无法进一步拆分」之一（金融/公用事业等单一商业模式）。
# ═══════════════════════════════════════════════════════════════════════
_BUILDUP_SEG_KW = (
    "分部", "分业务", "分产品", "分条线", "各业务线", "业务条线", "分部加总", "自下而上",
)
_BUILDUP_VOL_KW = ("销量", "出货量", "出货", "销售量", "交付量")
_BUILDUP_PRICE_KW = ("ASP", "单价", "均价", "售价", "单位价格")
_BUILDUP_AGG_KW = ("合计", "加总", "Σ", "汇总", "总营收")
_BUILDUP_SINGLE_ESCAPE = (
    "单一业务", "不可再拆", "无法进一步拆分", "仅单一", "单一产品线", "单一商业模式", "无法再拆",
)


def check_revenue_buildup(report_path: Path, text: str) -> List[str]:
    """P0-1g：§4.1.4 营收预测必须用『分部加总法（自下而上）』，
    拆到不可再拆的产品/业务条线（量×价×份额），禁止只拍整体增速。"""
    issues: List[str] = []
    if not TRADE_CODE_RE.search(report_path.name):
        return issues
    if "4.1.4" not in text or not any(k in text for k in ("盈利预测", "EPS", "归母净利")):
        return issues

    sec = _extract_section_414(text)
    # 单一业务豁免（金融/公用事业等）：显式声明即放行
    if any(e in sec for e in _BUILDUP_SINGLE_ESCAPE):
        return issues

    has_seg = any(k in sec for k in _BUILDUP_SEG_KW)
    has_vol = any(k in sec for k in _BUILDUP_VOL_KW)
    has_price = any(k in sec for k in _BUILDUP_PRICE_KW)
    has_agg = any(k in sec for k in _BUILDUP_AGG_KW)

    lacks: List[str] = []
    if not has_seg:
        lacks.append("缺『分部/分业务/分产品』拆解（营收未拆到产品/业务条线，疑似只拍整体增速）")
    if not (has_vol and has_price):
        miss = []
        if not has_vol:
            miss.append("销量/出货量")
        if not has_price:
            miss.append("ASP/单价")
        lacks.append(f"分部缺『量×价』驱动列（缺：{', '.join(miss)}；每条业务须有量、价两个可观测驱动）")
    if not has_agg:
        lacks.append("缺『合计/Σ 加总行』（分部营收须 Σ 加总成总营收，整体增速为加总后反算值）")

    if lacks:
        issues.append(
            "[GATE0·分部加总] §4.1.4 营收预测未采用『分部加总法（自下而上）』："
            + "；".join(lacks)
            + "。请把营收拆到不可再拆的产品/业务条线，每条用『销量 × ASP（×份额）』驱动、末行 Σ 加总成总营收，"
            "整体营收增速只能是加总后的反算校验值，禁止直接拍一个整体增速。"
            "建议同步在 assumptions.yaml 填 `L4_income_statement.revenue_segments`，让 forecast_engine 按分部加总复算。"
            "若公司确为单一不可再拆业务（金融/公用事业等），请在正文显式声明『单一业务/不可再拆』以豁免。"
        )
    return issues


# ═══════════════════════════════════════════════════════════════════════
# P1-A：方法论指纹校验（把"关键词存在"升级为"方法论真被用上"）
# ───────────────────────────────────────────────────────────────────────
# 痛点：旧门禁只查"护城河/ROIC/Beneish/PEG/DCF"等词是否出现，LLM 可以"只提
#       名词"凑过门禁，而方法论文件里的分析手法实际没落地。本函数为每个关键
#       方法定义"指纹"——要求同时具备【方法名 + 关键数值/维度 + 结论落点】三件，
#       缺则判定为"疑似只引名词、方法未真正展开"。
# 分级（呼应 P1-1 思路）：
#   · 护城河多维 / ROIC×WACC 价值判断 —— 长线报告的核心方法，深度缺失 → 硬 FAIL
#   · Beneish 分值 / PEG 数值 / DCF 参数 —— 深度缺失 → 软 WARN（不阻断 PASS）
#   · 短线/超短线不做深度基本面，整体跳过（避免误伤）。
_ROIC_VAL_RE = re.compile(r"(?:ROIC|投入资本回报率)[^\n%]{0,80}?(\d+(?:\.\d+)?)\s*%")
_WACC_VAL_RE = re.compile(r"(?:WACC|加权平均资本成本|资本成本)[^\n%]{0,60}?(\d+(?:\.\d+)?)\s*%")
_MSCORE_VAL_RE = re.compile(r"M[\s\-]?Score[^\n]{0,20}?(-?\d+(?:\.\d+)?)")
_PEG_VAL_RE = re.compile(r"PEG[^\n]{0,18}?(\d+(?:\.\d+)?)")
_MOAT_DIM_WORDS = [
    "成本优势", "成本领先", "成本曲线", "网络效应", "转换成本", "迁移成本",
    "无形资产", "品牌", "专利", "特许", "牌照", "规模效应", "规模优势",
]
_MOAT_VERDICT_WORDS = [
    "宽护城河", "窄护城河", "无护城河", "护城河宽", "护城河窄", "护城河深",
    "护城河强", "护城河弱", "护城河评级", "护城河类型", "护城河趋势", "护城河变宽", "护城河变窄",
]


def check_methodology_fingerprints(
    text: str, style: Optional[str], report_type: str
) -> Tuple[List[str], List[str]]:
    """返回 (fail_items, warn_items)。仅对长线（swing/long/full）的
    汇总决策报告与基本面报告生效；其余直接返回空。"""
    fails: List[str] = []
    warns: List[str] = []
    if style not in {"swing", "long", "full"} or report_type not in {"trade", "fundamental"}:
        return fails, warns

    # —— 护城河：出现“护城河”则要求 ≥2 个来源维度 + 宽/窄/强弱判定（FAIL） ——
    if "护城河" in text:
        dim_hits = sum(1 for d in _MOAT_DIM_WORDS if d in text)
        has_verdict = contains_any(text, _MOAT_VERDICT_WORDS)
        if dim_hits < 2 or not has_verdict:
            _lack = []
            if dim_hits < 2:
                _lack.append(f"来源维度仅命中 {dim_hits}/5（成本/网络/转换/无形/规模，需≥2）")
            if not has_verdict:
                _lack.append("缺宽/窄/无 或 强/弱 的护城河判定")
            fails.append(
                "[方法论指纹·护城河] 出现“护城河”但未真正展开多维分析：" + "；".join(_lack) +
                "。护城河方法要求列出至少 2 个来源维度并给出宽窄/强弱判定，禁止只写“有护城河”。"
            )

    # —— ROIC vs WACC：要求 ROIC 数值 + WACC 对照 + 价值创造/毁灭结论（FAIL） ——
    if contains_any(text, ["ROIC", "投入资本回报率"]):
        has_roic_val = bool(_ROIC_VAL_RE.search(text))
        has_wacc = ("WACC" in text) or ("资本成本" in text)
        has_verdict = contains_any(text, [
            "价值创造", "价值毁灭", "创造价值", "毁灭价值", "超额回报", "超额收益率",
            "利差", "ROIC>WACC", "ROIC＞WACC", "ROIC<WACC", "ROIC＜WACC",
            "高于资本成本", "低于资本成本", "高于 WACC", "低于 WACC",
        ])
        if not (has_roic_val and has_wacc and has_verdict):
            _lack = []
            if not has_roic_val:
                _lack.append("ROIC 具体数值(x%)")
            if not has_wacc:
                _lack.append("WACC/资本成本对照")
            if not has_verdict:
                _lack.append("价值创造/毁灭结论")
            fails.append(
                "[方法论指纹·ROIC] ROIC 增长质量验证未真正展开，缺：" + "、".join(_lack) +
                "。要求形如“ROIC≈XX% vs WACC≈YY% → 价值创造/毁灭”，禁止只提 ROIC 名词。"
            )

    # —— Beneish M-Score：提及则要求具体分值或八因子明细（WARN） ——
    if contains_any(text, ["M-Score", "Beneish"]):
        if not (_MSCORE_VAL_RE.search(text) or contains_any(text, ["八因子", "8因子", "八项指标", "8 因子"])):
            warns.append(
                "[方法论指纹·Beneish] 提及 Beneish/M-Score 但未给出具体分值或八因子明细，"
                "疑似只引名词。要求形如“M-Score = -2.31（< -1.78，低造假概率）”。"
            )

    # —— PEG：出现则要求具体数值（WARN） ——
    if "PEG" in text and not _PEG_VAL_RE.search(text):
        warns.append(
            "[方法论指纹·PEG] 出现 PEG 但无具体数值，要求形如“PEG≈0.8（<1 低估）”，禁止空列名词。"
        )

    # —— DCF：使用 DCF 则要求 WACC 数值 + 永续增长率 g（WARN；已声明不适用则跳过） ——
    if "DCF" in text and not contains_any(text, ["不适用DCF", "不适用 DCF", "跳过DCF", "跳过 DCF", "未采用DCF", "未采用 DCF"]):
        has_wacc_val = bool(_WACC_VAL_RE.search(text))
        has_g = contains_any(text, ["永续增长", "永续增长率", "终值增长", "g=", "g ="])
        if not (has_wacc_val and has_g):
            warns.append(
                "[方法论指纹·DCF] DCF 估值缺关键参数披露（WACC 数值 / 永续增长率 g），"
                "要求显式列出 WACC=X%、永续 g=Y%，便于复核与逆向估值。"
            )

    return fails, warns


def validate_report(
    report_path: Path,
    style: Optional[str],
    require_supply_demand: bool = False,
    check_companion: bool = True,
) -> Dict:
    text = report_path.read_text(encoding="utf-8")
    lines = text.splitlines()
    headings = collect_headings(lines)

    detected_style = detect_style(text)
    style = style or detected_style
    if style not in STYLE_LABELS:
        raise ValueError("无法识别交易风格，请通过 --style 显式指定")

    issues: List[str] = []
    warnings: List[str] = []

    # v24：决策稿是否启用 faces-split（不合稿）架构——影响三处「强制 §4 内聚」逻辑：
    #   ① 结构闸（长报告必须含 §4.1.1-4.6 深度章节）→ 改由独立深稿承载，决策稿跳过；
    #   ② 合稿内聚闸 check_merge_cohesion → 换成 check_faces_split_navigation；
    #   ③ §4 字数地板 BLOCK 升级 → 不适用（§4 仅结论速览）。
    _faces_split = is_faces_split_report(text)

    # ═══════════════════════════════════════════════════════════════════
    # 报告类型识别（六个面框架 v2）
    # ═══════════════════════════════════════════════════════════════════
    # 本函数只处理两种"完整门禁"报告（五个面单面报告走轻量路径，在 main 提前分流）：
    #   trade（汇总决策报告）：含买/卖/止盈止损等操作指令，文件名 交易决策报告_
    #   fundamental（基本面报告）：卖方深度研究风格，无操作指令，文件名 基本面_
    # 二者章节结构不同，需要独立校验路径。
    first_h2 = None
    first_h1 = None
    for h_idx, h_level, h_text in headings:
        if h_level == 1 and first_h1 is None:
            first_h1 = h_text
        if h_level == 2 and first_h2 is None:
            first_h2 = h_text
        if first_h1 is not None and first_h2 is not None:
            break
    # v1.22 修复：HEADING_RE 只匹配 ##~#####，不收集 H1（#）。
    # 而基本面报告往往用 `# 广发证券（000776）基本面深度研究报告` 作大标题，
    # 因此在 headings 之外补一次源文本扫描，找首个 `# ` 行。
    if first_h1 is None:
        _h1_re = re.compile(r"^#\s+(.+)$")
        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue
            m1 = _h1_re.match(stripped)
            if m1:
                first_h1 = m1.group(1).strip()
                break
            # 遇到 ## 级标题就停（说明文档没用 H1 作大标题）
            if stripped.startswith("## "):
                break
    h1_text = first_h1 if first_h1 else ""
    h2_text = first_h2.lstrip("#").strip() if first_h2 else ""
    title_text = h1_text or h2_text
    title_candidate = f"{h1_text} {h2_text}"
    # 基本面报告判定：标题命中 基本面(研究/深度/分析)，或文件名前缀为 基本面_
    is_fundamental_research = bool(
        re.search(r"基本面(研究|深度|分析|报告)", title_candidate) or
        re.search(r"基本面深度研究", title_candidate) or
        is_fundamental_report_name(report_path.name)
    )
    report_type = "fundamental" if is_fundamental_research else "trade"

    # v25：faces-split 无回落路径——Intent-1 铁律#1 强制先出 6 份分面深稿、决策稿只做
    #   结论速览+导航，不存在任何合法场景下决策稿该走已删除的 v23 合稿内聚口径。
    #   缺架构标记 100% 是"忘记写标记"这一种格式错误：记一条结构性硬 FAIL 并把 _faces_split
    #   强制置 True，从而让下面所有 `not _faces_split` 的旧口径分支整体失活——
    #   避免同一份报告同时报"缺标记"+"缺4.1.1-4.6整段正文"等一堆冗余噪音，
    #   agent 只需看到一条清晰指令："补一行标记"，而不会被引导去做 10x 的无谓返工。
    if report_type == "trade" and not _faces_split:
        issues.append(
            "[GATE0·架构标记缺失] 决策稿顶部缺 `<!-- INTENT1_ARCH: faces-split -->` 标记。"
            "Intent-1 决策稿一律采用 faces-split（不合稿）架构，六面正文承载在 6 份分面深稿、"
            "决策稿只放结论速览+`[详见：面名]`导览。修复方式：在决策稿文件第 1 行加上该标记后重跑门禁，"
            "**不要**尝试把六面正文逐字搬进决策稿 §4.1~§4.6（那是已删除的 v23 口径，不再被支持）。"
        )
        _faces_split = True

    # ═══════════════════════════════════════════════════════════════════
    # GATE 0: 格式前置检查 — 标题格式、元信息、标题层级
    # ═══════════════════════════════════════════════════════════════════

    # 0-1. 标题格式：根据报告类型判定
    if first_h1 or first_h2:
        if report_type == "fundamental":
            # 基本面报告
            if not re.search(r"基本面(研究|深度|分析|报告)", title_text):
                issues.append(
                    f"[GATE0] 标题格式不符：基本面报告首个 ## 标题为「{title_text[:50]}」，"
                    f"应含「基本面研究/深度研究/分析报告」字样"
                )
            elif not re.search(r"\d{6}", title_text):
                issues.append(f"[GATE0] 标题缺少股票代码：「{title_text[:50]}」应包含6位代码如（300308）")
        else:
            # 汇总决策报告（交易决策报告）
            if "交易决策报告" not in title_text:
                issues.append(f"[GATE0] 标题格式不符：首个 ## 标题为「{title_text[:50]}」，应为「[标的名称]（[代码]）交易决策报告」")
            elif not re.search(r"\d{6}", title_text):
                issues.append(f"[GATE0] 标题缺少股票代码：「{title_text[:50]}」应包含6位代码如（300308）")
    else:
        issues.append("[GATE0] 未发现 ## 级别标题，报告结构严重异常")

    # 0-2. 元信息行：标题后必须有至少 3 个 **XX**: YY 格式行
    # v1.22 修复：基本面报告常用 blockquote（> **XX**：）传递元信息，因此允许 `> ` 前缀
    meta_line_re = re.compile(r"^(?:>\s*)?\*\*[^*]+\*\*\s*[:：]")
    # v1.22 修复：HEADING_RE 只匹配 ##~#####；基本面报告用 H1（#）当大标题，
    # 因此元信息识别用单独的 H1+H2 正则覆盖
    _meta_h1_h2_re = re.compile(r"^#{1,2}\s+")
    meta_count = 0
    found_title = False
    for line in lines:
        stripped = line.strip()
        if not found_title:
            if _meta_h1_h2_re.match(stripped):
                found_title = True
            continue
        # 标题后的行：空行跳过，遇到 --- 或下一个标题则停止
        if not stripped:
            continue
        if stripped == "---":
            break
        if HEADING_RE.match(stripped) or _meta_h1_h2_re.match(stripped):
            break
        if meta_line_re.match(stripped):
            meta_count += 1
    # 基本面报告无操作指令，元信息要求降为 2（核心结论/适用对象/数据截止时间，至少 2 行）
    # 因为基本面报告标的+代码+报告类型已在 H1 大标题（含『基本面研究/深度研究/分析报告』）中体现
    if report_type == "fundamental":
        min_meta = 2
    else:
        min_meta = 4 if style in {"swing", "long", "full"} else 3
    if meta_count < min_meta:
        if report_type == "fundamental":
            hint = "（建议含 报告类型/数据截止时间/实时价格）"
        else:
            hint = "（需含交易风格/风险等级/数据截止时间/实时价格）"
        issues.append(f"[GATE0] 元信息不足：标题后仅 {meta_count} 个 **XX**: 行，最低要求 {min_meta}{hint}")

    # 0-3. 标题层级嵌套
    level_dist = {2: 0, 3: 0, 4: 0, 5: 0}
    for _, lv, _ in headings:
        if lv in level_dist:
            level_dist[lv] += 1
    if style in {"swing", "long", "full"}:
        # 波段/中长线：必须有 ###(h3) + ####(h4) + #####(h5) 三级
        if level_dist[3] == 0:
            issues.append("[GATE0] 缺少 ### 级别标题（章标题）")
        if level_dist[4] == 0:
            issues.append("[GATE0] 缺少 #### 级别标题（节标题 4.1~4.7）")
        # 基本面报告使用 1.1/2.1 等 H4 编号体系，不强制 H5
        # v1.32：五段式存在性改按"编号语义"判定（4.1.1~4.1.5 任意 # 层级均算存在），
        # 不再用 level_dist[5]==0 这种对 # 字面数量敏感的脆弱判据。
        # v24：faces-split 不合稿架构下，五段式基本面承载在【基本面深稿】里，决策稿只做结论速览，故跳过。
        if report_type == "trade" and not _faces_split and not all(
            has_numbered_heading(lines, f"4.1.{_k}") for _k in range(1, 6)
        ):
            issues.append("[GATE0] 缺少五段式基本面子节（4.1.1~4.1.5）— 波段/中长线必须使用五段式基本面（编号语义校验，容忍标题 # 层级）")
    elif style in {"short", "ultra_short"}:
        # 短线/超短线：至少有 ###(h3)
        if level_dist[3] == 0:
            issues.append("[GATE0] 缺少 ### 级别标题（章标题）")

    for heading in COMMON_REQUIRED + COMMON_SUBSECTIONS:
        if not _subsection_present(lines, text, heading):
            # 基本面报告不需要汇总决策报告的章节模板
            if report_type == "fundamental":
                continue
            issues.append(f"缺少必备章节/标题：{heading}")

    # 风险预案章节：兼容新/旧结构（二选一即可） — 仅汇总决策报告要求
    if report_type == "trade" and not any(alt in text for alt in RISK_PLAN_ALTERNATIVES):
        issues.append(
            f"缺少风险预案章节：需包含 `### 六、风险预案`（6章结构）或 `### 七、风险预案`（7章结构，含五综合研判）之一"
        )

    if report_type == "trade" and not _faces_split and style in {"swing", "long", "full"}:
        for heading in LONG_FORM_SUBSECTIONS:
            if not _subsection_present(lines, text, heading):
                issues.append(f"缺少长报告必备章节：{heading}")

        section_411 = get_section_text(lines, headings, "##### 4.1.1")
        section_412 = get_section_text(lines, headings, "##### 4.1.2")
        section_413 = get_section_text(lines, headings, "##### 4.1.3")
        section_414 = get_section_text(lines, headings, "##### 4.1.4")
        section_415 = get_section_text(lines, headings, "##### 4.1.5")

        section_map = {
            "4.1.1": section_411,
            "4.1.2": section_412,
            "4.1.3": section_413,
            "4.1.4": section_414,
            "4.1.5": section_415,
            "4.2": get_section_text(lines, headings, "#### 4.2"),
            "4.3": get_section_text(lines, headings, "#### 4.3"),
            "4.4": get_section_text(lines, headings, "#### 4.4"),
            "4.5": get_section_text(lines, headings, "#### 4.5"),
            "4.6": get_section_text(lines, headings, "#### 4.6"),
            "4.7": get_section_text(lines, headings, "#### 4.7"),
        }

        for section_id, section_text in section_map.items():
            if section_text is None:
                issues.append(f"缺少章节：{section_id}")

        if section_411 is not None:
            issues.extend(validate_section_411(section_411))
        if section_412 is not None:
            issues.extend(validate_section_412(section_412))
        if section_413 is not None:
            issues.extend(validate_section_413(section_413))
        if section_414 is not None:
            issues.extend(validate_section_414(section_414, require_supply_demand))
        if section_415 is not None:
            issues.extend(validate_section_415(section_415))

        for generic_id in ["4.2", "4.3", "4.4", "4.5", "4.6", "4.7"]:
            section_text = section_map.get(generic_id)
            if section_text is not None:
                issues.extend(validate_generic_section(generic_id, section_text))

    if require_supply_demand and report_type == "trade" and not _faces_split and style in {"swing", "long", "full"}:
        section_text = get_section_text(lines, headings, "##### 4.1.4")
        if section_text is None:
            issues.append("要求供需利润预测，但缺少章节：##### 4.1.4")
        else:
            required_groups = [
                ("供给", ["供给", "供需"]),
                ("需求", ["需求", "收入侧"]),
                ("收入", ["收入"]),
                ("成本", ["成本", "成本侧"]),
                ("费用", ["费用", "费用率"]),
                ("净利润", ["净利润"]),
            ]
            missing = missing_semantic_keywords(section_text, required_groups)
            if missing:
                issues.append(
                    "4.1.4 未完整覆盖供给-需求-收入-成本-费用-净利润联动：缺少 "
                    + ", ".join(missing)
                )

    if report_type == "trade":
        if "收益风险比" not in text:
            issues.append("缺少收益风险比推导")
        if "综合胜率" not in text:
            issues.append("缺少综合胜率评估")
        # v24 faces-split：基本面深度要素（护城河/ROIC/商业模式/造假预警/客户集中度/资本配置/
        #   国际对标/研发）承载在【基本面深稿】里，决策稿不内聚 → 这些关键词在「决策稿+基本面深稿」
        #   合并文本中扫描即视为达标；旧架构(合稿)下 _fund_scan 恒等于 text，行为完全不变。
        _fund_scan = text
        if _faces_split:
            _fd = _read_fundamental_draft_text(report_path)
            if _fd:
                _fund_scan = text + "\n" + _fd
        if style in {"swing", "long", "full"} and "护城河" not in _fund_scan:
            issues.append("长报告缺少护城河评估")
        if style in {"swing", "long", "full"} and not contains_any(_fund_scan, ["ROIC", "投入资本回报率"]):
            issues.append("长报告缺少ROIC增长质量验证")
        if style in {"swing", "long", "full"} and not contains_any(_fund_scan, ["商业模式", "收入模式", "价值主张"]):
            issues.append("长报告缺少商业模式分析")
        if style in {"swing", "long", "full"} and not contains_any(_fund_scan, ["M-Score", "Beneish", "财务造假", "造假预警"]):
            issues.append("长报告缺少财务造假预警排查")

        # ── v16：基本面深度要素硬化（让整体报告 §4.1 覆盖独立基本面深度报告的强制要素）──
        # 命题：独立基本面报告之所以更深，在于它强制覆盖 2.0~2.7 的"客户集中度/资本配置/
        # 国际对标/研发管线/业务条线量价"等要素，而整体报告旧门禁完全不查这些。v16 把其中
        # 通用性最强的两项升级为长报告硬 FAIL，行业差异大的两项作软 WARN 引导。
        # v24：faces-split 下这些要素在基本面深稿，故用 _fund_scan（合并文本）扫描。
        if style in {"swing", "long", "full"}:
            if not contains_any(_fund_scan, ["客户集中度", "前五大客户", "前五名客户", "大客户占比", "第一大客户", "客户依赖"]):
                issues.append(
                    "[基本面深度·BLOCK] §4.1 缺少客户集中度/大客户依赖分析"
                    "（前五大客户占比 + 对手方议价/绑定关系）——独立基本面深度报告的强制要素，不得省略"
                )
            if not contains_any(_fund_scan, ["资本配置", "ROIIC", "增量资本回报", "再投资回报", "回购", "分红率"]):
                issues.append(
                    "[基本面深度·BLOCK] §4.1 缺少管理层资本配置评估"
                    "（ROIIC vs WACC / 回购·分红·再投资的历史效率 / 激励对齐）——独立基本面深度报告的强制要素"
                )
            if not contains_any(_fund_scan, ["国际对标", "海外龙头", "全球龙头", "对标", "海外可比", "可比公司"]):
                warnings.append(
                    "[基本面深度·WARN] §4.1 未见国际对标/海外龙头对标——有国际龙头的行业"
                    "（科技/医药/消费/新能源/半导体）应拉取 1-2 家海外龙头 8-12 季 P&L 做实证锚点"
                )
            if not contains_any(_fund_scan, ["研发费用率", "研发投入", "研发占比", "研发管线", "在研项目", "专利"]):
                warnings.append(
                    "[基本面深度·WARN] §4.1 未见研发投入/管线分析——科技/医药/制造类为强制要素，其余行业建议补充"
                )

        # ── v16：关键市场数据防臆造·溯源覆盖软门禁 ──
        # 病根复盘（本次整改根因）：PE/主力资金/增减持/股价 等市场数据最易被凭记忆或估算
        # "臆造"——历史失误含 PE-TTM 误写 141（实际 102）、虚构 550 万股减持（实际 64 万股）、
        # 主力资金流向写反符号（实际 +126 亿写成 -59 亿）。这类数据【必须】来自当次落盘的
        # FinancialData/{code}_*.json（quote/fund_flow/insider_trading）并在正文标注来源编号。
        # 门禁此处做"溯源覆盖"软校验：高风险数字存在但通篇无任何来源标注 → WARN（硬纪律见团队铁律 10）。
        _HIGH_FAB_RISK = ["PE-TTM", "市盈率", "市净率", "主力净流入", "主力资金", "北向", "减持", "增持", "质押率", "融资余额"]
        if report_type == "trade" and contains_any(text, _HIGH_FAB_RISK):
            has_citation = ("<sup>" in text) or ("来源" in text) or ("信源" in text)
            if not has_citation:
                warnings.append(
                    "[防臆造·WARN] 正文含 PE/资金/增减持/质押 等高臆造风险市场数据，但通篇未见来源标注"
                    "（<sup>编号</sup> 或 来源/信源 列）。这类数据必须来自当次落盘 FinancialData/{code}_*.json"
                    "（quote/fund_flow/insider_trading）并逐处标注来源，严禁凭记忆/经验/估算填写（参见团队铁律 10）。"
                )

        # ── v1.29 新增：出口管制/实体清单漏采核查（防方向性政策风险遗漏）──
        # 病根复盘（历史校准根因）：政策面漏采美国 1260H 清单 / BIS 实体清单，对高度依赖
        #   海外（尤其北美算力/半导体/光通信）客户的标的，遗漏出口管制风险会系统性高估确定性。
        #   机器无法判定标的"是否应在清单上"，但能识别"明显有海外/北美敞口却完全不提清单核查"
        #   这一高危盲区——触发即 WARN，提示主动核查 BIS Entity List / NDAA 1260H / 制裁清单。
        _US_EXPOSURE = ["海外营收", "出海", "北美客户", "美国市场", "海外客户", "海外收入", "出口占比", "境外营收", "北美市场", "海外业务占比"]
        _EXPORT_CONTROL = ["实体清单", "entity list", "出口管制", "1260H", "1260h", "BIS", "制裁清单", "管制清单", "未被列入清单", "未列入实体清单"]
        if report_type == "trade" and contains_any(text, _US_EXPOSURE) and not contains_any(text, _EXPORT_CONTROL):
            warnings.append(
                "[政策面·出口管制漏采·WARN] 正文显示标的有显著海外/北美敞口（海外营收/出海/北美客户等），"
                "但全文未见任何『实体清单 / BIS Entity List / NDAA 1260H / 出口管制 / 制裁清单』核查。"
                "对依赖海外（尤其北美算力/半导体/光通信/AI 芯片）客户的标的，遗漏出口管制核查会系统性高估"
                "政策确定性（历史事故：漏采美国 1260H 清单致政策风险方向性低估）。请在政策面显式核查并给出"
                "『是否被列入 / 潜在被列入概率 / 一旦列入的营收冲击测算』结论，无论结论是否利空都须明确写出依据与信源。"
            )

        # ── v1.29 新增：约数技术指标臆造抽检（精确计算量禁用约数）──
        # 病根复盘（历史校准根因）：技术面深稿出现 MA~1330、RSI≈55 等约数写法。均线/RSI/MACD/
        #   KDJ/BOLL 等是由收盘价确定性计算出的【精确量】（可由 technical_indicator_calc.py 算到
        #   小数点后两位），写成约数/概数即说明未真正计算、而是凭记忆或目测"估"出来的——这是
        #   典型臆造。支撑/压力位用"附近/一线"是合理的（区间概念），但指标值必须精确。
        if report_type == "trade":
            _approx_ind = re.findall(
                r"(?:MA\s?\d{1,3}|EMA\s?\d{1,3}|RSI\s?\d{0,2}|MACD|KDJ|[KDJ]\s?值|DIF|DEA|BOLL|布林|乖离率?|CCI|WR|威廉)"
                r"[^。；\n]{0,6}(?:约|大约|≈|~|左右)\s*[-+]?\d",
                text,
            )
            if _approx_ind:
                _sample = "、".join(dict.fromkeys(_approx_ind))[:120]
                warnings.append(
                    "[防臆造·约数技术指标·WARN] 检测到技术指标疑似用约数/概数表达（如：" + _sample + "…）。"
                    "MA/EMA/RSI/MACD/KDJ/BOLL/乖离率/CCI/WR 等是由价量确定性计算的【精确量】，必须给出精确数值"
                    "（technical_indicator.py 可算到 0.01），写成『约/≈/~/左右+数字』即暴露未真正计算、凭目测估值的臆造。"
                    "支撑/压力『区间』可用『一线/附近』，但指标本身禁止约数。请落盘并回填精确值。"
                )

    # ═══════════════════════════════════════════════════════════════════
    # v7 新增：五、综合研判三子模块强制 — 仅汇总决策报告（交易决策）
    # ═══════════════════════════════════════════════════════════════════
    if report_type == "trade" and style in {"swing", "long", "full"}:
        section_five = get_section_text(lines, headings, "### 五、")
        if section_five is None:
            issues.append("缺少「### 五、综合研判」主章节")
        else:
            # 5.1 六维加权胜率推导表：需含 权重/分数/加权 三列或其等价表述
            if not contains_any(section_five, ["六维加权", "六维胜率", "加权胜率", "6维加权"]):
                issues.append("五、综合研判 5.1 缺少六维加权胜率推导表")
            elif not (contains_any(section_five, ["权重"]) and contains_any(section_five, ["加权"])):
                issues.append("五、综合研判 5.1 的六维加权胜率表缺少'权重'或'加权贡献'列")

            # 5.2 收益风险比量化表：需含 入场价 / 止盈 / 止损 / 赔率
            has_rr_core = contains_any(section_five, ["入场", "建仓"]) and contains_any(section_five, ["止盈"]) and contains_any(section_five, ["止损"])
            if not has_rr_core:
                issues.append("五、综合研判 5.2 收益风险比表缺少'入场价/止盈/止损'价位要素")

            # 5.3 三情景概率×目标价矩阵：需含 乐观/中性/悲观（允许别名）+ 概率 + 目标价
            scenario_aliases = {
                "乐观": ["乐观", "乐观情景", "bull"],
                "中性": ["中性", "中性情景", "基准", "基准情景", "base"],
                "悲观": ["悲观", "悲观情景", "bear"],
            }
            missing_scenarios = [
                canon for canon, aliases in scenario_aliases.items()
                if not contains_any(section_five, aliases)
            ]
            if missing_scenarios:
                issues.append(f"五、综合研判 5.3 三情景矩阵缺少情景：{', '.join(missing_scenarios)}")
            if not contains_any(section_five, ["概率", "加权期望", "期望收益"]):
                issues.append("五、综合研判 5.3 三情景表缺少'概率/期望收益'量化列")

            # 五、章节字数硬下限（适当宽松，此章节是汇总性质）
            # P1-1：降级为软建议（不阻断 PASS）
            sec_five_len = effective_length(section_five)
            if sec_five_len < 400:
                issues.append(
                    f"[字数建议] 五、综合研判 篇幅偏短：当前效率字数 {sec_five_len}，建议 ≥400（还差约 {400 - sec_five_len} 字）"
                )

            # v7.1 Q-1：章节互引硬化 ——
            # 五、综合研判作为收束章节，必须至少引用 2 个 4.x 子章节（将离散评分收束为综合结论）
            # 有效引用形式：4.1.3 / 4.1.4 / 4.1.5 / 4.2 / 4.3 / 4.4 / 4.5 / 4.6 / 4.7 等
            five_refs = len(re.findall(r"4\.\d(?:\.\d)?", section_five))
            if five_refs < 2:
                issues.append(
                    f"五、综合研判 章节互引不足：仅引用 {five_refs} 次 4.x 子章节"
                    "（应在各维度评级行中明示依据章节，如'基本面 ★★★ — 见 4.1.3/4.1.4'）"
                )

    # ═══════════════════════════════════════════════════════════════════
    # v7 新增：4.1.4 三情景利润预测表的核心字段检查 — 仅汇总决策报告
    # ═══════════════════════════════════════════════════════════════════
    if report_type == "trade" and style in {"swing", "long", "full"}:
        section_414 = get_section_text(lines, headings, "##### 4.1.4")
        if section_414:
            scenario_aliases_414 = {
                "乐观": ["乐观", "bull"],
                "中性": ["中性", "基准", "base"],
                "悲观": ["悲观", "bear"],
            }
            missing_414 = [
                canon for canon, aliases in scenario_aliases_414.items()
                if not contains_any(section_414, aliases)
            ]
            if missing_414:
                issues.append(f"章节 4.1.4 三情景利润预测缺少情景：{', '.join(missing_414)}")

    # ═══════════════════════════════════════════════════════════════════
    # v7 新增：4.7 研报辩证解读必须引用至少 2 家具体券商 + 列出评级 — 仅汇总决策报告
    # ═══════════════════════════════════════════════════════════════════
    if report_type == "trade" and style in {"swing", "long", "full"}:
        section_47 = get_section_text(lines, headings, "#### 4.7")
        if section_47:
            broker_patterns = [
                "证券", "国金", "中信", "中金", "海通", "国泰君安", "华泰", "招商", "广发", "申万",
                "民生", "东吴", "兴业", "国信", "安信", "东方", "方正", "长江", "银河", "中银",
                "山西", "群益", "国元", "华鑫", "开源", "华龙", "华创", "浙商", "中泰", "天风",
                "光大", "平安", "西南", "西部", "东兴", "信达", "国海", "长城", "太平洋",
            ]
            broker_hits = sum(1 for kw in broker_patterns if kw in section_47)
            if broker_hits < 2:
                issues.append(
                    "章节 4.7 研报辩证缺少至少 2 家具体券商的观点引用"
                    "（应列出券商名称+评级+目标价+核心观点的对比表）"
                )
            if not contains_any(section_47, ["买入", "增持", "中性", "减持", "卖出"]):
                issues.append("章节 4.7 缺少机构评级分布（买入/增持/中性/减持/卖出）")
            if not contains_any(section_47, ["辩证", "多空", "分歧", "预期差", "一致预期陷阱"]):
                issues.append("章节 4.7 缺少辩证分析段（多空观点对比 / 预期差识别 / 一致预期陷阱审视）")

            # v7.1 Q-5：辩证不仅要"有辩证段"，还要真正包含**反向观点**（避免"多数看多我也看多"同质化）
            reverse_view_kws = ["看空", "质疑", "反向", "高估", "泡沫", "过度乐观", "悲观情景",
                                "警示", "风险提示", "利空", "反对", "陷阱"]
            if not contains_any(section_47, reverse_view_kws):
                issues.append(
                    "章节 4.7 辩证流于表面：未出现任何反向观点/风险警示关键词"
                    "（应包含至少一段质疑主流观点的内容，如'市场一致看多的隐忧在于...'、"
                    "'若 XX 预期证伪将导致...'等）"
                )

    # ═══════════════════════════════════════════════════════════════════
    # v7 新增：附录数据信源汇总表 + 脚标引用 — 两类报告都适用
    # ═══════════════════════════════════════════════════════════════════
    if style in {"swing", "long", "full"}:
        if not contains_any(text, ["附录：数据信源", "附录:数据信源", "数据信源汇总", "信源汇总表"]):
            issues.append("缺少「附录：数据信源汇总表」（5 列：编号/名称/类型/URL或获取方式/时效）")
        # v27 脚标定义完整性硬校验：正文引用的 [^srcN] 必须能在信源表/定义块里找到对应编号，
        # 否则 md2html_report.py 转换后必然裸露成 C3 泄漏（此前只能靠事后临时脚本才能发现）。
        missing_refs = _undefined_footnote_refs(text)
        if missing_refs:
            issues.append(
                f"脚标引用缺定义（正文引用 {('、'.join('[^' + r + ']' for r in missing_refs))} "
                "但附录信源汇总表未收录对应编号，HTML 转换后必然裸露泄漏；须在信源表补行）"
            )
        # 基本面报告允许用纯文本来源标注，脚标硬性要求降为 0（仅做 warning）
        sup_count = len(re.findall(r"<sup>\d+</sup>", text))
        if report_type == "trade":
            # v24 faces-split：密集脚标承载在 6 份深稿（由 check_stage_a_face_drafts 校验脚标密度），
            #   决策稿只需对核心结论/目标价等关键数字挂源，故下限放宽到 2。
            _min_sup = 2 if _faces_split else 5
            if sup_count < _min_sup:
                issues.append(
                    f"脚标引用不足：正文中 <sup>N</sup> 脚标仅 {sup_count} 个，至少 {_min_sup} 个"
                    "（关键数据应在来源列或文字中标注脚标编号，与附录一一对应）"
                )

        # v7.1 Q-2：表格「来源」列覆盖度（汇总决策报告硬约束 5 张；基本面报告硬约束 5 张 + 软建议 ≥8 张）
        # 兼容 "来源/数据来源/信源/数据出处/预测来源/资料来源/信源类型/信源名称" 等多种表头写法
        # 子串匹配：单元格内含任一关键词即可
        source_hits = len(re.findall(
            r"[|｜][^|｜]*(?:来源|信源|数据出处|资料出处)[^|｜]*[|｜]", text))
        # v24 faces-split：带来源列的重型数据表多在深稿，决策稿放宽到 2。
        min_source = (2 if _faces_split else 5)  # 两类报告硬约束统一为 5（faces-split 决策稿 2）
        if source_hits < min_source:
            issues.append(
                f"数据表「来源」列覆盖不足：仅检测到 {source_hits} 张表包含来源列，至少需要 {min_source} 张"
                "（关键财务/估值/资金表必须含来源列，便于结论追溯）"
            )
        elif report_type == "fundamental" and source_hits < 8:
            warnings.append(
                f"基本面报告建议至少 8 张表带来源列（当前 {source_hits} 张），"
                "便于卖方研究的可追溯性。"
            )

    # ═══════════════════════════════════════════════════════════════════
    # v7 新增：首屏操作指令表（汇总决策报告必填，5 场景 × 4 列） — 仅汇总决策报告
    # ═══════════════════════════════════════════════════════════════════
    if report_type == "trade" and style in {"swing", "long", "full"}:
        section_one = get_section_text(lines, headings, "### 一、")
        if section_one:
            required_scenarios = ["已持仓", "止损", "止盈"]
            missing_ops = [s for s in required_scenarios if s not in section_one]
            if missing_ops:
                issues.append(
                    f"一、核心结论 缺少首屏 5 场景操作指令表的场景：{', '.join(missing_ops)}"
                    "（🟢已持仓者/🟡短线观望/🟡波段观望/🔴止损触发/🟢止盈分批 五行）"
                )
            # P0-1b：§一核心结论应锚定后文章节（弱信号 WARN，不阻断）——防"空抛结论"。
            # 生成顺序不可机器验证，故只作软提示：要求结论显式标注依据章节，间接逼结论挂靠论证。
            one_refs = len(re.findall(r"(?:见|详见|参见)\s*(?:第?[一二三四五六]|4\.\d|5\.\d)", section_one))
            if one_refs == 0:
                warnings.append(
                    "[决策固化·锚点] 一、核心结论未出现对后文章节的显式锚点（如'见 4.1.3 / 详见五'）。"
                    "建议每条结论标注依据章节，便于核验结论确由六面分析推导（路径A 写作纪律：先写四/五章、得出结论，再前置§一）。"
                )

    # ═══════════════════════════════════════════════════════════════════
    # 检查 .md 保持纯文字 + HTML（按需可选，未生成不阻断交付）
    # ═══════════════════════════════════════════════════════════════════
    if style in {"swing", "long", "full"}:
        # 1) .md 不应含 ![](...) 图片引用（图表只在 HTML 内）
        img_refs = re.findall(r"!\[[^\]]*\]\([^)]+\)", text)
        if img_refs:
            issues.append(
                f".md 文件中检测到 {len(img_refs)} 条 ![](...) 图片引用行 → "
                ".md 必须保持纯文字，图表由 md2html_report.py 在 HTML 内实时生成"
            )

        # 2) 同名 .html 文件为「按需可选产物」（v1.27）：默认只交付 .md，
        #    用户明确要求可视化版时才生成 HTML。未生成不再计入 issues（不阻断交付），
        #    仅作 warning 提示；若已生成则照常校验 SVG 数量与渲染结构。
        try:
            _html_path = report_path.with_suffix(".html")
            if not _html_path.exists():
                warnings.append(
                    f"HTML 可视化报告未生成（按需可选，不影响交付）：{_html_path.name} → "
                    f"如用户需要可视化版，运行 `python scripts/md2html_report.py "
                    f"{report_path}` 生成 HTML（图表会自动内嵌）"
                )
            else:
                try:
                    html_text = _html_path.read_text(encoding="utf-8")
                    svg_count = len(re.findall(r"<svg\b", html_text))
                    # 所有报告（汇总决策 / 基本面 等）均由 md2html_report.py 渲染：
                    # 图表以 inline <svg> 实时注入。SVG 数量阈值仅对汇总决策报告（trade）生效——
                    # 单面基本面深稿图表较少（多以 [[table:KEY]] 结构化数据表呈现），不强制 SVG 下限。
                    if svg_count < 5 and report_type == "trade":
                        issues.append(
                            f"HTML 报告内嵌图表不足：当前 {svg_count} 个 <svg>，至少需要 5 个。"
                            "请检查 FinancialData/ 数据是否齐全"
                        )
                except Exception:
                    pass
        except Exception:
            pass

    # ═══════════════════════════════════════════════════════════════════
    # v1.26 新增：HTML 渲染结构自检（不限报告类型/风格，汇总决策报告与基本面报告均生效）
    # 背景：v1.24（hero 大标题错位成章节名/泛词）、v1.25（图表目录位置错到第一章之后）
    #       两类错误都是 md2html_report.py 渲染结构出错，但旧门禁只查 HTML 是否存在 +
    #       SVG 数量，从不校验渲染结构，只能靠人眼发现。此处加结构断言形成自动防线。
    # ═══════════════════════════════════════════════════════════════════
    try:
        _html_path = report_path.with_suffix(".html")
        if _html_path.exists():
            _h = _html_path.read_text(encoding="utf-8")
            _m_hero = re.search(
                r'<header[^>]*class="[^"]*hero[^"]*"[^>]*>.*?<h1[^>]*>(.*?)</h1>',
                _h, re.DOTALL | re.IGNORECASE,
            )
            # 1) hero 大标题必须是报告主标题：不得是章节名(§开头/含"摘要 Snapshot")，且须含 6 位股票代码
            if _m_hero:
                _hero = re.sub(r"<[^>]+>", "", _m_hero.group(1)).strip()
                if _hero.startswith("§") or "摘要 Snapshot" in _hero:
                    issues.append(
                        f"[HTML] hero 大标题错位：当前为「{_hero[:40]}」疑似章节名，应为报告主标题。"
                        "请检查 md2html_report.py 的 extract_title（须取文档首个 H1/H2 报告标题）"
                    )
                elif not re.search(r"\d{6}", _hero):
                    issues.append(
                        f"[HTML] hero 大标题缺股票代码：当前为「{_hero[:40]}」疑为泛词，"
                        "应为「标的名（代码）…报告」。请检查 md2html_report.py 的 extract_title"
                    )
            # 2) <title> 应与 hero 大标题一致
            _m_title = re.search(r"<title>(.*?)</title>", _h, re.DOTALL | re.IGNORECASE)
            if _m_hero and _m_title:
                _t = _m_title.group(1).strip()
                _hero2 = re.sub(r"<[^>]+>", "", _m_hero.group(1)).strip()
                if _t and _hero2 and _t != _hero2:
                    warnings.append(f"[HTML] <title>「{_t[:30]}」与 hero 大标题「{_hero2[:30]}」不一致")
            # 3) 图表目录（若有）必须位于第一个章节标题之前（紧随目录、第一章之前）
            _main_pos = _h.find("<main")
            if _main_pos >= 0:
                _scope = _h[_main_pos:]
                _fig_pos = _scope.find("图表目录")
                _m_sec = re.search(r"<h[1-5][^>]*\sid=", _scope, re.IGNORECASE)
                _sec_pos = _m_sec.start() if _m_sec else -1
                if _fig_pos >= 0 and _sec_pos >= 0 and _fig_pos > _sec_pos:
                    issues.append(
                        "[HTML] 图表目录位置错位：出现在第一个章节标题之后，应置于正文最前"
                        "（目录之后、第一章之前）。请检查 md2html_report.py 的图表目录注入逻辑"
                    )
    except Exception:
        pass

    # ═══════════════════════════════════════════════════════════════════
    # 三段式 emoji 全局下限（配合章节级配对检查）
    # ═══════════════════════════════════════════════════════════════════
    data_count = text.count("📊")
    analysis_count = text.count("🔍")
    conclusion_count = text.count("📌")
    if report_type == "trade" and style in {"swing", "long", "full"}:
        # v7 调高阈值：基准报告均在 13-20 之间，之前 11 过低
        # v24 faces-split：海量三段式承载在 6 份深稿，决策稿（结论+速览+综合研判）下限放宽到 4。
        _min_seg = 4 if _faces_split else 13
        if data_count < _min_seg:
            issues.append(f"📊 数据段数量偏少：当前 {data_count}，建议至少 {_min_seg}")
        if analysis_count < _min_seg:
            issues.append(f"🔍 推导段数量偏少：当前 {analysis_count}，建议至少 {_min_seg}")
        if conclusion_count < _min_seg:
            issues.append(f"📌 结论段数量偏少：当前 {conclusion_count}，建议至少 {_min_seg}")

    table_count = count_tables(lines)
    # v24 faces-split：重型数据表在深稿，决策稿（5 场景指令/收益风险比/三情景/六维加权等）下限放宽到 5。
    _min_tables = (5 if _faces_split else 15)
    if style in {"swing", "long", "full"} and table_count < _min_tables:
        issues.append(f"表格数量偏少：当前 {table_count}，建议至少 {_min_tables}")

    # ═══════════════════════════════════════════════════════════════════
    # v7 新增：整份报告效率字数硬下限（终极防空壳）
    # ═══════════════════════════════════════════════════════════════════
    total_eff_len = effective_length(text)
    # v28 faces-split：决策稿不内聚六面正文——深度已完全交给六份分面深稿的
    #   STAGE_A_MIN_EFF_LEN 硬把关（check_stage_a_face_drafts / check_single_face_draft，
    #   基本10832/政策7950/资金6568/筹码5619/技术6322/消息6372 字，远高于旧决策稿门槛）。
    #   决策稿本身是"结论+导航"薄层（§一结论/§二风险红线/§三大盘/§四速览/§五研判/§六计划/
    #   §七预案），用裸字数卡它既冗余又可能误伤——表格密集、散文少的决策稿效率字数天然偏低。
    #   已有结构性门禁足够防空壳：表格数≥5 / 📊🔍📌三段式各≥4 / 六面导览逐面须
    #   [详见：面名]（check_faces_split_navigation）/ 收益风险比与综合胜率关键词 /
    #   §五综合研判三子模块（六维加权表/收益风险比表/三情景表）必填。
    #   故 faces-split 架构下（v25 起对 trade 报告无回落路径，恒为 True）整篇字数硬下限直接跳过；
    #   仅保留给理论上仍可能存在的非 faces-split 旧合稿口径（MIN_EFFECTIVE_LENGTH，当前无实际触发路径）。
    min_required = 0 if _faces_split else MIN_EFFECTIVE_LENGTH.get(style, 0)
    if min_required and total_eff_len < min_required:
        issues.append(
            f"报告整体深度不足：效率字数 {total_eff_len} < 门槛 {min_required}"
            f"（风格 = {STYLE_LABELS[style]}）。请按各章节补全数据表、推导段、三段式结构，"
            f"参考 templates/intent1_full_report.md 纯格式骨架 + references/faces/<face>.md §⭐ 报告输出规范，以及历史优秀报告（如 交易决策报告_300308_中际旭创.md）。"
        )

    if "CNBC" not in text and "FT" not in text and "Financial Times" not in text and "Reuters" not in text and "路透" not in text:
        warnings.append("未检测到国际权威财经媒体引用，建议复核宏观层信源")
    if "OutputReport" in text:
        warnings.append("报告正文中不建议写入路径性说明，可保留在流程层")

    # ═══════════════════════════════════════════════════════════════════
    # v1.8 新增：基本面研究报告"纯净度"硬约束（2026-05 用户反馈驱动）
    # ─────────────────────────────────────────────────────────────────
    # 用户痛点：① 基本面研究报告不应混入技术面/资金面/筹码面/市场情绪/研报辩证
    #          /大盘与板块/综合研判六维/精准交易计划等汇总决策报告专属章节
    #         ② 附录"数据信源汇总表"只能放真实外部信源，不能放
    #          "本报告预测/本报告分析/本报告假设/本报告测算/本报告计算"等
    #          本报告自身产出的条目——这些应改为正文章节交叉引用 [详见§X.X]
    # ═══════════════════════════════════════════════════════════════════
    if report_type == "fundamental":
        # ── 硬约束 纯净度-1：禁用章节清单（汇总决策报告六步流程章节不允许进基本面报告）─
        forbidden_section_patterns = [
            (r"^#{2,4}\s+.*?(?:大盘.*?(?:板块|环境)|板块.*?环境).*$",
             "基本面报告禁用章节「大盘与板块环境」（属汇总决策辅助层；基本面报告只做卖方深度研究）"),
            (r"^#{2,5}\s+.*?(?:4\.3\s*技术面|技术面[（(].*?[执执执][行][+＋].*?审视|技术面（执行)",
             "基本面报告禁用章节「4.3 技术面」（属技术面专项；基本面研究不做技术分析）"),
            (r"^#{2,5}\s+.*?(?:4\.4\s*资金面|资金面[（(].*?(?:验证|审视))",
             "基本面报告禁用章节「4.4 资金面」（属资金面专项）"),
            (r"^#{2,5}\s+.*?(?:4\.5\s*筹码面|筹码面[（(].*?(?:持仓|稳定))",
             "基本面报告禁用章节「4.5 筹码面」（属筹码面专项）"),
            (r"^#{2,5}\s+.*?(?:4\.6\s*市场情绪|市场情绪[（(].*?(?:择时|逆向))",
             "基本面报告禁用章节「4.6 市场情绪」（属消息面专项）"),
            (r"^#{2,5}\s+.*?(?:4\.7\s*研报辩证|研报辩证解读)",
             "基本面报告禁用章节「4.7 研报辩证」（属消息面专项；基本面报告可在结论部分简述一致预期偏差，无需单列章节）"),
            (r"^#{2,4}\s+.*?(?:五、综合研判|综合研判[（(]六维)",
             "基本面报告禁用章节「五、综合研判（六维收束）」（属汇总决策六维加权专章）"),
            (r"^#{2,4}\s+.*?精准交易计划",
             "基本面报告禁用章节「精准交易计划」（属汇总决策专属；基本面报告只回答估值定价，不下达交易指令）"),
            (r"^#{2,4}\s+.*?核心结论与操作指令",
             "基本面报告禁用章节「核心结论与操作指令」（属汇总决策首屏；基本面报告首屏应为「核心结论」+「投资概览卡」，不含操作指令）"),
        ]
        for pat, msg in forbidden_section_patterns:
            if re.search(pat, text, re.M):
                issues.append(f"[纯净度-1 章节] {msg}")

        # ── 硬约束 纯净度-2：附录信源汇总表"类型列"禁用伪信源 ──────
        # 真信源类型白名单（可扩展）；任何"本报告X / 内部测算"类型一律 FAIL
        forbidden_type_keywords = [
            "本报告预测", "本报告分析", "本报告假设", "本报告测算",
            "本报告计算", "本报告 DCF", "本报告 dcf", "本报告综合分析",
            "本报告敏感性", "本报告目标价", "本报告独立",
        ]
        # 探测附录区段
        appendix_match = re.search(
            r"(?:###?\s*附录[:：]?\s*数据信源.*?$|附录[:：].*?信源.*?$)([\s\S]*?)(?=^#{1,4}\s|\Z)",
            text, re.M,
        )
        if appendix_match:
            appendix_text = appendix_match.group(1)
            # 表格数据行（不要把表头算进去）
            row_lines = [
                ln for ln in appendix_text.splitlines()
                if ln.startswith("|") and not re.match(r"^\|[\s:\-\|]+\|$", ln.strip())
                and "编号" not in ln and "信源" not in ln.split("|")[1] if "|" in ln
            ]
            bad_rows = []
            for ln in row_lines:
                for kw in forbidden_type_keywords:
                    if kw in ln:
                        # 提取该行第一列作为可定位标识
                        first_cell = ln.split("|")[1].strip() if "|" in ln else ln[:30]
                        bad_rows.append(f"{first_cell} · 含禁用字眼「{kw}」")
                        break
            if bad_rows:
                issues.append(
                    "[纯净度-2 附录信源] 附录「数据信源汇总表」混入了非外部信源条目，"
                    f"共 {len(bad_rows)} 行：" + "; ".join(bad_rows[:5])
                    + (" …" if len(bad_rows) > 5 else "")
                    + "。请把这些行从附录表移除，并在正文中改用「[详见§X.X]」"
                    "形式交叉引用本报告自身的预测/假设/分析结论。"
                )

        # ── 硬约束 纯净度-3：附录表新格式（首列含 HTML 注释锚点）─────
        # 新规范：附录信源表的数据行首列应形如 `1<!--src1-->` 或 `1` 配合脚注 ID 标签，
        # 不应直接把 `src1` 等内部锚点 ID 作为给读者看的可见序号。
        if appendix_match:
            visible_src_id_rows = re.findall(
                r"^\|\s*(src[a-zA-Z0-9_\-]+)\s*\|", appendix_match.group(1), re.M
            )
            if visible_src_id_rows:
                issues.append(
                    f"[纯净度-3 附录信源可读性] 附录表首列直接展示了 {len(visible_src_id_rows)} 个"
                    "内部锚点 ID（如 `src1` / `src5`）。请改为 `| 1<!--src1--> |` 形式："
                    "首列对读者只显示阅读序号 1/2/3，HTML 注释中的 srcN 仅作脚注跳转锚点。"
                )

        # ── 硬约束 纯净度-4：D 类伪信源黑名单（v1.9 新增，2026-05 用户反馈）──
        # 用户痛点：前份报告附录混入 Wind/Bloomberg/IDC/Omdia/LightCounting 等
        #          AI 完全不可达的付费数据库，构成"伪信源"诚信事故。
        # 黑名单覆盖：付费金融终端 + 付费行业研究机构 + 付费咨询公司报告。
        # 触发范围：附录表所有列（信源名称/类型/URL/备注）+ 正文常见话术。
        D_CLASS_BLACKLIST_KEYWORDS = [
            # 付费金融终端
            "Wind", "wind", "万得", "Bloomberg", "bloomberg", "彭博",
            "Refinitiv", "路透 Eikon", "Eikon", "FactSet", "factset",
            "CapitalIQ", "Capital IQ", "S&P Global", "iFinD", "ifind",
            "同花顺 iFinD", "同花顺iFinD",
            "Choice 金融终端", "Choice金融终端", "聚源数据", "朝阳永续",
            # 付费行业研究机构
            "IDC", "Omdia", "LightCounting", "lightcounting",
            "Gartner", "gartner", "Counterpoint", "Dell'Oro", "DellOro",
            "Yole", "yole", "IHS Markit", "Strategy Analytics",
            "Canalys", "TrendForce 付费", "IC Insights",
            "SEMI 付费", "CINNO Research", "CINNO研究",
            # 付费咨询公司内部报告
            "McKinsey 内部", "Bain 内部", "BCG 内部", "Deloitte 内部",
            "Frost & Sullivan", "Forrester", "Euromonitor",
        ]
        # 黑名单 URL 域名（纯净度-5 子规则）
        D_CLASS_BLACKLIST_DOMAINS = [
            "wind.com.cn", "bloomberg.com/professional", "bloomberg.com/terminal",
            "refinitiv.com", "factset.com", "capitaliq.com", "spglobal.com",
            "ihs.com", "ihsmarkit.com",
            "idc.com", "omdia.com", "lightcounting.com", "gartner.com",
            "counterpointresearch.com", "delloro.com", "yole.fr",
            "strategyanalytics.com", "canalys.com",
            "icinsights.com", "cinno.com.cn",
            "mckinsey.com/our-insights/proprietary",
            "frost.com", "forrester.com", "euromonitor.com",
        ]

        if appendix_match:
            appendix_text_b4 = appendix_match.group(1)
            b4_hits = []
            for ln in appendix_text_b4.splitlines():
                if not ln.startswith("|"):
                    continue
                if re.match(r"^\|[\s:\-\|]+\|$", ln.strip()):
                    continue
                # 跳过表头行（含"编号""信源名称""类型"等表头关键词）
                if "信源名称" in ln or "编号" in ln or "时效" in ln:
                    continue
                ln_lower = ln.lower()
                # 关键词命中（区分大小写敏感与不敏感的混合判断）
                for kw in D_CLASS_BLACKLIST_KEYWORDS:
                    # 完整词边界匹配，避免误伤（如 "ChoiceGoods" 不应命中 "Choice"）
                    if kw.lower() in ln_lower:
                        # 进一步降低误伤：要求该 kw 不被 "公开""新闻""转引""转载""仅作背景"包裹
                        # 如果上下文是"权威媒体转引（原始来源：IDC，未独立验证）"则放过
                        if re.search(r"转引[自自].{0,30}" + re.escape(kw), ln) or \
                           re.search(r"原始来源[：:].{0,20}" + re.escape(kw) + r".{0,30}未[独独]立验证", ln):
                            continue
                        first_cell = ln.split("|")[1].strip() if "|" in ln else ln[:30]
                        b4_hits.append(f"{first_cell} · 命中 D 类黑名单关键词「{kw}」")
                        break
                # 域名命中（B-5）
                for dom in D_CLASS_BLACKLIST_DOMAINS:
                    if dom in ln_lower:
                        first_cell = ln.split("|")[1].strip() if "|" in ln else ln[:30]
                        b4_hits.append(f"{first_cell} · URL 命中 D 类黑名单域名「{dom}」")
                        break
            if b4_hits:
                issues.append(
                    "[纯净度-4/纯净度-5 D 类伪信源黑名单] 附录「数据信源汇总表」混入了 AI 完全不可达的"
                    f"付费数据库（Wind/Bloomberg/IDC/Omdia/LightCounting/Gartner 等），共 {len(b4_hits)} 行：\n  - "
                    + "\n  - ".join(b4_hits[:8])
                    + (f"\n  - …(还有 {len(b4_hits)-8} 行省略)" if len(b4_hits) > 8 else "")
                    + "\n  处置：① 删除该行；② 若数字必要，则按【铁律二·二手回溯一手原则】"
                    "搜索权威媒体（证券时报/财联社/中证报/上证报/新华社/证券日报/新浪财经/36氪）转引报道并替换 URL；"
                    "③ 若找不到任何 A/B 类信源，则整段删除该论述。"
                )

        # ── 硬约束 纯净度-4-Body：正文话术中"根据 Wind/Bloomberg/IDC..."拦截 ──
        body_b4_patterns = [
            (r"根据\s*Wind\s*[数数据据预预测测一一致致预预期期]", "Wind"),
            (r"根据\s*Bloomberg", "Bloomberg"),
            (r"根据\s*IDC\s*[预预测测数数据据]", "IDC"),
            (r"根据\s*Omdia", "Omdia"),
            (r"根据\s*LightCounting", "LightCounting"),
            (r"根据\s*Gartner", "Gartner"),
            (r"根据\s*Counterpoint", "Counterpoint"),
            (r"根据\s*Dell['’]?Oro", "Dell'Oro"),
            (r"根据\s*Yole", "Yole"),
            (r"Wind\s*一致预期", "Wind 一致预期（应改为东方财富一致预期）"),
        ]
        body_b4_hits = []
        for pat, kw in body_b4_patterns:
            for m in re.finditer(pat, text):
                # 取上下文 30 字定位
                start = max(0, m.start() - 20)
                end = min(len(text), m.end() + 30)
                snippet = text[start:end].replace("\n", " ")
                body_b4_hits.append(f"「{kw}」 ··· {snippet}")
        if body_b4_hits:
            issues.append(
                f"[纯净度-4 正文伪信源话术] 正文中出现「根据 D 类付费数据库」表述 {len(body_b4_hits)} 处：\n  - "
                + "\n  - ".join(body_b4_hits[:6])
                + (f"\n  - …(还有 {len(body_b4_hits)-6} 处省略)" if len(body_b4_hits) > 6 else "")
                + "\n  处置：必须替换为权威媒体转引 URL（证券时报/财联社/中证报/新华社等），"
                "或东方财富公开页面（一致预期、卖方研报汇总），找不到则整段删除。"
            )






    # ═══════════════════════════════════════════════════════════════════
    # v8 新增：汇总决策报告交付集合 + 行业信源硬门禁
    #   仅对 trade（交易决策报告_）类型生效；可经 --no-companion-check 关闭。
    # v9 新增：场景化数据信源连带门禁（风险红线 P0 + 股权激励/海外可比/期权情绪/
    #   机构持股明细 内容触发），同一开关控制。
    # ═══════════════════════════════════════════════════════════════════
    if report_type == "trade" and check_companion:
        issues.extend(check_companion_deliverables(report_path))
        # v17 GATE0·分面深写前置：6 份单面深稿必须存在/达深度/含摘要卡（阶段A 机器执法）
        issues.extend(check_stage_a_face_drafts(report_path))
        # v25：不再有"合稿内聚 vs 不合稿导航"二选一分支——faces-split 是唯一口径。
        #   若标记原本缺失，_faces_split 已在上方被强制置 True（并已记一条硬 FAIL），
        #   这里统一只跑 check_faces_split_navigation，不再触碰已删除的 check_merge_cohesion。
        issues.extend(check_faces_split_navigation(report_path, text))
        issues.extend(check_scenario_data_sources(report_path, text))
        # P0-1：决策固化产物 GATE0（quant_scorer all → {stem}_decision.json）
        issues.extend(check_decision_artifact(report_path, text))
        # P0-1c（v1.23）：三情景 EPS/净利 必须锚定 forecast.json，禁止手工锚定脱钩
        #   （校验对象是决策稿 §5/三情景【目标价】，始终留在决策稿，用 text）
        issues.extend(check_forecast_scenario_consistency(report_path, text))
        # P0-1d（v1.24 ②）：总体结论目标价必须可由 forecast 三档目标价推导（自上而下可推导性）
        issues.extend(check_conclusion_derivability(report_path, text))
        # P0-1e/f/g（§4.1.4 深度盈利预测三件套）：faces-split 下 §4.1.4 已移入【基本面深稿】，
        #   故对深稿正文取证；旧架构仍对决策稿 text 取证。
        _fc_text = _read_fundamental_draft_text(report_path) if _faces_split else None
        _fund_src = _fc_text if (_faces_split and _fc_text) else text
        # P0-1e（v1.24 ③）：§4.1.4 必须显式披露关键假设（数值+依据+信源），便于用户核查
        issues.extend(check_assumptions_disclosure(report_path, _fund_src))
        # P0-1f（v1.25 ④）：§4.1.4 必须展示「假设 → EPS」逐行计算推导链，禁止只摆终值
        issues.extend(check_forecast_derivation_chain(report_path, _fund_src))
        # P0-1g（v1.25 ⑤）：§4.1.4 营收必须用分部加总法（自下而上），禁止只拍整体增速
        issues.extend(check_revenue_buildup(report_path, _fund_src))

    # ═══════════════════════════════════════════════════════════════════
    # 汇总每章节的"深度快照"（用于 gate 产物 + Agent 定位缺口）
    # ═══════════════════════════════════════════════════════════════════
    section_snapshot = {}
    for sec_id, prefix in [
        ("4.1.1", "##### 4.1.1"), ("4.1.2", "##### 4.1.2"), ("4.1.3", "##### 4.1.3"),
        ("4.1.4", "##### 4.1.4"), ("4.1.5", "##### 4.1.5"),
        ("4.2", "#### 4.2"), ("4.3", "#### 4.3"), ("4.4", "#### 4.4"),
        ("4.5", "#### 4.5"), ("4.6", "#### 4.6"), ("4.7", "#### 4.7"),
        ("五、", "### 五、"),
    ]:
        sec_text = get_section_text(lines, headings, prefix)
        if sec_text:
            section_snapshot[sec_id] = {
                "effective_length": effective_length(sec_text),
                "table_count": count_tables(sec_text.splitlines()),
                "has_data": "📊" in sec_text,
                "has_analysis": "🔍" in sec_text,
                "has_conclusion": "📌" in sec_text,
                # v29：faces-split 决策稿 §4.1.x/4.2~4.6 仅承载"六面结论速览"，深度已移交
                #   6 份分面深稿，故这些小节不再暴露字数目标（置 None）——避免快照表以"门槛/还差X字"
                #   诱导撰写者复制粘贴凑字数（历史"文字堆砌"顽疾的制度性诱因）。收束节(§4.7/§五)不受影响。
                "min_required_length": (
                    None
                    if (_faces_split and sec_id in (
                        "4.1.1", "4.1.2", "4.1.3", "4.1.4", "4.1.5",
                        "4.2", "4.3", "4.4", "4.5", "4.6"))
                    else SECTION_MIN_EFFECTIVE_LEN.get(sec_id)
                ),
                "min_required_tables": SECTION_MIN_TABLES.get(sec_id),
            }
        else:
            section_snapshot[sec_id] = None

    # ═══════════════════════════════════════════════════════════════════
    # v12 准确性提升六机制·软门禁集成（仅追加 warnings，不阻断 pass）
    # ─────────────────────────────────────────────────────────────────────
    # 建议 1 追问链 / 2 拆分树 / 3 预注册 / 4 数字一致性 /
    # 建议 5 对手方论证 / 6 置信度校准
    # 每项独立 validator 返回 List[str]，统一前缀 [v12·xxx]。
    # 任一 validator 失败不影响主流程；仅记录一条 [v12·loader] 警告。
    # 仅在"完整报告模式"（trade 或 fundamental）下启用；轻量路径在 main 中分流，不会进入本函数。
    # ═══════════════════════════════════════════════════════════════════
    try:
        import sys as _sys
        _script_dir = Path(__file__).resolve().parent
        if str(_script_dir) not in _sys.path:
            _sys.path.insert(0, str(_script_dir))
        v12_validators = []
        try:
            from why_chain_validator import validate_why_chain
            v12_validators.append(("why_chain", validate_why_chain))
        except Exception as _e:
            warnings.append(f"[v12·loader] why_chain_validator 加载失败: {_e}")
        try:
            from decomposition_tree_validator import validate_decomposition_tree
            v12_validators.append(("decomposition_tree", validate_decomposition_tree))
        except Exception as _e:
            warnings.append(f"[v12·loader] decomposition_tree_validator 加载失败: {_e}")
        try:
            from prereg_validator import validate_prereg
            v12_validators.append(("prereg", validate_prereg))
        except Exception as _e:
            warnings.append(f"[v12·loader] prereg_validator 加载失败: {_e}")
        try:
            from numeric_consistency_auditor import audit_numeric_consistency
            v12_validators.append(("numeric_consistency", audit_numeric_consistency))
        except Exception as _e:
            warnings.append(f"[v12·loader] numeric_consistency_auditor 加载失败: {_e}")
        try:
            from devils_advocate_validator import validate_devils_advocate
            v12_validators.append(("devils_advocate", validate_devils_advocate))
        except Exception as _e:
            warnings.append(f"[v12·loader] devils_advocate_validator 加载失败: {_e}")
        try:
            from calibration_review import validate_calibration
            v12_validators.append(("calibration", validate_calibration))
        except Exception as _e:
            warnings.append(f"[v12·loader] calibration_review 加载失败: {_e}")

        # ── v13 准确性深化六机制（P0-P2）──────────────────────────
        # P0-6 信源引用交叉 / P1-1 资本配置 / P1-5 逆向估值 /
        # P1-3 行为偏差 / P2-2 竞争演化 / P2-8 数据时效
        try:
            from source_citation_auditor import audit_source_citation
            v12_validators.append(("source_citation", audit_source_citation))
        except Exception as _e:
            warnings.append(f"[v13·loader] source_citation_auditor 加载失败: {_e}")
        try:
            from capital_allocation_validator import validate_capital_allocation
            v12_validators.append(("capital_allocation", validate_capital_allocation))
        except Exception as _e:
            warnings.append(f"[v13·loader] capital_allocation_validator 加载失败: {_e}")
        try:
            from reverse_valuation_validator import validate_reverse_valuation
            v12_validators.append(("reverse_valuation", validate_reverse_valuation))
        except Exception as _e:
            warnings.append(f"[v13·loader] reverse_valuation_validator 加载失败: {_e}")
        try:
            from behavioral_bias_detector import detect_behavioral_bias
            v12_validators.append(("behavioral_bias", detect_behavioral_bias))
        except Exception as _e:
            warnings.append(f"[v13·loader] behavioral_bias_detector 加载失败: {_e}")
        try:
            from competition_evolution_validator import validate_competition_evolution
            v12_validators.append(("competition_evolution", validate_competition_evolution))
        except Exception as _e:
            warnings.append(f"[v13·loader] competition_evolution_validator 加载失败: {_e}")
        try:
            from data_freshness_auditor import audit_data_freshness
            v12_validators.append(("data_freshness", audit_data_freshness))
        except Exception as _e:
            warnings.append(f"[v13·loader] data_freshness_auditor 加载失败: {_e}")
        try:
            from forecast_quality_validator import validate_forecast_quality
            v12_validators.append(("forecast_quality", validate_forecast_quality))
        except Exception as _e:
            warnings.append(f"[v14·loader] forecast_quality_validator 加载失败: {_e}")
        try:
            from resonance_divergence_validator import validate_resonance_divergence
            v12_validators.append(("resonance_divergence", validate_resonance_divergence))
        except Exception as _e:
            warnings.append(f"[v14·loader] resonance_divergence_validator 加载失败: {_e}")
        try:
            from cross_face_reconciliation_validator import validate_cross_face_reconciliation
            v12_validators.append(("cross_face_reconciliation", validate_cross_face_reconciliation))
        except Exception as _e:
            warnings.append(f"[v15·loader] cross_face_reconciliation_validator 加载失败: {_e}")

        # P1-2：短线/超短线报告跳过中长线专属的重型软门禁（逆向估值/竞争演化/资本配置），
        #        避免对短周期决策过度工程；波段及以上周期仍全量执行。
        #        注意：cross_face_reconciliation（分面深写交叉勾稽）属【全周期统一流水线】门禁，
        #        全部交易风格都走阶段A/B/C，故【不】纳入重型跳过集合（v15 · 全周期统一）。
        _HEAVY_VALIDATORS = {"reverse_valuation", "competition_evolution", "capital_allocation"}
        if style in {"ultra_short", "short"}:
            _skipped = [n for n, _ in v12_validators if n in _HEAVY_VALIDATORS]
            v12_validators = [(n, f) for (n, f) in v12_validators if n not in _HEAVY_VALIDATORS]
            if _skipped:
                warnings.append(
                    f"[v12·分级] {STYLE_LABELS.get(style, style)}报告已跳过中长线专属重型软门禁："
                    f"{'、'.join(_skipped)}（逆向估值/竞争演化/资本配置仅对波段及以上周期强制）。"
                )

        for _name, _fn in v12_validators:
            try:
                _ws = _fn(report_path)
                if _ws:
                    warnings.extend(_ws)
            except Exception as _e:
                warnings.append(f"[v12·{_name}] 校验器抛异常（已跳过，不阻断）: {_e}")
    except Exception as _e:
        warnings.append(f"[v12·loader] 软门禁集成总入口失败: {_e}")

    # P1-A：方法论指纹校验——把"只提名词凑过门禁"变成可被识别的不合格。
    #   护城河多维 / ROIC×WACC 价值判断深度缺失 → FAIL；Beneish/PEG/DCF → WARN。
    try:
        _fp_fails, _fp_warns = check_methodology_fingerprints(text, style, report_type)
        issues.extend(_fp_fails)
        warnings.extend(_fp_warns)
    except Exception as _e:
        warnings.append(f"[方法论指纹·loader] 指纹校验异常（已跳过，不阻断）: {_e}")

    # v29 反复制粘贴填充闸：整段近似重复 → 硬 FAIL。
    #   针对"同一段万能分析被套进多个主题不同小节"的历史顽疾，从结果侧拦截堆砌行为。
    try:
        issues.extend(check_duplicate_paragraphs(text))
    except Exception as _e:
        warnings.append(f"[反重复·loader] 段落查重异常（已跳过，不阻断）: {_e}")

    # P1-1 / v16 / v23：把 [字数建议] 前缀的"纯字数偏短"项分两档处理——
    #   ① 基本面五段式 §4.1.1~4.1.5【硬 FAIL】（基本面是"买不买"的根，深度 ≥ 独立基本面报告）；
    #   ② v23：§4.2~4.6（政策/技术/资金/筹码/消息）也升级为【硬 FAIL】——这五个面章节是该面
    #      深度的"唯一最终承载"，深度须 = Intent-2 单面报告。旧版把它们降级为软建议是
    #      "合稿章节可薄片化为深稿 1/7"的制度漏洞根源（用户投诉"合稿没把深稿纳入、只写几句详见"）。
    #      字数地板已重锚定到"同面深稿核心正文下限 × 0.90"（见 SECTION_MIN_EFFECTIVE_LEN），故不再放水。
    #   ③ 仅 §4.7/五、综合研判/行业信源等收束/附属节字数不足仍为软建议（避免无价值凑字数）。
    _FUND_DEPTH_SECTIONS = ("4.1.1", "4.1.2", "4.1.3", "4.1.4", "4.1.5",
                            "4.2", "4.3", "4.4", "4.5", "4.6")
    _len_advice = [i for i in issues if str(i).startswith("[字数建议]")]
    if _len_advice:
        def _is_fundamental_len(item: str) -> bool:
            # v24 faces-split：§4.1-4.6 深度已移交独立深稿（由 check_stage_a_face_drafts 把关其深度），
            #   决策稿 §四仅承载「六面结论速览」，不应再把 §4.x 字数建议升级为阻断 PASS 的 BLOCK。
            if _faces_split:
                return False
            return any(f"章节 {s} " in str(item) for s in _FUND_DEPTH_SECTIONS)
        # 面章节字数项：去掉 [字数建议] 软前缀、改写为硬 FAIL 文案，留在 issues 阻断 PASS
        _fund_hard = []
        for it in _len_advice:
            if _is_fundamental_len(it):
                _fund_hard.append(
                    str(it).replace(
                        "[字数建议]", "[面深度·BLOCK]"
                    ).replace(
                        "字数不足不阻断 PASS（要素/表格/三段式齐全即视为达标），但建议补充数据表/推导段以提升深度。",
                        "v23：该面章节是其深度的唯一最终承载，须 = Intent-2 单面报告深度（字数地板 = "
                        "同面深稿核心正文下限 × 0.90）。字数不达标【阻断 PASS】，必须把对应深稿核心正文逐字内聚进来"
                        "（只增不减），不得概括/引用/薄片化。",
                    )
                )
        _soft = [i for i in _len_advice if not _is_fundamental_len(i)]
        if _faces_split:
            # v29：faces-split 决策稿 §4.x 只做结论速览、无字数目标——相关字数建议直接丢弃，
            #   既不升级为 BLOCK 也不作为 warning 打印，从源头消除"还差 X 字"对撰写者的凑字数诱导。
            _soft = [
                i for i in _soft
                if not any(f"章节 {s} " in str(i) for s in _FUND_DEPTH_SECTIONS)
            ]
        issues = [i for i in issues if not str(i).startswith("[字数建议]")] + _fund_hard
        warnings.extend(_soft)

    return {
        "report": str(report_path),
        "style": style,
        "style_label": STYLE_LABELS[style],
        "effective_length": total_eff_len,
        "min_required_length": min_required,
        "table_count": table_count,
        "md_image_refs": len(re.findall(r"!\[[^\]]*\]\([^)]+\)", text)),
        "html_exists": report_path.with_suffix(".html").exists(),
        "marker_counts": {
            "data": data_count,
            "analysis": analysis_count,
            "conclusion": conclusion_count,
        },
        "section_snapshot": section_snapshot,
        "pass": len(issues) == 0,
        "issues": issues,
        "warnings": warnings,
    }


# ═════════════════════════════════════════════════════════════════════
# v7 新增：Gate 产物 —— 生成人类可读的 _gate_result.md
# Agent 在对话中必须贴出此 md 的首屏作为"已跑门禁"的承诺证据
# ═════════════════════════════════════════════════════════════════════
def render_gate_markdown(result: Dict) -> str:
    status = "✅ PASS" if result["pass"] else "❌ FAIL"
    lines_md: List[str] = []
    lines_md.append(f"# 报告质量门禁结果 — {status}")
    lines_md.append("")
    lines_md.append(f"- **报告路径**: `{result['report']}`")
    lines_md.append(f"- **交易风格**: {result['style_label']} (`{result['style']}`)")
    lines_md.append(
        f"- **效率字数**: {result['effective_length']}"
        f"{' / 门槛 ' + str(result['min_required_length']) if result.get('min_required_length') else ''}"
    )
    lines_md.append(f"- **表格数**: {result['table_count']}")
    lines_md.append(f"- **.md 图片引用**: {result.get('md_image_refs', 0)} 条（要求为 0：.md 纯文字，图表只存在于 HTML 内）")
    lines_md.append(f"- **HTML 文件**: {'✅ 已生成' if result.get('html_exists') else '❌ 未生成'}")
    mc = result["marker_counts"]
    lines_md.append(f"- **三段式标记**: 📊 {mc['data']}  🔍 {mc['analysis']}  📌 {mc['conclusion']}")
    lines_md.append("")

    # 章节深度快照表
    lines_md.append("## 章节深度快照")
    lines_md.append("")
    lines_md.append("| 章节 | 效率字数 | 门槛 | 表格数 | 门槛 | 📊 | 🔍 | 📌 |")
    lines_md.append("|------|---------|------|--------|------|----|----|----|")
    for sec_id, snap in result.get("section_snapshot", {}).items():
        if snap is None:
            lines_md.append(f"| {sec_id} | ❌ 缺失 | — | — | — | — | — | — |")
            continue
        eff = snap["effective_length"]
        eff_req = snap["min_required_length"] or "—"
        eff_flag = "❌" if snap["min_required_length"] and eff < snap["min_required_length"] else "✅"
        tbl = snap["table_count"]
        tbl_req = snap["min_required_tables"] or "—"
        tbl_flag = "❌" if snap["min_required_tables"] and tbl < snap["min_required_tables"] else "✅"
        lines_md.append(
            f"| {sec_id} | {eff_flag} {eff} | {eff_req} | {tbl_flag} {tbl} | {tbl_req} | "
            f"{'✅' if snap['has_data'] else '❌'} | "
            f"{'✅' if snap['has_analysis'] else '❌'} | "
            f"{'✅' if snap['has_conclusion'] else '❌'} |"
        )
    lines_md.append("")

    # Issues
    if result["issues"]:
        lines_md.append(f"## ❌ 待补项（{len(result['issues'])} 条，全部修复后方可交付）")
        lines_md.append("")
        for i, issue in enumerate(result["issues"], 1):
            lines_md.append(f"{i}. {issue}")
        lines_md.append("")

    # Warnings
    if result["warnings"]:
        # 拆分软门禁 WARN（v12 六机制 + v13 六机制）与普通 warnings
        # 凡含「·WARN]」或以 [v12·/[v13· 前缀，统一归入软门禁列表。
        v12_warns = [
            w for w in result["warnings"]
            if str(w).startswith("[v12·") or str(w).startswith("[v13·") or "·WARN]" in str(w)
            or str(w).startswith((
                "[追问链·", "[拆分树·", "[预注册·", "[数字一致性·", "[对手方论证·", "[校准·",
                "[信源引用·", "[资本配置·", "[逆向估值·", "[行为偏差·", "[竞争演化·", "[数据时效·",
                "[方法论指纹·",
            ))
        ]
        other_warns = [w for w in result["warnings"] if w not in v12_warns]

        if other_warns:
            lines_md.append(f"## ⚠️ 警告（{len(other_warns)} 条，建议关注但不阻塞）")
            lines_md.append("")
            for w in other_warns:
                lines_md.append(f"- {w}")
            lines_md.append("")

        if v12_warns:
            n_soft = len(v12_warns)
            lines_md.append(f"## 🧭 软门禁 WARN 列表（准确性提升机制·v12+v13·{n_soft} 条·不阻塞 PASS）")
            lines_md.append("")
            # P0-7 累积升级提示：WARN 越多越醒目，不升级为 FAIL，但强制让信息显眼
            if n_soft >= 10:
                lines_md.append(f"> 🔴 **严重提示：本次软门禁 WARN 高达 {n_soft} 条（≥10）**。")
                lines_md.append("> 这强烈意味着分析在「追问深度 / 数字自洽 / 对手方论证 / 信源时效」等多个维度存在系统性欠缺。")
                lines_md.append("> **强烈建议在交付前逐条核实并尽量消除**——软门禁不阻断交付，但持续高 WARN = 结论可靠性存疑。")
                lines_md.append("")
            elif n_soft >= 5:
                lines_md.append(f"> 🟠 **注意：本次软门禁 WARN 共 {n_soft} 条（≥5）**，建议交付前重点排查下列高频问题。")
                lines_md.append("")
            lines_md.append("> 覆盖 v12（追问链 / 最小不可拆单元 / 预注册 / 数字一致性 / 对手方论证 / 置信度校准）")
            lines_md.append("> 与 v13（信源引用交叉 / 管理层资本配置 / 逆向估值 / 行为偏差 / 竞争演化 / 数据时效）。")
            lines_md.append("> 每一条都指向『结论是否经得起追问』。Agent 应在交付前自查并尽量消除。")
            lines_md.append("")
            for w in v12_warns:
                lines_md.append(f"- {w}")
            lines_md.append("")

    if result["pass"]:
        lines_md.append("---")
        lines_md.append("")
        lines_md.append("✅ 所有硬门禁通过，可以交付。Agent 请在对话中输出**轻量交付三件套**：")
        lines_md.append("")
        lines_md.append("1. **快速结论盒子**（汇总决策 ≤500 字 · 四问齐全 / 专项及兜底 ≤800 字 · 正面回答核心问题）")
        lines_md.append("2. **Gate 摘要一行**（如\"✅ Gate PASS · 字数 xxxx · 表格 xx 张\"）")
        lines_md.append("3. **双文件路径**（.md + .html），提示\"完整报告请打开 .md 或 .html 查看，图表在 HTML 内\"")
        lines_md.append("")
        lines_md.append("⚠️ **不要再重复贴完整报告正文**——文件已落盘，对话重复输出浪费时间和 tokens（旧规则已废弃）。")
    else:
        lines_md.append("---")
        lines_md.append("")
        lines_md.append("❌ **BLOCK — 不得交付**。按上述待补项逐条补写 Markdown 报告后，重新运行本 checker。")
        lines_md.append("")
        lines_md.append("**唯一合法动作是补写，不允许删减章节/降低深度。**")

    return "\n".join(lines_md) + "\n"


# --------------------------------------------------------------------------- #
# 三表预测专项校验（--check-3statement）
# --------------------------------------------------------------------------- #
# 与 references/faces/基本面.md §⭐ 报告输出规范 · §4.1.4B 节、§基本面原则 5 对齐：
#   - 勾稽 6 条硬约束：规则 1/3/5/6 必须 PASS；规则 2/4 FAIL 项必须说明原因
#   - 5 项预测假设自洽性校验必须全部解释
#   - 报告必须包含完整三表（IS/BS/CF）+ 比率矩阵 + 杜邦分解
_THREE_STMT_KEYWORDS = {
    "4.1.4B 章节": [r"4\.1\.4B", r"2\.4B", r"三表预测", r"卖方级"],
    "利润表预测": [r"利润表", r"营业收入.*预测|营收.*预测"],
    "资产负债表预测": [r"资产负债表", r"总资产|股东权益"],
    "现金流量表预测": [r"现金流量表|经营活动现金流|OCF"],
    "比率矩阵": [r"比率矩阵|财务比率|ROE|ROIC"],
    "杜邦分解": [r"杜邦|DuPont"],
    "勾稽规则 1 股东权益": [r"股东权益.*勾稽|权益.*期末.*期初|规则\s*1"],
    "勾稽规则 3 固定资产滚动": [r"固定资产.*滚动|CapEx.*折摊|规则\s*3"],
    "勾稽规则 5 资产恒等式": [r"资产\s*=\s*负债|会计恒等式|规则\s*5"],
    "勾稽规则 6 所得税": [r"实际税率|所得税.*勾稽|规则\s*6"],
    "5 项自洽性校验": [r"自洽|自洽性|预测假设.*校验|数据.*自洽|比率.*异常"],
    # --- 基本面卖方研报对标新增章节（P0 + P1 + P2）---
    # v1.3（2026-05）：放宽匹配，识别报告正文里**真实存在的研报章节标题**，不再要求附录里
    # 写"详见 §X"这类元说明锚点（避免污染最终交付报告）
    "基本面·公司概况 2.0（P0-A）": [
        r"2\.0\s*公司概况",
        r"发展历程时间轴|发展历程：|发展历程\s*\(",
        r"股权结构.*实际控制人|实际控制人.*股权结构|股权结构.*稳定|股权结构.*集中",
    ],
    "基本面·业务条线矩阵 2.3-B1（P0-B）": [
        r"2\.3-?B1",
        r"业务条线.*营收.*毛利率|分业务条线.*营收|业务条线拆解",
        r"业务条线汇总|分业务.*营收结构|核心业务线",
    ],
    "基本面·条线独立驱动 2.4B.1-A（P0-C）": [
        r"2\.4B\.1-?A",
        r"条线独立驱动|分业务条线.*驱动假设",
        r"核心驱动|驱动力分析|驱动因素|增长驱动",
            r"条线.*推导|营收.*驱动|增长驱动|分业务.*预测.*矩阵",
],
    "基本面·可比横比 2.5-B1（P1-D）": [
        r"2\.5-?B1",
        r"可比公司横向对标|可比公司横比|A\s*股.*港股.*美股",
        r"国际对标|对标公司|对标.*核心竞争对手",
            r"可比公司.*估值对比|可比.*横比|估值与定价.*可比",
],
    "基本面·国际对标 2.6（P1-E，可选）": [
        r"2\.6\s*国际对标|国际对标.*全球龙头",
        r"国际对标.*核心竞争对手|国际对标.*全维度|国际对标[:：]",
    ],
    "基本面·研发管线 2.3-B2（P1-F，科技/医药强制）": [
        r"2\.3-?B2",
        r"在研管线|研发.*管线|研发人员占比",
        r"在研项目|研发投入.*强度|研发投入.*结构|研发支出资本化|研发储备",
    ],
    "基本面·一页压缩三表 2.4B.9（P1-G）": [
        r"2\.4B\.9",
        r"一页压缩.*三表|压缩版三表",
    ],
    "基本面·出海/关税敏感性 2.7（P2-I，海外>=15%强制）": [
        r"2\.7\s*出海|出海.*关税|关税.*敏感性",
    ],
    "基本面· DCF 参数披露 2.5-B2（P2-J）": [
        r"2\.5-?B2",
        r"DCF.*参数.*披露|WACC.*永续增长|DCF.*全参数",
    ],
    # --- 对标中信建投《科大讯飞》底稿新增的 3 项 基本面硬要素（v1.1，2026-05）---
    "基本面·高频外部信源 2.3-B3（P0 强制）": [
        r"2\.3-?B3",
        r"高频外部信源|外部信源跟踪矩阵|信息优势.*外部变量",
    ],
    "基本面·分业务 E+1~E+3 预测矩阵 2.4B.0（P0-D 强制）": [
        r"2\.4B\.0",
        r"分业务.*E\+1.*E\+3|分业务条线.*完整预测|分业务.*营收.*毛利率.*预测矩阵",
        r"卖方一致预期.*对比|与卖方一致预期差异",
    ],
    "基本面·敏感性矩阵 2.4B.10（P0-E 强制）": [
        r"2\.4B\.10",
        r"收入.*盈利敏感性|双变量.*5\s*[×x]\s*5|单变量弹性系数|弹性系数.*敏感性排名",
            r"敏感性.*矩阵|单变量弹性|双变量.*5.*5",
],
    # --- 对标深度报告/ 目录下 54 篇真实卖方研报范式新增的 4 项 v1.4 卖方对标三件套（2026-05）---
    # 落地依据：references/faces/基本面.md §1.F.8 + §⭐ 报告输出规范 · §基本面一/§2.0.5/§2.5-B0
    # 注意：每条规则的所有 pattern 都使用 ALL（即必须全部命中）；正则收紧到只匹配标题/关键短语，避免被普通"估值/预测"字眼蒙混。
    "基本面·首页投资概览卡（v1.4，P0 强制）": [
        # v1.5 更新：不再依赖内部章节号，改为匹配内容实质
        r"投资概览卡|核心财务摘要表|投资摘要卡",  # 必须出现明确的章节名
        r"营收\s*YoY|营业收入.*同比增速|归母.*YoY",  # 必须出现 YoY 行
    ],
    "基本面·研发产出与核心 IP（v1.5 更新）": [
        # v1.5 更新：删除 r"2\.0\.5" 旧内部代号依赖，改为匹配内容实质
        r"研发产出|核心.*IP|研发.*项目储备|技术平台清单",  # 必须有 IP/项目储备表
        r"累计.*(授权|有效).*专利|PCT\s*国际专利|集成电路.*布图设计",  # 必须有专利相关硬指标
    ],
    "基本面· PE/PB-Band 历史估值带（v1.5 更新）": [
        # v1.5 更新：删除 r"2\.5-?B0" 旧内部代号依赖，改为匹配内容实质
        r"PE-?Band|PB-?Band|PE.*分位|PB.*分位|历史.*估值带",  # 必须明确出现 PE-Band / PB-Band / 历史分位
        r"历史.*分位|历史百分位|分位线.*中位",  # 必须含分位描述
    ],
    "基本面·可比公司预测来源标注（v1.5 更新）": [
        # v1.5 更新：保留内容检测，删除旧内部代号依赖
        r"预测来源",  # 必须有"预测来源"列名
        r"本(报告|所).*独立预测|自家覆盖.*预测",  # 必须标注"本报告独立预测"
        r"Wind.*一致预期|iFinD.*一致预期|携宁.*一致预期|Bloomberg.*consensus",  # 必须标注第三方一致预期
    ],
    # --- v1.7 新增（P2-14）：评级说明 + 风险等级标准化模块 ---
    # v1.7.3（2026-05）：删除"分析师承诺"伪资质表述，改为「AI 生成声明 + 非投资建议」诚实声明
    "基本面·评级说明与风险等级（v1.7，P2-14）": [
        r"投资评级体系|评级说明|买入.*增持.*中性.*减持",  # 必须明示评级体系
        r"风险等级|R[1-5]\s*[低中高]?风险|波动率.*配置",  # 必须明示风险等级
        r"AI\s*生成声明|AI\s*自动生成|大语言模型|非持牌分析师",  # 必须明示 AI 生成属性
        r"免责声明|非投资建议|不构成.*投资建议|盈亏自负",  # 必须含免责声明
    ],
    # --- v1.7 新增（P1-8）：可比公司估值 PEG + 业务对标列 ---
    "基本面·可比公司 PEG/业务对标列（v1.7，P1-8）": [
        r"PEG[\s\|（(]",  # 必须有 PEG 列
        r"主营业务对标|业务对标|业务相似度",  # 必须有业务对标列
    ],

}




def check_three_statement_coverage(report_path: Path) -> Dict:
    """
    校验 Markdown 报告中三表预测章节（4.1.4B）的关键要素覆盖度。

    返回结构：
      {
        "pass": bool,
        "coverage": {要素: bool, ...},
        "missing": [要素列表],
        "notes": [提示]
      }
    """
    if not report_path.exists():
        return {
            "pass": False,
            "coverage": {},
            "missing": [],
            "notes": [f"报告文件不存在: {report_path}"],
        }

    text = report_path.read_text(encoding="utf-8", errors="replace")
    coverage: Dict[str, bool] = {}
    # v1.4 卖方对标三件套：4 项需要"全部 pattern 命中"才算通过，避免被普通字眼蒙混；其余 v1.0-v1.1 规则沿用"任一命中即通过"
    _STRICT_ALL_KEYS = {
        "基本面·首页投资概览卡（v1.4，P0 强制）",
        "基本面·研发产出与核心 IP（v1.5 更新）",
        "基本面· PE/PB-Band 历史估值带（v1.5 更新）",
        "基本面·可比公司预测来源标注（v1.5 更新）",
        "基本面·评级说明与风险等级（v1.7，P2-14）",
        "基本面·可比公司 PEG/业务对标列（v1.7，P1-8）",
    }
    for key, patterns in _THREE_STMT_KEYWORDS.items():
        if key in _STRICT_ALL_KEYS:
            coverage[key] = all(re.search(p, text, re.IGNORECASE) for p in patterns)
        else:
            coverage[key] = any(re.search(p, text, re.IGNORECASE) for p in patterns)

    # 可选 / 条件强制章节：缺失不计入 missing（不阻断门禁），仅降级为 notes 告警
    # 说明：
    #   - 国际对标 2.6、出海/关税 2.7 对纯内需业务可跳过
    #   - 研发管线 2.3-B2 对非科技/医药行业可简化
    _OPTIONAL_KEYS = {
        "基本面·国际对标 2.6（P1-E，可选）",
        "基本面·研发管线 2.3-B2（P1-F，科技/医药强制）",
        "基本面·出海/关税敏感性 2.7（P2-I，海外>=15%强制）",
        # v1.4 新增：研发产出 2.0.5 仅对科技/制造/创新药强制；纯消费/金融/资源公司可降级为软告警
        "基本面·研发产出与核心 IP（v1.5 更新）",
    }

    missing = [k for k, ok in coverage.items() if not ok and k not in _OPTIONAL_KEYS]
    notes: List[str] = []
    # 软告警：可选章节缺失不阻断门禁
    soft_notes: List[str] = []
    for k in _OPTIONAL_KEYS:
        if not coverage.get(k, False):
            soft_notes.append(f"可选章节未覆盖：{k}（若适用业务特征请补充）")

    # 硬约束：规则 1/3/5/6 必须存在；规则 2/4 如被明示 FAIL 则必须附带"原因/说明/解释"
    if re.search(r"规则\s*2.*FAIL|经营现金流.*勾稽.*FAIL", text):
        if not re.search(r"规则\s*2[\s\S]{0,400}?(原因|说明|解释|due to|because)", text):
            notes.append("规则 2（经营现金流间接法）FAIL 但未见原因说明")
    if re.search(r"规则\s*4.*FAIL|货币资金.*FAIL", text):
        if not re.search(r"规则\s*4[\s\S]{0,400}?(原因|说明|解释|due to|because)", text):
            notes.append("规则 4（货币资金变动软校验）FAIL 但未见原因说明")

    # ── v1.6 新增：三表预测行数硬规则（P1-7）─────────────────────
    # 基本面研究报告的 §2.4B.2 / 2.4B.3 / 2.4B.4 必须完整渲染：
    #   - 利润表预测 ≥ 20 行
    #   - 资产负债表预测 ≥ 18 行
    #   - 现金流量表预测 ≥ 12 行
    # 不得以"已合并/见 xlsx"借口删减。该规则避免 LLM 偷懒输出只有 5-6 行的简化表。
    def _count_first_table_rows_after(anchor_re: str) -> Optional[int]:
        """从匹配 anchor 的章节标题开始向下扫描，找到第一个 Markdown 表格的数据行数。
        数据行 = 以 `|` 开头 + 非分隔行（不含 `---`）。"""
        m = re.search(anchor_re, text)
        if not m:
            return None
        rest = text[m.end():]
        # 截断到下一个 #### 或 ##### 章节
        end_m = re.search(r"\n#{3,5}\s", rest)
        if end_m:
            rest = rest[: end_m.start()]
        in_table = False
        row_count = 0
        for line in rest.splitlines():
            stripped = line.strip()
            is_table_line = stripped.startswith("|") and stripped.endswith("|")
            if is_table_line:
                # 跳过分隔行 | --- | --- |
                if re.match(r"^\|[\s:\-]+\|", stripped):
                    in_table = True
                    continue
                # 跳过表头（首行非分隔但下一行才是分隔）— 通过 in_table 累计逻辑解决
                if in_table:
                    row_count += 1
                else:
                    # 表头行（在分隔行之前出现）
                    pass
            elif in_table:
                # 表格结束（首个表格）
                break
        return row_count

    three_stmt_row_specs = [
        (r"####\s*2\.4B\.2\s+利润表预测", 20, "利润表预测（§2.4B.2）"),
        (r"####\s*2\.4B\.3\s+资产负债表预测", 18, "资产负债表预测（§2.4B.3）"),
        (r"####\s*2\.4B\.4\s+现金流量表预测", 12, "现金流量表预测（§2.4B.4）"),
    ]
    for anchor, min_rows, label in three_stmt_row_specs:
        actual_rows = _count_first_table_rows_after(anchor)
        if actual_rows is None:
            # 章节缺失由其他校验路径覆盖，此处不重复报错
            continue
        if actual_rows < min_rows:
            notes.append(
                f"[INSUFFICIENT_3STATEMENT_ROWS] {label} 仅 {actual_rows} 行 < 最低 {min_rows} 行；"
                f"R5-A 强制要求完整渲染全部模板行，空值用 `—` 占位但行必须保留"
            )
    # ───────────────────────────────────────────────────────



    # --- v1.1 新增：基本面报告交付物"三位一体"完整性校验（Markdown + HTML + xlsx 底稿）---
    # 中信建投《科大讯飞》底稿启发：基本面深度报告必须附带 xlsx 配套底稿。
    # 校验规则：若报告为基本面报告（即含 2.4B 章节），则同目录下应存在基本面底稿 xlsx。
    # 约定文件名形如 `基本面_{code}_{简称}_底稿.xlsx` 或同名 `_底稿.xlsx` / `_worksheet.xlsx`。
    is_b_class_report = bool(re.search(r"2\.4B|4\.1\.4B|三表预测|基本面.*深度", text)) or report_type == "fundamental"
    if is_b_class_report:
        report_stem = report_path.stem  # e.g. "基本面_002230_科大讯飞"
        report_dir = report_path.parent
        candidates = [
            report_dir / f"{report_stem}_底稿.xlsx",
            report_dir / f"{report_stem}_worksheet.xlsx",
            report_dir / f"{report_stem.replace('_report', '')}_底稿.xlsx",
        ]
        # 也接受同目录任何形如 *_底稿.xlsx 或 基本面_*.xlsx
        try:
            glob_hits = (list(report_dir.glob("*_底稿.xlsx"))
                         + list(report_dir.glob("基本面_*.xlsx")))
        except Exception:
            glob_hits = []
        xlsx_found = any(p.exists() for p in candidates) or len(glob_hits) > 0
        if not xlsx_found:
            soft_notes.append(
                "基本面深度报告未发现配套 .xlsx 基本面底稿（期望路径：{}_底稿.xlsx）。"
                "请运行 `python scripts/fundamental_worksheet_builder.py --code <xxx> --name <xxx> "
                "--all-sheets --out OutputReport/基本面_<xxx>_<简称>_底稿.xlsx` 生成底稿。"
                .format(report_stem)
            )
        else:
            # v1.2 升级：不仅校验 Sheet 数，还校验非空 cell 密度（防"只造模板不填数据"假阳性）
            try:
                from openpyxl import load_workbook
                xlsx_path = glob_hits[0] if glob_hits else next(p for p in candidates if p.exists())
                wb_check = load_workbook(xlsx_path, read_only=True, data_only=False)
                sheet_names = [s for s in wb_check.sheetnames if s.upper() != "README"]
                sheet_count = len(sheet_names)

                # --- 校验 1：Sheet 数 ---
                if sheet_count < 10:
                    notes.append(
                        f"【底稿校验】{xlsx_path.name} 仅 {sheet_count} 个 Sheet，"
                        f"低于卖方深度研报 10 Sheet 基线"
                    )

                # --- 校验 2：总非空 cell 数（对标中信建投 4650 的 40% = 1800 最低线）---
                total_nonempty = 0
                sheet_nonempty: Dict[str, int] = {}
                for sn in sheet_names:
                    ws = wb_check[sn]
                    cnt = 0
                    for row in ws.iter_rows():
                        for c in row:
                            if c.value is not None and str(c.value).strip():
                                cnt += 1
                    sheet_nonempty[sn] = cnt
                    total_nonempty += cnt

                MIN_TOTAL_NONEMPTY = 1800  # 中信建投 4650 的 ~40%
                if total_nonempty < MIN_TOTAL_NONEMPTY:
                    notes.append(
                        f"【底稿校验】{xlsx_path.name} 总非空 cell {total_nonempty}，"
                        f"低于最低基线 {MIN_TOTAL_NONEMPTY}（对标中信建投 4650）。"
                        f"这意味着底稿大部分是空模板——请使用 `fundamental_worksheet_builder.py v2.0 "
                        f"--data-source FinancialData/{{code}}_fundamental.md` 自动回填真实数据"
                    )

                # --- 校验 3：关键 Sheet 的单表非空基线 ---
                # （名字匹配：允许部分字符命中，容忍 Sheet 名略有出入）
                key_sheet_baselines = {
                    "营收情况": 200,   # 年度 + 季度时序，中信建投 1105
                    "分业务营收": 80,  # 条线 × 年份，中信建投 245
                    "毛利&净利": 60,
                    "偿债": 120,       # 偿债+现金流+CAPEX 合计，中信建投分散
                    "业绩预告": 80,
                    "股东": 50,
                }
                weak_sheets: List[str] = []
                for key_prefix, min_cnt in key_sheet_baselines.items():
                    for sn in sheet_names:
                        if key_prefix in sn:
                            if sheet_nonempty[sn] < min_cnt:
                                weak_sheets.append(
                                    f"{sn}（{sheet_nonempty[sn]} < {min_cnt}）"
                                )
                            break
                if weak_sheets:
                    soft_notes.append(
                        f"【底稿校验】以下关键 Sheet 数据密度偏低，建议补充回填："
                        f"{'; '.join(weak_sheets)}"
                    )

                # --- 校验 4：填充率（总非空 / 总使用 cell）---
                total_cells = 0
                for sn in sheet_names:
                    ws = wb_check[sn]
                    total_cells += ws.max_row * ws.max_column
                if total_cells > 0:
                    fill_rate = total_nonempty / total_cells
                    if fill_rate < 0.20:
                        soft_notes.append(
                            f"【底稿校验】整体填充率 {fill_rate:.1%}（非空 {total_nonempty} / 使用 {total_cells}），"
                            f"偏低；卖方底稿典型填充率 30-50%"
                        )
            except Exception as e:
                soft_notes.append(f"【底稿校验】底稿存在但读取失败：{e}")

    # --- v1.5 新增：Markdown/HTML 净化交付硬规则（black-list 反向检测）---
    # 规则依据：references/delivery_spec.md §第五部分 报告作为成品的清洁度铁律
    # 目标：阻止工程内部代号 / 版本号 / 写作元说明 泄漏到最终交付报告
    LEAKED_PATTERNS = [
        # 1) 带字母后缀的工程内部分节号（只检查出现在章节标题行的情况，避免误报）
        (r"^#+\s.*\b(?:2\.[0-9]+-B[0-9]+|2\.4B\.[0-9]+|4\.1\.4B|2\.3-B[0-9])\b",
         "LEAKED_INTERNAL_SECTION_CODE",
         "章节标题含内部字母后缀分节号（如 2.5-B0 / 2.4B.0 / 4.1.4B）——最终交付报告应使用阿拉伯数字连续编号（如 5.1 / 4.1 / 4.2）"),
        # 2) 工程内部等级代号出现在标题括号中
        (r"^#+\s.*[（(]\s*(?:P[0-2]-[A-Z]|B\s*类强制|B\s*类推荐)[^）)]*[）)]",
         "LEAKED_INTERNAL_RANK_TAG",
         "标题括号中含工程内部等级代号（P0-X / P1-X / P2-X / 基本面强制 / 基本面推荐）——应删除"),
        # 3) 版本号出现在标题或区块引用中
        (r"^[#>][^\n]{0,120}\b(?:v1\.[0-9]+\s*新增|2026-05\s*新增)\b",
         "LEAKED_VERSION_TAG",
         "标题/引用中含版本号标记（v1.X 新增 / 2026-05 新增）——应删除，版本号只留在工程内部"),
        # 4) 写作元说明型区块引用（以"设计意图/对标范式/硬规则验证"开头的引用块）
        (r"^>\s*\*{0,2}(?:设计意图|对标范式|硬规则验证|本节.*对标|严格遵守.*卖方研报)",
         "LEAKED_META_NARRATIVE",
         "存在写作元说明区块引用（'设计意图/对标范式/硬规则验证/本节对标 XX 范式'）——应整段删除"),
        # 5) Checker 锚点元注释
        (r"标准化章节锚点|三位一体\s*B\s*类合规校验标识",
         "LEAKED_CHECKER_ANCHOR",
         "正文含 Checker 元锚点标识（标准化章节锚点 / 三位一体基本面合规校验标识）——应删除"),
    ]
    for pat, code, desc in LEAKED_PATTERNS:
        hits = re.findall(pat, text, re.MULTILINE)
        if hits:
            notes.append(f"[{code}] {desc}；命中 {len(hits)} 处，示例：{str(hits[0])[:80]}")

    # --- v1.5 新增：投资评级量化卡（正向检测）---
    # 规则依据：delivery_spec.md §4.2 投资评级量化卡（汇总决策 + 基本面深度 §一 核心结论 必含）
    if is_b_class_report:
        rating_keywords = [
            r"投资评级\s*[:：]?\s*(?:买入|增持|中性|减持)",  # 卖方四档评级之一
            r"目标价\s*[:：]?\s*\d+(?:\.\d+)?",  # 目标价（精确到元或元.角分）
            r"预期收益率|隐含涨跌幅|时间窗口\s*[:：]?\s*\d+\s*[~-]\s*\d+",  # 预期收益/时间窗口
        ]
        rating_hit = sum(1 for p in rating_keywords if re.search(p, text))
        if rating_hit < 2:  # 至少命中 2/3 才算通过（容错"预期收益率"可选）
            soft_notes.append(
                "[MISSING_QUANT_RATING] 基本面报告缺失投资评级量化卡（应含：评级∈{买入/增持/中性/减持} + 目标价 + 预期收益率/时间窗口）"
            )

    # --- v1.5 新增：大客户/供应链依赖度专章（条件强制）---
    # 规则：若报告提及海外营收 ≥30% 或 Top1 客户 ≥20%，必须有大客户依赖度专章
    if is_b_class_report:
        # 探测是否"海外营收高" 或 "大客户集中"
        overseas_trigger = bool(re.search(r"海外(?:营收|收入).*?([4-9][0-9]|[3-9][0-9]\.?\d*)\s*%|海外占比.*?([3-9][0-9])\s*%", text))
        topn_trigger = bool(re.search(r"Top\s*1.*?([2-9][0-9])\s*%|前\s*1\s*大客户.*?([2-9][0-9])\s*%", text))
        if overseas_trigger or topn_trigger:
            has_client_chapter = bool(re.search(
                r"大客户|主要客户.*结构|客户.*依赖度|Top\s*[135].*客户.*占比|前\s*[135]\s*大客户",
                text
            ))
            has_client_table = bool(re.search(
                r"\|\s*Top\s*1\s*\||\|.*客户名称.*\|.*营收占比.*\|",
                text
            ))
            if not (has_client_chapter and has_client_table):
                soft_notes.append(
                    "[SOFT_MISSING_CLIENT_CHAPTER] 公司海外营收高或大客户集中，但未见「大客户与供应链依赖度专章」（含 Top 1/3/5 客户 + 营收占比 + 合作年限 + 长协 + 切换成本表）"
                )

    # 硬 notes 阻断 PASS；软 notes 仅提示
    result_pass = (len(missing) == 0) and (len(notes) == 0)
    return {
        "pass": result_pass,
        "coverage": coverage,
        "missing": missing,
        "notes": notes,
        "soft_notes": soft_notes,
    }


def render_three_statement_gate(result: Dict) -> str:
    """把 check_three_statement_coverage 结果渲染为人类可读 Markdown。"""
    lines: List[str] = []
    status = "✅ PASS" if result["pass"] else "❌ FAIL"
    lines.append(f"# 三表预测章节门禁：{status}\n")
    lines.append("## 要素覆盖")
    for key, ok in result.get("coverage", {}).items():
        mark = "✅" if ok else "❌"
        lines.append(f"- {mark} {key}")
    if result.get("missing"):
        lines.append("\n## 缺失要素")
        for m in result["missing"]:
            lines.append(f"- {m}")
    if result.get("notes"):
        lines.append("\n## 告警 / 未说明 FAIL")
        for n in result["notes"]:
            lines.append(f"- ⚠ {n}")
    if result.get("soft_notes"):
        lines.append("\n## 软提示（可选章节，未阻断）")
        for n in result["soft_notes"]:
            lines.append(f"- ℹ {n}")
    if not result["pass"]:
        lines.append("\n**修复指引**：参考 `references/faces/基本面.md §⭐ 报告输出规范 · §4.1.4B` 节要求；"
                     "若勾稽规则 2/4 FAIL，必须在报告中明示原因（如大额票据贴现、并购现金并入等）。")
    return "\n".join(lines) + "\n"


# ═════════════════════════════════════════════════════════════════════
# 六个面框架（v2）：政策/资金/筹码/技术/消息 五个面单面报告的轻量门禁
# 设计依据 references/six_dimension_framework.md §七.2：
#   这五个面的单面报告（Intent-2 专项）不强制三件套/三表/五段式，
#   只验三条底线 —— ① 面名+6位代码；② ★评级；③ 数据可核验（来源/脚标/URL ≥3）。
# ═════════════════════════════════════════════════════════════════════
def check_dimension_report(report_path: Path) -> Dict:
    name = report_path.name
    if name.startswith("基本面_"):
        dim = "基本面"
    else:
        dim = next((p.rstrip("_") for p in DIMENSION_REPORT_PREFIXES if name.startswith(p)), "面")
    issues: List[str] = []
    warnings: List[str] = []
    soft: List[str] = []
    if not report_path.exists():
        return {"report": str(report_path), "report_type": "dimension", "dimension": dim,
                "pass": False, "issues": [f"报告文件不存在: {report_path}"],
                "warnings": [], "soft_notes": []}
    text = report_path.read_text(encoding="utf-8", errors="replace")
    # ① 标题含面名 + 6 位代码
    if not re.search(r"^#{1,3}\s+.*" + re.escape(dim), text, re.M):
        warnings.append(f"标题未显式包含面名「{dim}」（建议形如 `# XX（000000）{dim}分析报告`）")
    if not re.search(r"\d{6}", text):
        issues.append("报告未出现 6 位股票代码")
    # ② ★评级结论
    if not (re.search(r"★", text) or contains_any(text, ["看多", "看空", "中性"])):
        issues.append(f"缺少 {dim} ★评级结论（★★★ 看多 / ★★☆ 中性 / ★☆☆ 看空）")
    # ③ 数据可核验：来源 / 脚标 / URL 合计 ≥3 处
    src_hits = (len(re.findall(r"来源", text)) + len(re.findall(r"<sup>\d+</sup>", text))
                + len(re.findall(r"https?://", text)))
    if src_hits < 3:
        issues.append(f"数据可核验性不足：来源/脚标/URL 合计仅 {src_hits} 处（要求 ≥3 处，便于一手追溯）")
    # 软提示：三段式骨架
    if not ("📊" in text and "🔍" in text and "📌" in text):
        soft.append("建议采用 📊数据 → 🔍推导 → 📌结论 三段式骨架（见 delivery_spec.md §4.1 三段式书写规范）")
    # ── v22 不变式根治：Intent-2 单面报告与 Intent-1 分面深稿共用同一契约 + 同一校验函数 ──
    # 这是"单独出某面报告 的深度 ≡ 交易决策报告里该面分面深稿 的深度"在代码层面的硬保证：
    # 二者都跑 _eval_single_face_draft(同一 face_contract)，仅【摘要卡】是 Intent-1 阶段A
    # 专用的交接结构、Intent-2 单面报告不强制（不属于"深度"维度）。
    if dim in STAGE_A_FACES:
        ev = _eval_single_face_draft(text, dim)
        if ev["too_thin"]:
            issues.append(
                f"[单面深度] {ev['too_thin']}（机器执法阈值取自 references/faces/{dim}.md 的 "
                f"face_contract·min_eff_len，与 Intent-1 分面深稿同源）。"
            )
        if ev["defects"]:
            issues.append(
                f"[结构/分析纵深] {('、'.join(ev['defects']))}。须严格套用 references/faces/{dim}.md "
                "§⭐报告输出规范及其 face_contract 的必填表格/小节/脚标/required_elements——"
                "本面单面报告与交易决策报告中的本面分面深稿执行同一套契约，深度必须一致。"
            )
    return {"report": str(report_path), "report_type": "dimension", "dimension": dim,
            "pass": len(issues) == 0, "issues": issues, "warnings": warnings, "soft_notes": soft}


def render_dimension_gate(result: Dict) -> str:
    status = "✅ PASS" if result["pass"] else "❌ FAIL"
    dim = result.get("dimension", "面")
    out = [
        f"# {dim}单面报告门禁 — {status}",
        "",
        f"- **报告路径**: `{result['report']}`",
        f"- **报告类型**: {dim}单面报告（Intent-2 专项 · 契约校验，与 Intent-1 分面深稿同源同深度）",
        "",
    ]
    if result["issues"]:
        out.append("## 阻断项 FAIL")
        out += [f"- ❌ {i}" for i in result["issues"]]
    if result.get("warnings"):
        out.append("\n## 告警")
        out += [f"- ⚠ {w}" for w in result["warnings"]]
    if result.get("soft_notes"):
        out.append("\n## 软提示（不阻断）")
        out += [f"- ℹ {n}" for n in result["soft_notes"]]
    if result["pass"]:
        out.append("\n> 门禁通过：已达该面 face_contract 契约深度（字数/表格/小节/脚标/信源/分析纵深），与 Intent-1 分面深稿同源等深。")
    return "\n".join(out) + "\n"


def _basic_draft_needs_triple_set(text: str) -> bool:
    """基本面**分面深稿**是否触发三件套 GATE0（独立于汇总报告的 4.1.4 口径）。

    分面深稿（`分面深稿_基本面_…md`）用的是**自身章节编号**（如 §2.4 盈利预测 /
    §2.4.2 三情景盈利预测），**不含汇总报告的 "4.1.4"**——故 `needs_trade_fundamental_gate`
    （为 `交易决策报告_` 汇总报告设计、键于 "4.1.4"）对深稿永远不触发。本判据按深稿口径
    识别：只要深稿真正做了**卖方级三情景盈利预测**（含「盈利预测」节 + 三情景/三档/
    forecast_engine 任一标志）即触发。超短/短线**排雷级**基本面深稿不做三情景预测、无此标志
    → 不触发，避免误伤。同时兼容汇总口径（含 4.1.4）以防深稿沿用汇总编号。
    """
    if not text:
        return False
    if needs_trade_fundamental_gate(text):  # 兼容含 4.1.4 的（汇总口径）深稿
        return True
    return ("盈利预测" in text) and any(
        kw in text for kw in ("三情景", "三档", "forecast_engine")
    )


def basic_face_draft_triple_set_issues(draft_path: Path) -> List[str]:
    """v1.32 问题2：把『财务预测三件套 GATE0』前移到**基本面分面深稿**阶段。

    设计动机（用户投诉 #2）：旧版三件套 GATE0 只在 `基本面_` 成稿报告与 `交易决策报告_`
    汇总阶段触发；而 Intent-1 流水线里基本面是先写 `分面深稿_基本面_…md`，该深稿走
    `--single-face 基本面` 检查点时**不触发**三件套，导致"盈利预测对不对"被一路拖到
    汇总报告阶段才暴露——既违反"财务预测是基本面研究的一部分"的职责归属，又催生了
    "汇总阶段临时回调假设硬凑过线"的数字游戏。本函数让基本面深稿在单面检查点即承担
    三件套 GATE0：① 三件套产物齐备；② forecast 三档 EPS/净利非空；③ 命门假设非待填
    + 带依据/信源 + 无数字游戏红旗；④ base 偏离一致预期 ≤50%；⑤ base 不失控外推。

    复用与汇总阶段**同一组底层校验函数**（语义单一可信源，不会两处漂移）。
    forecast.json 按**深稿 stem**定位（`{draft_stem}_forecast.json`）——即基本面深稿阶段
    须用 `forecast_engine.py {code} --report-name 分面深稿_基本面_{code}_{简称}_{ts}` 产出。
    返回 FAIL 问题清单（空=通过）。仅波段/中长线基本面（深稿含 4.1.4 盈利预测）触发；
    超短/短线排雷级基本面无 4.1.4 → 直接跳过，不误伤。"""
    issues: List[str] = []
    try:
        text = draft_path.read_text(encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        return issues
    if not _basic_draft_needs_triple_set(text):  # 深稿无三情景盈利预测（排雷级）→ 不触发
        return issues
    m = FUNDAMENTAL_CODE_RE.search(draft_path.name)
    if not m:
        return issues
    code = m.group(1)
    ws_root = draft_path.resolve().parent.parent
    fdata = ws_root / "FinancialData"
    outrep = draft_path.resolve().parent
    stem = draft_path.stem
    fcst = outrep / f"{stem}_forecast.json"
    yaml_path = fdata / f"{code}_assumptions.yaml"

    # ① 三件套产物存在
    for label, p, cmd in [
        (f"FinancialData/{code}_historical.xlsx", fdata / f"{code}_historical.xlsx",
         f"historical_data_collector.py {code} --force"),
        (f"FinancialData/{code}_assumptions.yaml", yaml_path,
         f"assumptions_yaml_generator.py --code {code}"),
        (f"{stem}_forecast.json", fcst,
         f"forecast_engine.py {code} --report-name {stem} --current-price <现价>"),
    ]:
        if not p.exists():
            issues.append(f"[GATE0·三件套] 缺 {label} —— 基本面深稿阶段必须先跑三件套（{cmd}）")

    # ② forecast 三档情景完整
    if fcst.exists():
        ok, why = _forecast_scenarios_complete(fcst)
        if not ok:
            issues.append(f"[GATE0·三件套] forecast.json 三档情景不完整：{why}")

    # ③ 命门假设：非待填 + 带依据/信源 + 无数字游戏红旗
    if yaml_path.exists():
        unf, fields = _assumptions_critical_unfilled(yaml_path)
        if unf:
            issues.append(f"[GATE0·三件套] 命门假设仍为'待填'：{', '.join(fields)}（直接决定 EPS 预测）")
        for s in _assumptions_sources_missing(yaml_path):
            issues.append(f"[GATE0·假设依据] {s}")
        ng = _assumptions_number_game_redflags(yaml_path)
        if ng:
            issues.append(
                "[GATE0·数字游戏] 假设依据出现『为过线而设数字』红旗：" + "；".join(ng[:4])
                + ("…" if len(ng) > 4 else "")
                + "——禁止用门禁阈值倒推数字，须改回年报/季报/一致预期的研究口径推导"
            )

    # ④ base 偏离一致预期 > 50% → FAIL
    if fcst.exists():
        div = _base_scenario_vs_consensus_divergence(fcst)
        if div.get("status") == "fail":
            issues.append(
                f"[GATE0·信源校准] base 档严重背离市场一致预期（{div['detail']}），"
                f"|EPS 偏离| > {int(_BASE_VS_CONSENSUS_FAIL_PCT)}% —— base 须是『最可能情形』，"
                f"把保守判断移 bear、激进移 bull，分歧在 §5.x 逆向估值显式说明依据后仍校准回最可能情形"
            )

    # ⑤ base 失控外推 → FAIL
    if fcst.exists():
        try:
            obj = json.loads(fcst.read_text(encoding="utf-8"))
            sanity = obj.get("revenue_sanity", []) or []
        except Exception:  # noqa: BLE001
            sanity = []
        base_runaway = [w for w in sanity if w.get("type") == "runaway" and w.get("scenario") == "base"]
        if base_runaway:
            issues.append(
                "[GATE0·失控外推] base 档营收失控外推：" + "；".join(w.get("msg", "") for w in base_runaway)
                + "——须补 L1_industry.tam_yi + L2_company.market_share_ceiling_pct 自上而下天花板，或核减分部增速使隐含总增速逐年收敛"
            )
    return issues


def main() -> None:
    parser = argparse.ArgumentParser(description="校验股票交易决策报告的结构、关键内容覆盖和三段式完整性")
    parser.add_argument("report", help="待校验的 Markdown 报告路径")
    parser.add_argument(
        "--style",
        choices=list(STYLE_LABELS.keys()),
        help="显式指定风格：ultra_short/short/swing/long/full",
    )
    parser.add_argument(
        "--require-supply-demand",
        action="store_true",
        help="要求 4.1.4 完整覆盖供给-需求-收入-成本-费用-净利润联动预测（成长/制造业建议开启）",
    )
    parser.add_argument(
        "--no-companion-check",
        action="store_true",
        dest="no_companion_check",
        help="关闭汇总决策报告的『行业信源文件 + 场景化数据信源(风险红线P0/股权激励/海外可比/期权情绪/机构持股)』连带门禁"
             "（默认开启；v11 起不再连带校验 6 份独立面报告，Intent-1 直接出单份深稿）。",
    )
    parser.add_argument(
        "--check-3statement",
        action="store_true",
        dest="check_3statement",
        help="仅对报告做三表预测章节（4.1.4B）的要素覆盖 + 勾稽 FAIL 原因说明校验；与主门禁独立。",
    )
    parser.add_argument(
        "--check-derivation",
        action="store_true",
        dest="check_derivation",
        help="联动推导链审计（R1 裸数字率 / R2 假设链 / R3 三档情景 / R4 证伪条件 / R5 三表勾稽）。"
             "深度基本面报告（路径前缀 基本面_ 且含三表预测）会自动启用。",
    )
    parser.add_argument(
        "--check-articulation",
        action="store_true",
        dest="check_articulation",
        help="联动三表勾稽差额校验（R5）。需配合 --code。",
    )
    parser.add_argument(
        "--code",
        dest="code",
        default=None,
        help="股票代码（用于推导链审计中的三表勾稽差额计算）",
    )
    parser.add_argument(
        "--bare-rate-threshold",
        type=float,
        default=0.10,
        help="R1 裸数字率阈值（默认 0.10 = 10%%）",
    )
    parser.add_argument(
        "--articulation-threshold",
        type=float,
        default=0.005,
        help="R5 三表勾稽差额阈值（默认 0.005 = 0.5%%）",
    )
    parser.add_argument(
        "--emit-gate",
        nargs="?",
        const="__AUTO__",
        default=None,
        help=(
            "产出人类可读的 _gate_result.md 门禁报告。"
            "默认路径：与报告同目录的 _gate_result.md；也可显式传路径。"
        ),
    )
    parser.add_argument(
        "--format",
        choices=["json", "gate", "both"],
        default="json",
        help="stdout 输出格式：json（默认）/ gate（人类可读）/ both（两者都输出）",
    )
    parser.add_argument(
        "--single-face",
        dest="single_face",
        default=None,
        choices=list(STAGE_A_FACES),
        help="v21 单面检查点：把 positional report 当作某一面的『分面深稿_{面}_…md』，"
             "只对该单份深稿跑 GATE0 单面纵深校验（字数/表格/小节/★/信源表/脚标/门禁自检/分析纵深），"
             "PASS→exit 0，FAIL→exit 1 并打印待补清单。用于阶段A『每面写完即卡』的回环。",
    )
    args = parser.parse_args()

    # ── v21 单面检查点分支：与主门禁独立，写完一个面立即校验该单份深稿 ──
    if args.single_face:
        draft_path = Path(args.report)
        passed, issues = check_single_face_draft(draft_path, args.single_face)
        # v1.32 问题2：财务预测三件套 GATE0 前移到「基本面分面深稿」阶段。
        # 只要基本面深稿含 4.1.4（即承担交易决策的盈利预测纵深），单面检查点即触发
        # 三件套门禁（历史→假设→预测 + 信源/数字游戏/一致预期偏离/外推护栏），
        # 不再等到交易决策汇总报告阶段才暴露问题。
        if args.single_face == "基本面":
            ts_issues = basic_face_draft_triple_set_issues(draft_path)
            if ts_issues:
                passed = False
                issues = list(issues) + ts_issues
        if passed:
            print(f"[GATE0·单面] PASS — {args.single_face}：{draft_path.name} 已达合格单面专项报告纵深。")
            sys.exit(0)
        print(f"[GATE0·单面] FAIL — {args.single_face}：{draft_path.name} 未达标，待补：")
        for it in issues:
            print(f"  - {it}")
        print("请按清单当场补深，复跑本命令直到 PASS，再写下一个面 / 汇总报告。")
        sys.exit(1)


    report_path = Path(args.report)

    # ════════════════════════════════════════════════════════════════════
    # 六个面框架分流：单面报告 → 轻量门禁
    #   - 政策/资金/筹码/技术/消息 五个面：一律走轻量门禁
    #   - 基本面（轻量，即 基本面_ 前缀且无三表/盈利预测深度标记）：同走轻量门禁
    #   - 基本面（深度，即 基本面_ 前缀且含三表预测）：留给下方主门禁 + GATE0
    # 轻量门禁不强制三件套/三表/五段式，只验 面名+代码 / ★评级 / 数据可核验。
    # ════════════════════════════════════════════════════════════════════
    _is_light_fundamental = (
        report_path.name.startswith("基本面_")
        and not needs_deep_fundamental_gate(report_path)
    )
    if is_dimension_single_report(report_path.name) or _is_light_fundamental:
        dim_result = check_dimension_report(report_path)
        if args.format in ("json", "both"):
            print(json.dumps(dim_result, ensure_ascii=False, indent=2))
        if args.format in ("gate", "both"):
            print(render_dimension_gate(dim_result))
        if args.emit_gate is not None:
            gate_out = (report_path.parent / "_gate_result.md"
                        if args.emit_gate == "__AUTO__" else Path(args.emit_gate))
            gate_out.write_text(render_dimension_gate(dim_result), encoding="utf-8")
            print(f"[gate] written: {gate_out}", file=sys.stderr)
        sys.exit(0 if dim_result["pass"] else 1)



    # ════════════════════════════════════════════════════════════════════
    # GATE0 · 基本面深度三件套前置预检（v1.22 工程层硬拦截）
    # ────────────────────────────────────────────────────────────────────
    # 触发条件（v10 · 详见 references/six_dimension_framework.md §二.1）：
    #   - 基本面_ 前缀 ：无条件触发（基本面=深度唯一形态，不再区分深度/轻量）
    #   （任何 `基本面_` 单面报告都必须备齐三件套，否则 FAIL）
    # 校验对象：
    #   ① FinancialData/{code}_historical.xlsx       缺即 FAIL
    #   ② FinancialData/{code}_assumptions.yaml      缺即 FAIL
    #   ③ OutputReport/{report_stem}_forecast.json   缺即 FAIL
    #   ④ OutputReport/{report_stem}_audit_sidecar.json  缺仅 WARN
    # 行为：缺失任一即 exit 1，不进入主门禁；stdout JSON + stderr 人读 + _gate_result.md 三处同步落盘
    # 设计目的：堵住"LLM 跳过 historical→assumptions→forecast 三件套直接拍脑袋写 EPS/三表"的口子
    # ════════════════════════════════════════════════════════════════════
    # v1.23 修复：① 漏洞 A — 触发不再限 `基本面_`，`交易决策报告_` 含 4.1.4 盈利预测者一并触发；
    #            ② 漏洞 B — 不止查"文件存在"，追加"内容校验"（forecast 三档情景非空 / assumptions 命门未待填）。
    _gate0_text = ""
    try:
        _gate0_text = report_path.read_text(encoding="utf-8")
    except Exception:  # noqa: BLE001
        pass
    _trigger_b = needs_deep_fundamental_gate(report_path)            # 基本面_ 单面深度报告
    _trigger_trade = needs_trade_fundamental_gate(_gate0_text)       # Intent-1 汇总决策报告含 4.1.4
    if _trigger_b or _trigger_trade:
        _code_match = (FUNDAMENTAL_CODE_RE.search(report_path.name)
                       or TRADE_CODE_RE.search(report_path.name))
        if _code_match:
            _bcode = _code_match.group(1)
            _ws_root = report_path.resolve().parent.parent  # OutputReport 的上级 = 工作区根
            _fdata = _ws_root / "FinancialData"
            _outrep = report_path.resolve().parent  # OutputReport/
            _stem = report_path.stem
            _fcst_path = _outrep / f"{_stem}_forecast.json"
            _yaml_path = _fdata / f"{_bcode}_assumptions.yaml"
            _gate0_checks = [
                ("historical_xlsx", _fdata / f"{_bcode}_historical.xlsx",
                 f"python scripts/historical_data_collector.py {_bcode} --force"),
                ("assumptions_yaml", _yaml_path,
                 f"python scripts/assumptions_yaml_generator.py --code {_bcode}"),
                ("forecast_json", _fcst_path,
                 f"python scripts/forecast_engine.py {_bcode} --report-name {_stem} --current-price <现价>"),
            ]
            _gate0_missing = [(k, p, cmd) for (k, p, cmd) in _gate0_checks if not p.exists()]
            _audit_sidecar = _outrep / f"{_stem}_audit_sidecar.json"
            _gate0_warn = not _audit_sidecar.exists()

            # ── 内容校验（v1.23 漏洞 B）：文件在还不够，必须"真填真算" ──
            _gate0_content: List[tuple] = []
            if _fcst_path.exists():
                _ok_fc, _why_fc = _forecast_scenarios_complete(_fcst_path)
                if not _ok_fc:
                    _gate0_content.append((
                        f"forecast.json 三档情景不完整：{_why_fc}",
                        f"先把 FinancialData/{_bcode}_assumptions.yaml 的 L4 营收增速/毛利率三档假设填实，"
                        f"再重跑 forecast_engine.py {_bcode} --report-name {_stem}",
                    ))
            if _yaml_path.exists():
                _unf, _unf_fields = _assumptions_critical_unfilled(_yaml_path)
                if _unf:
                    _gate0_content.append((
                        f"assumptions.yaml 命门假设仍为'待填'：{', '.join(_unf_fields)}（直接决定 EPS 预测）",
                        f"编辑 FinancialData/{_bcode}_assumptions.yaml，为 L4_income_statement.{{revenue_growth_pct, gross_margin_pct}} "
                        f"填入 bull/base/bear 三档数值 + comment 推导依据 + source 信源",
                    ))
                # ── v1.32 问题1①：命门假设必须带 comment 推导依据 + source 信源（禁裸填数字）──
                _src_miss = _assumptions_sources_missing(_yaml_path)
                if _src_miss:
                    _gate0_content.append((
                        "assumptions.yaml 命门假设缺『推导依据/信源』（裸填数字）：" + "；".join(_src_miss),
                        f"为这些假设补上 comment（自下而上量价/利润率推导逻辑）与 source（年报/季报/一致预期/研报 + 时效）；"
                        f"财务预测不是数字游戏——每个驱动 EPS 的数字都须可追溯到调研依据，而非凭空设定",
                    ))
                # ── v1.32 问题1②：识别"为过线而填数字"的门禁导向红旗表述 → FAIL ──
                _ng = _assumptions_number_game_redflags(_yaml_path)
                if _ng:
                    _gate0_content.append((
                        "assumptions.yaml 假设依据出现『为通过门禁而设数字』红旗：" + "；".join(_ng[:6])
                        + ("…" if len(_ng) > 6 else ""),
                        "禁止用『门禁/过线/凑/卡线/阈值以内』等论证假设——这证明数字是倒推门禁阈值而非源自调研。"
                        "请删除这些表述，回到最新年报/季报『分部量价』+ 一致预期重新推导该假设，依据写成研究口径"
                        "（如『Q1 实际出货量年化 × ASP 年降幅 → 分部增速 X%』），再重跑 forecast_engine.py",
                    ))

            # ── v1.28 信源校准合理性：base 档 vs 市场一致预期偏离过大拦截 ──
            #   设计：base 档应是"最可能情形"，不得无理由严重背离 29 家券商一致预期中位数；
            #   只在极端背离（EPS 偏离 > 50%）才 FAIL，30%~50% 仅 WARN（尊重"分析师可有据偏离"）。
            _base_div_warn = None
            if _fcst_path.exists():
                _div = _base_scenario_vs_consensus_divergence(_fcst_path)
                if _div["status"] == "fail":
                    _gate0_content.append((
                        f"base（中性）档严重背离市场一致预期（{_div['detail']}），|EPS 偏离| > {int(_BASE_VS_CONSENSUS_FAIL_PCT)}%"
                        f"——base 档应是『最可能情形』却被填成准 bull/bear 区；这是历史事故"
                        f"（把保守/激进单边判断误塞 base 档、目标价与现价大幅背离仍 PASS）的机器兜底",
                        f"重校准 FinancialData/{_bcode}_assumptions.yaml：base 档量价/毛利率锚定『最近一期实际季报年化 + 一致预期』"
                        f"（base 通常落在一致预期 ±{int(_BASE_VS_CONSENSUS_FAIL_PCT)}% 内），把极端单边判断移到 bull/bear 档；"
                        f"若确为有据的逆向判断，须在 §5.x 逆向估值/「我vs市场」显式呈现分歧依据后，仍把 base 校准回最可能情形；"
                        f"再重跑 forecast_engine.py {_bcode} --report-name {_stem}",
                    ))
                elif _div["status"] == "warn":
                    _base_div_warn = _div["detail"]
            if _base_div_warn:
                sys.stderr.write(
                    f"\n[GATE0·信源校准·WARN] base 档与一致预期偏离处于 "
                    f"{int(_BASE_VS_CONSENSUS_WARN_PCT)}%~{int(_BASE_VS_CONSENSUS_FAIL_PCT)}% 区间：{_base_div_warn}。"
                    f"建议在 §5.x 逆向估值/「我vs市场」预期差表显式说明分歧依据（非阻断，不影响 PASS）。\n"
                )

            # ── v1.30 失控外推哨兵：消费 forecast.json["revenue_sanity"]（来自 forecast_engine P0）──
            #   base 档 runaway（营收增速逐年加速且 >50% 不收敛）= 失控外推，FAIL（300308 病灶：
            #   base year_2 营收 +278%、year_3 EPS 399 元这类纯数学失控，此前无人拦截）；
            #   bull/bear 档 runaway 或天花板回切 ceiling_clip 仅 WARN（已自动修正/极端档可较陡）。
            if _fcst_path.exists():
                try:
                    _fc_obj = json.loads(_fcst_path.read_text(encoding="utf-8"))
                    _sanity = _fc_obj.get("revenue_sanity", []) or []
                except Exception:  # noqa: BLE001
                    _sanity = []
                _base_runaway = [w for w in _sanity if w.get("type") == "runaway" and w.get("scenario") == "base"]
                _other_warn = [w for w in _sanity
                               if (w.get("type") == "runaway" and w.get("scenario") != "base") or w.get("type") == "ceiling_clip"]
                if _base_runaway:
                    _msgs = "；".join(w.get("msg", "") for w in _base_runaway)
                    _gate0_content.append((
                        f"base（中性）档营收呈失控外推（{_msgs}）——base 应是『最可能情形』，"
                        f"恒定/加速复利使远期营收/EPS 脱离行业天花板（历史事故：300308 base year_2 营收 +278%、year_3 EPS 399 元仍 PASS）",
                        f"在 FinancialData/{_bcode}_assumptions.yaml 补『自上而下天花板』："
                        f"L1_industry.tam_yi（行业 TAM 折亿元）+ L2_company.market_share_ceiling_pct（市占率上限），"
                        f"或直接 L2_company.company_revenue_ceiling_yi（公司营收天花板，亿元）；"
                        f"并核对分部 volume_growth_pct/price_change_pct 是否高估，使隐含总增速逐年收敛；再重跑 forecast_engine.py {_bcode}",
                    ))
                if _other_warn:
                    _omsgs = "；".join(w.get("msg", "") for w in _other_warn)
                    sys.stderr.write(
                        f"\n[GATE0·失控/天花板·WARN] {_omsgs}（bull/bear 失控外推或天花板已自动回切，非阻断；"
                        f"建议复核分部增速假设是否过乐观）。\n"
                    )

            if _gate0_missing or _gate0_content:
                _trigger_label = "Intent-1 汇总决策报告（含 4.1.4 盈利预测）" if _trigger_trade and not _trigger_b else "基本面深度报告"
                _gate0_obj = {
                    "gate": "GATE0_基本面三件套前置预检",
                    "pass": False,
                    "trigger": _trigger_label,
                    "missing": [{"key": k, "path": str(p)} for (k, p, _c) in _gate0_missing],
                    "content_fail": [{"issue": i} for (i, _c) in _gate0_content],
                    "remediation": [c for (_k, _p, c) in _gate0_missing] + [c for (_i, c) in _gate0_content],
                    "audit_sidecar_warn": _gate0_warn,
                    "rationale": "凡含盈利预测的报告（基本面深稿 或 Intent-1 含 4.1.4）必须先跑 historical→assumptions→forecast 三件套，"
                                 "且 assumptions 命门假设须填实、forecast 三档情景 EPS/净利须非空；跳过/留空即 GATE0 FAIL（v1.23 内容级硬拦截）。"
                                 "v1.28 追加『base 信源校准合理性』硬校验：base 档预测 EPS 与市场一致预期中位数偏离 > 50% 即 FAIL——"
                                 "base 档须是最可能情形，禁止把保守/激进单边判断误塞 base 档（保守归 bear、乐观归 bull、分歧用 §5.x 逆向估值显式呈现）。"
                                 "v1.30 追加『远期一致性 + 防失控外推』：base 档 EPS 偏离一致预期的 year_1/year_2 较差者 > 50% 即 FAIL；"
                                 "base 档营收增速逐年加速且 >50% 不收敛（revenue_sanity.runaway）即 FAIL——须补 L1/L2 自上而下天花板或核减分部增速",
                }
                # stdout 输出 JSON
                print(json.dumps(_gate0_obj, ensure_ascii=False, indent=2))
                # stderr 输出人读
                sys.stderr.write("\n========== GATE0 · 基本面三件套前置预检 FAIL ==========\n")
                sys.stderr.write(f"  触发类型：{_trigger_label}\n")
                for (k, p, cmd) in _gate0_missing:
                    sys.stderr.write(f"\n  [缺文件] {k}\n    路径：{p}\n    补救：{cmd}\n")
                for (issue, cmd) in _gate0_content:
                    sys.stderr.write(f"\n  [内容FAIL] {issue}\n    补救：{cmd}\n")
                sys.stderr.write("\n  设计：含盈利预测的报告必须真跑三件套且真填真算——禁止 LLM 手算/锚定 EPS 冒充'自下而上预测'（v1.23）\n")
                sys.stderr.write("======================================================\n\n")
                # ── v1.31 根因诊断：FAIL 时自动解剖 base 档分部加总，引导 LLM 算无遗策排查 ──
                _diag_md = ""
                if _fcst_path.exists():
                    _diag_md = _render_forecast_diagnostic(_fcst_path)
                if _diag_md:
                    sys.stderr.write(_diag_md + "\n")
                # 同步落盘 _gate_result.md
                _gate_md = _outrep / "_gate_result.md"
                _md_lines = ["# GATE0 · 基本面三件套前置预检 — **FAIL**", "", f"> 触发类型：{_trigger_label}", ""]
                _md_lines += [f"- 缺文件：`{p}`（补救：`{cmd}`）" for (_k, p, cmd) in _gate0_missing]
                _md_lines += [f"- 内容FAIL：{issue}（补救：{cmd}）" for (issue, cmd) in _gate0_content]
                _md_lines += ["", "> 设计：含盈利预测的报告必须真跑三件套且真填真算（v1.23 内容级硬拦截）"]
                if _diag_md:
                    _md_lines += ["", "---", "", _diag_md]
                _gate_md.write_text("\n".join(_md_lines) + "\n", encoding="utf-8")
                sys.exit(1)

    # 三表预测专项校验：独立分支，不触发主门禁
    if args.check_3statement:
        ts_result = check_three_statement_coverage(report_path)
        if args.format in ("json", "both"):
            print(json.dumps(ts_result, ensure_ascii=False, indent=2))
        if args.format in ("gate", "both"):
            print(render_three_statement_gate(ts_result))
        if args.emit_gate is not None:
            gate_out = (
                report_path.parent / "_gate_3statement.md"
                if args.emit_gate == "__AUTO__"
                else Path(args.emit_gate)
            )
            gate_out.write_text(render_three_statement_gate(ts_result), encoding="utf-8")
            print(f"[gate-3statement] written: {gate_out}", file=sys.stderr)
        sys.exit(0 if ts_result["pass"] else 1)

    result = validate_report(
        report_path=report_path,
        style=args.style,
        require_supply_demand=args.require_supply_demand,
        check_companion=not args.no_companion_check,
    )

    # Stdout 输出
    if args.format in ("json", "both"):
        print(json.dumps(result, ensure_ascii=False, indent=2))
    if args.format in ("gate", "both"):
        print(render_gate_markdown(result))

    # --emit-gate 落地为文件
    if args.emit_gate is not None:
        if args.emit_gate == "__AUTO__":
            gate_out = report_path.parent / "_gate_result.md"
        else:
            gate_out = Path(args.emit_gate)
        gate_out.write_text(render_gate_markdown(result), encoding="utf-8")
        # 用 stderr 提示，不污染 stdout JSON
        print(f"[gate] written: {gate_out}", file=sys.stderr)

        # P0-7：把本次软门禁 WARN 总数记入校准台账，便于跨报告追踪质量趋势
        try:
            from calibration_review import record_warn_count
            _soft = [
                w for w in result.get("warnings", [])
                if str(w).startswith("[v12·") or str(w).startswith("[v13·") or "·WARN]" in str(w)
            ]
            record_warn_count(report_path, len(_soft), note="auto from report_quality_checker")
        except Exception as _e:
            print(f"[gate] record_warn_count 跳过: {_e}", file=sys.stderr)

    # ────────────────────────────────────────────────────────────────────
    # 基本面双门禁联动：推导链审计
    # 触发条件：① 显式 --check-derivation；或 ② 深度基本面报告（含三表/盈利预测）
    # ────────────────────────────────────────────────────────────────────
    is_b_report = needs_deep_fundamental_gate(report_path)
    run_derivation = args.check_derivation or is_b_report
    derivation_pass = True
    if run_derivation:
        try:
            from derivation_chain_auditor import run_audit, render_human
            # 自动从文件名识别股票代码（如 基本面_300308_中际旭创.md → 300308）
            code = args.code
            if not code:
                m = re.search(r"_(\d{6})_", report_path.name)
                if m:
                    code = m.group(1)
            skip_articulation = not (args.check_articulation or bool(code))
            audit = run_audit(
                report_path=report_path,
                code=code,
                bare_rate_threshold=args.bare_rate_threshold,
                articulation_threshold=args.articulation_threshold,
                skip_articulation=skip_articulation,
            )
            human = render_human(audit)
            print("\n" + "=" * 60, file=sys.stderr)
            print("【基本面双门禁】推导链审计:", audit.overall, file=sys.stderr)
            print("=" * 60, file=sys.stderr)
            print(human, file=sys.stderr)
            # 同步落盘
            deriv_out = report_path.parent / "_derivation_gate.md"
            deriv_out.write_text(human, encoding="utf-8")
            print(f"[gate-derivation] written: {deriv_out}", file=sys.stderr)
            # 严格按 derivation_chain_auditor.py 顶部 docstring 第 51-54 行的退出码契约：
            #   PASS  → 0（所有硬约束通过）
            #   FAIL  → 1（任一硬约束 FAIL，阻断交付）
            #   WARN  → 2（仅软约束/披露级问题，不阻断交付但需在报告中说明）
            # 因此本聚合门禁也将 WARN 视为通过——前提是报告作者已在正文中显式披露
            # 这些 WARN 项的来源（路径依赖 / 四舍五入 / 重述等）。FAIL 才硬阻断。
            derivation_pass = (audit.overall in ("PASS", "WARN"))
            if audit.overall == "WARN":
                print(
                    "[gate-derivation] 推导链审计为 WARN：报告未含硬约束违例，但存在 "
                    "需在正文披露的软性差额；按契约不阻断交付，请确认报告中已说明 WARN 项来源。",
                    file=sys.stderr,
                )
        except ImportError as e:
            print(f"[WARN] 推导链审计器不可用: {e}", file=sys.stderr)
        except Exception as e:
            print(f"[WARN] 推导链审计执行异常: {e}", file=sys.stderr)
            derivation_pass = False  # 异常视为失败，避免误放行

    # 退出码：主门禁 PASS 且 推导链门禁 PASS（含 WARN）才视为整体 PASS
    overall_pass = result["pass"] and derivation_pass
    sys.exit(0 if overall_pass else 1)


if __name__ == "__main__":
    main()
