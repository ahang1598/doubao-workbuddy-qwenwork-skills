#!/usr/bin/env python3
"""
excel_report.py — 生成票据匹配结果 Excel 对照表

定位: **本地产物**。不上传 COS、不进 `UiReq`、不参与呼起 UI。
      (因此 proto 无需新增 `excel_url` 字段)

CLI:
  python excel_report.py --input '<JSON>'
  python excel_report.py --input-file <path/to/input.json>

入参 JSON:
  {
    "items_file": "./_tmp/session_x/items.json",   // title_normalizer.py 落盘的全量结果
    "output_path": "./票据匹配结果.xlsx",           // 可选, 默认工作区根目录固定文件名
    "auto_install": true                            // 可选, 默认 true; 缺 openpyxl 时自动装
  }

  也支持直接传 items(小批量调试用):
  { "items": [ { ...MatchItem... } ], "output_path": "..." }

出参 JSON (stdout):
  {
    "success": true,
    "output_path": "D:/work/票据匹配结果.xlsx",
    "row_count": 20,
    "matched_count": 14,
    "failed_count": 6,
    "overwritten": true
  }

═══════════════════════════════════════════════════════════════════════════
表头(需求方给定, **顺序不得改, 不得增删列**)
═══════════════════════════════════════════════════════════════════════════
  是否上传 | 票据抬头 | 开票金额（元）| 项目名称 | 匹配状态 | 说明 / 原因 | 申请编号

| 列              | 来源                    | 取值                |
|-----------------|-------------------------|-------------------------------|
| 是否上传        | `status`                | `1`→是 / `0`→否|
| 票据抬头| `title`                 | 原样                |
| 开票金额（元）  | `amount`(**分**)        | **分→元**, 两位小数           |
| 项目名称| `project_name_list`     | 多个用 `、` 连接, 空数组留空  |
| 匹配状态        | `match_status`          | `1`→匹配成功 / `2`→未匹配     |
| 说明 / 原因     | `match_status_reason`   | 原样(五种文案之一)            |
| 申请编号        | `application_number`    | 未匹配时留空                  |

**排序: 未匹配(match_status=2)全部排在已匹配之前**, 组内按上传顺序(seq)。
  理由: Excel 的使用场景是离线核对**待处理项**, 未匹配的才是要动手的部分。
  这与 UI 两列表的呈现逻辑**不同**, 不要试图统一。

═══════════════════════════════════════════════════════════════════════════
⚠️ 金额: 这是全链路**唯一**用「元」的地方
═══════════════════════════════════════════════════════════════════════════
分 → 元的转换**必须用整数运算**(`c // 100` + `c % 100`), **严禁浮点除法**:
  33000 / 100 在部分平台会得到 329.99999999999994, 票据金额出这种误差是事故。

> 这**不违反**「严禁自己做元→分换算」铁律 —— 那条禁的是**匹配入参方向**
> (元→分, 必须用 amount_conversion.py 返回的 value_cents)。
> 这里是**反方向**(分→元)、属于**展示层**、且同样由脚本完成, 不是 LLM 心算。

═══════════════════════════════════════════════════════════════════════════
文件位置与覆盖
═══════════════════════════════════════════════════════════════════════════
  工作区根目录 / 票据匹配结果.xlsx     ← 固定文件名, 用户可见

- 每轮重匹配后**覆盖**同一个文件
- ⛔ 不得带时间戳 / 轮次号后缀(会产生一堆垃圾文件)
- ⛔ 不得放在 `_tmp/`(那是临时目录, 用户找不到; 且临时目录会被清理)
- 生成后**必须**在对话中告知用户文件路径, 不得静默生成
- ⛔ 表格内容**不得**逐行打进agent 上下文(token 治理)
"""
import argparse
import contextlib
import io
import json
import os
import subprocess
import sys
from typing import List, Optional

DEFAULT_FILENAME = "票据匹配结果.xlsx"


# ---------------------------------------------------------------------------
# stdout 纯净化（⚠️ 2026-08-10 实测事故, 别删）
# ---------------------------------------------------------------------------
# 契约是「stdout 只有一行 JSON」。第三方库 import / 安装期间往 stdout 打日志会
# 让调用方 `json.loads(stdout)` 直接崩(已在 pdf_to_images.py 实测复现:
# `import fitz` 打 deprecation 警告)。处理期间的 stdout 一律转stderr。
def _run_quiet(fn, *args, **kwargs):
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        result = fn(*args, **kwargs)
    noise = buf.getvalue()
    if noise.strip():
        sys.stderr.write(noise)
        sys.stderr.flush()
    return result

HEADERS = [
    "是否上传",
    "票据抬头",
    "开票金额（元）",
    "项目名称",
    "匹配状态",
    "说明 / 原因",
    "申请编号",
]

COLUMN_WIDTHS = [10, 40, 16, 30, 12, 34, 22]

MATCHED = 1
FAILED = 2


# ---------------------------------------------------------------------------
# openpyxl 探测与自助补齐(对齐 pdf_to_images.py 的模式)
# ---------------------------------------------------------------------------
def _probe_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:
        return False


def _try_auto_install() -> tuple:
    cmds = [
        [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"],
        [sys.executable, "-m", "pip", "install", "--quiet", "--user", "openpyxl"],
    ]
    logs = []
    for cmd in cmds:
        try:
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
            logs.append(
                f"$ {' '.join(cmd)}\n"
                f"  rc={proc.returncode} out={proc.stdout.strip()[:200]} "
                f"err={proc.stderr.strip()[:200]}"
            )
            if proc.returncode == 0 and _probe_openpyxl():
                return True, "\n".join(logs)
        except Exception as e:  # noqa: BLE001
            logs.append(f"$ {' '.join(cmd)}\n  exception={e}")
    return False, "\n".join(logs)


# ---------------------------------------------------------------------------
# 取值映射
# ---------------------------------------------------------------------------
def cents_to_yuan(cents) -> str:
    """分 → 元, 两位小数。**整数运算, 严禁浮点除法。**"""
    try:
        c = int(cents)
    except (TypeError, ValueError):
        return ""
    if c < 0:
        c = 0
    return f"{c // 100}.{c % 100:02d}"


def _upload_flag(status) -> str:
    return "是" if int(status or 0) == 1 else "否"


def _match_status_text(match_status) -> str:
    return "匹配成功" if int(match_status or FAILED) == MATCHED else "未匹配"


def _project_names(names) -> str:
    if not names:
        return ""
    if isinstance(names, str):
        return names
    return "、".join(str(n) for n in names if str(n or "").strip())


def _row_of(item: dict) -> list:
    match_status = int(item.get("match_status") or FAILED)
    application_number = str(item.get("application_number") or "")
    if match_status != MATCHED:
        # 未匹配时申请编号必须留空 —— 防止把落败前占过的编号写进表
        application_number = ""
    return [
        _upload_flag(item.get("status")),
        str(item.get("title") or ""),
        cents_to_yuan(item.get("amount")),
        _project_names(item.get("project_name_list")),
        _match_status_text(match_status),
        str(item.get("match_status_reason") or ""),
        application_number,
    ]


def sort_items(items: List[dict]) -> List[dict]:
    """未匹配在前, 已匹配在后; 组内按上传顺序(seq,缺失则按原顺序)。"""
    decorated = []
    for idx, item in enumerate(items):
        match_status = int(item.get("match_status") or FAILED)
        group = 0 if match_status != MATCHED else 1
        seq = item.get("seq")
        seq = int(seq) if isinstance(seq, int) or (isinstance(seq, str) and seq.isdigit()) else idx
        decorated.append((group, seq, idx, item))
    decorated.sort(key=lambda t: (t[0], t[1], t[2]))
    return [d[3] for d in decorated]


# ---------------------------------------------------------------------------
# 写 xlsx
# ---------------------------------------------------------------------------
def write_workbook(items: List[dict], output_path: str) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ordered = sort_items(items)

    wb = Workbook()
    ws = wb.active
    ws.title = "票据匹配结果"

    ws.append(HEADERS)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS[col - 1]

    for item in ordered:
        ws.append(_row_of(item))

    ws.freeze_panes = "A2"

    directory = os.path.dirname(os.path.abspath(output_path))
    if directory:
        os.makedirs(directory, exist_ok=True)
    existed = os.path.exists(output_path)
    wb.save(output_path)  # 覆盖式写入, 固定文件名

    matched = sum(1 for i in ordered if int(i.get("match_status") or FAILED) == MATCHED)
    return {
        "row_count": len(ordered),
        "matched_count": matched,
        "failed_count": len(ordered) - matched,
        "overwritten": existed,
    }


def _load_items(input_data: dict) -> tuple:
    items = input_data.get("items")
    if isinstance(items, list):
        return items, None
    path = input_data.get("items_file")
    if not path:
        return None, "必须提供 items_file 或 items"
    if not os.path.exists(path):
        return None, f"items_file 不存在: {path}"
    with open(path, "r", encoding="utf-8") as f:
        loaded = json.load(f)
    if isinstance(loaded, dict):
        loaded = loaded.get("items")
    if not isinstance(loaded, list):
        return None, f"items_file 内容不含 items 数组: {path}"
    return loaded, None


def process(input_data: dict) -> dict:
    items, err = _load_items(input_data)
    if err:
        return {"success": False, "error": err}
    if not items:
        return {"success": False, "error": "items 为空, 无可写入的匹配结果"}

    output_path = str(input_data.get("output_path") or "").strip() or os.path.join(
        os.getcwd(), DEFAULT_FILENAME
    )

    install_log = ""
    if not _probe_openpyxl():
        if not bool(input_data.get("auto_install", True)):
            return {
                "success": False,
                "error": "缺少 openpyxl 且auto_install=false",
                "fix_hint": f"{sys.executable} -m pip install openpyxl",
            }
        ok, install_log = _try_auto_install()
        if not ok:
            return {
                "success": False,
                "error": "openpyxl 安装失败, 无法生成 Excel 对照表",
                "attempted_remediation": install_log,
                "fix_hint": f"{sys.executable} -m pip install --user openpyxl",
            }

    try:
        stats = write_workbook(items, output_path)
    except Exception as e:  # noqa: BLE001
        return {"success": False, "error": f"写 xlsx 失败: {e}", "output_path": output_path}

    out = {"success": True, "output_path": os.path.abspath(output_path)}
    out.update(stats)
    if install_log:
        out["install_log"] = install_log
    return out


def _load_input(args) -> dict:
    if args.input_file:
        with open(args.input_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return json.loads(args.input)


def main():
    # ⚠️ Windows 实测: stdout 为管道时用 locale 编码(GBK), 中文 JSON 会写成 GBK 字节,
    #    UTF-8 读取方直接 UnicodeDecodeError。必须显式改 UTF-8。
    for _stream in (sys.stdout, sys.stderr):
        try:
            _stream.reconfigure(encoding="utf-8")
        except Exception:  # noqa: BLE001
            pass

    parser = argparse.ArgumentParser()
    parser.add_argument("--input", help="JSON 格式的入参")
    parser.add_argument("--input-file", dest="input_file", help="入参 JSON 文件路径")
    args = parser.parse_args()

    if not args.input and not args.input_file:
        print(json.dumps({"success": False, "error": "必须提供 --input 或 --input-file"}), flush=True)
        sys.exit(1)

    try:
        input_data = _load_input(args)
    except (json.JSONDecodeError, OSError) as e:
        print(
            json.dumps({"success": False, "error": f"入参非合法 JSON: {e}"}, ensure_ascii=False),
            flush=True,
        )
        sys.exit(1)

    result = _run_quiet(process, input_data)
    print(json.dumps(result, ensure_ascii=False), flush=True)
    sys.exit(0 if result.get("success") else 1)


if __name__ == "__main__":
    main()
