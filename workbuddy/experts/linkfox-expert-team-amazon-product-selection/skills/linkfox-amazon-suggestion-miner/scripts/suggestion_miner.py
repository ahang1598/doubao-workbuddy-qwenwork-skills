#!/usr/bin/env python3
"""
Amazon 搜索建议词挖掘工具 (Suggestion Miner) v2.3

基于 www.amazon.com/suggestions API，完整自动化「手动下拉框挖词技巧」。

对应手动技巧 → 模式映射：
  字母/数字顺推法   → az (后缀) / az_prefix (前缀) / numbers (数字规格)
  疑问词/介词拓展法 → expand (模板含 for/with/how/what 等)
  空格间隙插入法    → gap
  场景/人群/材质法  → expand (丰富模板)
  季节/节日扩展法   → expand
  滚雪球反向扩展    → reverse / deep

用法:
  # 批量扩展（介词+场景+材质+季节等模板）
  python3 suggestion_miner.py --seed "fan" --mode expand --rounds 2

  # A-Z 后缀扫描
  python3 suggestion_miner.py --seed "wireless charger" --mode az

  # A-Z 前缀扫描
  python3 suggestion_miner.py --seed "dog toy" --mode az_prefix

  # 数字/规格拓展
  python3 suggestion_miner.py --seed "dog toy" --mode numbers

  # 空格间隙插入（多词种子词效果最好）
  python3 suggestion_miner.py --seed "dog toy" --mode gap

  # 逆向检索（滚雪球）
  python3 suggestion_miner.py --seed "wireless charger" --mode reverse --top-n 30

  # 深度递归
  python3 suggestion_miner.py --seed "fan" --mode deep --depth 2 --top-n 5

  # 输出到Excel (xlsx) — 默认交付格式
  python3 suggestion_miner.py --seed "fan" --mode expand --xlsx /root/suggestions_fan.xlsx
"""

import argparse
import json
import sqlite3
import csv
import time
import sys
import os
import random
import functools
import urllib.parse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    from googletrans import Translator as GTranslator
    HAS_GOOGLETRANS = True
except ImportError:
    HAS_GOOGLETRANS = False

try:
    from linkfox_paths import resolve_data_path
    HAS_LINKFOX = True
except ImportError:
    HAS_LINKFOX = False


# ============================================================
# API 调用核心
# ============================================================

API_URL = "https://www.amazon.com/suggestions"
MID_US = "ATVPDKIKX0DER"

# ============================================================
# 多站点配置 (23个Amazon全球站点)
# ============================================================
AMAZON_MARKETS = {
    # Group 1: plain-mid=1 (北美 - 共享基础设施)
    "US":  {"mid": "ATVPDKIKX0DER",  "domain": "amazon.com",      "locale": "en_US",  "plain_mid": 1,  "currency": "USD", "lang": "en"},
    "CA":  {"mid": "A2EUQ1WTGCTBG2",  "domain": "amazon.ca",      "locale": "en_CA",  "plain_mid": 1,  "currency": "CAD", "lang": "en"},
    "MX":  {"mid": "A1MK81F0CSZALV",  "domain": "amazon.com.mx",  "locale": "es_MX",  "plain_mid": 1,  "currency": "MXN", "lang": "es"},
    # Group 2: plain-mid=3 (欧洲/亚洲/非洲)
    "DE":  {"mid": "A1PA6795UKMFR9",  "domain": "amazon.de",       "locale": "de_DE",  "plain_mid": 3,  "currency": "EUR", "lang": "de"},
    "FR":  {"mid": "A13V1IB3VIYZZH",  "domain": "amazon.fr",       "locale": "fr_FR",  "plain_mid": 3,  "currency": "EUR", "lang": "fr"},
    "IT":  {"mid": "APJ6JRA9NG5V4",   "domain": "amazon.it",       "locale": "it_IT",  "plain_mid": 3,  "currency": "EUR", "lang": "it"},
    "ES":  {"mid": "A1RKKUPIHCS9HS",  "domain": "amazon.es",       "locale": "es_ES",  "plain_mid": 3,  "currency": "EUR", "lang": "es"},
    "IN":  {"mid": "A21TJRUUN4KGV",   "domain": "amazon.in",       "locale": "en_IN",  "plain_mid": 3,  "currency": "INR", "lang": "en"},
    "NL":  {"mid": "A1805IZSGDJ6NL",  "domain": "amazon.nl",       "locale": "nl_NL",  "plain_mid": 3,  "currency": "EUR", "lang": "nl"},
    "AE":  {"mid": "A2VIGQ35RCS4UG",  "domain": "amazon.ae",       "locale": "en_AE",  "plain_mid": 3,  "currency": "AED", "lang": "en"},
    "SA":  {"mid": "A17E79B4JQC1RO",  "domain": "amazon.sa",       "locale": "ar_SA",  "plain_mid": 3,  "currency": "SAR", "lang": "ar"},
    "PL":  {"mid": "A1C3SOZRARQ6R3",  "domain": "amazon.pl",       "locale": "pl_PL",  "plain_mid": 3,  "currency": "PLN", "lang": "pl"},
    "BE":  {"mid": "AMEN7PMS3EDWL",   "domain": "amazon.com.be",   "locale": "fr_BE",  "plain_mid": 3,  "currency": "EUR", "lang": "fr"},
    "EG":  {"mid": "ARBP9OOSHTCHU",   "domain": "amazon.eg",       "locale": "ar_EG",  "plain_mid": 3,  "currency": "EGP", "lang": "ar"},
    "IE":  {"mid": "A28R2DX0Q7XYYQ",  "domain": "amazon.ie",       "locale": "en_IE",  "plain_mid": 3,  "currency": "EUR", "lang": "en"},
    "ZA":  {"mid": "A3T9NJ3R2K10RK",  "domain": "amazon.co.za",    "locale": "en_ZA",  "plain_mid": 3,  "currency": "ZAR", "lang": "en"},
    "SE":  {"mid": "A2NODRKZP88ZSS",  "domain": "amazon.se",       "locale": "sv_SE",  "plain_mid": 3,  "currency": "SEK", "lang": "sv"},
    # Group 3: no plain-mid
    "JP":  {"mid": "A1VC38T7YXB528",  "domain": "amazon.co.jp",    "locale": "ja_JP",  "plain_mid": None, "currency": "JPY", "lang": "ja"},
    # Group 4: 浏览器辅助 (curl被反爬拦截)
    "UK":  {"mid": "A1F83G8C2ARO7P",  "domain": "amazon.co.uk",    "locale": "en_GB",  "plain_mid": 3,  "currency": "GBP", "lang": "en", "browser_required": True},
    "AU":  {"mid": "A39IBJ37ZK2V1A",  "domain": "amazon.com.au",   "locale": "en_AU",  "plain_mid": 3,  "currency": "AUD", "lang": "en", "browser_required": True},
    "BR":  {"mid": "A2Q3Y263D006CC",  "domain": "amazon.com.br",    "locale": "pt_BR",  "plain_mid": 3,  "currency": "BRL", "lang": "pt", "browser_required": True},
    "SG":  {"mid": "A19VAU5T5FV4E7",  "domain": "amazon.sg",       "locale": "en_SG",  "plain_mid": 3,  "currency": "SGD", "lang": "en", "browser_required": True},
    "TR":  {"mid": "A33GVJZQ3S32I2",  "domain": "amazon.tr",       "locale": "tr_TR",  "plain_mid": 3,  "currency": "TRY", "lang": "tr", "browser_required": True},
}

def get_market_config(market_code):
    """获取站点配置，支持 'US'/'DE'/'JP' 等简码或完整域名"""
    if market_code in AMAZON_MARKETS:
        return AMAZON_MARKETS[market_code]
    # Try matching by domain
    for code, config in AMAZON_MARKETS.items():
        if market_code in config['domain']:
            return config
    return AMAZON_MARKETS["US"]  # Default to US

# ============================================================
# 种子词自动翻译 (Google Translate + 内置词典)
# ============================================================

# 内置常见品类词典（优先级最高，覆盖googletrans可能的误翻）
BUILTIN_TRANSLATIONS = {
    "feather duster": {"de": "Staubwedel", "fr": "plumeau", "it": "spolverino", "es": "plumero",
                       "ja": "ハタキ/ほこり取り", "nl": "stofkwast", "pl": "pędzel do kurzu",
                       "sv": "dammvippa", "pt": "espanador", "ar": "منفضة ريش", "tr": "tüy süpürge"},
    "ostrich feather duster": {"de": "Staubwedel Straußenfedern", "fr": "plumeau plumes d'autruche",
                                "ja": "ハタキ ダチョウ", "es": "plumero de plumas de avestruz"},
    "fan": {"de": "Ventilator", "fr": "ventilateur", "it": "ventilatore", "es": "ventilador",
            "ja": "扇風機", "nl": "ventilator", "pl": "wentylator", "sv": "fläkt",
            "pt": "ventilador", "ar": "مروحة", "tr": "fan"},
    "vacuum cleaner": {"de": "Staubsauger", "fr": "aspirateur", "ja": "掃除機", "es": "aspiradora"},
    "power bank": {"de": "Powerbank", "fr": "batterie externe", "ja": "モバイルバッテリー", "es": "batería externa"},
    "phone case": {"de": "Handyhülle", "fr": "coque téléphone", "ja": "スマホケース", "es": "fundas móvil"},
    "LED light": {"de": "LED Licht", "fr": "lampe LED", "ja": "LEDライト", "es": "luz LED"},
    "water bottle": {"de": "Wasserflasche", "fr": "gourde", "ja": "水筒", "es": "botella de agua"},
    "backpack": {"de": "Rucksack", "fr": "sac à dos", "ja": "バックパック", "es": "mochila"},
    "hair dryer": {"de": "Haartrockner", "fr": "sèche-cheveux", "ja": "ヘアドライヤー", "es": "secador de pelo"},
}

# Google Translate 语言代码映射
GT_LANG_MAP = {"de": "de", "fr": "fr", "it": "it", "es": "es", "ja": "ja",
               "nl": "nl", "pl": "pl", "sv": "sv", "pt": "pt", "ar": "ar",
               "tr": "tr", "en": "en", "zh": "zh-cn", "ko": "ko", "th": "th"}

_translator_instance = None

def translate_seed(seed, market_code, verbose=False):
    """
    将英文种子词翻译为目标站点语言，返回所有候选词。
    四层优先级：内置词典（含同义词）→ Google Translate → 英文原词保底
    
    Returns:
        list[tuple]: [(候选词, 来源), ...]，如 [('Staubwedel','builtin'), ('feather duster','original')]
    """
    market = get_market_config(market_code)
    target_lang = market.get('lang', 'en')
    
    # 英语站点不需要翻译
    if target_lang == 'en':
        return [(seed, 'original')]
    
    candidates = []
    seen = set()
    
    # Layer 1: 内置词典（含同义词候选）
    seed_lower = seed.lower().strip()
    if seed_lower in BUILTIN_TRANSLATIONS:
        trans_dict = BUILTIN_TRANSLATIONS[seed_lower]
        if target_lang in trans_dict:
            val = trans_dict[target_lang]
            # 支持多个候选词（用 / 分隔）
            for word in val.split('/'):
                word = word.strip()
                if word and word.lower() not in seen:
                    seen.add(word.lower())
                    candidates.append((word, 'builtin'))
                    if verbose:
                        print(f"  📖 内置词典: '{seed}' → [{market_code}/{target_lang}] '{word}'")
    
    # Layer 2: Google Translate（补充同义词）
    if HAS_GOOGLETRANS:
        try:
            global _translator_instance
            if _translator_instance is None:
                _translator_instance = GTranslator()
            gt_lang = GT_LANG_MAP.get(target_lang, target_lang)
            result = _translator_instance.translate(seed, dest=gt_lang)
            translated = result.text.strip().rstrip('.')
            if translated.lower() not in seen:
                seen.add(translated.lower())
                candidates.append((translated, 'google'))
                if verbose:
                    print(f"  🌐 Google翻译: '{seed}' → [{market_code}/{target_lang}] '{translated}'")
        except Exception as e:
            if verbose:
                print(f"  ⚠️ Google翻译失败: {e}")
    
    # Layer 3: 英文原词保底（很多非英语站点用户也搜英文，尤其DE/JP）
    if seed.lower() not in seen:
        seen.add(seed.lower())
        candidates.append((seed, 'original'))
        if verbose:
            print(f"  🔄 保留英文原词: '{seed}' [{market_code}]（当地用户可能也搜此词）")
    
    return candidates if candidates else [(seed, 'original')]

DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Accept": "application/json",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": "https://www.amazon.com/",
    "Origin": "https://www.amazon.com",
}

def _make_headers(market_config):
    """根据站点配置动态生成请求头"""
    domain = market_config['domain']
    locale = market_config['locale']
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        "Accept": "application/json",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": locale,
        "Referer": f"https://www.{domain}/",
        "Origin": f"https://www.{domain}",
    }

def _make_params(prefix, session_id=None, request_id=None, market_code="US"):
    """构造 API 请求参数（支持多站点）"""
    if not session_id:
        session_id = f"135-{os.urandom(4).hex()}-{os.urandom(4).hex()}"
    if not request_id:
        request_id = os.urandom(8).hex()
    
    market = get_market_config(market_code)
    
    params_dict = {
        "limit": 11,
        "prefix": prefix,
        "suggestion-type": ["WIDGET", "KEYWORD"],
        "page-type": "Gateway",
        "alias": "aps",
        "site-variant": "desktop",
        "version": 3,
        "event": "onkeypress",
        "wc": "",
        "lop": market['locale'],
        "last-prefix": "\x00",
        "avg-ks-time": 5,
        "fb": 1,
        "predicted_text_accepted": "",
        "estoken": "",
        "session-id": session_id,
        "request-id": request_id,
        "mid": market['mid'],
        "client-info": "search-ui",
    }
    
    # Only add plain-mid if it's defined for this market
    if market.get('plain_mid') is not None:
        params_dict["plain-mid"] = market['plain_mid']
    
    return urllib.parse.urlencode(params_dict, doseq=True), market


# ============================================================
# 随机抖动 + 错误重试装饰器（优化版）
# ============================================================
def with_jitter(base_delay=0.5, jitter_ratio=0.6, max_retries=3, backoff_base=1.5):
    """
    请求间隔随机抖动 + 错误重试装饰器（方法二优化版）

    正常请求：base_delay + 随机抖动
    失败请求：指数退避 + 抖动，最多重试 max_retries 次

    可通过 kwargs 运行时覆盖：
        delay=0.8, jitter_ratio=0.4, max_retries=5
    """
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            delay = kwargs.pop('delay', base_delay)
            jitter = kwargs.pop('jitter_ratio', jitter_ratio)
            retries = kwargs.pop('max_retries', max_retries)
            verbose = kwargs.get('verbose', False)

            last_result = None
            last_error = None

            for attempt in range(retries + 1):
                if attempt == 0:
                    actual = delay + random.uniform(0, delay * jitter)
                else:
                    backoff = (backoff_base ** attempt) * delay
                    actual = backoff + random.uniform(0, backoff * 0.3)
                    if verbose:
                        print(f"  🔄 第{attempt}次重试，退避等待 {actual:.2f}s ...")

                if verbose and attempt == 0:
                    print(f"  ⏳ 抖动等待 {actual:.2f}s (base={delay:.2f}, jitter={jitter:.1f})")

                time.sleep(actual)

                try:
                    result = func(*args, **kwargs)
                except Exception as e:
                    last_error = str(e)
                    if verbose:
                        print(f"  ✗ 请求异常: {e}")
                    if attempt == retries:
                        return {
                            'keywords': [], 'widget_items': [],
                            'market': kwargs.get('market_code', 'US'),
                            'raw': None, 'error': last_error
                        }
                    continue

                has_error = bool(result.get('error'))
                is_empty = (not result.get('keywords') and not result.get('widget_items'))

                if has_error or (is_empty and attempt < retries):
                    last_result = result
                    last_error = result.get('error', 'empty response')
                    if verbose and has_error:
                        print(f"  ⚠ 返回错误，准备重试: {last_error}")
                    if attempt == retries:
                        return result
                    continue

                return result

            return last_result or {
                'keywords': [], 'widget_items': [],
                'market': kwargs.get('market_code', 'US'),
                'raw': None, 'error': last_error or 'max retries exceeded'
            }
        return wrapper
    return decorator


@with_jitter(base_delay=0.5, jitter_ratio=0.6, max_retries=3, backoff_base=1.5)
def fetch_suggestions(prefix, session_id=None, verbose=False, market_code="US"):
    """
    获取单个前缀的搜索建议词。

    Args:
        prefix: 搜索前缀
        session_id: Amazon session ID
        verbose: 是否打印详细日志
        market_code: 站点简码 (US/DE/JP/UK 等)

    注意：delay / jitter_ratio / max_retries 由装饰器 with_jitter 处理，
          调用时可传 delay=xx, jitter_ratio=xx, max_retries=xx 覆盖。

    Returns:
        dict: {
            'keywords': [...],
            'widget_items': [...],
            'market': str,
            'raw': dict or None,
            'error': str (仅失败时)
        }
    """
    market = get_market_config(market_code)

    if market.get('browser_required'):
        if verbose:
            print(f"  ⚠️ 站点 {market_code} ({market['domain']}) 需浏览器辅助，curl可能被拦截")

    params, market = _make_params(prefix, session_id, market_code=market_code)
    domain = market['domain']
    api_url = f"https://www.{domain}/suggestions"
    url = f"{api_url}?{params}"

    headers = _make_headers(market)

    if verbose:
        print(f"  → [{market_code}] 请求前缀: '{prefix}' (domain: {domain})")

    try:
        if HAS_REQUESTS:
            resp = requests.get(url, headers=headers, timeout=15)
            if resp.status_code in (429, 503, 502, 500):
                raise RuntimeError(f"HTTP {resp.status_code}")
            data = resp.json()
        else:
            from urllib import request as urllib_request
            req = urllib_request.Request(url, headers=headers)
            with urllib_request.urlopen(req, timeout=15) as response:
                data = json.loads(response.read().decode('utf-8'))
    except Exception as e:
        if verbose:
            print(f"  ✗ 请求失败: {e}")
        return {'keywords': [], 'widget_items': [], 'market': market_code, 'raw': None, 'error': str(e)}

    # 原来的 time.sleep 已移到装饰器 with_jitter 中

    keywords = []
    widget_items = []

    keywords = []
    widget_items = []
    
    suggestions = data.get('suggestions', [])
    for i, s in enumerate(suggestions):
        if s.get('suggType') == 'KeywordSuggestion':
            keywords.append({
                'keyword': s['value'],
                'rank': i + 1,
                'sugg_type': s.get('suggType', ''),
                'candidate_source': s.get('candidateSources', ''),
                'strategy_id': s.get('strategyId', ''),
            })
        elif s.get('suggType') == 'WidgetSuggestion':
            # Widget 包含子分类建议词
            widget = s
            widget_title = widget.get('metadata', {}).get('title', '')
            for item in widget.get('widgetItems', []):
                meta = item.get('metadata', {})
                text = meta.get('text', '')
                # 拼回完整关键词（从 search_url 中提取）
                link = meta.get('link_url', '')
                # link 格式: /s?k=clip+fan+for+bedroom&...
                # 提取 k= 参数值作为完整关键词
                full_kw = ''
                if link:
                    kw_match = urllib.parse.parse_qs(urllib.parse.urlparse(link).query).get('k', [])
                    if kw_match:
                        full_kw = kw_match[0].replace('+', ' ')
                
                widget_items.append({
                    'keyword': text,
                    'full_keyword': full_kw,
                    'image_url': meta.get('image_url', ''),
                    'search_url': link,
                    'widget_title': widget_title,
                })
    
    if verbose:
        print(f"  ✓ 返回: {len(keywords)} 条关键词 + {len(widget_items)} 条Widget子类")
    
    return {'keywords': keywords, 'widget_items': widget_items, 'raw': data}


# ============================================================
# 前缀变体生成
# ============================================================

# 默认精简模板（约 15 个，速度快，覆盖核心手动技巧）
SEED_PREFIX_TEMPLATES_CORE = [
    # 介词拓展（最高价值）
    "{seed} for",
    "{seed} with",
    "{seed} without",
    # 高频场景/人群
    "{seed} for women",
    "{seed} for men",
    "{seed} for travel",
    "{seed} for home",
    # 属性前置
    "best {seed}",
    "cheap {seed}",
    "casual {seed}",
    "long {seed}",
    "short {seed}",
    # 季节/场景
    "{seed} summer",
    "{seed} gift",
    # 词序倒换
    "{seed2}",
]

# 完整模板（约 50 个，覆盖全部手动技巧，速度慢，用 --full-templates 开启）
SEED_PREFIX_TEMPLATES_FULL = [
    # 介词拓展
    "{seed} for",
    "{seed} with",
    "{seed} without",
    "{seed} in",
    "{seed} on",
    "{seed} to",
    "{seed} from",
    # 场景/人群
    "{seed} for men",
    "{seed} for women",
    "{seed} for kids",
    "{seed} for baby",
    "{seed} for beginners",
    "{seed} for seniors",
    "{seed} for travel",
    "{seed} for home",
    "{seed} for office",
    "{seed} for bedroom",
    "{seed} for car",
    "{seed} for outdoor",
    "{seed} for indoor",
    "{seed} for apartment",
    # 材质/功能/痛点（部分品类适用）
    "quiet {seed}",
    "portable {seed}",
    "small {seed}",
    "large {seed}",
    "lightweight {seed}",
    "foldable {seed}",
    "adjustable {seed}",
    "durable {seed}",
    # 属性前置
    "best {seed}",
    "cheap {seed}",
    "top {seed}",
    "premium {seed}",
    "casual {seed}",
    "elegant {seed}",
    "long {seed}",
    "short {seed}",
    # 季节/节日
    "{seed} christmas",
    "{seed} halloween",
    "{seed} summer",
    "{seed} winter",
    "{seed} birthday",
    "{seed} gift",
    # 疑问词
    "how to {seed}",
    "what is {seed}",
    "best {seed} for",
    # 词序倒换
    "{seed2}",
]

# 默认使用精简版
SEED_PREFIX_TEMPLATES = SEED_PREFIX_TEMPLATES_CORE

# 多语言修饰词模板（按站点语言适配）
# 格式: {lang: [前缀模板列表]}
# 对于非英语站点，修饰词也要用当地语言
LOCALIZED_PREFIX_TEMPLATES = {
    "de": [
        "{seed} für",
        "{seed} mit",
        "{seed} ohne",
        "kleiner {seed}",
        "großer {seed}",
        "bester {seed}",
        "günstiger {seed}",
        "{seed} für bedroom",  # 一些德用户也会用英文
        "{seed} für zu Hause",
        "{seed} für Auto",
        "{seed} lang",
        "{seed} ausziehbar",
    ],
    "fr": [
        "{seed} pour",
        "{seed} avec",
        "{seed} sans",
        "petit {seed}",
        "grand {seed}",
        "meilleur {seed}",
        "{seed} pas cher",
        "{seed} pour la maison",
        "{seed} pour voiture",
        "{seed} long",
        "{seed} extensible",
    ],
    "it": [
        "{seed} per",
        "{seed} con",
        "piccolo {seed}",
        "grande {seed}",
        "migliore {seed}",
        "{seed} economico",
        "{seed} per casa",
        "{seed} lungo",
    ],
    "es": [
        "{seed} para",
        "{seed} con",
        "pequeño {seed}",
        "grande {seed}",
        "mejor {seed}",
        "{seed} barato",
        "{seed} para casa",
        "{seed} para coche",
        "{seed} largo",
        "{seed} extensible",
    ],
    "ja": [
        "{seed} おすすめ",
        "{seed} 静音",
        "{seed} 小型",
        "{seed} 大型",
        "{seed} 長い",
        "{seed} 伸縮",
        "{seed} 車用",
        "{seed} 家庭用",
        "{seed} オフィス用",
        "{seed} 充電式",
    ],
    "nl": [
        "{seed} voor",
        "{seed} met",
        "kleine {seed}",
        "grote {seed}",
        "beste {seed}",
        "goedkope {seed}",
        "{seed} voor huis",
        "{seed} lang",
    ],
    "pl": [
        "{seed} do",
        "{seed} z",
        "mały {seed}",
        "duży {seed}",
        "najlepszy {seed}",
        "tani {seed}",
        "{seed} do domu",
        "{seed} do samochodu",
    ],
    "sv": [
        "{seed} för",
        "{seed} med",
        "liten {seed}",
        "stor {seed}",
        "bäst {seed}",
        "billig {seed}",
        "{seed} för hemmet",
        "{seed} lång",
    ],
    "pt": [
        "{seed} para",
        "{seed} com",
        "pequeno {seed}",
        "grande {seed}",
        "melhor {seed}",
        "{seed} barato",
        "{seed} para casa",
        "{seed} para carro",
    ],
    "ar": [
        "{seed} لـ",
        "{seed} مع",
        "أفضل {seed}",
        "{seed} رخيص",
        "{seed} صغير",
        "{seed} كبير",
        "{seed} للمنزل",
        "{seed} للسيارة",
    ],
    "tr": [
        "{seed} için",
        "{seed} ile",
        "küçük {seed}",
        "büyük {seed}",
        "en iyi {seed}",
        "ucuz {seed}",
        "{seed} ev için",
        "{seed} araba için",
    ],
}

def generate_prefixes(seed, mode='expand', market_code="US", full_templates=False):
    """根据种子词和模式生成前缀列表（自动适配当地语言修饰词）
    
    对应手动下拉框技巧：
    - expand: 介词/场景/人群/材质/功能/季节/疑问词 模板扩展
    - gap:    空格间隙插入法（多词种子词中间插入热门修饰词）
    
    full_templates=True 时使用完整模板（约50个），默认只用精简核心模板（约15个）
    """
    prefixes = []
    market = get_market_config(market_code)
    lang = market.get('lang', 'en')
    
    # 基础：种子词本身
    prefixes.append(seed)
    
    if mode in ('expand', 'deep'):
        # 选择模板：完整版 or 精简版
        if lang != 'en' and lang in LOCALIZED_PREFIX_TEMPLATES:
            templates = LOCALIZED_PREFIX_TEMPLATES[lang]
        else:
            templates = SEED_PREFIX_TEMPLATES_FULL if full_templates else SEED_PREFIX_TEMPLATES_CORE
        
        for template in templates:
            if template == "{seed2}":
                words = seed.split()
                if len(words) >= 2:
                    reversed_seed = ' '.join(words[::-1])
                    prefixes.append(reversed_seed)
                    prefixes.append(f"{words[-1]} {words[0]}")
            else:
                prefixes.append(template.format(seed=seed))
    
    # 空格间隙插入法：只在 gap 模式启用（不再默认混入 expand，避免前缀爆炸）
    if mode == 'gap':
        words = seed.split()
        if len(words) >= 2:
            # 精简版修饰词（约12个），避免过多请求
            common_mid_modifiers = [
                "casual", "elegant", "floral", "long", "short", "maxi",
                "cotton", "linen", "plus size", "sleeveless", "boho", "vintage"
            ]
            if full_templates:
                common_mid_modifiers = [
                    "casual", "elegant", "floral", "long", "short", "maxi", "mini",
                    "cotton", "linen", "silk", "plus size", "sleeveless", "boho",
                    "vintage", "beach", "party", "wedding", "work", "summer", "flowy"
                ]
            for mid in common_mid_modifiers:
                inserted = f"{words[0]} {mid} {' '.join(words[1:])}"
                prefixes.append(inserted)
                if len(words) > 2:
                    inserted2 = f"{' '.join(words[:-1])} {mid} {words[-1]}"
                    prefixes.append(inserted2)

    # 去重
    unique = []
    seen = set()
    for p in prefixes:
        if p.lower() not in seen:
            seen.add(p.lower())
            unique.append(p)
    
    return unique


# ============================================================
# 各模式执行逻辑
# ============================================================

def mode_expand(seed, rounds=3, delay=0.5, jitter_ratio=0.6, verbose=True, market_code="US", full_templates=False):
    """模式2：批量扩展（默认精简模板，--full-templates 开启完整版）"""
    all_keywords = {}  # keyword -> info dict
    all_widget_items = []
    prefixes_tried = 0
    
    prefixes = generate_prefixes(seed, mode='expand', market_code=market_code, full_templates=full_templates)
    
    if verbose:
        print(f"\n🔍 种子词: '{seed}'")
        print(f"   前缀数: {len(prefixes)}")
        print(f"   执行轮次: {rounds}")
    
    for round_num in range(1, rounds + 1):
        if verbose:
            print(f"\n--- 第 {round_num} 轮 ---")
        
        for prefix in prefixes:
            result = fetch_suggestions(prefix, delay=delay, jitter_ratio=jitter_ratio, verbose=verbose, market_code=market_code)
            prefixes_tried += 1
            
            for kw_info in result['keywords']:
                kw = kw_info['keyword']
                if kw.lower() not in {k.lower() for k in all_keywords}:
                    kw_info['source'] = 'autocomplete'
                    kw_info['prefix'] = prefix
                    all_keywords[kw] = kw_info
            
            for wi in result['widget_items']:
                full_kw = wi['full_keyword']
                if full_kw and full_kw.lower() not in {k.lower() for k in all_keywords}:
                    all_keywords[full_kw] = {
                        'keyword': full_kw,
                        'source': 'widget',
                        'prefix': prefix,
                        'rank': 0,
                        'sugg_type': 'WidgetSuggestion',
                        'candidate_source': '',
                    }
                all_widget_items.append(wi)
        
        # 如果还有后续轮次，可以从结果中生成新前缀
        if round_num < rounds:
            # 取前5个高频词作为新前缀
            top_words = list(all_keywords.keys())[:5]
            new_prefixes = []
            for w in top_words:
                if w != seed and len(w.split()) <= 4:
                    new_prefixes.append(w)
            prefixes = new_prefixes
            if verbose:
                print(f"\n   → 新增前缀: {new_prefixes}")
    
    # 分类统计
    question_style = [k for k in all_keywords if any(k.lower().startswith(w) for w in ['what','how','which','is','are','can','do','why','best','should'])]
    rufus_long = [k for k in all_keywords if len(k.split()) > 6]
    
    output = {
        'seed': seed,
        'mode': 'expand',
        'total_keywords': len(all_keywords),
        'total_widget_items': len(all_widget_items),
        'keywords': list(all_keywords.values()),
        'widget_items': all_widget_items,
        'stats': {
            'rounds_executed': rounds,
            'prefixes_tried': prefixes_tried,
            'raw_suggestions': len(all_keywords),
            'question_style': len(question_style),
            'rufus_style_long': len(rufus_long),
        }
    }
    return output


def mode_az(seed, delay=0.3, verbose=True, market_code="US"):
    """模式3：A-Z 字母扫描扩展

    在种子词后依次输入空格+a/b/c.../z，记录每个字母组合的长尾词。
    共 26 个前缀，每个返回 ~10 条建议词，去重后通常 150-260 条。
    """
    import string
    all_keywords = {}
    all_widget_items = []
    prefixes_tried = 0

    prefixes = [f"{seed} {ch}" for ch in string.ascii_lowercase]

    if verbose:
        print(f"\n🔤 A-Z 字母扫描种子词: '{seed}'")
        print(f"   前缀数: {len(prefixes)} (a-z)")

    for prefix in prefixes:
        result = fetch_suggestions(prefix, delay=delay, verbose=verbose, market_code=market_code)
        prefixes_tried += 1

        for kw_info in result['keywords']:
            kw = kw_info['keyword']
            if kw.lower() not in {k.lower() for k in all_keywords}:
                kw_info['source'] = 'az_scan'
                kw_info['prefix'] = prefix
                all_keywords[kw] = kw_info

        for wi in result['widget_items']:
            full_kw = wi['full_keyword']
            if full_kw and full_kw.lower() not in {k.lower() for k in all_keywords}:
                all_keywords[full_kw] = {
                    'keyword': full_kw,
                    'source': 'widget',
                    'prefix': prefix,
                    'rank': 0,
                    'sugg_type': 'WidgetSuggestion',
                    'candidate_source': '',
                }
            all_widget_items.append(wi)

    output = {
        'seed': seed,
        'mode': 'az',
        'total_keywords': len(all_keywords),
        'total_widget_items': len(all_widget_items),
        'keywords': list(all_keywords.values()),
        'widget_items': all_widget_items,
        'stats': {
            'prefixes_tried': prefixes_tried,
            'raw_suggestions': len(all_keywords),
        }
    }
    return output


def mode_az_prefix(seed, delay=0.3, verbose=True, market_code="US"):
    """A-Z 前缀法（对应手动技巧：字母前缀法）

    把字母放在种子词前面：a {seed}、b {seed} ... z {seed}
    适合挖属性词、品牌词、场景前置词。
    """
    import string
    all_keywords = {}
    all_widget_items = []
    prefixes_tried = 0

    prefixes = [f"{ch} {seed}" for ch in string.ascii_lowercase]

    if verbose:
        print(f"\n🔤 A-Z 前缀扫描种子词: '{seed}'")
        print(f"   前缀数: {len(prefixes)} (a-z + seed)")

    for prefix in prefixes:
        result = fetch_suggestions(prefix, delay=delay, verbose=verbose, market_code=market_code)
        prefixes_tried += 1

        for kw_info in result['keywords']:
            kw = kw_info['keyword']
            if kw.lower() not in {k.lower() for k in all_keywords}:
                kw_info['source'] = 'az_prefix'
                kw_info['prefix'] = prefix
                all_keywords[kw] = kw_info

        for wi in result['widget_items']:
            full_kw = wi['full_keyword']
            if full_kw and full_kw.lower() not in {k.lower() for k in all_keywords}:
                all_keywords[full_kw] = {
                    'keyword': full_kw,
                    'source': 'widget',
                    'prefix': prefix,
                    'rank': 0,
                    'sugg_type': 'WidgetSuggestion',
                    'candidate_source': '',
                }
            all_widget_items.append(wi)

    output = {
        'seed': seed,
        'mode': 'az_prefix',
        'total_keywords': len(all_keywords),
        'total_widget_items': len(all_widget_items),
        'keywords': list(all_keywords.values()),
        'widget_items': all_widget_items,
        'stats': {
            'prefixes_tried': prefixes_tried,
            'raw_suggestions': len(all_keywords),
        }
    }
    return output


def mode_numbers(seed, delay=0.3, verbose=True, market_code="US"):
    """数字拓展法（对应手动技巧：数字顺推法）

    在种子词后依次加常见数字/单位，挖出尺寸、数量、包装规格等长尾。
    例：dog toy 1 / dog toy 3 pack / dog toy 6 inch
    """
    all_keywords = {}
    all_widget_items = []
    prefixes_tried = 0

    # 常见数字 + 单位组合
    number_suffixes = [
        "1", "2", "3", "4", "5", "6", "7", "8", "9", "10",
        "12", "24", "50", "100",
        "1 pack", "2 pack", "3 pack", "4 pack", "6 pack", "12 pack",
        "1 set", "2 set", "3 set",
        "6 inch", "8 inch", "10 inch", "12 inch",
        "mm", "cm", "oz", "lb",
        "small", "medium", "large", "xl", "xxl",
    ]
    prefixes = [f"{seed} {suf}" for suf in number_suffixes]

    if verbose:
        print(f"\n🔢 数字/规格拓展种子词: '{seed}'")
        print(f"   前缀数: {len(prefixes)}")

    for prefix in prefixes:
        result = fetch_suggestions(prefix, delay=delay, verbose=verbose, market_code=market_code)
        prefixes_tried += 1

        for kw_info in result['keywords']:
            kw = kw_info['keyword']
            if kw.lower() not in {k.lower() for k in all_keywords}:
                kw_info['source'] = 'numbers'
                kw_info['prefix'] = prefix
                all_keywords[kw] = kw_info

        for wi in result['widget_items']:
            full_kw = wi['full_keyword']
            if full_kw and full_kw.lower() not in {k.lower() for k in all_keywords}:
                all_keywords[full_kw] = {
                    'keyword': full_kw,
                    'source': 'widget',
                    'prefix': prefix,
                    'rank': 0,
                    'sugg_type': 'WidgetSuggestion',
                    'candidate_source': '',
                }
            all_widget_items.append(wi)

    output = {
        'seed': seed,
        'mode': 'numbers',
        'total_keywords': len(all_keywords),
        'total_widget_items': len(all_widget_items),
        'keywords': list(all_keywords.values()),
        'widget_items': all_widget_items,
        'stats': {
            'prefixes_tried': prefixes_tried,
            'raw_suggestions': len(all_keywords),
        }
    }
    return output


def mode_gap(seed, delay=0.5, jitter_ratio=0.6, verbose=True, market_code="US", full_templates=False):
    """空格间隙插入法（对应手动技巧：空格间隙插入法）

    对多词种子词，在词中间插入热门修饰词，模拟“把光标移到两个词中间敲空格”。
    例：dog toy → dog chew toy / dog plush toy / dog rope toy
    """
    all_keywords = {}
    all_widget_items = []
    prefixes_tried = 0

    prefixes = generate_prefixes(seed, mode='gap', market_code=market_code, full_templates=full_templates)

    if verbose:
        print(f"\n↔️ 空格间隙插入法种子词: '{seed}'")
        print(f"   前缀数: {len(prefixes)}")

    for prefix in prefixes:
        result = fetch_suggestions(prefix, delay=delay, verbose=verbose, market_code=market_code)
        prefixes_tried += 1

        for kw_info in result['keywords']:
            kw = kw_info['keyword']
            if kw.lower() not in {k.lower() for k in all_keywords}:
                kw_info['source'] = 'gap'
                kw_info['prefix'] = prefix
                all_keywords[kw] = kw_info

        for wi in result['widget_items']:
            full_kw = wi['full_keyword']
            if full_kw and full_kw.lower() not in {k.lower() for k in all_keywords}:
                all_keywords[full_kw] = {
                    'keyword': full_kw,
                    'source': 'widget',
                    'prefix': prefix,
                    'rank': 0,
                    'sugg_type': 'WidgetSuggestion',
                    'candidate_source': '',
                }
            all_widget_items.append(wi)

    output = {
        'seed': seed,
        'mode': 'gap',
        'total_keywords': len(all_keywords),
        'total_widget_items': len(all_widget_items),
        'keywords': list(all_keywords.values()),
        'widget_items': all_widget_items,
        'stats': {
            'prefixes_tried': prefixes_tried,
            'raw_suggestions': len(all_keywords),
        }
    }
    return output


def mode_reverse(seed, top_n=100, delay=0.3, verbose=True, market_code="US"):
    """逆向检索扩展（对应手动技巧：滚雪球反向扩展 + 前置修饰）

    第一步：对种子词做 A-Z 字母扫描，拿到后缀衍生词。
    第二步：从 A-Z 结果中提取高频单词，用 "{word} {seed}" 作为前缀再次请求。
    模拟手动把光标移到最前面看系统推荐的前置修饰词。
    """
    from collections import Counter

    # ---------- Step 1: A-Z 扫描 ----------
    az_result = mode_az(seed, delay=delay, verbose=verbose, market_code=market_code)
    az_keywords = az_result['keywords']

    # 把 A-Z 扫描结果先收进来
    all_keywords = {}
    all_widget_items = list(az_result.get('widget_items', []))
    for kw_info in az_keywords:
        kw_info['source'] = 'az_scan'
        all_keywords[kw_info['keyword']] = kw_info

    # ---------- Step 2: 提取高频词 ----------
    seed_words = set(seed.lower().split())
    word_counter = Counter()
    for kw_info in az_keywords:
        for word in kw_info['keyword'].lower().split():
            if word not in seed_words and len(word) >= 2:
                word_counter[word] += 1

    top_words = [w for w, _ in word_counter.most_common(top_n)]

    if verbose:
        print(f"\n🔁 逆向检索种子词: '{seed}'")
        print(f"   A-Z 扫描获得 {len(az_keywords)} 条关键词")
        print(f"   提取高频词 Top-{len(top_words)}: {top_words}")

    # ---------- Step 3: 用高频词做前置检索 ----------
    prefixes_tried = az_result['stats']['prefixes_tried']

    for word in top_words:
        prefix = f"{word} {seed}"
        result = fetch_suggestions(prefix, delay=delay, verbose=verbose, market_code=market_code)
        prefixes_tried += 1

        for kw_info in result['keywords']:
            kw = kw_info['keyword']
            if kw.lower() not in {k.lower() for k in all_keywords}:
                kw_info['source'] = 'reverse'
                kw_info['prefix'] = prefix
                all_keywords[kw] = kw_info

        for wi in result['widget_items']:
            full_kw = wi['full_keyword']
            if full_kw and full_kw.lower() not in {k.lower() for k in all_keywords}:
                all_keywords[full_kw] = {
                    'keyword': full_kw,
                    'source': 'widget',
                    'prefix': prefix,
                    'rank': 0,
                    'sugg_type': 'WidgetSuggestion',
                    'candidate_source': '',
                }
            all_widget_items.append(wi)

    output = {
        'seed': seed,
        'mode': 'reverse',
        'total_keywords': len(all_keywords),
        'total_widget_items': len(all_widget_items),
        'keywords': list(all_keywords.values()),
        'widget_items': all_widget_items,
        'stats': {
            'az_scan_keywords': len(az_keywords),
            'reverse_prefixes': len(top_words),
            'prefixes_tried': prefixes_tried,
            'raw_suggestions': len(all_keywords),
            'top_words': top_words,
        }
    }
    return output


def mode_deep(seed, depth=2, top_n=5, delay=0.3, verbose=True, market_code="US"):
    """模式4：深度递归扩展（单次查询 + 逐轮递归）

    第一轮：用种子词本身做单次查询，拿到 ~10 条建议词。
    第二轮：取第一轮结果的 Top-N 高频词，逐一作为新前缀再次查询。
    第三轮及以后：对上一轮结果继续取 Top-N 递归扩展。
    共 depth 轮，每轮 top_n 个前缀。
    """
    all_keywords = {}
    all_widget_items = []
    prefixes_tried = 0

    if verbose:
        print(f"\n🕳️ 深度递归扩展种子词: '{seed}'")
        print(f"   递归深度: {depth} 轮")
        print(f"   每轮取Top-N: {top_n}")

    # 第 1 轮：单次查询种子词
    if verbose:
        print(f"\n--- 第 1 轮 (种子词查询) ---")
        print(f"   前缀: ['{seed}']")

    result = fetch_suggestions(seed, delay=delay, verbose=verbose, market_code=market_code)
    prefixes_tried += 1

    for kw_info in result['keywords']:
        kw = kw_info['keyword']
        if kw.lower() not in {k.lower() for k in all_keywords}:
            kw_info['source'] = 'autocomplete'
            kw_info['prefix'] = seed
            kw_info['depth'] = 1
            all_keywords[kw] = kw_info

    for wi in result['widget_items']:
        full_kw = wi['full_keyword']
        if full_kw and full_kw.lower() not in {k.lower() for k in all_keywords}:
            all_keywords[full_kw] = {
                'keyword': full_kw,
                'source': 'widget',
                'prefix': seed,
                'rank': 0,
                'sugg_type': 'WidgetSuggestion',
                'candidate_source': '',
                'depth': 1,
            }
        all_widget_items.append(wi)

    # 第 2~depth 轮：递归扩展
    for d in range(2, depth + 1):
        # 取上一轮 depth=d-1 中 Top-N 关键词作为新前缀
        candidates = [k for k, info in all_keywords.items()
                      if info.get('depth') == d - 1
                      and len(k.split()) <= 5
                      and k.lower() != seed.lower()]
        next_prefixes = candidates[:top_n]

        if not next_prefixes:
            if verbose:
                print(f"\n   第{d}轮: 无更多前缀可用，停止递归")
            break

        if verbose:
            print(f"\n--- 第 {d} 轮 (递归深度={d}) ---")
            print(f"   新前缀 ({len(next_prefixes)}): {next_prefixes}")

        for prefix in next_prefixes:
            result = fetch_suggestions(prefix, delay=delay, verbose=verbose, market_code=market_code)
            prefixes_tried += 1

            for kw_info in result['keywords']:
                kw = kw_info['keyword']
                if kw.lower() not in {k.lower() for k in all_keywords}:
                    kw_info['source'] = 'autocomplete_deep'
                    kw_info['prefix'] = prefix
                    kw_info['depth'] = d
                    all_keywords[kw] = kw_info

            for wi in result['widget_items']:
                full_kw = wi['full_keyword']
                if full_kw and full_kw.lower() not in {k.lower() for k in all_keywords}:
                    all_keywords[full_kw] = {
                        'keyword': full_kw,
                        'source': 'widget',
                        'prefix': prefix,
                        'rank': 0,
                        'sugg_type': 'WidgetSuggestion',
                        'candidate_source': '',
                        'depth': d,
                    }
                all_widget_items.append(wi)

    # 统计
    question_style = [k for k in all_keywords if any(k.lower().startswith(w) for w in ['what','how','which','is','are','can','do','why','best','should'])]

    output = {
        'seed': seed,
        'mode': 'deep',
        'total_keywords': len(all_keywords),
        'total_widget_items': len(all_widget_items),
        'keywords': list(all_keywords.values()),
        'widget_items': all_widget_items,
        'stats': {
            'depth': depth,
            'prefixes_tried': prefixes_tried,
            'raw_suggestions': len(all_keywords),
            'question_style': len(question_style),
        }
    }
    return output


# ============================================================
# 输出处理
# ============================================================

def save_json(data, path):
    """保存为JSON"""
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\n✅ JSON 已保存: {path}")


def save_csv(data, path):
    """保存为CSV"""
    keywords = data.get('keywords', [])
    widget_items = data.get('widget_items', [])
    
    with open(path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        
        # Keywords section
        writer.writerow(['=== KEYWORDS ==='])
        writer.writerow(['keyword', 'source', 'prefix', 'rank', 'sugg_type', 'candidate_source', 'depth'])
        for kw in keywords:
            writer.writerow([
                kw.get('keyword', ''),
                kw.get('source', ''),
                kw.get('prefix', ''),
                kw.get('rank', ''),
                kw.get('sugg_type', ''),
                kw.get('candidate_source', ''),
                kw.get('depth', ''),
            ])
        
        # Widget section
        writer.writerow([])
        writer.writerow(['=== WIDGET ITEMS ==='])
        writer.writerow(['keyword', 'full_keyword', 'widget_title', 'image_url', 'search_url'])
        for wi in widget_items:
            writer.writerow([
                wi.get('keyword', ''),
                wi.get('full_keyword', ''),
                wi.get('widget_title', ''),
                wi.get('image_url', ''),
                wi.get('search_url', ''),
            ])
    
    print(f"\n✅ CSV 已保存: {path}")


def save_xlsx(data, path):
    """保存为 Excel (xlsx) — 多 Sheet 结构化输出

    Sheet 结构:
      1. 摘要          — 挖掘统计信息
      2. 关键词        — 所有关键词明细
      3. Widget分类词   — Widget 分类卡片词
      4. 问句式关键词   — 问句风格搜索词（如有）
    """
    if not HAS_OPENPYXL:
        print(f"\n❌ 需要 openpyxl 库: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    def style_header(ws, col_count):
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.freeze_panes = 'A2'

    def style_data_cells(ws, row_count, col_count):
        for row_idx in range(2, row_count + 2):
            for col_idx in range(1, col_count + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = cell_align
                cell.border = thin_border

    def auto_width(ws, col_count, min_w=10, max_w=60):
        for col_idx in range(1, col_count + 1):
            max_len = min_w
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    # CJK chars take more width
                    display_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
                    if display_len > max_len:
                        max_len = display_len
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_w)

    # ---- Sheet 1: 摘要 ----
    ws_summary = wb.active
    ws_summary.title = '摘要'

    summary_rows = [
        ['种子词', data.get('seed', '')],
        ['模式', data.get('mode', '')],
        ['站点', data.get('market', 'US')],
        ['关键词总数', data.get('total_keywords', 0)],
        ['Widget分类词数', data.get('total_widget_items', 0)],
        ['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ]

    stats = data.get('stats', {})
    if stats:
        summary_rows.append([])
        summary_rows.append(['--- 模式统计 ---', ''])
        for k, v in stats.items():
            if isinstance(v, list):
                summary_rows.append([k, ', '.join(str(item) for item in v)])
            else:
                summary_rows.append([k, v])

    for row_idx, row_data in enumerate(summary_rows, 1):
        if len(row_data) == 2:
            label, value = row_data
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)
        elif len(row_data) == 1:
            ws_summary.cell(row=row_idx, column=1, value=row_data[0]).font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 22
    ws_summary.column_dimensions['B'].width = 60

    # ---- Sheet 2: 关键词 ----
    ws_kw = wb.create_sheet('关键词')
    kw_headers = ['关键词', '来源', '前缀', '排名', '建议类型', '候选来源', '递归深度']
    for col_idx, h in enumerate(kw_headers, 1):
        ws_kw.cell(row=1, column=col_idx, value=h)

    keywords = data.get('keywords', [])
    for row_idx, kw in enumerate(keywords, 2):
        ws_kw.cell(row=row_idx, column=1, value=kw.get('keyword', ''))
        ws_kw.cell(row=row_idx, column=2, value=kw.get('source', ''))
        ws_kw.cell(row=row_idx, column=3, value=kw.get('prefix', ''))
        ws_kw.cell(row=row_idx, column=4, value=kw.get('rank', ''))
        ws_kw.cell(row=row_idx, column=5, value=kw.get('sugg_type', ''))
        ws_kw.cell(row=row_idx, column=6, value=kw.get('candidate_source', ''))
        ws_kw.cell(row=row_idx, column=7, value=kw.get('depth', ''))

    style_header(ws_kw, len(kw_headers))
    style_data_cells(ws_kw, len(keywords), len(kw_headers))
    auto_width(ws_kw, len(kw_headers))

    # ---- Sheet 3: Widget分类词 ----
    ws_widget = wb.create_sheet('Widget分类词')
    widget_headers = ['分类标签', '完整关键词', 'Widget标题', '图片URL', '搜索URL']
    for col_idx, h in enumerate(widget_headers, 1):
        ws_widget.cell(row=1, column=col_idx, value=h)

    widget_items = data.get('widget_items', [])
    for row_idx, wi in enumerate(widget_items, 2):
        ws_widget.cell(row=row_idx, column=1, value=wi.get('keyword', ''))
        ws_widget.cell(row=row_idx, column=2, value=wi.get('full_keyword', ''))
        ws_widget.cell(row=row_idx, column=3, value=wi.get('widget_title', ''))
        ws_widget.cell(row=row_idx, column=4, value=wi.get('image_url', ''))
        ws_widget.cell(row=row_idx, column=5, value=wi.get('search_url', ''))

    style_header(ws_widget, len(widget_headers))
    style_data_cells(ws_widget, len(widget_items), len(widget_headers))
    auto_width(ws_widget, len(widget_headers))

    # ---- Sheet 4: 问句式关键词 (如有) ----
    question_kws = [kw for kw in keywords
                    if any(kw.get('keyword', '').lower().startswith(w)
                           for w in ['what', 'how', 'which', 'is', 'are', 'can', 'do', 'why', 'best', 'should'])]
    if question_kws:
        ws_q = wb.create_sheet('问句式关键词')
        q_headers = ['关键词', '来源', '前缀', '排名']
        for col_idx, h in enumerate(q_headers, 1):
            ws_q.cell(row=1, column=col_idx, value=h)
        for row_idx, kw in enumerate(question_kws, 2):
            ws_q.cell(row=row_idx, column=1, value=kw.get('keyword', ''))
            ws_q.cell(row=row_idx, column=2, value=kw.get('source', ''))
            ws_q.cell(row=row_idx, column=3, value=kw.get('prefix', ''))
            ws_q.cell(row=row_idx, column=4, value=kw.get('rank', ''))
        style_header(ws_q, len(q_headers))
        style_data_cells(ws_q, len(question_kws), len(q_headers))
        auto_width(ws_q, len(q_headers))

    wb.save(path)
    print(f"\n✅ Excel 已保存: {path}")
    print(f"   Sheet: 摘要 / 关键词 ({len(keywords)} 条) / Widget分类词 ({len(widget_items)} 条)" +
          (f" / 问句式关键词 ({len(question_kws)} 条)" if question_kws else ""))


def save_sqlite(data, db_path, seed=None, mode=None):
    """保存到SQLite"""
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS suggestion_keywords (
        keyword TEXT,
        source TEXT,
        prefix TEXT,
        rank INT,
        sugg_type TEXT,
        candidate_source TEXT,
        seed TEXT,
        mode TEXT,
        depth INT,
        market TEXT,
        translated_seed TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        PRIMARY KEY (keyword, market)
    )''')
    
    # 迁移：为旧表添加 market/translated_seed 列
    try:
        c.execute("ALTER TABLE suggestion_keywords ADD COLUMN market TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE suggestion_keywords ADD COLUMN translated_seed TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE suggestion_widget_items ADD COLUMN market TEXT")
    except sqlite3.OperationalError:
        pass
    
    c.execute('''CREATE TABLE IF NOT EXISTS suggestion_widget_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        keyword TEXT,
        full_keyword TEXT,
        widget_title TEXT,
        image_url TEXT,
        search_url TEXT,
        seed TEXT,
        market TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    market_code = data.get('market', 'US')
    translated_seed = data.get('translated_seed', seed)
    
    for kw in data.get('keywords', []):
        c.execute('''INSERT OR REPLACE INTO suggestion_keywords 
            (keyword, source, prefix, rank, sugg_type, candidate_source, seed, mode, depth, market, translated_seed)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (kw.get('keyword'), kw.get('source'), kw.get('prefix'), 
             kw.get('rank'), kw.get('sugg_type'), kw.get('candidate_source'),
             seed or data.get('seed'), mode or data.get('mode'), kw.get('depth'),
             market_code, translated_seed))
    
    for wi in data.get('widget_items', []):
        c.execute('''INSERT INTO suggestion_widget_items 
            (keyword, full_keyword, widget_title, image_url, search_url, seed, market)
            VALUES (?, ?, ?, ?, ?, ?, ?)''',
            (wi.get('keyword'), wi.get('full_keyword'), wi.get('widget_title'),
             wi.get('image_url'), wi.get('search_url'), seed or data.get('seed'), market_code))
    
    conn.commit()
    conn.close()
    print(f"\n✅ SQLite 已保存: {db_path}")
    print(f"   suggestion_keywords: {len(data.get('keywords', []))} 条")
    print(f"   suggestion_widget_items: {len(data.get('widget_items', []))} 条")




def widget_deep_expand(main_result, original_seed, delay=0.3, verbose=True, market_code="US", max_widget_seeds=10):
    """把 Widget 卡片子分类标签作为修饰词拼接到原种子词上做二次扩展。"""
    widget_items = main_result.get('widget_items', [])
    if not widget_items:
        return main_result

    widget_labels = []
    seen_labels = set()
    for wi in widget_items:
        label = wi.get('keyword', '').strip()
        if label and label.lower() not in seen_labels:
            seen_labels.add(label.lower())
            widget_labels.append(label)

    if not widget_labels:
        return main_result

    widget_labels = widget_labels[:max_widget_seeds]

    if verbose:
        print(f"\n🔮 Widget 二次扩展：{len(widget_labels)} 个子分类标签 → {widget_labels}")

    new_prefixes = [f"{original_seed} {label}" for label in widget_labels]

    all_keywords = {}
    for k in main_result.get('keywords', []):
        all_keywords[k['keyword'].lower()] = k
    all_widget_items = list(main_result.get('widget_items', []))
    prefixes_tried = 0
    new_count = 0

    for prefix in new_prefixes:
        result = fetch_suggestions(prefix, delay=delay, verbose=verbose, market_code=market_code)
        prefixes_tried += 1

        for kw_info in result['keywords']:
            kw = kw_info['keyword']
            if kw.lower() not in all_keywords:
                kw_info['source'] = 'widget_deep'
                kw_info['prefix'] = prefix
                all_keywords[kw.lower()] = kw_info
                new_count += 1

        for wi in result['widget_items']:
            full_kw = wi.get('full_keyword', '')
            if full_kw and full_kw.lower() not in all_keywords:
                all_keywords[full_kw.lower()] = {
                    'keyword': full_kw,
                    'source': 'widget',
                    'prefix': prefix,
                    'rank': 0,
                    'sugg_type': 'WidgetSuggestion',
                    'candidate_source': '',
                }
                new_count += 1
            all_widget_items.append(wi)

    main_result['keywords'] = list(all_keywords.values())
    main_result['widget_items'] = all_widget_items
    main_result['total_keywords'] = len(all_keywords)
    main_result['total_widget_items'] = len(all_widget_items)

    stats = main_result.get('stats', {})
    stats['widget_deep_prefixes'] = prefixes_tried
    stats['widget_deep_labels'] = widget_labels
    stats['widget_deep_new'] = new_count
    main_result['stats'] = stats

    if verbose:
        print(f"   Widget 二次扩展完成：新增 {new_count} 条关键词，总计 {len(all_keywords)} 条")

    return main_result


def print_summary(data):
    """打印结果摘要"""
    print(f"\n{'='*60}")
    print(f"📊 搜索建议词挖掘结果")
    print(f"{'='*60}")
    print(f"  种子词: {data['seed']}")
    print(f"  模式: {data['mode']}")
    
    stats = data.get('stats', {})
    if stats:
        for k, v in stats.items():
            print(f"  {k}: {v}")
    
    print(f"\n  关键词总数: {data['total_keywords']}")
    print(f"  Widget子类: {data['total_widget_items']}")
    
    # 分类展示
    keywords = data.get('keywords', [])
    question_style = [k for k in keywords if any(k['keyword'].lower().startswith(w) for w in ['what','how','which','is','are','can','do','why','best','should'])]
    regular = [k for k in keywords if k not in question_style]
    
    if regular:
        print(f"\n  🔍 普通关键词 ({len(regular)}条):")
        for k in regular[:20]:
            src_marker = "🆕widget" if k.get('source') == 'widget' else ""
            depth_marker = f"[D{k.get('depth','-')}]" if k.get('depth') else ""
            print(f"    {k['keyword']:<50} {src_marker} {depth_marker}")
        if len(regular) > 20:
            print(f"    ... 还有 {len(regular) - 20} 条")
    
    if question_style:
        print(f"\n  🤖 问句式 ({len(question_style)}条):")
        for k in question_style[:10]:
            print(f"    {k['keyword']}")
        if len(question_style) > 10:
            print(f"    ... 还有 {len(question_style) - 10} 条")
    
    widget_items = data.get('widget_items', [])
    if widget_items:
        print(f"\n  📦 Widget分类卡片 ({len(widget_items)}条):")
        for wi in widget_items[:15]:
            print(f"    {wi.get('keyword',''):>12} → {wi.get('full_keyword','')}")
        if len(widget_items) > 15:
            print(f"    ... 还有 {len(widget_items) - 15} 条")


# ============================================================
# CLI 入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description='Amazon 搜索建议词挖掘工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    
    # 必选参数
    parser.add_argument('--seed', '-s', help='种子词（必填）')

    # 模式
    parser.add_argument('--mode', '-m', 
                        choices=['expand', 'az', 'az_prefix', 'numbers', 'gap', 'reverse', 'deep'],
                        default='expand', 
                        help='运行模式: expand(介词/场景/材质模板) | az(A-Z后缀) | az_prefix(A-Z前缀) | numbers(数字规格) | gap(空格间隙插入) | reverse(逆向滚雪球) | deep(深度递归)')
    
    # 模式参数
    parser.add_argument('--rounds', '-r', type=int, default=2, help='扩展轮次(expand模式)')
    parser.add_argument('--full-templates', action='store_true',
                        help='使用完整前缀模板（约50个，更全但更慢）；默认使用精简核心模板（约15个）')
    parser.add_argument('--depth', '-d', type=int, default=2, help='递归深度(deep模式)')
    parser.add_argument('--top-n', '-t', type=int, default=None, help='每层取Top-N前缀(deep模式默认5) / 逆向检索取Top-N高频词(reverse模式默认100)')
    
    # 请求参数
    parser.add_argument('--delay', type=float, default=0.5,
                        help='基础请求间隔秒数（默认0.5，成功时自动加随机抖动）')
    parser.add_argument('--jitter', type=float, default=0.6,
                        help='抖动比例（默认0.6，最多额外增加 base_delay 的 60%%）')
    parser.add_argument('--max-retries', type=int, default=3,
                        help='失败最大重试次数（默认3，使用指数退避）')
    parser.add_argument('--verbose', '-v', action='store_true', help='详细输出')
    parser.add_argument('--no-widget-deep', action='store_true',
                        help='禁用Widget卡片子分类词二次扩展（默认启用）')
    parser.add_argument('--max-widget-seeds', type=int, default=10,
                        help='Widget二次扩展最多取多少个子分类标签（默认10）')
    
    # 多站点
    market_list = ','.join(sorted(AMAZON_MARKETS.keys()))
    parser.add_argument('--market', '-M', default='US',
                        help=f'Amazon站点简码 (默认US): {market_list}')
    parser.add_argument('--markets', help='多站点批量(逗号分隔)，如 US,DE,JP')
    parser.add_argument('--auto-translate', '-T', action='store_true',
                        help='自动将英文种子词翻译为当地语言（内置词典+Google翻译）')
    parser.add_argument('--translations', 
                        help='手动指定各站点翻译词(格式: DE:Staubwedel,JP:羽毛掃き)')
    
    # 输出
    parser.add_argument('--output', '-o', help='JSON输出文件路径')
    parser.add_argument('--csv', help='CSV输出文件路径')
    parser.add_argument('--xlsx', help='Excel (xlsx) 输出文件路径')
    parser.add_argument('--db', help='SQLite数据库路径')
    
    args = parser.parse_args()

    # 验证参数
    if not args.seed:
        parser.error(f"{args.mode}模式需要 --seed")
    
    # 确定要跑的站点列表
    manual_translations = {}
    if args.translations:
        for pair in args.translations.split(','):
            if ':' in pair:
                mk_code, trans_word = pair.strip().split(':', 1)
                manual_translations[mk_code.strip().upper()] = trans_word.strip()
    
    if args.markets:
        target_markets = [m.strip().upper() for m in args.markets.split(',')]
        # 验证每个站点
        invalid = [m for m in target_markets if m not in AMAZON_MARKETS]
        if invalid:
            parser.error(f"未知站点: {invalid}。可用: {market_list}")
    else:
        target_markets = [args.market.upper()]
        if target_markets[0] not in AMAZON_MARKETS:
            parser.error(f"未知站点: {target_markets[0]}。可用: {market_list}")
    
    # 执行
    verbose = args.verbose
    all_results = []
    
    for mk in target_markets:
        if verbose and len(target_markets) > 1:
            print(f"\n🌐 ===== 站点: {mk} ({AMAZON_MARKETS[mk]['domain']}) =====")
        
        # 自动翻译种子词到当地语言（返回多候选词列表）
        seed_or_prefix = args.seed
        
        # 优先使用手动指定的翻译
        if mk in manual_translations:
            candidates = [(manual_translations[mk], 'manual')]
            if verbose or args.auto_translate:
                print(f"  ✏️ 手动翻译: [{mk}] '{seed_or_prefix}' → '{manual_translations[mk]}'")
        else:
            candidates = translate_seed(seed_or_prefix, mk, verbose=verbose or args.auto_translate)
        
        # 对每个候选词都执行一轮挖掘
        for local_seed, trans_source in candidates:
            if len(candidates) > 1 and (verbose or args.auto_translate):
                print(f"\n  🔎 候选词: '{local_seed}' (来源: {trans_source})")
            
            if args.mode == 'expand':
                result = mode_expand(local_seed, rounds=args.rounds, delay=args.delay,
                                    jitter_ratio=getattr(args, 'jitter', 0.6),
                                    verbose=verbose, market_code=mk,
                                    full_templates=args.full_templates)
            elif args.mode == 'az':
                result = mode_az(local_seed, delay=args.delay, verbose=verbose, market_code=mk)
            elif args.mode == 'az_prefix':
                result = mode_az_prefix(local_seed, delay=args.delay, verbose=verbose, market_code=mk)
            elif args.mode == 'numbers':
                result = mode_numbers(local_seed, delay=args.delay, verbose=verbose, market_code=mk)
            elif args.mode == 'gap':
                result = mode_gap(local_seed, delay=args.delay,
                                 jitter_ratio=getattr(args, 'jitter', 0.6),
                                 verbose=verbose, market_code=mk,
                                 full_templates=args.full_templates)
            elif args.mode == 'reverse':
                result = mode_reverse(local_seed, top_n=args.top_n or 100, delay=args.delay, verbose=verbose, market_code=mk)
            elif args.mode == 'deep':
                result = mode_deep(local_seed, depth=args.depth, top_n=args.top_n or 5,
                                  delay=args.delay, verbose=verbose, market_code=mk)
            
            result['market'] = mk
            result['market_domain'] = AMAZON_MARKETS[mk]['domain']
            result['market_currency'] = AMAZON_MARKETS[mk]['currency']
            result['original_seed'] = seed_or_prefix
            result['translated_seed'] = local_seed
            result['translation_source'] = trans_source
            all_results.append(result)
    
    # Widget 卡片子分类词二次扩展
    if not args.no_widget_deep:
        for result in all_results:
            widget_deep_expand(result, args.seed, delay=args.delay, verbose=verbose,
                             market_code=result.get('market', 'US'),
                             max_widget_seeds=args.max_widget_seeds)

    # 打印摘要（单站点直接打印，多站点逐个打印）
    for result in all_results:
        print_summary(result)
    
    # 保存输出
    main_result = all_results[0] if len(all_results) == 1 else {
        'seed': args.seed,
        'mode': args.mode,
        'markets': target_markets,
        'results': all_results,
        'total_keywords': sum(r['total_keywords'] for r in all_results),
        'total_widget_items': sum(r['total_widget_items'] for r in all_results),
    }
    
    if args.output:
        save_json(main_result, args.output)

    if args.csv:
        save_csv(main_result, args.csv)

    if args.xlsx:
        save_xlsx(main_result, args.xlsx)

    if args.db:
        for result in all_results:
            save_sqlite(result, args.db, seed=args.seed, mode=args.mode)

    # 始终落盘到会话目录（LinkFox 三大硬规范 #3）
    if HAS_LINKFOX:
        _ts = time.time()
        _session_xlsx = resolve_data_path("linkfox-amazon-suggestion-miner", _ts, "xlsx")
        _session_json = resolve_data_path("linkfox-amazon-suggestion-miner", _ts, "json")
        save_xlsx(main_result, _session_xlsx)
        save_json(main_result, _session_json)

    # 默认：如果没指定任何输出文件且无 LinkFox 环境，自动保存 Excel
    if not args.output and not args.csv and not args.db and not args.xlsx and not HAS_LINKFOX:
        seed_name = args.seed.replace(' ', '_')
        mk_suffix = f"_{'+'.join(target_markets)}" if len(target_markets) > 1 or target_markets[0] != 'US' else ''
        default_path = f"/root/suggestion_{seed_name}{mk_suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        save_xlsx(main_result, default_path)


if __name__ == '__main__':
    main()