#!/usr/bin/env python3
"""
Amazon Widget 卡片专门挖掘工具

基于试跑经验，Widget 卡片（WidgetSuggestion）是 Amazon 推荐引擎返回的高价值分类卡片，
包含子分类标签、完整关键词、商品图片 URL 和搜索 URL。

核心策略：
  1. 多策略触发：a-z扫描 + 品类修饰词 + 数字单位后缀 + 介词扩展
  2. Widget 标签递归扩展：把 Widget 子分类标签拼回种子词做二次查询，发现嵌套分类
  3. 去重输出：按 Widget标题+标签去重，输出多 Sheet Excel

用法:
  python3 widget_miner.py --seed "Summer Dresses for Women" -v
  python3 widget_miner.py --seed "dog toy" --depth 3 --market US -v
  python3 widget_miner.py --seed "wireless charger" --max-labels 20 -v
"""

import argparse
import json
import time
import sys
import os
import random
import string
import urllib.parse
from datetime import datetime
from collections import defaultdict

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from linkfox_paths import resolve_data_path
    HAS_LINKFOX = True
except ImportError:
    HAS_LINKFOX = False


# ============================================================
# 多站点配置
# ============================================================
AMAZON_MARKETS = {
    "US": {"mid": "ATVPDKIKX0DER", "domain": "amazon.com", "locale": "en_US", "plain_mid": 1, "lang": "en"},
    "CA": {"mid": "A2EUQ1WTGCTBG2", "domain": "amazon.ca", "locale": "en_CA", "plain_mid": 1, "lang": "en"},
    "MX": {"mid": "A1MK81F0CSZALV", "domain": "amazon.com.mx", "locale": "es_MX", "plain_mid": 1, "lang": "es"},
    "DE": {"mid": "A1PA6795UKMFR9", "domain": "amazon.de", "locale": "de_DE", "plain_mid": 3, "lang": "de"},
    "FR": {"mid": "A13V1IB3VIYZZH", "domain": "amazon.fr", "locale": "fr_FR", "plain_mid": 3, "lang": "fr"},
    "IT": {"mid": "APJ6JRA9NG5V4", "domain": "amazon.it", "locale": "it_IT", "plain_mid": 3, "lang": "it"},
    "ES": {"mid": "A1RKKUPIHCS9HS", "domain": "amazon.es", "locale": "es_ES", "plain_mid": 3, "lang": "es"},
    "JP": {"mid": "A1VC38T7YXB528", "domain": "amazon.co.jp", "locale": "ja_JP", "plain_mid": None, "lang": "ja"},
    "UK": {"mid": "A1F83G8C2ARO7P", "domain": "amazon.co.uk", "locale": "en_GB", "plain_mid": 3, "lang": "en"},
    "AU": {"mid": "A39IBJ37ZK2V1A", "domain": "amazon.com.au", "locale": "en_AU", "plain_mid": 3, "lang": "en"},
    "BR": {"mid": "A2Q3Y263D006CC", "domain": "amazon.com.br", "locale": "pt_BR", "plain_mid": 3, "lang": "pt"},
    "IN": {"mid": "A21TJRUUN4KGV", "domain": "amazon.in", "locale": "en_IN", "plain_mid": 3, "lang": "en"},
    "NL": {"mid": "A1805IZSGDJ6NL", "domain": "amazon.nl", "locale": "nl_NL", "plain_mid": 3, "lang": "nl"},
    "AE": {"mid": "A2VIGQ35RCS4UG", "domain": "amazon.ae", "locale": "en_AE", "plain_mid": 3, "lang": "en"},
    "SA": {"mid": "A17E79B4JQC1RO", "domain": "amazon.sa", "locale": "ar_SA", "plain_mid": 3, "lang": "ar"},
    "PL": {"mid": "A1C3SOZRARQ6R3", "domain": "amazon.pl", "locale": "pl_PL", "plain_mid": 3, "lang": "pl"},
    "SE": {"mid": "A2NODRKZP88ZSS", "domain": "amazon.se", "locale": "sv_SE", "plain_mid": 3, "lang": "sv"},
    "SG": {"mid": "A19VAU5T5FV4E7", "domain": "amazon.sg", "locale": "en_SG", "plain_mid": 3, "lang": "en"},
    "TR": {"mid": "A33GVJZQ3S32I2", "domain": "amazon.tr", "locale": "tr_TR", "plain_mid": 3, "lang": "tr"},
}

def get_market_config(code):
    return AMAZON_MARKETS.get(code.upper(), AMAZON_MARKETS["US"])


# ============================================================
# API 调用（带防封：随机抖动 + 指数退避重试）
# ============================================================

def fetch_suggestions(prefix, market_code="US", delay=0.5, max_retries=3, verbose=False):
    """调 Amazon Suggestions API，返回 keywords + widget_items"""
    market = get_market_config(market_code)
    session_id = f"135-{os.urandom(4).hex()}-{os.urandom(4).hex()}"
    request_id = os.urandom(8).hex()

    params_dict = {
        "limit": 11,
        "prefix": prefix,
        "suggestion-type": ["WIDGET", "KEYWORD"],
        "page-type": "Gateway",
        "alias": "aps",
        "site-variant": "desktop",
        "version": 3,
        "event": "onkeypress",
        "wc": "",
        "lop": market['locale'],
        "last-prefix": "\x00",
        "avg-ks-time": 5,
        "fb": 1,
        "predicted-text-accepted": "",
        "estoken": "",
        "session-id": session_id,
        "request-id": request_id,
        "mid": market['mid'],
        "client-info": "search-ui",
    }
    if market.get('plain_mid') is not None:
        params_dict["plain-mid"] = market['plain_mid']

    params = urllib.parse.urlencode(params_dict, doseq=True)
    domain = market['domain']
    url = f"https://www.{domain}/suggestions?{params}"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        "Accept": "application/json",
        "Accept-Language": market['locale'],
        "Referer": f"https://www.{domain}/",
        "Origin": f"https://www.{domain}",
    }

    for attempt in range(max_retries + 1):
        try:
            if HAS_REQUESTS:
                resp = requests.get(url, headers=headers, timeout=15)
                data = resp.json()
            else:
                from urllib.request import urlopen, Request
                req = Request(url, headers=headers)
                with urlopen(req, timeout=15) as response:
                    data = json.loads(response.read().decode())

            # 随机抖动延迟
            actual_delay = delay + random.uniform(0, delay * 0.6)
            time.sleep(actual_delay)
            break

        except Exception as e:
            if attempt < max_retries:
                backoff = (1.5 ** attempt) * delay
                actual = backoff + random.uniform(0, backoff * 0.3)
                if verbose:
                    print(f"  ⏳ 重试 {attempt+1}/{max_retries}，等待 {actual:.2f}s ({e})")
                time.sleep(actual)
            else:
                if verbose:
                    print(f"  ✗ 请求失败: {e}")
                return {'keywords': [], 'widget_items': [], 'error': str(e)}

    keywords = []
    widget_items = []

    for i, s in enumerate(data.get('suggestions', [])):
        if s.get('suggType') == 'KeywordSuggestion':
            keywords.append({
                'keyword': s['value'],
                'rank': i + 1,
                'sugg_type': 'KeywordSuggestion',
                'candidate_source': s.get('candidateSources', ''),
            })
        elif s.get('suggType') == 'WidgetSuggestion':
            widget_title = s.get('metadata', {}).get('title', '')
            for item in s.get('widgetItems', []):
                meta = item.get('metadata', {})
                text = meta.get('text', '')
                link = meta.get('link_url', '')
                full_kw = ''
                if link:
                    kw_match = urllib.parse.parse_qs(urllib.parse.urlparse(link).query).get('k', [])
                    if kw_match:
                        full_kw = kw_match[0].replace('+', ' ')

                widget_items.append({
                    'keyword': text,
                    'full_keyword': full_kw,
                    'widget_title': widget_title,
                    'image_url': meta.get('image_url', ''),
                    'search_url': link,
                })

    if verbose:
        widget_count = len(widget_items)
        marker = f" 📦{widget_count}Widget" if widget_count else ""
        print(f"  ✓ {len(keywords)}关键词{marker} ← '{prefix[:50]}'")

    return {'keywords': keywords, 'widget_items': widget_items}


# ============================================================
# Widget 卡片挖掘核心
# ============================================================

# 品类修饰词（试跑经验：容易触发 Widget 卡片）
CATEGORY_MODIFIERS = [
    "casual", "long", "short", "maxi", "midi", "plus size", "floral",
    "formal", "boho", "cotton", "linen", "wrap", "sleeveless",
    "short sleeve", "long sleeve", "v neck", "halter", "strapless",
    "beach", "wedding", "party", "work", "travel", "vacation",
]

# 数字/单位后缀（试跑经验：意外触发大量 Widget 卡片）
UNIT_SUFFIXES = ["mm", "inch", "oz", "pack", "set", "1", "3", "6"]

# 介词（for/with 容易触发 Widget）
PREPOSITIONS = ["for", "with", "without"]


def generate_trigger_prefixes(seed):
    """生成多种触发 Widget 卡片的前缀组合"""
    prefixes = [seed]  # 种子词本身

    # 策略 1: a-z 后缀扫描
    for ch in string.ascii_lowercase:
        prefixes.append(f"{seed} {ch}")

    # 策略 2: 品类修饰词扩展
    for mod in CATEGORY_MODIFIERS:
        prefixes.append(f"{seed} {mod}")
        # 修饰词前置
        prefixes.append(f"{mod} {seed}")

    # 策略 3: 数字/单位后缀
    for unit in UNIT_SUFFIXES:
        prefixes.append(f"{seed} {unit}")

    # 策略 4: 介词 + a-z
    for prep in PREPOSITIONS:
        prefixes.append(f"{seed} {prep}")
        for ch in string.ascii_lowercase:
            prefixes.append(f"{seed} {prep} {ch}")

    # 去重
    seen = set()
    unique = []
    for p in prefixes:
        if p.lower() not in seen:
            seen.add(p.lower())
            unique.append(p)

    return unique


def extract_widget_labels(widgets, already_used=None):
    """从 Widget 卡片中提取去重的子分类标签"""
    if already_used is None:
        already_used = set()

    labels = []
    seen = set(already_used)
    for wi in widgets:
        label = wi.get('keyword', '').strip()
        if label and label.lower() not in seen:
            seen.add(label.lower())
            labels.append(label)

    return labels


def dedup_widgets(widgets):
    """按 Widget标题 + 子分类标签去重"""
    seen = set()
    unique = []
    for wi in widgets:
        key = f"{wi.get('widget_title', '').lower()}|{wi.get('keyword', '').lower()}"
        if key not in seen:
            seen.add(key)
            unique.append(wi)
    return unique


def mine_widgets(seed, market_code="US", max_depth=2, max_labels=15, delay=0.5, verbose=True):
    """专门挖掘 Widget 卡片

    Args:
        seed: 种子词
        market_code: 站点
        max_depth: 递归扩展深度（1=只扫描，2=+Widget标签二次扩展，3=+三轮）
        max_labels: 每轮最多取多少个 Widget 标签做二次扩展
        delay: 请求间隔基数
        verbose: 详细输出

    Returns:
        dict: 挖掘结果
    """
    all_widgets = []
    all_keywords = {}
    tried_prefixes = set()
    used_labels = set()
    stats = {'rounds': []}

    # === 第 1 轮：多策略触发 ===
    if verbose:
        print(f"\n🔍 Widget 卡片挖掘 — 种子词: '{seed}'")
        print(f"   站点: {market_code} | 深度: {max_depth} | 最大标签数: {max_labels}")

    round1_prefixes = generate_trigger_prefixes(seed)
    if verbose:
        print(f"\n--- 第 1 轮：多策略触发（{len(round1_prefixes)} 个前缀）---")

    round1_widgets = 0
    for prefix in round1_prefixes:
        if prefix.lower() in tried_prefixes:
            continue
        tried_prefixes.add(prefix.lower())

        result = fetch_suggestions(prefix, market_code=market_code, delay=delay, verbose=verbose)

        for wi in result['widget_items']:
            wi['round'] = 1
            wi['trigger_prefix'] = prefix
            all_widgets.append(wi)
            round1_widgets += 1

        for kw in result['keywords']:
            kw_key = kw['keyword'].lower()
            if kw_key not in all_keywords:
                kw['source'] = 'round1'
                kw['prefix'] = prefix
                all_keywords[kw_key] = kw

    round1_unique = len(dedup_widgets(all_widgets))
    stats['rounds'].append({
        'round': 1,
        'prefixes_tried': len(tried_prefixes),
        'widgets_found': round1_widgets,
        'widgets_unique': round1_unique,
        'keywords_found': len(all_keywords),
    })

    if verbose:
        print(f"\n📊 第 1 轮完成：{round1_widgets} 个 Widget 卡片（去重 {round1_unique}），{len(all_keywords)} 个关键词")

    # === 第 2 轮：Widget 标签二次扩展 ===
    if max_depth >= 2:
        labels = extract_widget_labels(all_widgets, used_labels)
        labels = labels[:max_labels]

        if verbose:
            print(f"\n--- 第 2 轮：Widget 标签二次扩展（{len(labels)} 个标签）---")
            print(f"   标签: {labels}")

        round2_widgets = 0
        for label in labels:
            used_labels.add(label.lower())
            new_prefix = f"{seed} {label}"
            if new_prefix.lower() in tried_prefixes:
                continue
            tried_prefixes.add(new_prefix.lower())

            result = fetch_suggestions(new_prefix, market_code=market_code, delay=delay, verbose=verbose)

            for wi in result['widget_items']:
                wi['round'] = 2
                wi['trigger_prefix'] = new_prefix
                all_widgets.append(wi)
                round2_widgets += 1

            for kw in result['keywords']:
                kw_key = kw['keyword'].lower()
                if kw_key not in all_keywords:
                    kw['source'] = 'round2_widget'
                    kw['prefix'] = new_prefix
                    all_keywords[kw_key] = kw

        total_unique = len(dedup_widgets(all_widgets))
        stats['rounds'].append({
            'round': 2,
            'labels_used': len(labels),
            'prefixes_tried': len(tried_prefixes),
            'widgets_found': round2_widgets,
            'widgets_unique': total_unique,
            'keywords_found': len(all_keywords),
        })

        if verbose:
            print(f"\n📊 第 2 轮完成：新增 {round2_widgets} 个 Widget 卡片（总计去重 {total_unique}），{len(all_keywords)} 个关键词")

    # === 第 3 轮（可选）：从第 2 轮 Widget 中提取新标签 ===
    if max_depth >= 3:
        round2_widgets_list = [w for w in all_widgets if w.get('round') == 2]
        new_labels = extract_widget_labels(round2_widgets_list, used_labels)
        new_labels = new_labels[:max_labels]

        if new_labels and verbose:
            print(f"\n--- 第 3 轮：嵌套标签扩展（{len(new_labels)} 个新标签）---")

        round3_widgets = 0
        for label in new_labels:
            used_labels.add(label.lower())
            new_prefix = f"{seed} {label}"
            if new_prefix.lower() in tried_prefixes:
                continue
            tried_prefixes.add(new_prefix.lower())

            result = fetch_suggestions(new_prefix, market_code=market_code, delay=delay, verbose=verbose)

            for wi in result['widget_items']:
                wi['round'] = 3
                wi['trigger_prefix'] = new_prefix
                all_widgets.append(wi)
                round3_widgets += 1

            for kw in result['keywords']:
                kw_key = kw['keyword'].lower()
                if kw_key not in all_keywords:
                    kw['source'] = 'round3_widget'
                    kw['prefix'] = new_prefix
                    all_keywords[kw_key] = kw

        total_unique = len(dedup_widgets(all_widgets))
        stats['rounds'].append({
            'round': 3,
            'labels_used': len(new_labels),
            'prefixes_tried': len(tried_prefixes),
            'widgets_found': round3_widgets,
            'widgets_unique': total_unique,
            'keywords_found': len(all_keywords),
        })

        if verbose:
            print(f"\n📊 第 3 轮完成：新增 {round3_widgets} 个 Widget 卡片（总计去重 {total_unique}），{len(all_keywords)} 个关键词")

    # 去重
    unique_widgets = dedup_widgets(all_widgets)

    # 按 Widget 标题分组
    widget_groups = defaultdict(list)
    for wi in unique_widgets:
        title = wi.get('widget_title', '其他')
        widget_groups[title].append(wi)

    result = {
        'seed': seed,
        'market': market_code,
        'mode': 'widget_mining',
        'total_widgets': len(unique_widgets),
        'total_keywords': len(all_keywords),
        'widget_groups': len(widget_groups),
        'widgets': unique_widgets,
        'keywords': list(all_keywords.values()),
        'widget_group_details': {title: len(items) for title, items in widget_groups.items()},
        'stats': {
            'rounds': stats['rounds'],
            'total_prefixes_tried': len(tried_prefixes),
            'total_widget_labels_used': len(used_labels),
        },
    }

    return result


# ============================================================
# 输出
# ============================================================

def save_json(data, path):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\n✅ JSON 已保存: {path}")


def save_xlsx(data, path):
    if not HAS_OPENPYXL:
        print(f"\n❌ 需要 openpyxl: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    def style_header(ws, col_count):
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.freeze_panes = 'A2'

    def auto_width(ws, col_count, min_w=10, max_w=60):
        for col_idx in range(1, col_count + 1):
            max_len = min_w
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    display_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
                    if display_len > max_len:
                        max_len = display_len
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_w)

    # Sheet 1: 摘要
    ws_summary = wb.active
    ws_summary.title = '摘要'
    summary_rows = [
        ['种子词', data.get('seed', '')],
        ['站点', data.get('market', 'US')],
        ['Widget卡片总数(去重)', data.get('total_widgets', 0)],
        ['Widget分类组数', data.get('widget_groups', 0)],
        ['关键词总数', data.get('total_keywords', 0)],
        ['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        [],
        ['--- 各轮统计 ---', ''],
    ]
    for r in data.get('stats', {}).get('rounds', []):
        summary_rows.append([f"第{r['round']}轮", ''])
        for k, v in r.items():
            if isinstance(v, list):
                v = ', '.join(str(i) for i in v)
            summary_rows.append([f"  {k}", v])

    summary_rows.append([])
    summary_rows.append(['--- Widget 分类组详情 ---', ''])
    for title, count in data.get('widget_group_details', {}).items():
        summary_rows.append([title, f"{count} 个子分类"])

    for row_idx, row_data in enumerate(summary_rows, 1):
        if len(row_data) == 2:
            ws_summary.cell(row=row_idx, column=1, value=row_data[0]).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=row_data[1])
        elif len(row_data) == 1:
            ws_summary.cell(row=row_idx, column=1, value=row_data[0]).font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 60

    # Sheet 2: Widget 分类卡片
    ws_widget = wb.create_sheet('Widget分类卡片')
    widget_headers = ['轮次', 'Widget标题', '子分类标签', '完整关键词', '触发前缀', '图片URL', '搜索URL']
    for col_idx, h in enumerate(widget_headers, 1):
        ws_widget.cell(row=1, column=col_idx, value=h)

    for row_idx, wi in enumerate(data.get('widgets', []), 2):
        ws_widget.cell(row=row_idx, column=1, value=wi.get('round', ''))
        ws_widget.cell(row=row_idx, column=2, value=wi.get('widget_title', ''))
        ws_widget.cell(row=row_idx, column=3, value=wi.get('keyword', ''))
        ws_widget.cell(row=row_idx, column=4, value=wi.get('full_keyword', ''))
        ws_widget.cell(row=row_idx, column=5, value=wi.get('trigger_prefix', ''))
        ws_widget.cell(row=row_idx, column=6, value=wi.get('image_url', ''))
        ws_widget.cell(row=row_idx, column=7, value=wi.get('search_url', ''))

    style_header(ws_widget, len(widget_headers))
    auto_width(ws_widget, len(widget_headers))

    # Sheet 3: 所有关键词
    ws_kw = wb.create_sheet('关键词')
    kw_headers = ['关键词', '来源', '触发前缀', '排名']
    for col_idx, h in enumerate(kw_headers, 1):
        ws_kw.cell(row=1, column=col_idx, value=h)

    for row_idx, kw in enumerate(data.get('keywords', []), 2):
        ws_kw.cell(row=row_idx, column=1, value=kw.get('keyword', ''))
        ws_kw.cell(row=row_idx, column=2, value=kw.get('source', ''))
        ws_kw.cell(row=row_idx, column=3, value=kw.get('prefix', ''))
        ws_kw.cell(row=row_idx, column=4, value=kw.get('rank', ''))

    style_header(ws_kw, len(kw_headers))
    auto_width(ws_kw, len(kw_headers))

    wb.save(path)
    print(f"✅ Excel 已保存: {path}")
    print(f"   Sheet: 摘要 / Widget分类卡片 ({len(data.get('widgets', []))} 条) / 关键词 ({len(data.get('keywords', []))} 条)")


def print_summary(data):
    print(f"\n{'='*60}")
    print(f"📊 Widget 卡片挖掘结果")
    print(f"{'='*60}")
    print(f"  种子词: {data['seed']}")
    print(f"  站点: {data['market']}")
    print(f"  Widget卡片(去重): {data['total_widgets']}")
    print(f"  Widget分类组: {data['widget_groups']}")
    print(f"  关键词总数: {data['total_keywords']}")

    print(f"\n  📦 Widget 分类组详情:")
    for title, count in data.get('widget_group_details', {}).items():
        print(f"    {title}: {count} 个子分类")

    print(f"\n  📋 各轮统计:")
    for r in data.get('stats', {}).get('rounds', []):
        print(f"    第{r['round']}轮: {r.get('widgets_unique', 0)} 个Widget, {r.get('keywords_found', 0)} 个关键词")

    # 展示前 20 个 Widget 卡片
    widgets = data.get('widgets', [])
    if widgets:
        print(f"\n  📦 Widget 卡片预览 (前20条):")
        for wi in widgets[:20]:
            label = wi.get('keyword', '')
            full_kw = wi.get('full_keyword', '')
            title = wi.get('widget_title', '')
            round_num = wi.get('round', '')
            print(f"    [R{round_num}] {label:<15} → {full_kw:<50} ({title})")
        if len(widgets) > 20:
            print(f"    ... 还有 {len(widgets) - 20} 条")


# ============================================================
# CLI 入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='Amazon Widget 卡片专门挖掘工具')
    parser.add_argument('--seed', '-s', required=True, help='种子词（必填）')
    parser.add_argument('--market', '-M', default='US', help='站点（默认US）')
    parser.add_argument('--depth', '-d', type=int, default=2, help='递归扩展深度（1=只扫描, 2=+标签二次扩展, 3=+三轮嵌套）')
    parser.add_argument('--max-labels', type=int, default=15, help='每轮最多取多少个Widget标签做扩展（默认15）')
    parser.add_argument('--delay', type=float, default=0.5, help='请求间隔基数（默认0.5秒）')
    parser.add_argument('--verbose', '-v', action='store_true', help='详细输出')
    parser.add_argument('--xlsx', help='Excel输出路径')
    parser.add_argument('--output', '-o', help='JSON输出路径')

    args = parser.parse_args()

    # 执行挖掘
    result = mine_widgets(
        seed=args.seed,
        market_code=args.market,
        max_depth=args.depth,
        max_labels=args.max_labels,
        delay=args.delay,
        verbose=args.verbose,
    )

    # 打印摘要
    print_summary(result)

    # 用户指定输出
    if args.output:
        save_json(result, args.output)
    if args.xlsx:
        save_xlsx(result, args.xlsx)

    # 始终落盘到会话目录
    if HAS_LINKFOX:
        ts = time.time()
        session_xlsx = resolve_data_path("linkfox-amazon-widget-miner", ts, "xlsx")
        session_json = resolve_data_path("linkfox-amazon-widget-miner", ts, "json")
        save_xlsx(result, session_xlsx)
        save_json(result, session_json)
    elif not args.output and not args.xlsx:
        # 无 LinkFox 环境的默认输出
        seed_name = args.seed.replace(' ', '_')
        default_path = f"/root/widget_cards_{seed_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        save_xlsx(result, default_path)


if __name__ == '__main__':
    main()
