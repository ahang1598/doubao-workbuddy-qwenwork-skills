#!/usr/bin/env python3
"""
尽调材料清单 Excel 生成器 v1.2.0
接受 JSON 格式的清单数据，输出 .xlsx 文件

用法:
    python excel_generator.py --input checklist_data.json --output 尽调材料清单.xlsx
    python excel_generator.py --demo

依赖: openpyxl >= 3.1.0
安装: pip install openpyxl
"""

import json
import sys
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── 样式常量 ──────────────────────────────────────────

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DATA_FONT = Font(name="微软雅黑", size=10, color="000000")

MUST_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RECOMMEND_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# 接收状态条件格式
RECEIVED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 已收-绿色
NOT_RECEIVED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 未收-红色
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 部分-黄色
REJECTED_FILL = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")  # 拒绝-蓝色

THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# v1.2.0: 列7→11
COL_WIDTHS = [10, 22, 35, 12, 10, 12, 12, 12, 12, 12, 28]
HEADERS = ["序号", "模块", "资料名称", "资料类型", "必需性", "提交阶段", "存档部门", "接收状态", "收件日期", "提供形式", "备注"]

DISCLAIMER = (
    "免责声明：本清单由AI辅助生成，仅供参考。"
    "律师应根据具体交易情况审核调整后发送目标公司。"
    "本清单不构成法律意见，不保证清单完备无遗漏。"
)

# 部门→模块映射（参考表，用于Phase 3 Step 3.3分配dept字段；
# Sheet4实际按item.dept字段分组，不直接使用此映射）
DEPT_MODULE_MAP = {
    "行政部": ["M1", "M3", "M9", "M12"],
    "财务部": ["M4"],
    "法务部": ["M2", "M3", "M6", "M8", "M10", "M11"],
    "人力资源部": ["M7"],
    "IT部": ["M5", "M11"],
    "业务部": ["M5", "M6"],
}


def _get_module_id(item):
    """从 item.id 提取模块号, 如 M1-001 → M1"""
    return item.get("id", "").split("-")[0] if "-" in item.get("id", "") else ""


def create_sheet1_main(wb, items):
    """Sheet1: 尽调材料清单（主表, v1.2.0 11列）"""
    ws = wb.active
    ws.title = "尽调材料清单"

    # 表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # 列宽
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 数据行
    for row_idx, item in enumerate(items, 2):
        values = [
            item.get("id", ""),
            item.get("module", ""),
            item.get("name", ""),
            item.get("type", ""),
            item.get("necessity", ""),
            item.get("stage", ""),
            item.get("dept", ""),
            item.get("receive_status", "未收"),
            item.get("receive_date", ""),
            item.get("provide_form", ""),
            item.get("remark", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = LEFT_ALIGN if col_idx in (3, 11) else CENTER_ALIGN
            cell.border = THIN_BORDER
        ws.row_dimensions[row_idx].height = 20

    last_row = len(items) + 1

    # 必需性条件格式（E列）
    necessity_range = f"E2:E{last_row}"
    ws.conditional_formatting.add(
        necessity_range,
        CellIsRule(operator="equal", formula=['"必须"'], fill=MUST_FILL),
    )
    ws.conditional_formatting.add(
        necessity_range,
        CellIsRule(operator="equal", formula=['"推荐"'], fill=RECOMMEND_FILL),
    )
    ws.conditional_formatting.add(
        necessity_range,
        CellIsRule(operator="equal", formula=['"可选"'], fill=OPTIONAL_FILL),
    )

    # v1.2.0: 接收状态条件格式（H列）
    status_range = f"H2:H{last_row}"
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"已收"'], fill=RECEIVED_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"未收"'], fill=NOT_RECEIVED_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"部分提供"'], fill=PARTIAL_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"拒绝提供"'], fill=REJECTED_FILL),
    )

    # v1.2.0: 接收状态下拉框数据验证
    dv = DataValidation(
        type="list",
        formula1='"已收,未收,部分提供,拒绝提供"',
        allow_blank=True,
    )
    dv.error = "请选择：已收/未收/部分提供/拒绝提供"
    dv.errorTitle = "输入错误"
    dv.prompt = "选择接收状态"
    dv.promptTitle = "接收状态"
    ws.add_data_validation(dv)
    dv.add(status_range)

    # 模块列合并相同值
    _merge_module_cells(ws, last_row)

    # 冻结首行
    ws.freeze_panes = "A2"

    # 打印设置
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "1:1"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # 末行追加免责声明
    disclaimer_row = last_row + 2
    ws.cell(row=disclaimer_row, column=1, value=DISCLAIMER).font = Font(
        name="微软雅黑", size=9, italic=True, color="666666"
    )
    ws.merge_cells(start_row=disclaimer_row, start_column=1, end_row=disclaimer_row, end_column=11)
    ws.cell(row=disclaimer_row, column=1).alignment = LEFT_ALIGN

    # v1.2.0: 仪表盘行（材料状态统计）
    dashboard_row = disclaimer_row + 2
    ws.cell(row=dashboard_row, column=1, value="📊 状态统计").font = Font(
        name="微软雅黑", size=10, bold=True, color="4472C4"
    )
    ws.merge_cells(start_row=dashboard_row, start_column=1, end_row=dashboard_row, end_column=4)
    dashboard_row += 1
    
    # 统计行
    total = len(items)
    must_count = sum(1 for item in items if item.get("necessity") == "必须")
    recommend_count = sum(1 for item in items if item.get("necessity") == "推荐")
    optional_count = sum(1 for item in items if item.get("necessity") == "可选")
    
    stats = [
        ("材料项总数", f"{total} 项", ""),
        ("必须", f"{must_count} 项", f"占比 {must_count*100//total if total else 0}%"),
        ("推荐", f"{recommend_count} 项", f"占比 {recommend_count*100//total if total else 0}%"),
        ("可选", f"{optional_count} 项", f"占比 {optional_count*100//total if total else 0}%"),
    ]
    for label, val, pct in stats:
        ws.cell(row=dashboard_row, column=1, value=label).font = Font(name="微软雅黑", size=9)
        ws.cell(row=dashboard_row, column=2, value=val).font = Font(name="微软雅黑", size=9)
        ws.cell(row=dashboard_row, column=3, value=pct).font = Font(name="微软雅黑", size=9, color="666666")
        dashboard_row += 1


def _merge_module_cells(ws, last_row):
    """合并模块列中相同值的连续单元格"""
    if last_row < 2:
        return
    start_row = 2
    current_value = ws.cell(row=2, column=2).value
    for row in range(3, last_row + 1):
        val = ws.cell(row=row, column=2).value
        if val != current_value:
            if row - 1 > start_row:
                ws.merge_cells(
                    start_row=start_row, start_column=2,
                    end_row=row - 1, end_column=2,
                )
            start_row = row
            current_value = val
    if last_row > start_row:
        ws.merge_cells(
            start_row=start_row, start_column=2,
            end_row=last_row, end_column=2,
        )


def create_sheet2_guide(wb, metadata):
    """Sheet2: 场景说明与填写指引"""
    ws = wb.create_sheet("场景说明与填写指引")

    # 标题
    ws.merge_cells("A1:K1")
    title_cell = ws.cell(row=1, column=1, value="尽调材料清单 — 场景说明与填写指引")
    title_cell.font = Font(name="微软雅黑", size=14, bold=True)
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    # 内容
    sections = [
        ("一、场景信息", [
            f"场景类型: {metadata.get('scenario_name', 'N/A')}",
            f"场景标识: {metadata.get('scenario_id', 'N/A')}",
            f"目标公司: {metadata.get('target_company', 'N/A')}",
            f"生成日期: {metadata.get('generated_date', 'N/A')}",
            f"材料项总数: {metadata.get('total_items', 'N/A')}",
            f"尽调深度: {metadata.get('dd_depth', '标准')}",
            f"清单版本: {metadata.get('checklist_version', '1')}",
        ]),
        ("二、必需性说明", [
            "必须(红底): 核心材料，缺失将影响尽调基本判断",
            "推荐(黄底): 重要材料，缺失可能影响尽调深度",
            "可选(灰底): 辅助材料，视交易具体情况决定是否索取",
        ]),
        ("三、接收状态说明（v1.2.0）", [
            "未收(红底): 尚未收到该材料",
            "已收(绿底): 已收到该材料",
            "部分提供(黄底): 仅提供部分内容",
            "拒绝提供(蓝底): 目标公司明确拒绝提供",
            "请在H列下拉框中选择状态，收件日期和提供形式同步更新",
        ]),
        ("四、提交阶段说明", [
            "第一阶段: 基础信息类（公司证照/股权/组织架构），快速建立公司画像",
            "第二阶段: 核心经营类（财务/资产/合同），深入了解经营状况",
            "第三阶段: 补充完善类（劳动/诉讼/合规/数据安全/保险），排查风险点",
        ]),
        ("五、填写指引", [
            "1. 律师发送前应审核清单的模块覆盖度和必需性合理性",
            "2. 可根据具体交易情况调整材料项和必需性标注",
            "3. 建议分三阶段逐步索取，避免一次性发全量清单",
            "4. 备注列标注'需原件核对'或'复印件加盖公章'",
            "5. 收到材料后在H列更新接收状态",
            "6. 如生成补充清单，建议在备注列标注'v2补充'",
        ]),
        ("六、免责声明", [DISCLAIMER]),
    ]

    row = 3
    for title, lines in sections:
        ws.cell(row=row, column=1, value=title).font = Font(name="微软雅黑", size=12, bold=True)
        row += 1
        for line in lines:
            ws.cell(row=row, column=1, value=line).font = Font(name="微软雅黑", size=10)
            ws.cell(row=row, column=1).alignment = LEFT_ALIGN
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 22
    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_sheet3_stages(wb, items):
    """Sheet3: 提交阶段分配（当材料项>30时生成）"""
    ws = wb.create_sheet("提交阶段分配")

    stages = {"第一阶段": [], "第二阶段": [], "第三阶段": []}
    for item in items:
        stage = item.get("stage", "")
        if stage in stages:
            stages[stage].append(item)

    stage_headers = ["序号", "资料名称", "模块", "必需性", "存档部门"]
    col_widths = [12, 40, 25, 12, 12]

    row = 1
    for stage_name in ["第一阶段", "第二阶段", "第三阶段"]:
        stage_items = stages.get(stage_name, [])
        if not stage_items:
            continue

        # 阶段标题
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        title_cell = ws.cell(row=row, column=1, value=f"── {stage_name}（{len(stage_items)}项）──")
        title_cell.font = Font(name="微软雅黑", size=12, bold=True, color="4472C4")
        title_cell.alignment = CENTER_ALIGN
        ws.row_dimensions[row].height = 28
        row += 1

        # 表头
        for col_idx, header in enumerate(stage_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
        row += 1

        # 数据
        for item in stage_items:
            values = [
                item.get("id", ""),
                item.get("name", ""),
                item.get("module", ""),
                item.get("necessity", ""),
                item.get("dept", ""),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.alignment = LEFT_ALIGN if col_idx == 2 else CENTER_ALIGN
                cell.border = THIN_BORDER
            ws.row_dimensions[row].height = 20
            row += 1

        row += 2  # 阶段间空行

    # 列宽
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_sheet4_depts(wb, items):
    """Sheet4: 部门分派视图（v1.2.0新增，当材料项>30时生成）"""
    ws = wb.create_sheet("部门分派视图")

    # 按部门分组
    dept_groups = {}
    for item in items:
        dept = item.get("dept", "未分配")
        if dept not in dept_groups:
            dept_groups[dept] = []
        dept_groups[dept].append(item)

    dept_headers = ["序号", "资料名称", "模块", "必需性", "提交阶段"]
    col_widths = [12, 40, 25, 12, 12]

    # 部门显示顺序
    dept_order = ["行政部", "财务部", "法务部", "人力资源部", "IT部", "业务部"]

    row = 1
    for dept_name in dept_order:
        dept_items = dept_groups.get(dept_name, [])
        if not dept_items:
            continue

        # 部门标题
        dept_icons = {"财务部": "💰", "法务部": "⚖️", "人力资源部": "👥", "IT部": "💻", "行政部": "🏢", "业务部": "📊"}
        icon = dept_icons.get(dept_name, "📋")
        
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        title_cell = ws.cell(row=row, column=1, value=f"{icon} {dept_name}（{len(dept_items)}项）")
        title_cell.font = Font(name="微软雅黑", size=12, bold=True, color="4472C4")
        title_cell.alignment = CENTER_ALIGN
        ws.row_dimensions[row].height = 28
        row += 1

        # 表头
        for col_idx, header in enumerate(dept_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
        row += 1

        # 数据
        for item in dept_items:
            values = [
                item.get("id", ""),
                item.get("name", ""),
                item.get("module", ""),
                item.get("necessity", ""),
                item.get("stage", ""),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.alignment = LEFT_ALIGN if col_idx == 2 else CENTER_ALIGN
                cell.border = THIN_BORDER
            ws.row_dimensions[row].height = 20
            row += 1

        row += 2  # 部门间空行

    # 列宽
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def generate_excel(data, output_path=None):
    """
    主函数：从 JSON 数据生成 Excel 文件

    Args:
        data: dict, 符合 output-interface.md JSON Schema 的清单数据
        output_path: str, 输出文件路径。如不指定则自动命名

    Returns:
        str: 生成的文件路径
    """
    metadata = data.get("metadata", {})
    items = data.get("items", [])

    if not items:
        print("WARNING: items 为空，生成空清单", file=sys.stderr)

    wb = Workbook()

    # Sheet1 主表
    create_sheet1_main(wb, items)

    # Sheet2 说明
    create_sheet2_guide(wb, metadata)

    # Sheet3 提交阶段分配（>30项时）
    if len(items) > 30:
        create_sheet3_stages(wb, items)
        # v1.2.0: Sheet4 部门分派视图
        create_sheet4_depts(wb, items)

    # 文件命名
    if output_path is None:
        scenario = metadata.get("scenario_id", "UNKNOWN")
        company = metadata.get("target_company", "公司")
        date_str = metadata.get("generated_date", datetime.now().strftime("%Y%m%d"))
        version = metadata.get("checklist_version", "v1")
        output_path = f"尽调材料清单_{scenario}_{company}_{date_str}_{version}.xlsx"

    wb.save(output_path)
    print(f"SUCCESS: Excel 文件已生成 → {output_path}")
    return output_path


def demo():
    """演示模式：使用示例数据生成 Excel"""
    demo_data = {
        "metadata": {
            "scenario_id": "S-PE",
            "scenario_name": "PE/VC投资",
            "target_company": "XX科技有限公司",
            "generated_date": datetime.now().strftime("%Y-%m-%d"),
            "total_items": 8,
            "modules_covered": 6,
            "dd_depth": "标准",
            "checklist_version": "v1.2.0-demo",
        },
        "items": [
            {"id": "M1-001", "module": "公司基本信息与历史沿革", "name": "营业执照副本", "type": "证照", "necessity": "必须", "stage": "第一阶段", "dept": "行政部", "receive_status": "未收", "receive_date": "", "provide_form": "", "remark": "需提供最新年检版本"},
            {"id": "M1-002", "module": "公司基本信息与历史沿革", "name": "公司章程（含修正案）", "type": "决议", "necessity": "必须", "stage": "第一阶段", "dept": "行政部", "receive_status": "未收", "receive_date": "", "provide_form": "", "remark": "含历次修正案"},
            {"id": "M2-001", "module": "股权结构与股东信息", "name": "股东名册", "type": "决议", "necessity": "必须", "stage": "第一阶段", "dept": "法务部", "receive_status": "未收", "receive_date": "", "provide_form": "", "remark": "最新股东名册"},
            {"id": "M2-014", "module": "股权结构与股东信息", "name": "对赌协议/估值调整协议", "type": "合同", "necessity": "必须", "stage": "第一阶段", "dept": "法务部", "receive_status": "未收", "receive_date": "", "provide_form": "", "remark": "PE场景核心文件"},
            {"id": "M5-001", "module": "资产与知识产权", "name": "专利证书及登记簿", "type": "证照", "necessity": "必须", "stage": "第二阶段", "dept": "IT部", "receive_status": "未收", "receive_date": "", "provide_form": "", "remark": "含已申请未授权"},
            {"id": "M7-001", "module": "劳动人事", "name": "员工名册", "type": "其他", "necessity": "推荐", "stage": "第三阶段", "dept": "人力资源部", "receive_status": "未收", "receive_date": "", "provide_form": "", "remark": "含在职/离职"},
            {"id": "M11-002", "module": "数据安全与个人信息保护", "name": "个人信息保护影响评估报告(PIA)", "type": "其他", "necessity": "推荐", "stage": "第三阶段", "dept": "法务部", "receive_status": "未收", "receive_date": "", "provide_form": "", "remark": "《个人信息保护法》第55/56条"},
            {"id": "M12-003", "module": "保险与风险管理", "name": "董监高责任保险(D&O)保单", "type": "证照", "necessity": "推荐", "stage": "第三阶段", "dept": "行政部", "receive_status": "未收", "receive_date": "", "provide_form": "", "remark": "上市公司核心尽调项"},
        ],
    }
    return generate_excel(demo_data, "尽调材料清单_DEMO_v1.2.0.xlsx")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="尽调材料清单 Excel 生成器 v1.2.0")
    parser.add_argument("--input", "-i", help="JSON 格式的清单数据文件路径")
    parser.add_argument("--output", "-o", help="输出 Excel 文件路径（可选，自动命名）")
    parser.add_argument("--demo", action="store_true", help="使用演示数据生成示例 Excel")

    args = parser.parse_args()

    if args.demo:
        demo()
    elif args.input:
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)
        generate_excel(data, args.output)
    else:
        parser.print_help()
        print("\n示例:")
        print("  python excel_generator.py --demo")
        print("  python excel_generator.py --input checklist.json --output 清单.xlsx")
        sys.exit(1)
