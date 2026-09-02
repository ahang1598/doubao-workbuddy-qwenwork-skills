#!/usr/bin/env python3
"""从风险 JSON 生成 Richee 人读风险清单 Excel（输出规范 v1.1.0 excel-human profile）。

用法:
  python build_risk_list.py <风险.json> <输出.xlsx>

布局: 第 1 行免责声明（合并单元格，OUT-COM-001）、第 2 行表头、第 3 行起数据。
列: # / 条款 / 风险点 / 风险等级 / 修改建议 / 依据来源 / 来源
视觉: 表头黑底白字(0A0D12)、细灰边框(E2E5EA)、偶数行斑马纹(F8F9FA)、
风险等级列浅状态底（高 FEF3F2/D92D20、中 FFFAEB/B54708、低 ECFDF3/039855）、
冻结表头并启用筛选、数据单元格上下居中自动换行、无 emoji（OUT-EXCEL-001/002/003）。
openpyxl 缺失时降级输出同名 .csv（首行同样写免责声明）并提示。
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

HEADERS = ["#", "条款", "风险点", "风险等级", "修改建议", "依据来源", "来源"]
DISCLAIMER = ("本文档由 AI 辅助生成，仅供参考，不构成正式法律意见，"
              "不能替代具有执业资格的律师出具的专业法律意见。")
COL_WIDTHS = [6, 24, 40, 10, 46, 28, 14]
LEVEL_STYLE = {
    "高": ("FEF3F2", "D92D20"),
    "中": ("FFFAEB", "B54708"),
    "低": ("ECFDF3", "039855"),
}
SOURCE_LABEL = {
    "local": "本地 AI 审查",
    "engine": "法大大审查引擎",
    "both": "融合",
}


def load_items(path: Path) -> tuple[list[dict], dict]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict) and "data" in payload and isinstance(payload["data"], dict):
        payload = payload["data"]
    return payload.get("items", []), payload.get("merge_summary", {})


def basis_text(item: dict) -> str:
    tag = str(item.get("basis_tag", "")).strip()
    detail = str(item.get("basis_detail", "")).strip()
    if tag and detail:
        return f"{tag}：{detail}"
    return tag or detail or "/"


def rows_from_items(items: list[dict]) -> list[list[str]]:
    rows = []
    for item in items:
        suggestion = str(item.get("suggestion", "")).strip()
        engine_suggestion = str(item.get("engine_suggestion", "")).strip()
        if engine_suggestion:
            suggestion = f"{suggestion}\n引擎建议：{engine_suggestion}"
        rows.append([
            str(item.get("index", "")),
            str(item.get("clause", "")).strip(),
            str(item.get("issue", "")).strip(),
            str(item.get("risk_level", "")).strip(),
            suggestion,
            basis_text(item),
            SOURCE_LABEL.get(str(item.get("source", "")).strip(), "本地 AI 审查"),
        ])
    return rows


def write_xlsx(rows: list[list[str]], output: Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "风险清单"

    thin = Side(style="thin", color="E2E5EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="0A0D12")
    header_font = Font(name="PingFang SC", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="PingFang SC", size=10, color="0A0D12")
    zebra_fill = PatternFill("solid", fgColor="F8F9FA")
    wrap = Alignment(vertical="center", wrap_text=True)

    last_col = get_column_letter(len(HEADERS))
    ws.merge_cells(f"A1:{last_col}1")
    disclaimer_cell = ws.cell(row=1, column=1, value=DISCLAIMER)
    disclaimer_cell.font = Font(name="PingFang SC", size=9, color="667085")
    disclaimer_cell.fill = PatternFill("solid", fgColor="F2F4F7")
    disclaimer_cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    header_row = 2
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[col - 1]

    level_col = HEADERS.index("风险等级") + 1
    for r, row in enumerate(rows, start=header_row + 1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = body_font
            cell.border = border
            cell.alignment = wrap
            if (r - header_row) % 2 == 0:
                cell.fill = zebra_fill
            if c == level_col and value in LEVEL_STYLE:
                bg, fg = LEVEL_STYLE[value]
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="PingFang SC", size=10, bold=True, color=fg)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(rows)}"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def write_csv_fallback(rows: list[list[str]], output: Path) -> Path:
    import csv
    csv_path = output.with_suffix(".csv")
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([DISCLAIMER])
        writer.writerow(HEADERS)
        writer.writerows(rows)
    return csv_path


def main() -> None:
    if len(sys.argv) < 3:
        print(json.dumps({
            "success": False,
            "error": "参数不足",
            "usage": "python build_risk_list.py <风险.json> <输出.xlsx>",
        }, ensure_ascii=False))
        sys.exit(1)

    source = Path(sys.argv[1])
    output = Path(sys.argv[2])

    try:
        items, merge_summary = load_items(source)
    except Exception as exc:
        print(json.dumps({"success": False, "error": f"读取风险 JSON 失败: {exc}"},
                         ensure_ascii=False))
        sys.exit(1)
    if not items:
        print(json.dumps({"success": False, "error": "风险 JSON 中无 items"},
                         ensure_ascii=False))
        sys.exit(1)

    rows = rows_from_items(items)
    try:
        write_xlsx(rows, output)
        result = {"success": True, "output": str(output), "rows": len(rows)}
    except ImportError:
        csv_path = write_csv_fallback(rows, output)
        result = {
            "success": True,
            "output": str(csv_path),
            "rows": len(rows),
            "warning": "openpyxl 缺失，已降级输出 CSV；如需 Excel 请 pip install openpyxl",
        }
    if merge_summary:
        result["merge_summary"] = merge_summary
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
