#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从两份 xlsx 数据源重建资源目录 reference 文件。

输入（references/data/）：
  - 法律人常用网站.xlsx          中国为主的常用法律网站（11 分类）
  - legaldatahunter_sources_analysis.xlsx  LegalDataHunter 全球法律数据源（1931 条）

输出（references/）：
  - sources-china.md     中国一手法源，按业务领域分类（剔除非法源类）
  - sources-global.md    全球高置信源，按地区→国家→数据类型分类
  - source-index-tableB.md  国家/地区→sources-global.md 锚点（供手写 source-index.md 引用）

复现：xlsx 更新后重跑 `python scripts/build_sources.py` 即可。
所有条目均保留原始 URL，确保可点击核验、禁止凭训练数据编造。
"""
import os
import openpyxl
from collections import defaultdict, OrderedDict

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL = os.path.dirname(HERE)
DATA = os.path.join(SKILL, "references", "data")
REF = os.path.join(SKILL, "references")

CHINA_XLSX = os.path.join(DATA, "法律人常用网站.xlsx")
LDH_XLSX = os.path.join(DATA, "legaldatahunter_sources_analysis.xlsx")

# ---- 中国源：剔除的非一手法源分类 ----
CHINA_DROP = {"效率工具", "法律学习", "数据报告查询"}
# 业务领域呈现顺序（一手法源优先）
CHINA_ORDER = [
    "司法案例", "法律规定与标准", "知识产权", "资本市场",
    "争议解决", "主体信息查询", "行业政府部门", "涉外港澳台相关",
]

# ---- 全球源：地区呈现顺序 ----
REGION_ORDER = ["国际/区域组织", "欧洲", "美洲", "亚洲", "大洋洲", "非洲"]

# 数据类型规范化：raw 片段 -> 短标签
TYPE_MAP = {
    "判例/裁判文书": "判例",
    "法律法规/规章": "法规",
    "学说/评论/指南": "学说",
    "议会记录/立法过程": "立法过程",
}
TYPE_ORDER = ["法规", "判例", "学说", "立法过程"]


def norm_types(raw):
    if not raw:
        return []
    parts = [p.strip() for p in str(raw).replace("；", ";").split(";") if p.strip()]
    out = []
    for p in parts:
        t = TYPE_MAP.get(p, p)
        if t not in out:
            out.append(t)
    return out


def md_escape(s):
    if s is None:
        return ""
    return str(s).replace("|", "\\|").replace("\n", " ").strip()


def stars(auth, commercial):
    # 认证要求 none + 可商用 => 五星，否则四星
    if (auth in (None, "", "none")) and commercial:
        return "⭐⭐⭐⭐⭐"
    return "⭐⭐⭐⭐"


# =========================================================
# 1) 中国源 -> sources-china.md
# =========================================================
def build_china():
    wb = openpyxl.load_workbook(CHINA_XLSX, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))[1:]
    by_cat = defaultdict(list)
    for r in rows:
        cat, name, desc, url, owner = (list(r) + [None] * 5)[:5]
        if not cat or cat in CHINA_DROP:
            continue
        if not name or not url:
            continue
        by_cat[cat].append((name, desc, url, owner))

    total = sum(len(v) for v in by_cat.values())
    cats = [c for c in CHINA_ORDER if c in by_cat] + \
           [c for c in by_cat if c not in CHINA_ORDER]

    lines = []
    lines.append("# 中国一手法律数据源目录（按业务领域）")
    lines.append("")
    lines.append("> 数据来源：`法律人常用网站.xlsx`（运营方整理）。本目录仅收录**一手/权威法源**")
    lines.append("> （已剔除效率工具、法律学习、数据报告查询等非法源类），共 **%d** 条，覆盖 %d 个业务领域。"
                 % (total, len(cats)))
    lines.append("> **援引规则**：下表 URL 为候选入口；运行时须经 WebFetch 实时核验并锚定具体法条/案号后方可引用，")
    lines.append("> 禁止凭训练数据直接给出法规或判例内容。详见 `verification-engine.md`。")
    lines.append("")
    lines.append("## 业务领域速查")
    lines.append("")
    lines.append("| 业务领域 | 条数 | 锚点 |")
    lines.append("|---|---|---|")
    for c in cats:
        anchor = "cn-" + c
        lines.append("| %s | %d | [#%s](#%s) |" % (c, len(by_cat[c]), anchor, anchor))
    lines.append("")
    lines.append("---")
    lines.append("")
    for c in cats:
        lines.append('<a name="cn-%s"></a>' % c)
        lines.append("## %s（%d）" % (c, len(by_cat[c])))
        lines.append("")
        lines.append("| 名称 | 简介/用途 | 链接（核验入口） | 运营主体 |")
        lines.append("|---|---|---|---|")
        for name, desc, url, owner in by_cat[c]:
            lines.append("| %s | %s | %s | %s |" % (
                md_escape(name), md_escape(desc), md_escape(url), md_escape(owner)))
        lines.append("")
    out = os.path.join(REF, "sources-china.md")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print("wrote %s  (%d sources, %d categories)" % (out, total, len(cats)))
    return total


# =========================================================
# 2) 全球源 -> sources-global.md  +  source-index-tableB.md
# =========================================================
def build_global():
    wb = openpyxl.load_workbook(LDH_XLSX, read_only=True, data_only=True)
    ws = wb["Sources"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    idx = {name: i for i, name in enumerate(hdr)}

    def g(r, name):
        return r[idx[name]] if name in idx else None

    # 过滤：高置信 + 可商用 + 有 URL
    useful = []
    for r in rows[1:]:
        if g(r, "置信级别") != "高":
            continue
        if g(r, "是否可商用") not in (True, "TRUE", "true"):
            continue
        if not g(r, "网址"):
            continue
        useful.append(r)

    # region -> country(code,name) -> [sources]
    tree = defaultdict(lambda: OrderedDict())
    for r in useful:
        region = g(r, "地区") or "其他"
        code = g(r, "国家/地区代码") or "XX"
        cname = g(r, "国家/地区名称") or code
        key = (code, cname)
        tree[region].setdefault(key, []).append(r)

    regions = [x for x in REGION_ORDER if x in tree] + \
              [x for x in tree if x not in REGION_ORDER]

    total = len(useful)
    ncountries = sum(len(v) for v in tree.values())

    lines = []
    lines.append("# 全球法律数据源目录（按地区 → 国家/地区 → 数据类型）")
    lines.append("")
    lines.append("> 数据来源：`legaldatahunter_sources_analysis.xlsx`（LegalDataHunter，抓取于 2026-05-21）。")
    lines.append("> 收录口径：**置信级别=高 ∧ 可商用 ∧ 有 URL**，共 **%d** 条，覆盖 **%d** 个国家/地区。"
                 % (total, ncountries))
    lines.append("> 本目录是离线快照；运行时国家代码与 Source ID 以 LDH 实时 discover 目录为准。")
    lines.append("> 数据类型：法规=法律法规/规章，判例=判例/裁判文书，学说=学说/评论/指南。")
    lines.append("> **援引规则**：URL 为官方候选入口；运行时须 WebFetch 核验可达 + 锚定法条号/案号后方可引用，")
    lines.append("> 禁止凭训练数据编造法规与判例。详见 `verification-engine.md`。")
    lines.append("")
    lines.append("## 地区速查")
    lines.append("")
    lines.append("| 地区 | 国家/地区数 | 源数 |")
    lines.append("|---|---|---|")
    for region in regions:
        rc = sum(len(v) for v in tree[region].values())
        lines.append("| %s | %d | %d |" % (region, len(tree[region]), rc))
    lines.append("")
    lines.append("---")
    lines.append("")

    tableB = []  # (region, code, cname, nsrc)

    for region in regions:
        lines.append("# 地区：%s" % region)
        lines.append("")
        countries = sorted(tree[region].items(), key=lambda kv: kv[0][1])
        for (code, cname), srcs in countries:
            tableB.append((region, code, cname, len(srcs)))
            lines.append('<a name="c-%s"></a>' % code)
            lines.append("## %s（%s）— %d 源" % (cname, code, len(srcs)))
            lines.append("")
            lines.append("| Source ID | 名称 | 类型 | 链接（核验入口） | 场景用途 | 质量 |")
            lines.append("|---|---|---|---|---|---|")
            # 同国内按 类型(法规>判例>学说) 再按名称排序
            def sortkey(r):
                ts = norm_types(g(r, "数据类型"))
                ti = min([TYPE_ORDER.index(t) for t in ts if t in TYPE_ORDER] or [99])
                return (ti, str(g(r, "网站名称") or ""))
            for r in sorted(srcs, key=sortkey):
                sid = g(r, "Source ID") or ""
                name = g(r, "网站名称") or ""
                types = "/".join(norm_types(g(r, "数据类型"))) or "—"
                url = g(r, "网址") or ""
                use = g(r, "场景用途") or ""
                star = stars(g(r, "认证要求"), g(r, "是否可商用"))
                lines.append("| %s | %s | %s | %s | %s | %s |" % (
                    md_escape(sid), md_escape(name), md_escape(types),
                    md_escape(url), md_escape(use), star))
            lines.append("")

    out = os.path.join(REF, "sources-global.md")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print("wrote %s  (%d sources, %d countries)" % (out, total, ncountries))

    # source-index 表 B（国家→锚点），按地区分组
    bl = []
    bl.append("<!-- 由 build_sources.py 生成；粘贴进 source-index.md 的「表 B」 -->")
    bl.append("")
    by_region = defaultdict(list)
    for region, code, cname, n in tableB:
        by_region[region].append((cname, code, n))
    for region in regions:
        bl.append("### %s" % region)
        bl.append("")
        items = sorted(by_region[region], key=lambda x: x[0])
        # 紧凑成一行行链接
        cells = ["[%s(%s)·%d](sources-global.md#c-%s)" % (cn, cd, n, cd)
                 for cn, cd, n in items]
        bl.append(" · ".join(cells))
        bl.append("")
    out2 = os.path.join(REF, "source-index-tableB.md")
    with open(out2, "w", encoding="utf-8") as f:
        f.write("\n".join(bl) + "\n")
    print("wrote %s" % out2)
    return total, ncountries


if __name__ == "__main__":
    cn = build_china()
    gt, gc = build_global()
    print("\nSUMMARY: china=%d  global=%d (countries=%d)" % (cn, gt, gc))
