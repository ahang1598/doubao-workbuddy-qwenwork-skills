#!/usr/bin/env python3
"""本地模式案件台账（xlsx）的读写脚本，替代原 Markdown 台账读写。

用法：
  python3 ledger_excel.py init    [--path cases/案件台账.xlsx]
  python3 ledger_excel.py read    --path cases/案件台账.xlsx [--field 案件编号 --value A001]
  python3 ledger_excel.py append  --path cases/案件台账.xlsx --json '{"案件编号":"A001",...}'
  python3 ledger_excel.py update  --path cases/案件台账.xlsx --field 案件编号 --value A001 --json '{"当前阶段":"一审",...}'

约定：
  - 表头为 20 项最小字段集（见冷启动访谈 references/aitable-template.md），容忍律师扩展列。
  - 每次写入前把上一版备份为 案件台账.backup.xlsx。
  - 文件被 Excel 占用导致写入失败时报错退出，不静默跳过。
  - append/update 未显式给「最后更新」时自动填当天日期。

退出码：0 成功；1 入参/数据错误；2 文件被占用或 IO 错误；3 表头结构缺 20 项字段；4 缺 openpyxl。
"""

import argparse
import datetime
import json
import os
import shutil
import sys

SHEET_NAME = "案件台账"

REQUIRED_FIELDS = [
    "案件编号", "案件名称", "案由", "我方当事人与地位", "对方当事人",
    "标的额", "审理机关", "案号", "主办人", "当前阶段",
    "风险等级", "下一期限与描述", "举证届满", "开庭日", "上诉期届满",
    "当前进展", "下一步动作", "结案日与结果", "最后更新", "备注",
]


def load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        print("缺少依赖 openpyxl，请先执行：pip install openpyxl", file=sys.stderr)
        sys.exit(4)
    return Workbook, load_workbook


def backup(path: str) -> None:
    if not os.path.exists(path):
        return
    backup_path = os.path.join(os.path.dirname(path) or ".", "案件台账.backup.xlsx")
    try:
        shutil.copy2(path, backup_path)
    except OSError as exc:
        print(f"备份失败（{exc}），为安全起见中止写入。", file=sys.stderr)
        sys.exit(2)


def open_sheet(path: str, create_if_missing: bool):
    """返回 (wb, ws, headers)。headers 为首行字段名列表。"""
    from openpyxl import Workbook
    _, load_workbook = load_openpyxl()
    if not os.path.exists(path):
        if not create_if_missing:
            print(f"台账文件不存在：{path}。请先运行 init 子命令创建。", file=sys.stderr)
            sys.exit(1)
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        for col, name in enumerate(REQUIRED_FIELDS, start=1):
            ws.cell(row=1, column=col, value=name)
        return wb, ws, list(REQUIRED_FIELDS)

    wb = load_workbook(path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    headers = [cell.value for cell in ws[1]]
    headers = [str(h) for h in headers if h is not None]
    missing = [f for f in REQUIRED_FIELDS if f not in headers]
    if missing:
        print(f"台账表头缺少必备字段：{', '.join(missing)}。请补齐表头后重试。", file=sys.stderr)
        sys.exit(3)
    return wb, ws, headers


def save(wb, path: str) -> None:
    try:
        wb.save(path)
    except PermissionError:
        print(
            f"写入失败：{path} 可能正被 Excel 或其他程序打开。"
            "请关闭文件后重试；本次写入未生效，上一版已备份。",
            file=sys.stderr,
        )
        sys.exit(2)
    except OSError as exc:
        print(f"写入失败（{exc}）。", file=sys.stderr)
        sys.exit(2)


def rows_as_dicts(ws, headers):
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        record = {}
        for idx, name in enumerate(headers):
            if idx < len(row):
                value = row[idx]
                if isinstance(value, datetime.datetime):
                    value = value.strftime("%Y-%m-%d")
                record[name] = value
        out.append(record)
    return out


def parse_json_arg(raw: str) -> dict:
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        print(f"--json 解析失败：{exc}", file=sys.stderr)
        sys.exit(1)
    if not isinstance(data, dict) or not data:
        print("--json 必须是非空对象。", file=sys.stderr)
        sys.exit(1)
    return data


def cmd_init(args):
    backup_path_exists = os.path.exists(args.path)
    if backup_path_exists:
        backup(args.path)
    wb, ws, headers = open_sheet(args.path, create_if_missing=True)
    save(wb, args.path)
    print(json.dumps({"status": "ok", "path": args.path, "headers": headers}, ensure_ascii=False))
    return 0


def cmd_read(args):
    wb, ws, headers = open_sheet(args.path, create_if_missing=False)
    rows = rows_as_dicts(ws, headers)
    if args.field and args.value:
        if args.field not in headers:
            print(f"筛选字段不在表头中：{args.field}", file=sys.stderr)
            return 1
        rows = [r for r in rows if str(r.get(args.field) or "").strip() == args.value.strip()]
    print(json.dumps(rows, ensure_ascii=False, indent=1))
    return 0


def cmd_append(args):
    data = parse_json_arg(args.json)
    backup(args.path)
    wb, ws, headers = open_sheet(args.path, create_if_missing=True)

    unknown = [k for k in data if k not in headers]
    if unknown:
        print(f"以下字段不在表头中，拒绝写入：{', '.join(unknown)}", file=sys.stderr)
        return 1

    if "案件编号" in data:
        existing = [
            r for r in rows_as_dicts(ws, headers)
            if str(r.get("案件编号") or "").strip() == str(data["案件编号"]).strip()
        ]
        if existing:
            print(f"案件编号已存在：{data['案件编号']}，建档请改用新编号，刷新请改用 update。", file=sys.stderr)
            return 1

    data.setdefault("最后更新", datetime.date.today().isoformat())
    next_row = ws.max_row + 1
    for col, name in enumerate(headers, start=1):
        if name in data:
            ws.cell(row=next_row, column=col, value=data[name])
    save(wb, args.path)
    print(json.dumps({"status": "ok", "action": "append", "row": next_row, "data": data}, ensure_ascii=False))
    return 0


def cmd_update(args):
    data = parse_json_arg(args.json)
    if not args.field or not args.value:
        print("update 必须提供 --field 与 --value 用于定位案件行。", file=sys.stderr)
        return 1

    wb, ws, headers = open_sheet(args.path, create_if_missing=False)
    if args.field not in headers:
        print(f"定位字段不在表头中：{args.field}", file=sys.stderr)
        return 1
    unknown = [k for k in data if k not in headers]
    if unknown:
        print(f"以下字段不在表头中，拒绝写入：{', '.join(unknown)}", file=sys.stderr)
        return 1

    matched_rows = []
    for row_idx in range(2, ws.max_row + 1):
        col_idx = headers.index(args.field) + 1
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        if cell_value is not None and str(cell_value).strip() == args.value.strip():
            matched_rows.append(row_idx)
    if not matched_rows:
        print(f"未找到 {args.field} = {args.value} 的案件行，请先确认定位信息。", file=sys.stderr)
        return 1
    if len(matched_rows) > 1:
        print(f"{args.field} = {args.value} 命中 {len(matched_rows)} 行，请改用更精确的定位（如案件编号）。", file=sys.stderr)
        return 1

    backup(args.path)
    data.setdefault("最后更新", datetime.date.today().isoformat())
    row_idx = matched_rows[0]
    for col, name in enumerate(headers, start=1):
        if name in data:
            ws.cell(row=row_idx, column=col, value=data[name])
    save(wb, args.path)
    print(json.dumps({"status": "ok", "action": "update", "row": row_idx, "data": data}, ensure_ascii=False))
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="本地模式案件台账（xlsx）读写")
    sub = ap.add_subparsers(dest="command", required=True)

    p_init = sub.add_parser("init", help="创建台账文件（含 20 项表头），已存在则校验表头")
    p_init.add_argument("--path", default="cases/案件台账.xlsx")
    p_init.set_defaults(func=cmd_init)

    p_read = sub.add_parser("read", help="读取台账，输出 JSON")
    p_read.add_argument("--path", required=True)
    p_read.add_argument("--field", help="筛选字段名")
    p_read.add_argument("--value", help="筛选值（精确匹配）")
    p_read.set_defaults(func=cmd_read)

    p_append = sub.add_parser("append", help="新增一行（建档）")
    p_append.add_argument("--path", required=True)
    p_append.add_argument("--json", required=True, help="字段 JSON 对象")
    p_append.set_defaults(func=cmd_append)

    p_update = sub.add_parser("update", help="刷新指定案件行的字段")
    p_update.add_argument("--path", required=True)
    p_update.add_argument("--field", help="定位字段名，如 案件编号")
    p_update.add_argument("--value", help="定位值")
    p_update.add_argument("--json", required=True, help="要刷新的字段 JSON 对象")
    p_update.set_defaults(func=cmd_update)

    args = ap.parse_args()
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
