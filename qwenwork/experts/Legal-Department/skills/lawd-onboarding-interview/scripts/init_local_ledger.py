#!/usr/bin/env python3
"""创建初始案件台账 Excel：cases/案件台账.xlsx，表头为 20 项最小字段集。

退出码：0 创建成功；1 文件已存在（保留原文件不覆盖）；2 依赖缺失或 IO 错误。
字段规格来源：references/aitable-template.md 第一节。
"""

import argparse
import os
import sys

SHEET_NAME = "案件台账"

FIELDS = [
    "案件编号", "案件名称", "案由", "我方当事人与地位", "对方当事人",
    "标的额", "审理机关", "案号", "主办人", "当前阶段",
    "风险等级", "下一期限与描述", "举证届满", "开庭日", "上诉期届满",
    "当前进展", "下一步动作", "结案日与结果", "最后更新", "备注",
]

DATE_FIELDS = {"举证届满", "开庭日", "上诉期届满", "最后更新"}


def main() -> int:
    ap = argparse.ArgumentParser(description="创建初始案件台账 Excel（20 列表头）")
    ap.add_argument("--path", default="cases/案件台账.xlsx", help="目标 xlsx 路径")
    args = ap.parse_args()
    path = args.path

    if os.path.exists(path):
        print(f"文件已存在：{path}，保留原文件不覆盖，直接沿用。")
        return 1

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("缺少依赖 openpyxl，请先执行：pip install openpyxl", file=sys.stderr)
        return 2

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for col, name in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        width = 12 if name in DATE_FIELDS else max(14, len(name) * 2 + 4)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}1"

    dirname = os.path.dirname(path)
    if dirname:
        os.makedirs(dirname, exist_ok=True)
    try:
        wb.save(path)
    except PermissionError:
        print(f"写入失败：{path} 可能被 Excel 或其他程序占用，请关闭后重试。", file=sys.stderr)
        return 2
    except OSError as exc:
        print(f"写入失败（{exc}）：请确认目录可写。", file=sys.stderr)
        return 2

    print(f"已创建：{path}（Sheet: {SHEET_NAME}，表头 {len(FIELDS)} 列）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
