#!/usr/bin/env python3
"""
extract_tables.py — PDF Table Extraction

Detects and extracts tables from PDF pages, exporting them as
CSV files or a multi-sheet Excel workbook.

For xlsx output, columns that look like plain numbers (optionally with
thousands separators, e.g. "287,000" / "12,580.35") are converted to
numeric dtype. Columns that consistently match a percent pattern
("23.43%") are converted to the fraction form (0.2343) and the Excel
number_format "0.00%" is applied post-write, so they still display as
"23.43%" to the user while supporting SUM/AVG.

Columns are left as strings when:
  - the column name matches an ID-like keyword (身份证 / 手机 / 银行卡 / 账号 …);
  - any value has a leading zero or exceeds 15 pure digits (float precision);
  - any value carries non-percent semantic markers: ¥/$/€, parenthetical
    negatives, unit suffixes (万/亿), dates, etc. — these require semantic
    interpretation that a rule cannot safely do and are left to the user
    or to a future LLM-driven schema step.

Disable inference entirely with `--no-infer-types`.

Usage:
    python scripts/extract_tables.py report.pdf --output tables/
    python scripts/extract_tables.py report.pdf --output tables/ --format xlsx
    python scripts/extract_tables.py report.pdf --output all_tables.xlsx --format xlsx --single-file
    python scripts/extract_tables.py report.pdf --pages 1-3 --output tables/
    python scripts/extract_tables.py report.pdf --output tables/ --format xlsx --no-infer-types
"""

from __future__ import annotations

import argparse
import csv
import importlib.util
import shutil
import sys
import tempfile
from pathlib import Path


SCRIPT_DIRECTORY = Path(__file__).resolve().parent
if str(SCRIPT_DIRECTORY) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIRECTORY))
from _cloud_runtime import CloudRuntimeError, resolve_qwenwork_cli, run_document_tool
from _execution_route import BackendFailure
from _semantic_route import execute_semantic_script


def _page_spec_to_indices(spec: str, total: int) -> list[int]:
    indices: set[int] = set()
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            lo, hi = part.split("-", 1)
            for p in range(int(lo), int(hi) + 1):
                if 1 <= p <= total:
                    indices.add(p - 1)
        else:
            p = int(part)
            if 1 <= p <= total:
                indices.add(p - 1)
    return sorted(indices)


def _clean_cell(value) -> str:
    """Normalize a table cell to a clean string."""
    if value is None:
        return ""
    return str(value).strip().replace("\n", " ")


# Column-name hints for ID-like fields that must stay string:
# - 身份证 / 银行卡: > 15 digits, float64 precision loss
# - 手机号 / 电话号码: 11 digits, Excel shows as 1.38E+10 without text format
FORCE_STRING_KEYWORDS = (
    "身份证", "证件号", "id_card", "idcard",
    "银行卡", "卡号", "账号", "account",
    "手机", "电话", "mobile", "phone", "tel",
)


def _column_should_stay_string(col_name: str) -> bool:
    name = str(col_name).lower().replace("_", "").replace(" ", "")
    for kw in FORCE_STRING_KEYWORDS:
        if kw.lower().replace("_", "") in name:
            return True
    return False


def _column_has_risky_values(values) -> bool:
    """Return True if any value would lose information when cast to a number."""
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        # Leading zero on a pure-digit string → would lose the "0"
        if s.startswith("0") and len(s) > 1 and s.isdigit():
            return True
        # Pure digits longer than 15 chars → exceeds float64 precision
        if s.isdigit() and len(s) > 15:
            return True
    return False


def _try_convert_numeric(series):
    """
    Strip thousands separators only, then attempt pd.to_numeric. Return
    (success_ratio, converted_series) where success_ratio is the fraction of
    non-empty values that parsed successfully.

    Conservative on purpose: values carrying semantic markers (%, ¥/$/€,
    parenthetical negatives, unit suffixes like "万") are NOT stripped, so
    they fail to_numeric and the column stays as string. This avoids silently
    discarding the % / currency / negative convention, which a pure rule can't
    decide correctly.
    """
    import pandas as pd

    as_str = series.astype(str).str.strip()
    cleaned = as_str.str.replace(",", "", regex=False)
    numeric = pd.to_numeric(cleaned, errors="coerce")

    non_empty_mask = as_str.replace("", pd.NA).notna()
    total = int(non_empty_mask.sum())
    if total == 0:
        return 0.0, series
    success = int(numeric[non_empty_mask].notna().sum()) / total
    return success, numeric


# Matches "23.43%", "-23.43%", "23%", "23,000%", "23.43 %" — all real percent values
_PERCENT_PATTERN = r"^-?[\d,]+(\.\d+)?\s*%$"


def _try_convert_percent(series):
    """
    Detect percent-format strings ("23.43%") and return fraction values
    (0.2343) so Excel can display them as percentages via number_format.
    Returns (success_ratio, converted_series).
    """
    import pandas as pd

    as_str = series.astype(str).str.strip()
    non_empty_mask = as_str.replace("", pd.NA).notna()
    total = int(non_empty_mask.sum())
    if total == 0:
        return 0.0, series

    match_mask = as_str.str.match(_PERCENT_PATTERN, na=False)
    success = int(match_mask.sum()) / total

    # Strip %, whitespace, commas, then to_numeric → divide by 100
    stripped = (as_str.str.replace("%", "", regex=False)
                .str.replace(",", "", regex=False)
                .str.strip())
    numeric = pd.to_numeric(stripped, errors="coerce") / 100.0
    return success, numeric


def _apply_percent_format(xlsx_path, sheet_percent_map):
    """
    Re-open the written xlsx and set number_format='0.00%' on each listed
    (sheet_name, [col_idx...]) pair. col_idx is 0-based DataFrame position.
    Row 1 (header) is skipped.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(str(xlsx_path))
    for sheet_name, col_indices in sheet_percent_map.items():
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        for col_idx in col_indices:
            col_letter = get_column_letter(col_idx + 1)
            for row_num in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row_num}"].number_format = "0.00%"
    wb.save(str(xlsx_path))


_WARNED_NO_ISETITEM = False


def _infer_column_types(df, threshold: float = 0.7):
    """
    Convert all-string columns to numeric when safe. Returns (df, percent_col_indices)
    where percent_col_indices is the 0-based positions of columns that were
    detected as percent (stored as fractions); callers should apply the Excel
    "0.00%" number_format on those columns so the original % reading is preserved.

    Rules (conservative, avoid false-positive ID conversion):
      1. Skip columns whose name matches an ID-like keyword.
      2. Skip columns that contain values with leading zeros or >15 digits.
      3. Try percent detection first — matching values are divided by 100 and
         the column index is collected for downstream format styling.
      4. Fall back to plain numeric (thousands-separator cleaning only).
      5. Only replace the column when >= threshold of non-empty values parse.

    Iterate by position to stay correct when column names are duplicated
    (common in balance-sheet style PDFs with side-by-side period columns).

    Requires pandas >= 2.0 for DataFrame.isetitem (positional, dtype-changing
    column replacement). On older pandas we skip type inference entirely and
    warn once — callers still get a valid string-typed DataFrame instead of
    an AttributeError crash.
    """
    global _WARNED_NO_ISETITEM
    if not hasattr(df, "isetitem"):
        if not _WARNED_NO_ISETITEM:
            import pandas as pd
            print(
                f"Warning: pandas=={pd.__version__} lacks DataFrame.isetitem "
                f"(added in 2.0). Skipping numeric/percent type inference; "
                f"xlsx columns will remain strings. Upgrade with: "
                f"pip install 'pandas>=2.0'",
                file=sys.stderr,
            )
            _WARNED_NO_ISETITEM = True
        return df, []

    percent_col_indices: list[int] = []
    for i in range(len(df.columns)):
        col_name = df.columns[i]
        series = df.iloc[:, i]
        if _column_should_stay_string(col_name):
            continue
        if _column_has_risky_values(series.tolist()):
            continue

        pct_ratio, pct_converted = _try_convert_percent(series)
        if pct_ratio >= threshold:
            df.isetitem(i, pct_converted)
            percent_col_indices.append(i)
            continue

        num_ratio, num_converted = _try_convert_numeric(series)
        if num_ratio >= threshold:
            # isetitem replaces by position and allows dtype change, which
            # plain .iloc[:, i] = ... refuses under pandas strict string dtype.
            df.isetitem(i, num_converted)
    return df, percent_col_indices


def _row_looks_like_data(row) -> bool:
    """
    Heuristic: the first row of a table is data (not a header) when any cell
    parses as a number. Header rows in real-world PDFs are column labels —
    almost always pure text. A numeric-looking first row usually means the
    table is a cross-page continuation, has no header, or the extractor
    missed the real header. In those cases the caller should treat all rows
    as data so type inference runs on row 0 too.
    """
    numeric_like = 0
    for cell in row:
        s = str(cell).strip()
        if not s:
            continue
        # Strip thousands separators and decimals; if what remains is digits
        # (optionally signed), the cell is a number.
        stripped = s.replace(",", "").replace(".", "").lstrip("-")
        if stripped and stripped.isdigit():
            numeric_like += 1
    return numeric_like > 0


def gather_tables(pdf_path: Path, page_indices: list[int] | None) -> list[dict]:
    """
    Return a list of table records, each with:
        page, table_index, headers, rows (list of lists of strings),
        has_detected_header (bool)
    """
    import pdfplumber

    collected: list[dict] = []
    with pdfplumber.open(str(pdf_path)) as doc:
        targets = page_indices if page_indices is not None else list(range(len(doc.pages)))
        for idx in targets:
            if idx >= len(doc.pages):
                continue
            page = doc.pages[idx]
            raw_tables = page.extract_tables() or []
            for tbl_idx, raw in enumerate(raw_tables):
                if not raw or not any(any(cell for cell in row) for row in raw):
                    continue  # skip empty tables
                cleaned = [[_clean_cell(cell) for cell in row] for row in raw]

                # Cross-page continuation tables have no header row — the
                # first row is already data. Detect this and avoid consuming
                # a data row as "headers"; otherwise the numeric values in
                # that row would never see _infer_column_types and would
                # land in xlsx as raw strings, poisoning downstream dtypes.
                if _row_looks_like_data(cleaned[0]):
                    headers = [f"col_{i + 1}" for i in range(len(cleaned[0]))]
                    data_rows = cleaned
                    has_detected_header = False
                else:
                    headers = cleaned[0]
                    data_rows = cleaned[1:] if len(cleaned) > 1 else []
                    has_detected_header = True

                collected.append({
                    "page": idx + 1,
                    "table_index": tbl_idx + 1,
                    "headers": headers,
                    "rows": data_rows,
                    "row_count": len(data_rows),
                    "col_count": len(headers),
                    "has_detected_header": has_detected_header,
                })
    return collected


def save_as_csv(tables: list[dict], output_dir: Path) -> list[Path]:
    """Write each table to its own CSV file."""
    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for tbl in tables:
        fname = output_dir / f"page{tbl['page']:03d}_table{tbl['table_index']:02d}.csv"
        with open(fname, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(tbl["headers"])
            writer.writerows(tbl["rows"])
        written.append(fname)
    return written


def save_as_xlsx_separate(tables: list[dict], output_dir: Path, infer_types: bool = True) -> list[Path]:
    """Write each table to its own XLSX file."""
    try:
        import pandas as pd
    except ImportError:
        print("Error: pandas required for XLSX output. Run: pip install pandas openpyxl",
              file=sys.stderr)
        sys.exit(1)

    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for tbl in tables:
        fname = output_dir / f"page{tbl['page']:03d}_table{tbl['table_index']:02d}.xlsx"
        df = pd.DataFrame(tbl["rows"], columns=tbl["headers"])
        percent_cols: list[int] = []
        if infer_types:
            df, percent_cols = _infer_column_types(df)
        df.to_excel(str(fname), index=False)
        if percent_cols:
            # pandas default sheet name when writing a single-sheet xlsx
            _apply_percent_format(fname, {"Sheet1": percent_cols})
        written.append(fname)
    return written


def save_as_xlsx_combined(tables: list[dict], output_path: Path, infer_types: bool = True) -> None:
    """Write all tables into one XLSX workbook, one sheet per table."""
    try:
        import pandas as pd
    except ImportError:
        print("Error: pandas required for XLSX output. Run: pip install pandas openpyxl",
              file=sys.stderr)
        sys.exit(1)

    sheet_percent_map: dict[str, list[int]] = {}
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        for tbl in tables:
            sheet = f"P{tbl['page']}_T{tbl['table_index']}"
            df = pd.DataFrame(tbl["rows"], columns=tbl["headers"])
            percent_cols: list[int] = []
            if infer_types:
                df, percent_cols = _infer_column_types(df)
            df.to_excel(writer, sheet_name=sheet, index=False)
            if percent_cols:
                sheet_percent_map[sheet] = percent_cols
    if sheet_percent_map:
        _apply_percent_format(output_path, sheet_percent_map)


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract tables from a PDF file.")
    parser.add_argument("pdf_file", help="Input PDF path")
    parser.add_argument("--output", "-o", required=True,
                        help="Output directory (CSV/xlsx separate) or file path (xlsx combined)")
    parser.add_argument("--format", choices=["csv", "xlsx"], default="csv",
                        help="Output format (default: csv)")
    parser.add_argument("--single-file", action="store_true",
                        help="Combine all tables into one XLSX workbook (xlsx format only)")
    parser.add_argument("--pages", help="Page range spec e.g. '1-5'")
    parser.add_argument("--no-infer-types", action="store_true",
                        help="Disable automatic numeric type inference for xlsx output")
    args = parser.parse_args()

    pdf_path = Path(args.pdf_file)
    if not pdf_path.exists():
        print(f"Error: file not found: {pdf_path}", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output)

    def local_ready() -> bool:
        modules = ["pdfplumber"]
        if args.format == "xlsx":
            modules.extend(["pandas", "openpyxl"])
        return all(importlib.util.find_spec(module) is not None for module in modules)

    def cloud_ready() -> bool:
        try:
            return resolve_qwenwork_cli(required=False) is not None
        except RuntimeError:
            return False

    def run_cloud() -> Path:
        flags: list[tuple[str, str | None]] = [("format", args.format)]
        if args.pages:
            flags.append(("pages", args.pages))
        if args.no_infer_types:
            flags.append(("no-infer-types", None))
        if args.single_file:
            flags.append(("single-file", None))
        try:
            if args.single_file:
                output_path.parent.mkdir(parents=True, exist_ok=True)
                with tempfile.TemporaryDirectory(
                    prefix="qwenwork-tables-", dir=output_path.parent,
                ) as temporary_directory:
                    cloud_directory = Path(temporary_directory) / "artifacts"
                    run_document_tool(
                        ("document", "pdf", "extract-tables"),
                        pdf_path,
                        save_path=cloud_directory,
                        flags=tuple(flags),
                    )
                    candidates = sorted(cloud_directory.glob("*.xlsx"))
                    if len(candidates) != 1:
                        raise RuntimeError("cloud table extraction did not return one workbook")
                    temporary_output = output_path.with_name(f".{output_path.name}.tmp")
                    shutil.copy2(candidates[0], temporary_output)
                    temporary_output.replace(output_path)
            else:
                run_document_tool(
                    ("document", "pdf", "extract-tables"),
                    pdf_path,
                    save_path=output_path,
                    flags=tuple(flags),
                )
        except CloudRuntimeError as exc:
            raise BackendFailure(exc.code, retryable=exc.retryable) from exc
        except RuntimeError as exc:
            raise BackendFailure("CLOUD_OPERATION_FAILED", retryable=True) from exc
        return output_path

    def valid_output(path: Path) -> bool:
        if args.single_file:
            return path.is_file() and path.stat().st_size > 0 and path.suffix.lower() == ".xlsx"
        return path.is_dir() and all(
            candidate.is_file() and candidate.suffix.lower() == "." + args.format
            for candidate in path.iterdir()
        )

    try:
        handled = execute_semantic_script(
            argv=[str(Path(__file__).resolve()), *sys.argv[1:]],
            local_ready=local_ready,
            cloud_ready=cloud_ready,
            run_cloud=run_cloud,
            local_result=lambda: output_path,
            validate=valid_output,
        )
        if handled:
            return
    except BackendFailure as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)

    page_indices = None
    if args.pages:
        import pdfplumber
        with pdfplumber.open(str(pdf_path)) as doc:
            total = len(doc.pages)
        page_indices = _page_spec_to_indices(args.pages, total)

    print(f"Scanning {pdf_path} for tables…")
    tables = gather_tables(pdf_path, page_indices)

    if not tables:
        print("No tables found in the specified pages.")
        return

    print(f"Found {len(tables)} table(s)")
    infer_types = not args.no_infer_types

    if args.format == "csv":
        written = save_as_csv(tables, output_path)
        for p in written:
            print(f"  Wrote {p}")
    elif args.format == "xlsx" and args.single_file:
        save_as_xlsx_combined(tables, output_path, infer_types=infer_types)
        print(f"  Combined workbook → {output_path}")
    else:
        written = save_as_xlsx_separate(tables, output_path, infer_types=infer_types)
        for p in written:
            print(f"  Wrote {p}")

    # Flag tables without a detectable header so the caller knows to borrow
    # headers from the preceding page when merging cross-page continuations.
    no_header = [t for t in tables if not t.get("has_detected_header", True)]
    if no_header:
        print(
            f"Note: {len(no_header)} table(s) have no detectable header row (likely "
            f"cross-page continuations); their column names are placeholders "
            f"col_1, col_2, … — merge with the preceding table to restore real headers:"
        )
        for t in no_header:
            print(f"  - page {t['page']} table {t['table_index']}")

    print(f"Done. {len(tables)} table(s) exported.")


if __name__ == "__main__":
    main()
