"""
Excel Formula Verification & Recalculation
Validates all formulas in an Excel workbook via LibreOffice headless mode.
Also flags derived values that were written as static numbers instead of formulas.
"""

import argparse
import json
import os
import platform
import re
import shutil
import subprocess
import sys
from pathlib import Path
from typing import List, Optional, Set, Tuple

_SHARED_SCRIPTS = os.path.abspath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "..", "shared-assets", "scripts")
)
if _SHARED_SCRIPTS not in sys.path:
    sys.path.insert(0, _SHARED_SCRIPTS)

from lo_runtime import get_soffice_env

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

# 表头命中这些词 = 该行/列是「由其他单元格推导出来的量」，本就该写公式。
# 中文按子串匹配（字符级无误伤）；英文必须整词匹配（\b）——rate/sum/rank 等短串
# 会误命中 Corporate/Summary/Frank 这类普通列名，误报会硬阻断正确交付，比漏报更贵。
# index/share 因歧义过大（Index Number / Share Class 常为原始数据）不入词表。
DERIVED_HEADER_CN = re.compile(
    r"率|比例|占比|占%|合计|总计|小计|总和|增长|增速|同比|环比|排名|排序|名次"
    r"|平均|均值|累计|累积|差额|净额|毛利|利润|周转|复合|贡献|权重|方差|标准差|相关"
    r"|回报|收益|折旧|摊销|敞口|估值|倍数|系数|得分|评分|指数|客单价|人效"
)
DERIVED_HEADER_EN = re.compile(
    r"\b(rate|ratio|growth|rank|avg|average|total|sum|margin|cagr|yoy|mom|qoq"
    r"|pct|percent|contribution|weight|variance|stdev|std_?dev|correl"
    r"|return|score|npv|irr|ebitda|eps|arpu|arppu|ltv|roi|roe|roa|wacc)\b",
    re.IGNORECASE,
)


def _is_derived_header(text: str) -> bool:
    return bool(DERIVED_HEADER_CN.search(text) or DERIVED_HEADER_EN.search(text))


# 纯记录型表头，即使命中上面的词也不算派生（避免误报原始数据列）
RAW_HEADER = re.compile(r"^(日期|时间|date|time|id|编号|名称|name|备注|remark)", re.IGNORECASE)


def is_libreoffice_available() -> bool:
    return shutil.which("soffice") is not None


def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _formula_text(v) -> str:
    """普通公式、ArrayFormula、DataTableFormula 统一提取公式串；非公式返回空串。

    openpyxl 3.1 把数组公式 / 数据表公式读成对象（不是 "=" 开头的字符串），
    只认字符串会让这些格整体漏出公式计数、错误扫描与不变量扫描。
    """
    if isinstance(v, str):
        return v if v.startswith("=") else ""
    if type(v).__name__ in ("ArrayFormula", "DataTableFormula"):
        text = getattr(v, "text", None)
        if isinstance(text, str) and text:
            return text if text.startswith("=") else "=" + text
        return "="  # 无文本的公式对象：至少按「存在公式」处理
    return ""


def _is_formula(v) -> bool:
    return bool(_formula_text(v))


def _normalize_sheet_names(raw: Optional[str]) -> Tuple[Set[str], List[str]]:
    names = []
    if raw:
        names = [s.strip() for s in raw.split(",") if s.strip()]
    seen = set()
    uniq = []
    for name in names:
        if name not in seen:
            seen.add(name)
            uniq.append(name)
    return set(uniq), uniq


def _sheet_has_formula(ws, cap_cells: int = 500000) -> Tuple[bool, bool]:
    """全表短路扫描公式是否存在。返回 (found, complete)。

    找到第一个公式立即返回；扫满 cap_cells 仍未找到则 complete=False——
    抽样式的「未见公式」不允许升级为 high（会对大表误判整表无公式并阻断）。
    """
    seen = 0
    for row in ws.iter_rows():
        for cell in row:
            if _is_formula(cell.value):
                return True, True
            seen += 1
            if seen >= cap_cells:
                return False, False
    return False, True


def detect_hardcode_suspects(filename: str, max_report: int = 12, static_source_sheets: Optional[Set[str]] = None) -> dict:
    """扫描「表头表明是派生量、但整行/整列没有一个公式」的区域。

    只报告有足够证据的区域（≥3 个静态数值且 0 个公式），避免噪声。
    分两级：sheet 内完全无公式 = high；sheet 有公式但个别派生区域没有 = low。
    被显式声明为历史/外部静态来源的 sheet 不参与 high_confidence 计数，降为 low。
    """
    static_source_sheets = set(static_source_sheets or ())
    try:
        wb = load_workbook(filename, data_only=False)
    except Exception as e:
        return {"status": "skipped", "reason": str(e)}

    suspects = []
    sheets_scanned = 0
    unknown_static_source_sheets = []
    try:
        workbook_sheets = set(wb.sheetnames)
        unknown_static_source_sheets = sorted(static_source_sheets - workbook_sheets)
        declared_static_source_sheets = workbook_sheets & static_source_sheets
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.max_row < 2 or ws.max_column < 1:
                continue
            sheets_scanned += 1
            rows = list(ws.iter_rows(max_row=min(ws.max_row, 2000)))
            # 公式存在性单独全表短路扫描——前 2000 行的截断样本只用于表头/列画像，
            # 不允许决定「整表无公式」这种会升 high 并阻断退出码的判定
            has_formula, scan_complete = _sheet_has_formula(ws)
            is_declared_static_source = name in declared_static_source_sheets
            level = "high" if (not has_formula and scan_complete and not is_declared_static_source) else "low"

            # 表头行：前 5 行里字符串最多的那行
            header_idx, best = 0, -1
            for i, r in enumerate(rows[:5]):
                cnt = sum(1 for c in r if isinstance(c.value, str) and c.value.strip())
                if cnt > best:
                    header_idx, best = i, cnt
            header = rows[header_idx]

            # 列方向：表头命中派生词，列内全是静态数值
            for c in header:
                text = str(c.value).strip() if c.value is not None else ""
                if not text or RAW_HEADER.match(text) or not _is_derived_header(text):
                    continue
                col = c.column
                vals = [r[col - 1].value for r in rows[header_idx + 1:] if len(r) >= col]
                nums = sum(1 for v in vals if _is_number(v))
                fmls = sum(1 for v in vals if _is_formula(v))
                if nums >= 3 and fmls == 0:
                    suspects.append({
                        "level": level, "sheet": name, "axis": "column",
                        "header": text[:40], "static_cells": nums,
                        "declared_static_source": is_declared_static_source,
                    })

            # 行方向（横向年份布局）：首列行头命中派生词，行内全是静态数值
            for r in rows[header_idx + 1:]:
                if not r:
                    continue
                text = str(r[0].value).strip() if r[0].value is not None else ""
                if not text or not _is_derived_header(text):
                    continue
                vals = [c.value for c in r[1:]]
                nums = sum(1 for v in vals if _is_number(v))
                fmls = sum(1 for v in vals if _is_formula(v))
                if nums >= 3 and fmls == 0:
                    suspects.append({
                        "level": level, "sheet": name, "axis": "row",
                        "header": text[:40], "static_cells": nums,
                        "declared_static_source": is_declared_static_source,
                    })
    finally:
        wb.close()

    high = [s for s in suspects if s["level"] == "high"]
    return {
        "status": "suspected" if suspects else "clean",
        "total_suspects": len(suspects),
        "high_confidence": len(high),
        "sheets_scanned": sheets_scanned,
        "details": suspects[:max_report],
        "static_source_sheets": sorted(static_source_sheets),
        "unknown_static_source_sheets": unknown_static_source_sheets,
    }


HARDCODE_HINT = """\
[hardcode] 检出 {n} 处「表头表明是推导值、但整行/整列没有一个公式」的区域{extra}：
{lines}
这些格子应写成引用其他单元格的 Excel 公式，而不是先算好数值再写入。
结果对 != 合规：写死的表没有联动能力，用户改一个输入，全表不会更新。
处置：把上列区域改写为公式（假设 / 输入项集中放一处，由公式引用），改完重跑本脚本。
唯一可写静态值的情况：历史真实数据、供人修改的输入假设、标注了来源的外部取数。
若确属这三类，无需改写。"""


def _emit_hardcode_hint(hc: dict) -> bool:
    """在 stderr 打印处置指引。返回是否存在高置信嫌疑。"""
    details = hc.get("details") or []
    if not details:
        return False
    lines = "\n".join(
        f"  - {d['sheet']} {'列' if d['axis'] == 'column' else '行'}「{d['header']}」"
        f"（{d['static_cells']} 个静态数值，0 公式）"
        + (" [declared static source → WARN only]" if d.get("declared_static_source") else "")
        for d in details
    )
    extra = ""
    if hc["total_suspects"] > len(details):
        extra = f"（仅列出前 {len(details)} 处，共 {hc['total_suspects']} 处）"
    print(HARDCODE_HINT.format(n=hc["total_suspects"], extra=extra, lines=lines), file=sys.stderr)
    return hc.get("high_confidence", 0) > 0


EXCEL_ERROR_VALUES = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}


def _scan_formula_error_values(filename: str) -> tuple[int, dict]:
    """只统计公式格重算后的精确错误值；备注文本里的错误码不算公式错误。"""
    formula_wb = load_workbook(filename, data_only=False)
    value_wb = load_workbook(filename, data_only=True)
    error_details = {err: [] for err in EXCEL_ERROR_VALUES}
    total_errors = 0
    try:
        for sheet_name in formula_wb.sheetnames:
            if sheet_name not in value_wb.sheetnames:
                continue
            formula_ws = formula_wb[sheet_name]
            value_ws = value_wb[sheet_name]
            for row in formula_ws.iter_rows():
                for formula_cell in row:
                    if not _is_formula(formula_cell.value):
                        continue
                    value = value_ws[formula_cell.coordinate].value
                    if isinstance(value, str) and value.strip() in EXCEL_ERROR_VALUES:
                        error = value.strip()
                        error_details[error].append(f"{sheet_name}!{formula_cell.coordinate}")
                        total_errors += 1
    finally:
        formula_wb.close()
        value_wb.close()
    return total_errors, error_details


def scan_errors_only(filename: str, static_source_sheets: Optional[Set[str]] = None) -> dict:
    """当 LibreOffice 不可用时，仅扫描文件中已有的公式错误（不触发重算）。"""
    try:
        total_errors, error_details = _scan_formula_error_values(filename)
        result = {
            "status": "skipped_no_libreoffice",
            "warning": "LibreOffice 未安装，公式未重算，仅扫描已有错误值",
            "total_errors": total_errors,
            "error_summary": {k: {"count": len(v), "locations": v[:20]} for k, v in error_details.items() if v},
            "hardcode": detect_hardcode_suspects(filename, static_source_sheets=static_source_sheets),
            "invariants": evaluate_formula_invariants(filename),
        }
        return result
    except Exception as e:
        # 解析失败不是「未重算」这种可降级状态：损坏 / 非 xlsx 文件必须以 error
        # 进入退出码，否则坏文件会打印异常后 exit 0 被当成通过
        return {"status": "invalid_input",
                "error": f"cannot parse workbook: {e}",
                "invariants": {"status": "skipped", "reason": str(e), "categories": {}}}

# 宏目录是 LibreOffice 自身的固定安装路径（Basic 宏只认这里），不是本脚本可
# 自选的缓存位置——不随 lo_runtime 的 shim 缓存一起迁移到 ~/.cache。
MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def has_gtimeout():
    try:
        subprocess.run(
            ["gtimeout", "--version"], capture_output=True, timeout=1, check=False
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def setup_libreoffice_macro():
    macro_dir = os.path.expanduser(
        MACRO_DIR_MACOS if platform.system() == "Darwin" else MACRO_DIR_LINUX
    )
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if (
        os.path.exists(macro_file)
        and "RecalculateAndSave" in Path(macro_file).read_text()
    ):
        return True

    if not os.path.exists(macro_dir):
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
            env=get_soffice_env(),
        )
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def _safe_number(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return None


CELL_REF_RE = re.compile(r"(?<![A-Z0-9_])(?:'([^']+)'!)?(\$?[A-Z]{1,3}\$?\d+)")
RANGE_FUNC_RE = re.compile(r"^=(SUM|AVERAGE)\((.+)\)$", re.IGNORECASE)
ARITH_ALLOWED_RE = re.compile(r"^[=0-9A-Z_+$\-*/().,: '\t]+$", re.IGNORECASE)
SENSITIVITY_RE = re.compile(r"sensitivity|敏感性|scenario", re.IGNORECASE)
TIEOUT_ROW_RE = re.compile(r"assets?|liabilit(?:y|ies)|equity|cash|net change|total", re.IGNORECASE)


def _split_sheet_and_ref(expr: str, current_sheet: str) -> tuple[str, str]:
    if "!" in expr:
        sheet, ref = expr.split("!", 1)
        sheet = sheet.strip().strip("'")
        return sheet, ref
    return current_sheet, expr


def _cell_value(ws_map, sheet_name: str, coord: str):
    ws = ws_map.get(sheet_name)
    if ws is None:
        raise KeyError(sheet_name)
    return ws[coord.replace("$", "")].value


def _iter_range_values(ws_map, sheet_name: str, range_expr: str):
    start_col, start_row, end_col, end_row = range_boundaries(range_expr.replace("$", ""))
    ws = ws_map[sheet_name]
    values = []
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            values.append(ws.cell(row=row, column=col).value)
    return values


def _normalize_formula_expr(expr: str, current_sheet: str):
    refs = []
    placeholder_map = {}
    counter = 0

    def repl(match):
        nonlocal counter
        sheet = match.group(1) or current_sheet
        ref = match.group(2)
        key = f"__REF_{counter}__"
        placeholder_map[key] = (sheet, ref)
        refs.append((sheet, ref))
        counter += 1
        return key

    replaced = CELL_REF_RE.sub(repl, expr)
    return replaced, placeholder_map, refs


def _eval_arithmetic_formula(expr: str, current_sheet: str, ws_map):
    body = expr.strip()
    if not body.startswith("="):
        return None
    inner = body[1:]
    if not ARITH_ALLOWED_RE.match(body):
        return None
    norm, placeholder_map, refs = _normalize_formula_expr(inner, current_sheet)
    values = {}
    for key, (sheet, ref) in placeholder_map.items():
        if ":" in ref:
            return None
        value = _safe_number(_cell_value(ws_map, sheet, ref))
        if value is None:
            return None
        values[key] = value
    try:
        import ast
        node = ast.parse(norm, mode="eval")
    except SyntaxError:
        return None

    def walk(n):
        if isinstance(n, ast.Expression):
            return walk(n.body)
        if isinstance(n, ast.Constant) and isinstance(n.value, (int, float)):
            return float(n.value)
        if isinstance(n, ast.UnaryOp) and isinstance(n.op, (ast.UAdd, ast.USub)):
            v = walk(n.operand)
            return v if isinstance(n.op, ast.UAdd) else -v
        if isinstance(n, ast.BinOp) and isinstance(n.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)):
            left = walk(n.left)
            right = walk(n.right)
            if isinstance(n.op, ast.Add):
                return left + right
            if isinstance(n.op, ast.Sub):
                return left - right
            if isinstance(n.op, ast.Mult):
                return left * right
            return left / right
        if isinstance(n, ast.Name) and n.id in values:
            return values[n.id]
        raise ValueError("unsupported arithmetic")

    try:
        return {"value": walk(node), "refs": [f"{s}!{r}" if s != current_sheet else r for s, r in refs]}
    except Exception:
        return None


def _eval_range_formula(expr: str, current_sheet: str, ws_map):
    match = RANGE_FUNC_RE.match(expr.strip())
    if not match:
        return None
    func = match.group(1).upper()
    target = match.group(2).strip()
    sheet_name, range_expr = _split_sheet_and_ref(target, current_sheet)
    if ":" not in range_expr:
        return None
    try:
        values = [_safe_number(v) for v in _iter_range_values(ws_map, sheet_name, range_expr)]
    except Exception:
        return None
    nums = [v for v in values if v is not None]
    if not nums:
        return None
    if func == "SUM":
        out = sum(nums)
    else:
        out = sum(nums) / len(nums)
    ref = f"{sheet_name}!{range_expr}" if sheet_name != current_sheet else range_expr
    return {"value": out, "refs": [ref]}


def evaluate_formula_invariants(filename: str, max_samples: int = 8) -> dict:
    try:
        wb_formula = load_workbook(filename, data_only=False)
        wb_values = load_workbook(filename, data_only=True)
    except Exception as exc:
        return {"status": "skipped", "reason": str(exc), "categories": {}}

    ws_formula = {ws.title: ws for ws in wb_formula.worksheets}
    ws_values = {ws.title: ws for ws in wb_values.worksheets}
    categories = {
        "arithmetic": {"supported": 0, "matched": 0, "mismatched": 0, "skipped": 0, "samples": []},
        "tieout": {"supported": 0, "matched": 0, "mismatched": 0, "skipped": 0, "samples": []},
        "sensitivity": {"supported": 0, "matched": 0, "mismatched": 0, "skipped": 0, "samples": []},
    }

    def add_sample(kind: str, sample: dict):
        if len(categories[kind]["samples"]) < max_samples:
            categories[kind]["samples"].append(sample)

    try:
        for sheet_name, ws in ws_formula.items():
            values_ws = ws_values.get(sheet_name)
            if values_ws is None:
                continue
            rows = list(ws.iter_rows())
            for row in rows:
                for cell in row:
                    if not _is_formula(cell.value):
                        continue
                    formula_text = _formula_text(cell.value)
                    result = _eval_range_formula(formula_text, sheet_name, ws_values) or _eval_arithmetic_formula(formula_text, sheet_name, ws_values)
                    if result is None:
                        categories["arithmetic"]["skipped"] += 1
                        continue
                    actual = _safe_number(values_ws[cell.coordinate].value)
                    if actual is None:
                        categories["arithmetic"]["skipped"] += 1
                        continue
                    expected = result["value"]
                    categories["arithmetic"]["supported"] += 1
                    tolerance = max(1e-6, abs(expected) * 1e-6)
                    if abs(actual - expected) <= tolerance:
                        categories["arithmetic"]["matched"] += 1
                    else:
                        categories["arithmetic"]["mismatched"] += 1
                        add_sample("arithmetic", {
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "expected": expected,
                            "actual": actual,
                            "refs": result["refs"],
                        })

            header = {c.column: str(c.value).strip() if c.value is not None else "" for c in ws[1]}
            label_col = 1 if ws.max_column >= 1 else None
            value_col = None
            for col, title in header.items():
                low = title.lower()
                if TIEOUT_ROW_RE.search(low):
                    label_col = col
                if value_col is None and low in {"value", "amount", "金额", "数值"}:
                    value_col = col
            if value_col is None and ws.max_column >= 2:
                value_col = 2
            if label_col and value_col:
                labels = {}
                for r in range(2, min(ws.max_row, 200) + 1):
                    label = ws.cell(row=r, column=label_col).value
                    value = _safe_number(values_ws.cell(row=r, column=value_col).value)
                    if isinstance(label, str) and value is not None:
                        labels[label.strip().lower()] = value
                def _pick(keywords, exclude=()):
                    # 只认「类别词 + 总计词」同现的标签：普通 KPI（Return on Assets /
                    # Liability Ratio）没有总计词，不再被当成资产负债表勾稽项误报；
                    # 明细行（Current Assets）也不再顶替总计行。exclude 做类别互斥——
                    # 「负债和所有者权益总计 / total liabilities and equity」是合计行
                    # （数值 = 资产总计），单类命中会让勾稽整体错位。
                    total_keys = ("total", "总计", "合计", "总额")
                    def ok(k):
                        return (any(w in k for w in keywords)
                                and any(t in k for t in total_keys)
                                and not any(x in k for x in exclude))
                    return next((v for k, v in labels.items() if ok(k)), None)

                assets = _pick(("asset", "资产"), exclude=("liabilit", "负债", "equity", "权益"))
                liabilities = _pick(("liabilit", "负债"), exclude=("equity", "权益"))
                equity = _pick(("equity", "权益"), exclude=("liabilit", "负债"))
                if assets is not None and liabilities is not None and equity is not None:
                    categories["tieout"]["supported"] += 1
                    expected = liabilities + equity
                    tolerance = max(1e-6, abs(expected) * 1e-6)
                    if abs(assets - expected) <= tolerance:
                        categories["tieout"]["matched"] += 1
                    else:
                        categories["tieout"]["mismatched"] += 1
                        add_sample("tieout", {
                            "sheet": sheet_name,
                            "rule": "assets_equals_liabilities_plus_equity",
                            "expected": expected,
                            "actual": assets,
                        })
                else:
                    categories["tieout"]["skipped"] += 1
            else:
                categories["tieout"]["skipped"] += 1

            # 首版 sensitivity invariant 只在能唯一定位 main output + baseline intersection 时才可信；
            # 当前没有稳定结构证据时一律 skipped，避免把「baseline != scenario」误当可靠自洽性。
            categories["sensitivity"]["skipped"] += 1
    finally:
        wb_formula.close()
        wb_values.close()

    mismatch_total = sum(categories[k]["mismatched"] for k in categories)
    return {
        "status": "warn" if mismatch_total else "clean",
        "categories": categories,
        "supported": sum(categories[k]["supported"] for k in categories),
        "mismatched": mismatch_total,
    }


def recalc(filename, timeout=30, static_source_sheets: Optional[Set[str]] = None):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    if not is_libreoffice_available():
        return scan_errors_only(filename, static_source_sheets=static_source_sheets)

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    cmd = [
        "soffice",
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin" and has_gtimeout():
        cmd = ["gtimeout", str(timeout)] + cmd

    # subprocess 自带超时兜底：无 gtimeout 的 macOS 上 soffice 挂死会拖垮整条校验链
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, env=get_soffice_env(),
                                timeout=timeout + 30)
    except subprocess.TimeoutExpired:
        return {"error": f"LibreOffice recalculation timed out after {timeout + 30}s"}

    if result.returncode == 124:
        return {"error": f"LibreOffice recalculation timed out after {timeout}s; cached formula values are not authoritative"}
    if result.returncode != 0:
        error_msg = result.stderr or "Unknown error during recalculation"
        # 只在明确命中 macro 配置特征时归类为 macro 问题，其余透传真实 stderr
        if "Module1" in error_msg and "RecalculateAndSave" in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        return {"error": error_msg}

    try:
        total_errors, error_details = _scan_formula_error_values(filename)

        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_count += 1
        wb_formulas.close()

        result["total_formulas"] = formula_count
        result["hardcode"] = detect_hardcode_suspects(filename, static_source_sheets=static_source_sheets)
        result["invariants"] = evaluate_formula_invariants(filename)
        return result

    except Exception as e:
        return {"error": str(e)}


def main():
    ap = argparse.ArgumentParser(
        description="重算并校验工作簿中的公式错误，同时扫描推导列/行的静态值硬编码嫌疑。"
    )
    ap.add_argument("excel_file")
    ap.add_argument(
        "timeout_seconds",
        nargs="?",
        type=int,
        default=30,
        help="重算超时秒数（保留旧 positional timeout 兼容）",
    )
    ap.add_argument(
        "--static-source-sheets",
        help="显式声明允许保留静态历史/外部数据的 sheet 名，逗号分隔、精确匹配；这些 sheet 的 hardcode suspect 降为 low/WARN",
    )
    args = ap.parse_args()

    static_source_sheets, static_source_sheet_list = _normalize_sheet_names(args.static_source_sheets)
    result = recalc(args.excel_file, args.timeout_seconds, static_source_sheets=static_source_sheets)
    hc = result.get("hardcode") or {}
    unknown_static_source_sheets = hc.get("unknown_static_source_sheets") or []
    if unknown_static_source_sheets:
        result.setdefault("error", f"Unknown --static-source-sheets: {', '.join(unknown_static_source_sheets)}")
    print(json.dumps(result, indent=2, ensure_ascii=False))

    if static_source_sheet_list:
        print(
            "[hardcode] 已声明静态来源 sheet：" + "、".join(static_source_sheet_list)
            + "——仅这些 sheet 的 hardcode suspect 会降为 WARN，其它 sheet 仍按原规则判定。",
            file=sys.stderr,
        )
    if unknown_static_source_sheets:
        print(
            "[hardcode] 未知的 --static-source-sheets：" + "、".join(unknown_static_source_sheets)
            + "——sheet 名必须精确匹配工作簿里的现有 sheet；修正参数后重跑。",
            file=sys.stderr,
        )

    has_high = _emit_hardcode_hint(hc) if hc.get("status") == "suspected" else False
    has_formula_errors = bool(result.get("total_errors")) or result.get("status") == "errors_found"
    has_runtime_error = bool(result.get("error"))
    # 3 = 校验发现问题（公式错误 / 重算失败 / 高置信硬编码嫌疑 / 静态来源点名错误，文件本身可用，非崩溃）；0 = 通过
    sys.exit(3 if (has_formula_errors or has_runtime_error or has_high) else 0)


if __name__ == "__main__":
    main()
