#!/usr/bin/env python3
"""
对单元格范围一次性应用样式（字体/底色/边框/对齐/数值格式/合并/条件格式）。

用法示例：
  python scripts/format_range.py file.xlsx Sheet1 A1 --end C1 --bold --bg-color FFFF00
  python scripts/format_range.py file.xlsx Sheet1 B2:D10 --number-format "0.0%" --align center --wrap
  python scripts/format_range.py file.xlsx Sheet1 A1:E20 --cond-format \
      '{"type":"color_scale","params":{"start_type":"min","start_color":"FFAA0000","end_type":"max","end_color":"FF00AA00"}}'
"""
import argparse
import json
from copy import copy

from openpyxl.styles import Color, PatternFill, Protection, Side
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
)

from _excel_utils import emit, emit_error, load_or_create_wb, parse_range, require_sheet


def _normalize_color(c: str) -> str:
    c = c.strip().upper().lstrip("#")
    return c if len(c) == 8 else f"FF{c}"


def _build_conditional_rule(spec: dict):
    rule_type = spec.get("type")
    params = dict(spec.get("params", {}))
    if rule_type == "color_scale":
        return ColorScaleRule(**params)
    if rule_type == "data_bar":
        return DataBarRule(**params)
    if rule_type == "icon_set":
        return IconSetRule(**params)
    if rule_type == "formula":
        return FormulaRule(**params)
    if rule_type == "cell_is":
        fill_spec = params.get("fill")
        if isinstance(fill_spec, dict):
            fg = _normalize_color(fill_spec.get("fgColor", "FFC7CE"))
            params["fill"] = PatternFill(start_color=fg, end_color=fg, fill_type="solid")
        return CellIsRule(**params)
    raise ValueError(f"Unsupported conditional format type: {rule_type}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("filepath")
    p.add_argument("sheet")
    p.add_argument("start", help="起始单元格，可写成 'A1' 或 'A1:C5'")
    p.add_argument("--end", help="结束单元格（start 用 'A1:C5' 时无需）")
    # 字体
    p.add_argument("--bold", action="store_true")
    p.add_argument("--italic", action="store_true")
    p.add_argument("--underline", action="store_true")
    p.add_argument("--font-size", type=int)
    p.add_argument("--font-color")
    p.add_argument("--font-name")
    # 填充与边框
    p.add_argument("--bg-color")
    p.add_argument("--border-style", choices=["thin", "medium", "thick", "dashed", "dotted", "double"])
    p.add_argument("--border-color")
    p.add_argument("--border-sides", help="要修改的边，逗号分隔：top,bottom,left,right；使用 --border-style 时必填")
    # 数值与对齐
    p.add_argument("--number-format")
    p.add_argument("--align", choices=["left", "center", "right", "justify"])
    p.add_argument("--vertical", choices=["top", "center", "bottom"])
    p.add_argument("--wrap", action="store_true")
    # 合并
    p.add_argument("--merge", action="store_true")
    # 保护
    p.add_argument("--protection", help='JSON 串，如 \'{"locked":true,"hidden":false}\'')
    # 条件格式
    p.add_argument("--cond-format", help="条件格式 JSON 串")
    args = p.parse_args()

    try:
        sr, sc, er, ec = parse_range(args.start, args.end)
    except ValueError as e:
        emit_error(str(e))

    wb = load_or_create_wb(args.filepath)
    try:
        ws = require_sheet(wb, args.sheet)
    except ValueError as e:
        emit_error(str(e))

    font_requested = any((args.bold, args.italic, args.underline, args.font_size, args.font_color, args.font_name))

    fill = None
    if args.bg_color:
        c = _normalize_color(args.bg_color)
        fill = PatternFill(start_color=c, end_color=c, fill_type="solid")

    border_side = None
    border_sides = []
    if args.border_style:
        if not args.border_sides:
            emit_error("--border-sides is required with --border-style; choose from top,bottom,left,right")
        border_sides = [part.strip().lower() for part in args.border_sides.split(",") if part.strip()]
        invalid_sides = sorted(set(border_sides) - {"top", "bottom", "left", "right"})
        if invalid_sides:
            emit_error(f"Invalid --border-sides: {','.join(invalid_sides)}")
        bc = _normalize_color(args.border_color or "000000")
        border_side = Side(style=args.border_style, color=Color(rgb=bc))
    elif args.border_sides:
        emit_error("--border-sides requires --border-style")

    alignment_requested = any((args.align, args.wrap, args.vertical))

    protect = None
    if args.protection:
        try:
            protect = Protection(**json.loads(args.protection))
        except Exception as e:
            emit_error(f"Invalid --protection JSON: {e}")

    merge_requested = args.merge and (er > sr or ec > sc)
    merge_value = None
    if merge_requested:
        non_empty = [
            ws.cell(row=r, column=c).value
            for r in range(sr, er + 1)
            for c in range(sc, ec + 1)
            if ws.cell(row=r, column=c).value not in (None, "")
        ]
        distinct = {(type(value).__name__, repr(value)) for value in non_empty}
        if len(distinct) > 1:
            emit_error("Refusing to merge: target range contains different non-empty values")
        merge_value = non_empty[0] if non_empty else None

    for r in range(sr, er + 1):
        for c in range(sc, ec + 1):
            cell = ws.cell(row=r, column=c)
            if font_requested:
                font = copy(cell.font)
                if args.bold:
                    font.bold = True
                if args.italic:
                    font.italic = True
                if args.underline:
                    font.underline = "single"
                if args.font_size:
                    font.size = args.font_size
                if args.font_name:
                    font.name = args.font_name
                if args.font_color:
                    font.color = Color(rgb=_normalize_color(args.font_color))
                cell.font = font
            if fill is not None:
                cell.fill = fill
            if border_side is not None:
                border = copy(cell.border)
                for side_name in border_sides:
                    setattr(border, side_name, border_side)
                cell.border = border
            if alignment_requested:
                alignment = copy(cell.alignment)
                if args.align:
                    alignment.horizontal = args.align
                if args.vertical:
                    alignment.vertical = args.vertical
                if args.wrap:
                    alignment.wrap_text = True
                cell.alignment = alignment
            if protect is not None:
                cell.protection = protect
            if args.number_format:
                cell.number_format = args.number_format

    range_str = f"{args.start}" if (args.end is None and ":" not in args.start) else (
        args.start if ":" in args.start else f"{args.start}:{args.end}"
    )

    if merge_requested:
        anchor = ws.cell(row=sr, column=sc)
        if anchor.value in (None, "") and merge_value is not None:
            anchor.value = merge_value
        for r in range(sr, er + 1):
            for c in range(sc, ec + 1):
                if (r, c) != (sr, sc):
                    ws.cell(row=r, column=c).value = None
        ws.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)

    if args.cond_format:
        try:
            spec = json.loads(args.cond_format)
            rule = _build_conditional_rule(spec)
            ws.conditional_formatting.add(range_str, rule)
        except Exception as e:
            emit_error(f"Failed to apply conditional format: {e}")

    try:
        wb.save(args.filepath)
    except Exception as e:
        emit_error(f"Failed to save workbook: {e}")

    emit({
        "status": "success",
        "range": range_str,
        "sheet": args.sheet,
        "merged": bool(merge_requested),
        "applied": {
            "font": bool(font_requested),
            "fill": fill is not None,
            "border": border_side is not None,
            "align": alignment_requested,
            "number_format": bool(args.number_format),
            "conditional_format": bool(args.cond_format),
        },
    })


if __name__ == "__main__":
    main()
