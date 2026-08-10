#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""将个人信息保护审计底稿CSV生成一个Excel工作簿。"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_FILES = [
    ("附件与证据目录", "附件与证据目录.csv"),
    ("事实与不确定性", "事实与不确定性.csv"),
    ("证据矛盾与未决事项", "证据矛盾与未决事项.csv"),
    ("信息分类", "信息分类.csv"),
    ("处理活动与角色", "处理活动与角色.csv"),
    ("处理情形与模块适用性", "处理情形与模块适用性.csv"),
    ("第三方委托与跨境链路", "第三方委托与跨境链路.csv"),
    ("AI系统与自动化决策", "AI系统与自动化决策清单.csv"),
    ("AI高风险专项分析", "AI高风险专项分析.csv"),
    ("法规文献数据库", "法规文献数据库.csv"),
    ("法规版本台账", "法规版本台账.csv"),
    ("适用规范矩阵", "适用规范矩阵.csv"),
    ("主要问题与整改", "主要问题与整改.csv"),
    ("审计程序点检", "审计程序点检.csv"),
    ("107项点检", "107项点检.csv"),
    ("审计分析链", "审计分析链.csv"),
]

CONCLUSION_COLORS = {
    "合规": "C6EFCE",
    "部分合规": "FFEB9C",
    "不合规": "FFC7CE",
    "无法判断": "D9EAF7",
    "证据不足": "D9EAF7",
    "不涉及": "E7E6E6",
    "重要风险": "F4CCCC",
    "一般风险": "FFF2CC",
    "轻微风险": "D9EAD3",
    "待确认风险": "D9EAF7",
    "已确认": "C6EFCE",
    "材料未提及": "E7E6E6",
    "证据冲突": "F4CCCC",
}

WIDE_HEADERS = {
    "已确认事实", "尚不能确认的事项", "需要补充的证据", "事实来源",
    "判断依据", "业务环境与关联能力", "判断前提", "需补充证据",
    "判断理由", "尚待确认", "审计证据", "适用要求及法律依据",
    "事实与合规要求", "风险等级及理由", "尚待核实事项", "整改目标",
    "具体措施", "验收标准", "整改证据", "法律要求", "差异或判断理由",
    "真实附件名称", "相关具体内容", "具体定位", "对本行为的实际要求",
    "触发条件", "例外或豁免", "履行所需证据", "权威出处", "事实基础",
    "具体要求", "后续行动",
    "适用场景", "重点条款或章节", "替代或衔接关系", "适用或排除理由",
    "具体规则内容", "证据及精确定位", "证据状态与证明力", "冲突或相反证据",
    "尚不能确认事项及影响", "待补数据与核验程序", "冲突或相反证据及管理层解释",
    "事实与规则差异", "要件分析", "规范名称及条款",
    "原子化事实", "样本与外推边界", "竞争性假设", "最小补证", "核验程序",
    "触发要素", "模型与数据流", "供应商及分包商实际行为", "测试环境样本与步骤",
    "实际结果及证据定位", "整改编号及措施",
    "上位依据编号及条款", "下位法或配套规范编号及条款", "衔接类型",
    "下位规范触发事实", "下位规范不适用理由", "层级权限冲突核验",
    "正文引用组合", "上位法基础引用", "下位法或配套规范补充引用",
    "上下位或配套衔接说明", "权限及冲突核验",
}


def read_csv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.reader(file))


def add_sheet(workbook: Workbook, title: str, rows: list[list[str]]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["暂无数据"])
        return

    for row in rows:
        sheet.append(row)

    thin = Side(style="thin", color="D1D5DB")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="微软雅黑", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if str(cell.value or "") in CONCLUSION_COLORS:
                color = CONCLUSION_COLORS[str(cell.value)]
                cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [str(cell.value or "") for cell in sheet[1]]
    for index, header in enumerate(headers, start=1):
        if header in WIDE_HEADERS:
            width = 42
        elif any(keyword in header for keyword in ("编号", "结论", "状态", "级别", "模块", "适用性")):
            width = 14
        else:
            width = 22
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.row_dimensions[1].height = 32
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def resolve_sources(input_path: Path) -> tuple[list[tuple[str, Path]], Path]:
    if input_path.is_dir():
        missing = [
            filename for _, filename in SHEET_FILES
            if not (input_path / filename).exists()
        ]
        if missing:
            raise SystemExit("缺少底稿CSV：" + "、".join(missing))
        sources = [
            (sheet_name, input_path / filename)
            for sheet_name, filename in SHEET_FILES
        ]
        return sources, input_path / "个人信息保护合规审计底稿.xlsx"
    return [(input_path.stem[:31], input_path)], input_path.with_suffix(".xlsx")


def main() -> None:
    parser = argparse.ArgumentParser(description="将审计底稿CSV生成Excel")
    parser.add_argument("input", help="包含审计底稿CSV的目录，或单个CSV文件")
    parser.add_argument("--output", help="输出Excel路径")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        raise SystemExit(f"输入不存在：{input_path}")

    sources, default_output = resolve_sources(input_path)
    if not sources:
        names = "、".join(filename for _, filename in SHEET_FILES)
        raise SystemExit(f"未找到底稿CSV。目录中应包含：{names}")

    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, source in sources:
        add_sheet(workbook, title, read_csv(source))

    output = Path(args.output) if args.output else default_output
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(f"已生成：{output}")
    print("工作表：" + "、".join(title for title, _ in sources))


if __name__ == "__main__":
    main()
