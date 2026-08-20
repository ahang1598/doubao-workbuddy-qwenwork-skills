"""XLSX full-workbook text dump via openpyxl.

Pipe-delimited rows, all sheets, all cells. The global text budget in
the calling skill remains the only safety net — see
``trace_llm_judge._extract_xlsx_content`` for the original design notes.
"""

from __future__ import annotations

import io
import logging
from typing import Any

logger = logging.getLogger(__name__)

_CELL_MAX_CHARS = 500


def _fmt_cell(value: Any) -> str:
    """Render an openpyxl cell value as compact, single-line text.

    Long cells are clipped to ``_CELL_MAX_CHARS``; embedded newlines and
    pipes are escaped so each row stays on one line in the
    pipe-delimited dump.
    """
    if value is None:
        return ""
    s = str(value)
    if len(s) > _CELL_MAX_CHARS:
        s = s[:_CELL_MAX_CHARS] + "…"
    return s.replace("\n", " ").replace("\r", " ").replace("|", "\\|")


def extract_xlsx_text(xlsx_bytes: bytes) -> tuple[str, list[dict]]:
    """Dump every sheet as ``--- Sheet: <name> --- \\n R<i>: c1 | c2 …``.

    Returns
    -------
    (text, sheets)
        ``sheets`` is a list of ``{"name", "rows", "cols"}`` per sheet.

    Falls back to ``("", [])`` when openpyxl is missing or the workbook
    can't be opened (legacy ``.xls`` binary format, password-protected,
    truncated bytes, etc.).
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed — xlsx extract skipped")
        return "", []

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        logger.warning("[xlsx] open failed: %s", exc)
        return "", []

    text_parts: list[str] = []
    sheets: list[dict] = []
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
        except Exception as exc:
            logger.debug("[xlsx] skip sheet %r: %s", sheet_name, exc)
            continue
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        sheets.append({"name": sheet_name, "rows": max_row, "cols": max_col})
        if max_row == 0 or max_col == 0:
            text_parts.append(f"--- Sheet: {sheet_name} (empty) ---")
            continue
        lines: list[str] = []
        for r_idx, row in enumerate(
            ws.iter_rows(
                min_row=1, max_row=max_row, min_col=1, max_col=max_col,
                values_only=True,
            ),
            start=1,
        ):
            cells = [_fmt_cell(v) for v in row]
            while cells and cells[-1] == "":
                cells.pop()
            lines.append(f"R{r_idx}: " + " | ".join(cells))
        header = f"--- Sheet: {sheet_name} (rows={max_row}, cols={max_col}) ---"
        text_parts.append(header + "\n" + "\n".join(lines))

    try:
        wb.close()
    except Exception:
        pass
    return "\n\n".join(text_parts), sheets
