#!/usr/bin/env python3
"""Materialize the formula-driven DCF workbook with Python and OpenPyXL only."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.properties import CalcProperties

SCRIPT_DIR = Path(__file__).resolve().parent
PACKAGE_ROOT = SCRIPT_DIR.parents[1]
sys.path.insert(0, str(SCRIPT_DIR))

from calculate_dcf import calculate  # noqa: E402


SCENARIOS = ("bear", "base", "bull")
SCENARIO_COLUMNS = {"bear": "B", "base": "C", "bull": "D"}
SCENARIO_ZH = {"bear": "悲观", "base": "基准", "bull": "乐观"}
VISIBLE_SHEETS = [
    "封面",
    "执行计划",
    "数据来源",
    "股本与市值桥",
    "估值假设",
    "经营预测",
    "DCF估值",
    "敏感性分析",
    "模型检查",
    "术语说明",
]
FORMULA_RANGES = [
    ("估值假设", "B22", "WACC calculation cell"),
    ("经营预测", "E5:I17", "forecast calculation block"),
    ("DCF估值", "E10:I12", "DCF forecast and discount block"),
    ("DCF估值", "E14:I14", "discounted FCFF block"),
    ("DCF估值", "B15:B29", "DCF valuation bridge"),
    ("敏感性分析", "C5:G9", "5x5 sensitivity grid"),
    ("敏感性分析", "C27:G41", "bear scenario forecast block"),
    ("敏感性分析", "B43:B52", "bear scenario valuation block"),
    ("敏感性分析", "C60:G74", "base scenario forecast block"),
    ("敏感性分析", "B76:B85", "base scenario valuation block"),
    ("敏感性分析", "C93:G107", "bull scenario forecast block"),
    ("敏感性分析", "B109:B118", "bull scenario valuation block"),
    ("敏感性分析", "D15:D17", "scenario summary outputs"),
    ("模型检查", "B3:B4", "formula-driven model status"),
]


def _is_number(value: Any) -> bool:
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(float(value))
    )


def _require_object(value: Any, field: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ValueError(f"{field} must be an object")
    return value


def _as_date(value: Any, field: str) -> date:
    if not isinstance(value, str):
        raise ValueError(f"{field} must be YYYY-MM-DD")
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(f"{field} must be YYYY-MM-DD") from exc


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _clear_values(ws: Any, cell_range: str) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.value = None


def _formula_refs(sheet: str, cell_range: str) -> list[str]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return [
        f"{sheet}!{get_column_letter(col)}{row}"
        for row in range(min_row, max_row + 1)
        for col in range(min_col, max_col + 1)
    ]


def _assert_formulas(wb: Any) -> tuple[list[str], dict[str, str]]:
    refs: list[str] = []
    formulas: dict[str, str] = {}
    for sheet, cell_range, label in FORMULA_RANGES:
        for ref in _formula_refs(sheet, cell_range):
            _, coordinate = ref.rsplit("!", 1)
            value = wb[sheet][coordinate].value
            if not isinstance(value, str) or not value.startswith("="):
                raise ValueError(
                    f"{label} contains a non-formula cell: {ref} value={value!r}"
                )
            refs.append(ref)
            formulas[ref] = value
    return refs, formulas


def _forecast_metric(row: dict[str, Any], field: str) -> float:
    if field == "ebit_margin":
        if _is_number(row.get("ebit_margin")):
            return float(row["ebit_margin"])
        if (
            _is_number(row.get("ebit"))
            and _is_number(row.get("revenue"))
            and row["revenue"] != 0
        ):
            return float(row["ebit"]) / float(row["revenue"])
    if field == "da_ratio":
        return float(row["da"]) / float(row["revenue"]) if row["revenue"] else 0.0
    if field == "capex_ratio":
        return (
            float(row["capex"]) / float(row["revenue"]) if row["revenue"] else 0.0
        )
    value = row.get(field)
    if not _is_number(value):
        raise ValueError(f"forecast field {field} must be numeric")
    return float(value)


def _number_or_blank(value: Any) -> float | str:
    return float(value) if _is_number(value) else ""


def _validate_inputs(
    normalized: dict[str, Any], calculated: dict[str, Any]
) -> dict[str, Any]:
    meta = _require_object(normalized.get("meta"), "meta")
    bridge = _require_object(normalized.get("equity_bridge"), "equity_bridge")
    scenarios = _require_object(normalized.get("scenarios"), "scenarios")
    historical = _require_object(
        normalized.get("historical_anchor"), "historical_anchor"
    )
    if not _is_number(historical.get("revenue")) or historical["revenue"] <= 0:
        raise ValueError("historical_anchor.revenue must be positive")
    for name in SCENARIOS:
        scenario = _require_object(scenarios.get(name), f"scenarios.{name}")
        forecast = scenario.get("forecast")
        if not isinstance(forecast, list) or len(forecast) != 5:
            raise ValueError(
                f"scenarios.{name}.forecast must contain exactly 5 periods "
                "for dcf-model-template.xlsx"
            )
        for index, row in enumerate(forecast):
            if not isinstance(row, dict):
                raise ValueError(f"scenarios.{name}.forecast[{index}] must be an object")
            for field in ("revenue", "da", "capex", "delta_nwc"):
                if not _is_number(row.get(field)):
                    raise ValueError(
                        f"scenarios.{name}.forecast[{index}].{field} must be numeric"
                    )
            _forecast_metric(row, "ebit_margin")
    base_periods = [row.get("period") for row in scenarios["base"]["forecast"]]
    for name in SCENARIOS:
        periods = [row.get("period") for row in scenarios[name]["forecast"]]
        if periods != base_periods:
            raise ValueError(f"scenarios.{name}.forecast periods must match base")
    if len(bridge.get("share_classes") or []) > 3:
        raise ValueError("template supports at most 3 share classes")
    review = bridge.get("corporate_action_review") or {}
    if len(review.get("actions") or []) > 3:
        raise ValueError("template supports at most 3 corporate-action rows")
    sensitivity = calculated.get("sensitivity") or {}
    if len(sensitivity.get("wacc_rates") or []) != 5:
        raise ValueError("calculated DCF must contain 5 sensitivity WACC rates")
    sensitivity_rows = sensitivity.get("rows") or []
    if (
        len(sensitivity_rows) != 5
        or any(len(row.get("per_share_values") or []) != 5 for row in sensitivity_rows)
    ):
        raise ValueError("calculated DCF must contain a complete 5x5 sensitivity grid")
    fresh = calculate(normalized)
    for name in SCENARIOS:
        expected = calculated.get("scenarios", {}).get(name)
        actual = fresh.get("scenarios", {}).get(name)
        if not expected or not actual:
            raise ValueError(f"calculated DCF is missing scenario {name}")
        for key in ("enterprise_value", "equity_value", "per_share_value"):
            if not _is_number(expected.get(key)) or abs(
                float(expected[key]) - float(actual[key])
            ) > 1e-8:
                raise ValueError(
                    f"normalized/calculated parity failed: {name}.{key}"
                )
    for row_index, row in enumerate(sensitivity_rows):
        fresh_row = fresh["sensitivity"]["rows"][row_index]["per_share_values"]
        for col_index, expected in enumerate(row["per_share_values"]):
            actual = fresh_row[col_index]
            if expected is None or actual is None:
                if expected != actual:
                    raise ValueError(
                        "normalized/calculated sensitivity null mismatch at "
                        f"{row_index + 1},{col_index + 1}"
                    )
            elif abs(float(expected) - float(actual)) > 1e-8:
                raise ValueError(
                    "normalized/calculated sensitivity parity failed at "
                    f"{row_index + 1},{col_index + 1}"
                )
    return fresh


def _populate(
    wb: Any,
    normalized: dict[str, Any],
    calculated: dict[str, Any],
    fresh: dict[str, Any],
) -> None:
    missing = [name for name in VISIBLE_SHEETS if name not in wb.sheetnames]
    if missing:
        raise ValueError("template is missing sheets: " + ", ".join(missing))
    meta = normalized["meta"]
    bridge = normalized["equity_bridge"]
    scenarios = normalized["scenarios"]
    historical_anchor = normalized["historical_anchor"]
    review = bridge.get("corporate_action_review") or {}
    base_periods = [row.get("period") for row in scenarios["base"]["forecast"]]

    execution = wb["执行计划"]
    sources_ws = wb["数据来源"]
    share_ws = wb["股本与市值桥"]
    assumptions = wb["估值假设"]
    operating = wb["经营预测"]
    dcf = wb["DCF估值"]
    sensitivity_ws = wb["敏感性分析"]
    checks = wb["模型检查"]

    statuses = ["已完成", "已完成", "已完成", "已完成", "待完成", "待完成"]
    notes = [
        meta.get("task_id") or "由标准化输入生成",
        f"来源记录 {len(normalized.get('sources') or [])} 条",
        "normalized-dcf.json 与 calculated-dcf.json 已通过确定性一致性检查",
        "已由 build_dcf_workbook.py 使用 OpenPyXL 填充",
        "必须另行运行 validate_delivery.py",
        "只有交付验证PASS后方可发布",
    ]
    for row, status, note in zip(range(4, 10), statuses, notes):
        execution[f"B{row}"] = status
        execution[f"H{row}"] = note

    _clear_values(sources_ws, "A4:J30")
    for row_index, src in enumerate((normalized.get("sources") or [])[:27], 4):
        fields = src.get("fields")
        sources_ws.cell(row_index, 1, src.get("source_id") or "")
        sources_ws.cell(
            row_index,
            2,
            ", ".join(str(v) for v in fields)
            if isinstance(fields, list)
            else (src.get("topic") or ""),
        )
        sources_ws.cell(row_index, 4, meta.get("units") or "")
        sources_ws.cell(
            row_index, 5, src.get("period_end") or src.get("as_of_date") or ""
        )
        sources_ws.cell(
            row_index, 6, src.get("source_type") or src.get("label") or ""
        )
        sources_ws.cell(
            row_index, 7, src.get("title") or src.get("source_name") or ""
        )
        sources_ws.cell(
            row_index, 8, src.get("url_or_path") or src.get("url") or ""
        )
        if src.get("public_date"):
            sources_ws.cell(
                row_index,
                9,
                _as_date(
                    src["public_date"],
                    f"sources.{src.get('source_id', row_index)}.public_date",
                ),
            )
        sources_ws.cell(row_index, 10, src.get("notes") or "")

    share_ws["B4"] = (
        _as_date(review["baseline_share_date"], "corporate_action_review.baseline_share_date")
        if review.get("baseline_share_date")
        else None
    )
    share_ws["B5"] = (
        _as_date(review["search_start_date"], "corporate_action_review.search_start_date")
        if review.get("search_start_date")
        else None
    )
    share_ws["B7"] = (
        "已完成"
        if review.get("no_unrecorded_actions_confirmed") is True
        and review.get("reviewed_through_date") == meta.get("valuation_date")
        else "未完成"
    )
    classes = bridge.get("share_classes") or []
    price_basis_ok = bool(classes) and all(
        item.get("price_basis") == "unadjusted_close"
        and item.get("shares_date") == meta.get("valuation_date")
        for item in classes
    )
    share_ws["B8"] = "已确认" if price_basis_ok else "未完成"
    share_ws["B9"] = bridge["diluted_shares"]
    _clear_values(share_ws, "A13:H15")
    _clear_values(share_ws, "J13:J15")
    _clear_values(share_ws, "L13:L15")
    for index, item in enumerate(classes):
        row = 13 + index
        values = [
            item.get("security_id"),
            item.get("exchange"),
            item.get("shares"),
            _as_date(item.get("shares_date"), f"share_classes.{index}.shares_date"),
            item.get("price"),
            _as_date(item.get("price_date"), f"share_classes.{index}.price_date"),
            item.get("currency"),
            item.get("fx_to_valuation_currency"),
        ]
        for col, value in enumerate(values, 1):
            share_ws.cell(row, col, value)
        share_ws.cell(row, 10, item.get("reference_market_cap"))
        share_ws.cell(
            row, 12, item.get("source_id") or item.get("market_cap_source_id") or ""
        )
    _clear_values(share_ws, "A20:J22")
    for index, action in enumerate((review.get("actions") or [])[:3]):
        row = 20 + index
        values = [
            action.get("security_id"),
            action.get("action_type"),
            _as_date(
                action.get("announcement_date"),
                f"corporate_actions.{index}.announcement_date",
            ),
            _as_date(
                action.get("effective_date"),
                f"corporate_actions.{index}.effective_date",
            ),
            action.get("before_shares"),
            action.get("change_shares"),
            action.get("after_shares"),
            "是" if action.get("applied_to_share_count") is True else "否",
            action.get("source_id") or "",
            action.get("notes") or "",
        ]
        for col, value in enumerate(values, 1):
            share_ws.cell(row, col, value)

    assumptions["B3"] = meta.get("company")
    assumptions["B4"] = meta.get("ticker") or ""
    assumptions["B5"] = "基准"
    assumptions["B6"] = _as_date(meta.get("valuation_date"), "meta.valuation_date")
    assumptions["B7"] = meta.get("currency")
    assumptions["B8"] = meta.get("units")
    components = normalized.get("wacc_components") or {}
    for row, field in zip(
        range(12, 22),
        (
            "risk_free_rate",
            "beta",
            "equity_risk_premium",
            "country_risk_premium",
            "size_premium",
            "other_equity_premium",
            "pre_tax_cost_of_debt",
            "marginal_tax_rate",
            "equity_weight",
            "debt_weight",
        ),
    ):
        assumptions[f"B{row}"] = _number_or_blank(components.get(field))
    base_wacc = float(fresh["scenarios"]["base"]["wacc"])
    assumptions["B23"] = (
        base_wacc if not components and meta.get("model_purpose") == "illustrative" else None
    )
    for row, field in zip(
        range(15, 25),
        (
            "cash",
            "non_operating_investments",
            "associates",
            "debt",
            "lease_liabilities",
            "unfunded_pension",
            "preferred_stock",
            "minority_interest",
            "other_claims",
            "diluted_shares",
        ),
    ):
        assumptions[f"G{row}"] = _number_or_blank(bridge.get(field))
    field_sources = normalized.get("field_sources") or {}
    bridge_sources = field_sources.get("equity_bridge") or {}
    assumptions["G25"] = bridge_sources.get("source_id") or (
        (bridge_sources.get("source_ids") or [""])[0]
    )

    _clear_values(operating, "B3:D17")
    operating["A3"] = f"{meta.get('currency', '')} / {meta.get('units', '')}"
    historicals = normalized.get("historicals") or []
    for offset, hist in enumerate(historicals[-3:], 4 - len(historicals[-3:])):
        col = get_column_letter(offset)
        operating[f"{col}3"] = hist.get("period") or "实际"
        operating[f"{col}5"] = hist.get("revenue")
        if col != "B":
            previous = get_column_letter(offset - 1)
            operating[f"{col}6"] = f'=IFERROR({col}5/{previous}5-1,"")'
        if _is_number(hist.get("ebit_margin")):
            operating[f"{col}7"] = hist["ebit_margin"]
            operating[f"{col}8"] = f"={col}5*{col}7"
        elif _is_number(hist.get("ebit")):
            operating[f"{col}8"] = hist["ebit"]
            operating[f"{col}7"] = f'=IFERROR({col}8/{col}5,"")'
        for target_row, field in (
            (9, "tax_rate"),
            (11, "da"),
            (12, "capex"),
            (13, "delta_nwc"),
            (14, "other_noncash"),
            (15, "other_investment"),
        ):
            if _is_number(hist.get(field)):
                operating[f"{col}{target_row}"] = hist[field]
        if _is_number(hist.get("ebit")) or _is_number(hist.get("ebit_margin")):
            operating[f"{col}10"] = f"={col}8*(1-{col}9)"
        if all(_is_number(hist.get(field)) for field in ("tax_rate", "da", "capex", "delta_nwc")):
            operating[f"{col}16"] = (
                f'={col}10+{col}11-{col}12-{col}13+IF({col}14="",0,{col}14)'
                f'-IF({col}15="",0,{col}15)'
            )
            operating[f"{col}17"] = f'=IFERROR({col}16/{col}5,"")'
    operating["D3"] = historical_anchor.get("period") or "最近实际"
    operating["D5"] = historical_anchor["revenue"]
    if _is_number(historical_anchor.get("ebit_margin")):
        operating["D7"] = historical_anchor["ebit_margin"]
        operating["D8"] = "=D5*D7"

    for index, period in enumerate(base_periods):
        col = get_column_letter(5 + index)
        operating[f"{col}3"] = period
        dcf[f"{col}8"] = period
        checks[f"A{6 + index}"] = f"{period} 企业自由现金流（FCFF）勾稽"
    for start in (23, 56, 89):
        for index, period in enumerate(base_periods):
            sensitivity_ws.cell(start + 2, 3 + index, period)

    for name in SCENARIOS:
        col = SCENARIO_COLUMNS[name]
        scenario = scenarios[name]
        previous_revenue = float(historical_anchor["revenue"])
        calc_scenario = fresh["scenarios"][name]
        for index, row in enumerate(scenario["forecast"]):
            period = base_periods[index]
            assumptions[f"A{27 + index}"] = f"{period} 收入增长"
            assumptions[f"A{32 + index}"] = f"{period} 息税前利润率（EBIT利润率）"
            assumptions[f"A{37 + index}"] = f"{period} 折旧与摊销/收入"
            assumptions[f"A{42 + index}"] = f"{period} 资本性支出/收入"
            assumptions[f"A{47 + index}"] = f"{period} 经营性净营运资本增加（ΔNWC）"
            assumptions[f"A{52 + index}"] = f"{period} 其他非现金调整"
            assumptions[f"A{57 + index}"] = f"{period} 其他经营性投资"
            assumptions[f"A{62 + index}"] = f"{period} 现金税率"
            assumptions[f"A{67 + index}"] = f"{period} 折现时点"
            assumptions[f"{col}{27 + index}"] = float(row["revenue"]) / previous_revenue - 1
            assumptions[f"{col}{32 + index}"] = _forecast_metric(row, "ebit_margin")
            assumptions[f"{col}{37 + index}"] = _forecast_metric(row, "da_ratio")
            assumptions[f"{col}{42 + index}"] = _forecast_metric(row, "capex_ratio")
            assumptions[f"{col}{47 + index}"] = row["delta_nwc"]
            assumptions[f"{col}{52 + index}"] = (
                row["other_noncash"] if _is_number(row.get("other_noncash")) else 0
            )
            assumptions[f"{col}{57 + index}"] = (
                row["other_investment"] if _is_number(row.get("other_investment")) else 0
            )
            assumptions[f"{col}{62 + index}"] = calc_scenario["forecast"][index]["tax_rate"]
            assumptions[f"{col}{67 + index}"] = calc_scenario["forecast"][index]["discount_time"]
            previous_revenue = float(row["revenue"])
        assumptions[f"{col}72"] = float(calc_scenario["wacc"]) - base_wacc
        assumptions[f"{col}73"] = scenario["terminal_growth"]
        assumptions[f"{col}74"] = calc_scenario["terminal_discount_time"]
        assumptions[f"{col}75"] = (
            scenario["terminal_fcff"] if _is_number(scenario.get("terminal_fcff")) else None
        )

    sensitivity = calculated["sensitivity"]
    for index, value in enumerate(sensitivity["wacc_rates"]):
        sensitivity_ws.cell(4, 3 + index, value)
    for index, row in enumerate(sensitivity["rows"]):
        sensitivity_ws.cell(5 + index, 2, row["terminal_growth"])

    checks["B4"] = (
        '=IF(COUNTIF(F6:F21,"错误")>0,"FAIL",'
        'IF(COUNTIF(F6:F21,"未完成")>0,"INCOMPLETE","PASS"))'
    )
    if wb.calculation is None:
        wb.calculation = CalcProperties()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"


def _build_support(
    workbook_path: Path,
    support_dir: Path,
    normalized: dict[str, Any],
    calculated: dict[str, Any],
    formula_refs: list[str],
    formula_map: dict[str, str],
) -> None:
    support_dir.mkdir(parents=True, exist_ok=True)
    workbook_hash = _sha256(workbook_path)
    base_name = calculated.get("base_scenario") or "base"
    base = calculated["scenarios"][base_name]
    key_outputs = [
        {
            "sheet": "DCF估值",
            "cell": "B21",
            "expected": base["enterprise_value"],
            "tolerance": 1e-6,
            "require_formula": True,
            "label": "企业价值",
        },
        {
            "sheet": "DCF估值",
            "cell": "B26",
            "expected": base["equity_value"],
            "tolerance": 1e-6,
            "require_formula": True,
            "label": "股权价值",
        },
        {
            "sheet": "DCF估值",
            "cell": "B28",
            "expected": base["per_share_value"],
            "tolerance": 1e-6,
            "require_formula": True,
            "label": "每股价值",
        },
    ]
    for index, name in enumerate(SCENARIOS, 15):
        key_outputs.append(
            {
                "sheet": "敏感性分析",
                "cell": f"D{index}",
                "expected": calculated["scenarios"][name]["per_share_value"],
                "tolerance": 1e-6,
                "require_formula": True,
                "label": f"{name}情景每股价值",
            }
        )
    for row_index, row in enumerate(calculated["sensitivity"]["rows"], 5):
        for col_index, value in enumerate(row["per_share_values"], 3):
            key_outputs.append(
                {
                    "sheet": "敏感性分析",
                    "cell": f"{get_column_letter(col_index)}{row_index}",
                    "expected": value,
                    "tolerance": 1e-6,
                    "require_formula": True,
                    "label": "敏感性每股价值",
                }
            )

    formula_ranges = [
        {"sheet": sheet, "range": cell_range, "label": label}
        for sheet, cell_range, label in FORMULA_RANGES
    ]
    contract = {
        "workflow": "dcf",
        "workbook": workbook_path.name,
        "workbook_sha256": workbook_hash,
        "required_sheets": VISIBLE_SHEETS,
        "formula_ranges": formula_ranges,
        "required_formulas": formula_refs,
        "required_formula_map": formula_map,
        "key_outputs": key_outputs,
        "check_status": {
            "sheet": "模型检查",
            "cell": "B3",
            "pass_value": "PASS",
            "require_formula": True,
        },
    }
    _write_json(support_dir / "workbook-contract.json", contract)

    field_sources = normalized.get("field_sources") or {}
    source_ids = [
        item.get("source_id")
        for item in normalized.get("sources") or []
        if item.get("source_id")
    ]
    review = normalized.get("equity_bridge", {}).get("corporate_action_review") or {}
    lineage = [
        {
            "field": "meta.valuation_date",
            "sheet": "估值假设",
            "cell_or_range": "B6",
            "role": "input",
            "source_ids": source_ids,
            "scenario": None,
            "notes": "估值基准日",
        },
        {
            "field": "wacc_components",
            "sheet": "估值假设",
            "cell_or_range": "B12:B22",
            "role": "formula",
            "source_ids": (field_sources.get("wacc") or {}).get("source_ids")
            or [(field_sources.get("wacc") or {}).get("source_id")],
            "scenario": "base",
            "notes": "WACC组成及计算单元格",
        },
        {
            "field": "equity_bridge.share_classes",
            "sheet": "股本与市值桥",
            "cell_or_range": "A12:L16",
            "role": "input",
            "source_ids": source_ids,
            "scenario": None,
            "notes": "分证券市值反向校验",
        },
        {
            "field": "equity_bridge.corporate_action_review",
            "sheet": "股本与市值桥",
            "cell_or_range": "A18:J22",
            "role": "input",
            "source_ids": review.get("source_ids") or [],
            "scenario": None,
            "notes": "公司行动滚存",
        },
        {
            "field": "scenarios",
            "sheet": "估值假设",
            "cell_or_range": "A26:D75",
            "role": "input",
            "source_ids": source_ids,
            "scenario": "bear/base/bull",
            "notes": "三情景经营与终值假设",
        },
        {
            "field": "enterprise_value",
            "sheet": "DCF估值",
            "cell_or_range": "B21",
            "role": "output",
            "source_ids": [],
            "scenario": "selected",
            "notes": "公式输出",
        },
        {
            "field": "equity_value",
            "sheet": "DCF估值",
            "cell_or_range": "B26",
            "role": "output",
            "source_ids": [],
            "scenario": "selected",
            "notes": "公式输出",
        },
        {
            "field": "per_share_value",
            "sheet": "DCF估值",
            "cell_or_range": "B28",
            "role": "output",
            "source_ids": [],
            "scenario": "selected",
            "notes": "公式输出",
        },
    ]
    for item in lineage:
        item["source_ids"] = [value for value in item["source_ids"] if value]
    _write_json(support_dir / "cell-lineage.json", lineage)

    build_audit = {
        "workbook": workbook_path.name,
        "sha256": workbook_hash,
        "engine": "Python/OpenPyXL",
        "artifact_tool_used": False,
        "formula_contract_status": "PASS",
        "required_formula_count": len(formula_refs),
        "required_sheets_status": "PASS",
        "normalized_calculated_parity": "PASS",
        "delivery_status_at_build": "INCOMPLETE",
        "notes": [
            "All key DCF, scenario and sensitivity derived cells are formulas.",
            "OpenPyXL writes formulas but does not calculate them; open in Excel/LibreOffice or run the package recalculation step.",
        ],
    }
    _write_json(support_dir / "workbook-build-audit.json", build_audit)
    _write_json(
        support_dir / "run-record.json",
        {
            "task_id": normalized.get("meta", {}).get("task_id")
            or workbook_path.stem,
            "workflows": ["dcf"],
            "generator": "scripts/dcf/build_dcf_workbook.py",
            "engine": "Python/OpenPyXL",
            "generated_at_utc": datetime.now(timezone.utc).isoformat(),
            "stages": {
                "scope_locked": "PASS",
                "input_calculation_parity": "PASS",
                "formula_materialized": "PASS",
                "artifact_verified": "PASS",
                "delivery_validated": "INCOMPLETE",
            },
            "model_status": "INCOMPLETE",
            "hard_failures": [],
            "warnings": [
                "Delivery validation has not yet been run; no valuation conclusion may be released."
            ],
        },
    )
    _write_json(
        support_dir / "artifact-manifest.json",
        {
            "hero": {
                "path": workbook_path.name,
                "type": "formula_workbook",
                "bytes": workbook_path.stat().st_size,
                "sha256": workbook_hash,
            },
            "support": [
                "workbook-contract.json",
                "cell-lineage.json",
                "workbook-build-audit.json",
                "run-record.json",
            ],
        },
    )


def build(
    normalized_path: Path,
    calculated_path: Path,
    output_path: Path,
    support_dir: Path | None = None,
    template_path: Path | None = None,
) -> tuple[Path, Path]:
    normalized = json.loads(normalized_path.read_text(encoding="utf-8"))
    calculated = json.loads(calculated_path.read_text(encoding="utf-8"))
    fresh = _validate_inputs(normalized, calculated)
    template_path = template_path or PACKAGE_ROOT / "assets/dcf/dcf-model-template.xlsx"
    if not template_path.is_file():
        raise FileNotFoundError(f"DCF template not found: {template_path}")
    wb = load_workbook(template_path, data_only=False)
    _populate(wb, normalized, calculated, fresh)
    formula_refs, formula_map = _assert_formulas(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    reopened = load_workbook(output_path, data_only=False, read_only=False)
    reopened_refs, reopened_formulas = _assert_formulas(reopened)
    reopened.close()
    if reopened_refs != formula_refs or reopened_formulas != formula_map:
        raise ValueError("formula persistence check failed after reopening workbook")

    support_dir = support_dir or output_path.with_name(f"{output_path.stem}-audit")
    _build_support(
        output_path,
        support_dir,
        normalized,
        calculated,
        formula_refs,
        formula_map,
    )
    return output_path, support_dir


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build a formula-driven DCF workbook with Python/OpenPyXL"
    )
    parser.add_argument("normalized_dcf", type=Path)
    parser.add_argument("calculated_dcf", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    parser.add_argument("--support-dir", type=Path)
    parser.add_argument("--template", type=Path)
    args = parser.parse_args()
    output, support = build(
        args.normalized_dcf,
        args.calculated_dcf,
        args.output_xlsx,
        support_dir=args.support_dir,
        template_path=args.template,
    )
    print(f"saved={output}")
    print(f"audit_dir={support}")


if __name__ == "__main__":
    main()
