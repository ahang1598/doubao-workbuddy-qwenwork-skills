#!/usr/bin/env python3
"""Audit workbook semantics against one compact model contract."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


CELL_RE = re.compile(r"^(?P<sheet>.+)!(?P<cell>\$?[A-Z]{1,3}\$?[1-9][0-9]*)$")
FORMULA_REF_RE = re.compile(
    r"(?:(?:'([^']+)'|([A-Za-z0-9_\u3400-\u9fff][A-Za-z0-9_\u3400-\u9fff .-]*))!)?"
    r"(\$?[A-Z]{1,3}\$?[1-9][0-9]*)"
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalize_ref(reference: str) -> str:
    match = CELL_RE.match(reference)
    if not match:
        raise ValueError(f"invalid cell reference: {reference}")
    return f"{match.group('sheet')}!{match.group('cell').replace('$', '')}"


def split_ref(reference: str) -> tuple[str, str]:
    normalized = normalize_ref(reference)
    return tuple(normalized.rsplit("!", 1))  # type: ignore[return-value]


def iter_contract_refs(book: Any, reference: str) -> list[str]:
    if "!" not in reference:
        raise ValueError(f"range requires sheet: {reference}")
    sheet, address = reference.rsplit("!", 1)
    sheet = sheet.strip("'")
    if sheet not in book.sheetnames:
        raise KeyError(f"missing sheet: {sheet}")
    if ":" not in address:
        return [f"{sheet}!{address.replace('$', '')}"]
    start_col, start_row, end_col, end_row = range_boundaries(address.replace("$", ""))
    return [
        f"{sheet}!{book[sheet].cell(row=row, column=column).coordinate}"
        for row in range(start_row, end_row + 1)
        for column in range(start_col, end_col + 1)
    ]


def formula_dependencies(formula: str, current_sheet: str, sheets: set[str]) -> set[str]:
    dependencies: set[str] = set()
    for match in FORMULA_REF_RE.finditer(formula):
        explicit_sheet = match.group(1) or match.group(2)
        sheet = explicit_sheet if explicit_sheet in sheets else current_sheet
        cell = match.group(3).replace("$", "")
        dependencies.add(f"{sheet}!{cell}")
    return dependencies


def build_graph(book: Any) -> dict[str, set[str]]:
    sheets = set(book.sheetnames)
    graph: dict[str, set[str]] = {}
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" and isinstance(cell.value, str):
                    graph[f"{sheet.title}!{cell.coordinate}"] = formula_dependencies(cell.value, sheet.title, sheets)
    return graph


def path_exists(graph: dict[str, set[str]], source: str, target: str) -> bool:
    source = normalize_ref(source)
    target = normalize_ref(target)
    stack = [target]
    visited: set[str] = set()
    while stack:
        current = stack.pop()
        if current == source:
            return True
        if current in visited:
            continue
        visited.add(current)
        stack.extend(graph.get(current, set()) - visited)
    return False


def cell(book: Any, reference: str) -> Any:
    sheet, address = split_ref(reference)
    if sheet not in book.sheetnames:
        raise KeyError(f"missing sheet: {sheet}")
    return book[sheet][address]


def numeric_value(book: Any, reference: str) -> float:
    value = cell(book, reference).value
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value)):
        raise ValueError(f"{reference} is not a finite numeric value")
    return float(value)


def close_enough(actual: float, expected: float, tolerance: float) -> bool:
    return abs(actual - expected) <= max(tolerance, tolerance * max(abs(actual), abs(expected), 1.0))


def make_check(check_id: str, passed: bool, message: str, **details: Any) -> dict[str, Any]:
    return {
        "check_id": check_id,
        "status": "PASS" if passed else "FAIL",
        "message": message,
        **details,
    }


def audit(workbook_path: Path, contract: dict[str, Any]) -> dict[str, Any]:
    checks: list[dict[str, Any]] = []
    errors: list[str] = []
    warnings: list[str] = []
    try:
        formula_book = load_workbook(workbook_path, data_only=False, read_only=False)
        value_book = load_workbook(workbook_path, data_only=True, read_only=False)
    except Exception as exc:
        return {
            "status": "FAIL",
            "artifact_sha256": sha256(workbook_path),
            "checks": [],
            "errors": [f"workbook cannot be opened: {exc}"],
            "warnings": [],
        }

    graph = build_graph(formula_book)
    requirements = contract.get("prompt_requirements")
    drivers = contract.get("drivers")
    if not isinstance(requirements, list) or not all(isinstance(item, str) and item for item in requirements):
        errors.append("prompt_requirements must be a non-empty string list")
        requirements = []
    if not isinstance(drivers, dict):
        errors.append("drivers must be an object")
        drivers = {}

    for requirement in requirements:
        config = drivers.get(requirement)
        passed = isinstance(config, dict)
        checks.append(make_check(f"prompt.{requirement}", passed, "prompt requirement mapped" if passed else "prompt requirement missing"))
        if not passed:
            continue
        source_refs: list[str] = []
        for declared in config.get("cells", []):
            try:
                source_refs.extend(iter_contract_refs(formula_book, declared))
            except (ValueError, KeyError) as exc:
                errors.append(str(exc))
        targets = config.get("downstream_outputs", [])
        mapped = bool(source_refs) and isinstance(targets, list) and bool(targets)
        if mapped:
            for target in targets:
                try:
                    mapped = mapped and any(path_exists(graph, source, target) for source in source_refs)
                except ValueError as exc:
                    errors.append(str(exc))
                    mapped = False
        checks.append(
            make_check(
                f"lineage.{requirement}",
                mapped,
                "driver reaches required output through formulas" if mapped else "driver is absent or does not reach required output",
                source_cells=source_refs,
                target_cells=targets,
            )
        )

    for item in contract.get("scenario_mappings", []):
        check_id = f"scenario.{item.get('id', 'unnamed')}"
        try:
            display = normalize_ref(item["display"])
            source = normalize_ref(item["source"])
            formula_cell = cell(formula_book, display)
            dependencies = graph.get(display, set())
            direct = source in dependencies
            scale = float(item.get("scale", 1))
            tolerance = float(item.get("tolerance", 1e-6))
            actual = numeric_value(value_book, display)
            expected = numeric_value(value_book, source) * scale
            passed = formula_cell.data_type == "f" and direct and close_enough(actual, expected, tolerance)
            checks.append(make_check(check_id, passed, "scenario mapping verified" if passed else "scenario mapping shifted or inconsistent", actual=actual, expected=expected, formula=formula_cell.value))
        except (KeyError, ValueError, TypeError) as exc:
            checks.append(make_check(check_id, False, str(exc)))

    for item in contract.get("formula_ranges", []):
        check_id = f"coverage.{item.get('id', 'unnamed')}"
        try:
            refs = iter_contract_refs(formula_book, item["range"])
            nonempty = [ref for ref in refs if cell(formula_book, ref).value is not None]
            formulas = [ref for ref in nonempty if cell(formula_book, ref).data_type == "f"]
            ratio = len(formulas) / len(nonempty) if nonempty else 0.0
            minimum = float(item.get("minimum_ratio", 1))
            passed = bool(nonempty) and ratio + 1e-12 >= minimum
            checks.append(make_check(check_id, passed, "formula coverage verified" if passed else "derived range contains hardcoded values", actual=ratio, expected=minimum))
        except (KeyError, ValueError, TypeError) as exc:
            checks.append(make_check(check_id, False, str(exc)))

    for item in contract.get("identities", []):
        check_id = f"identity.{item.get('id', 'unnamed')}"
        try:
            identity_type = item["type"]
            tolerance = float(item.get("tolerance", 1e-6))
            if identity_type == "sum":
                actual = numeric_value(value_book, item["total"])
                expected = sum(numeric_value(value_book, ref) for ref in item["components"])
            elif identity_type == "per_share":
                actual = numeric_value(value_book, item["per_share"])
                expected = numeric_value(value_book, item["equity_value"]) / numeric_value(value_book, item["shares"]) * float(item.get("scale", 1))
            elif identity_type == "fx":
                actual = numeric_value(value_book, item["quote_value"])
                base = numeric_value(value_book, item["base_value"])
                fx = numeric_value(value_book, item["fx"])
                expected = base / fx if item.get("operation", "divide") == "divide" else base * fx
            elif identity_type == "linear":
                actual = numeric_value(value_book, item["output"])
                expected = float(item.get("constant", 0)) + sum(
                    numeric_value(value_book, term["cell"]) * float(term.get("coefficient", 1))
                    for term in item["terms"]
                )
            else:
                raise ValueError(f"unsupported identity type: {identity_type}")
            output_ref = {
                "sum": item.get("total"),
                "per_share": item.get("per_share"),
                "fx": item.get("quote_value"),
                "linear": item.get("output"),
            }[identity_type]
            formula_ok = not item.get("formula_required", True) or cell(formula_book, output_ref).data_type == "f"
            passed = formula_ok and close_enough(actual, expected, tolerance)
            checks.append(make_check(check_id, passed, "dimensional identity verified" if passed else "dimensional identity mismatch", actual=actual, expected=expected, difference=actual - expected, tolerance=tolerance))
        except (KeyError, ValueError, TypeError, ZeroDivisionError) as exc:
            checks.append(make_check(check_id, False, str(exc)))

    reverse = contract.get("reverse_dcf")
    if isinstance(reverse, dict):
        try:
            solved = cell(formula_book, reverse["solved_variable"])
            formula_required = reverse.get("formula_required", True)
            market = numeric_value(value_book, reverse["market_value"])
            recalculated = numeric_value(value_book, reverse["recalculated_value"])
            tolerance = float(reverse.get("tolerance", 1e-6))
            passed = (not formula_required or solved.data_type == "f") and close_enough(market, recalculated, tolerance)
            checks.append(make_check("reverse_dcf.closed_loop", passed, "reverse DCF closes to market value" if passed else "reverse DCF is hardcoded or does not close", market_value=market, recalculated_value=recalculated, residual=recalculated - market, solved_formula=solved.value if solved.data_type == "f" else None))
        except (KeyError, ValueError, TypeError) as exc:
            checks.append(make_check("reverse_dcf.closed_loop", False, str(exc)))

    check_cells: list[str] = []
    for declared in contract.get("model_check_cells", []):
        try:
            check_cells.extend(iter_contract_refs(formula_book, declared))
        except (KeyError, ValueError) as exc:
            errors.append(str(exc))
    for reference in check_cells:
        target = cell(formula_book, reference)
        formula_driven = target.data_type == "f"
        checks.append(make_check(f"model_check.{reference}", formula_driven, "model check is formula-driven" if formula_driven else "static or blank status is not audit evidence"))

    cached_errors: list[dict[str, str]] = []
    for sheet in value_book.worksheets:
        for row in sheet.iter_rows():
            for target in row:
                if isinstance(target.value, str) and target.value.startswith("#"):
                    cached_errors.append({"cell": f"{sheet.title}!{target.coordinate}", "value": target.value})
    checks.append(make_check("workbook.cached_errors", not cached_errors, "no cached formula errors" if not cached_errors else "cached formula errors found", errors=cached_errors))

    failed = [item for item in checks if item["status"] == "FAIL"]
    errors.extend(item["message"] for item in failed)
    return {
        "status": "FAIL" if errors else "PASS",
        "artifact_sha256": sha256(workbook_path),
        "checks": checks,
        "errors": errors,
        "warnings": warnings,
        "summary": {"passed": len(checks) - len(failed), "failed": len(failed), "total": len(checks)},
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("contract", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    contract = json.loads(args.contract.read_text(encoding="utf-8"))
    result = audit(args.workbook.resolve(), contract)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
