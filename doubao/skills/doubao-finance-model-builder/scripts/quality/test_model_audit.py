#!/usr/bin/env python3

from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


MODULE_PATH = Path(__file__).with_name("audit_model.py")
SPEC = importlib.util.spec_from_file_location("audit_model", MODULE_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
SPEC.loader.exec_module(MODULE)


def make_workbook(static_pass: bool = False) -> Path:
    path = Path(tempfile.mkdtemp()) / "model.xlsx"
    workbook = Workbook()
    forecast = workbook.active
    forecast.title = "经营预测"
    forecast["C5"] = 100
    forecast["C6"] = 2
    forecast["C7"] = "=C5*C6"
    valuation = workbook.create_sheet("DCF估值")
    valuation["C5"] = "=经营预测!C7"
    checks = workbook.create_sheet("模型检查")
    checks["D5"] = "PASS" if static_pass else '=IF(DCF估值!C5>0,"PASS","FAIL")'
    workbook.save(path)
    return path


def base_contract() -> dict:
    return {
        "prompt_requirements": ["volume", "unit_revenue"],
        "drivers": {
            "volume": {"cells": ["经营预测!C5"], "downstream_outputs": ["DCF估值!C5"]},
            "unit_revenue": {"cells": ["经营预测!C6"], "downstream_outputs": ["DCF估值!C5"]},
        },
        "model_check_cells": ["模型检查!D5"],
    }


class UnifiedModelAuditTests(unittest.TestCase):
    def test_complete_formula_path_passes(self) -> None:
        result = MODULE.audit(make_workbook(), base_contract())
        self.assertEqual(result["status"], "PASS")

    def test_missing_prompt_driver_fails(self) -> None:
        contract = base_contract()
        contract["prompt_requirements"].append("capacity_utilization")
        result = MODULE.audit(make_workbook(), contract)
        self.assertEqual(result["status"], "FAIL")
        self.assertTrue(any(item["check_id"] == "prompt.capacity_utilization" for item in result["checks"]))

    def test_static_pass_fails(self) -> None:
        result = MODULE.audit(make_workbook(static_pass=True), base_contract())
        self.assertEqual(result["status"], "FAIL")
        self.assertTrue(any("static or blank" in item["message"] for item in result["checks"]))

    def test_hardcoded_driver_does_not_reach_output(self) -> None:
        contract = base_contract()
        contract["drivers"]["volume"]["downstream_outputs"] = ["模型检查!E5"]
        result = MODULE.audit(make_workbook(), contract)
        self.assertEqual(result["status"], "FAIL")

    def test_per_share_identity_catches_fx_adjusted_share_count(self) -> None:
        path = make_workbook()
        from openpyxl import load_workbook

        workbook = load_workbook(path)
        sheet = workbook["DCF估值"]
        sheet["E2"] = 1000
        sheet["E3"] = 100
        sheet["E4"] = 11
        workbook.save(path)
        contract = base_contract()
        contract["identities"] = [{
            "id": "per_share",
            "type": "per_share",
            "equity_value": "DCF估值!E2",
            "shares": "DCF估值!E3",
            "per_share": "DCF估值!E4",
            "tolerance": 1e-9,
        }]
        result = MODULE.audit(path, contract)
        self.assertEqual(result["status"], "FAIL")

    def test_formula_coverage_catches_static_derived_range(self) -> None:
        contract = base_contract()
        contract["formula_ranges"] = [{"id": "forecast", "range": "经营预测!C5:C7", "minimum_ratio": 0.5}]
        result = MODULE.audit(make_workbook(), contract)
        self.assertEqual(result["status"], "FAIL")

    def test_reverse_dcf_hardcoded_solution_fails(self) -> None:
        path = make_workbook()
        from openpyxl import load_workbook

        workbook = load_workbook(path)
        reverse = workbook.create_sheet("反向DCF")
        reverse["C4"] = 100
        reverse["C11"] = 0.02
        reverse["C12"] = 100
        workbook.save(path)
        contract = base_contract()
        contract["reverse_dcf"] = {
            "solved_variable": "反向DCF!C11",
            "market_value": "反向DCF!C4",
            "recalculated_value": "反向DCF!C12",
            "tolerance": 0.01,
            "formula_required": True,
        }
        result = MODULE.audit(path, contract)
        self.assertEqual(result["status"], "FAIL")

    def test_scenario_mapping_shift_fails(self) -> None:
        path = make_workbook()
        from openpyxl import load_workbook

        workbook = load_workbook(path)
        summary = workbook.create_sheet("执行摘要")
        summary["C5"] = "=DCF估值!C5"
        workbook.save(path)
        contract = base_contract()
        contract["scenario_mappings"] = [{
            "id": "base",
            "display": "执行摘要!C5",
            "source": "DCF估值!D5",
            "scale": 1,
        }]
        result = MODULE.audit(path, contract)
        self.assertEqual(result["status"], "FAIL")


if __name__ == "__main__":
    unittest.main(verbosity=2)
