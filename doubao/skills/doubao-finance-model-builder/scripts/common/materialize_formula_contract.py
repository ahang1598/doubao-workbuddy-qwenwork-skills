#!/usr/bin/env python3
"""Materialize formula templates from stable field IDs into a frozen workbook layout."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PLACEHOLDER_RE = re.compile(r"\{([a-z][a-z0-9_.-]*)\}")


def split_address(address: str) -> tuple[str, str]:
    if "!" not in address:
        raise ValueError(f"地址缺少工作表：{address}")
    sheet, coordinate = address.rsplit("!", 1)
    return sheet.strip("'").replace("''", "'"), coordinate.replace("$", "").upper()


def absolute_reference(address: str) -> str:
    sheet, coordinate = split_address(address)
    match = re.fullmatch(r"([A-Z]{1,3})(\d+)", coordinate)
    if not match:
        raise ValueError(f"不是单一单元格地址：{address}")
    escaped = sheet.replace("'", "''")
    return f"'{escaped}'!${match.group(1)}${match.group(2)}"


def compile_template(template: str, cell_map: dict[str, str]) -> str:
    if not isinstance(template, str) or not template.startswith("="):
        raise ValueError("formula_template必须以=开头")

    def replace(match: re.Match[str]) -> str:
        field_id = match.group(1)
        if field_id not in cell_map:
            raise ValueError(f"公式引用未映射字段：{field_id}")
        return absolute_reference(str(cell_map[field_id]))

    compiled = PLACEHOLDER_RE.sub(replace, template)
    unresolved = PLACEHOLDER_RE.findall(compiled)
    if unresolved:
        raise ValueError(f"公式仍有未解析字段：{', '.join(unresolved)}")
    return compiled


def materialize(template_path: Path, output_path: Path, contract: dict[str, Any]) -> dict[str, str]:
    fields = contract.get("fields") or {}
    cell_map = contract.get("cell_map") or {}
    workbook = load_workbook(template_path, data_only=False, read_only=False)
    formulas: dict[str, str] = {}
    try:
        for field_id, specification in fields.items():
            if (specification or {}).get("kind") not in {"formula", "output_formula"}:
                continue
            template = (specification or {}).get("formula_template")
            if not template:
                raise ValueError(f"{field_id}缺少formula_template")
            if field_id not in cell_map:
                raise ValueError(f"{field_id}缺少cell_map映射")
            sheet_name, coordinate = split_address(str(cell_map[field_id]))
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"{field_id}引用不存在的工作表：{sheet_name}")
            formula = compile_template(str(template), cell_map)
            workbook[sheet_name][coordinate] = formula
            formulas[field_id] = formula
        calculation = getattr(workbook, "calculation", None)
        if calculation is not None:
            calculation.calcMode = "auto"
            calculation.fullCalcOnLoad = True
            calculation.forceFullCalc = True
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()
    return formulas


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("template")
    parser.add_argument("formula_contract")
    parser.add_argument("output")
    args = parser.parse_args()
    contract = json.loads(Path(args.formula_contract).read_text(encoding="utf-8"))
    formulas = materialize(Path(args.template), Path(args.output), contract)
    print(json.dumps({"status": "PASS", "formula_count": len(formulas), "formulas": formulas}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
