#!/usr/bin/env python3
"""Shared helpers for deterministic, formula-first finance workbooks."""

from __future__ import annotations

import hashlib
import json
import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
DEFAULT_FONT = "Arial Unicode MS"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def new_workbook(title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "封面"
    ws["A1"] = title
    ws["A1"].font = Font(name=DEFAULT_FONT, size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws.merge_cells("A1:F2")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A4"] = "本工作簿由确定性生成器创建；黄色为输入，黑色为公式。"
    ws["A4"].font = Font(name=DEFAULT_FONT)
    configure_print(ws)
    return wb


def configure_print(ws: Any) -> None:
    """Keep each compact model sheet on one reviewable A4 page."""
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape" if ws.max_column > 6 else "portrait"
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def style_sheet(ws: Any, widths: dict[int, float] | None = None) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.font = copy(cell.font)
            cell.font = Font(
                name=DEFAULT_FONT,
                size=cell.font.sz,
                bold=cell.font.bold,
                italic=cell.font.italic,
                color=cell.font.color,
            )
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(name=DEFAULT_FONT, color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for col, width in (widths or {1: 24, 2: 18, 3: 18, 4: 18, 5: 18}).items():
        ws.column_dimensions[get_column_letter(col)].width = width
    configure_print(ws)


def mark_inputs(ws: Any, refs: list[str]) -> None:
    for ref in refs:
        ws[ref].fill = INPUT_FILL
        ws[ref].font = Font(name=DEFAULT_FONT, color="0000FF")


def save_contract(
    path: Path,
    workflow: str,
    workbook: Path,
    required_formulas: list[str],
    required_sheets: list[str],
    parity: list[dict[str, Any]],
    check_status: dict[str, Any] | None = None,
) -> None:
    formula_book = load_workbook(workbook, data_only=False, read_only=False)
    formula_map: dict[str, str] = {}
    for ref in sorted(set(required_formulas)):
        sheet, coordinate = ref.rsplit("!", 1)
        value = formula_book[sheet][coordinate].value
        if not isinstance(value, str) or not value.startswith("="):
            raise ValueError(f"required formula was not materialized: {ref}")
        formula_map[ref] = value
    row_columns: dict[tuple[str, int], list[int]] = {}
    for ref in sorted(set(required_formulas)):
        sheet, coordinate = ref.rsplit("!", 1)
        match = re.fullmatch(r"([A-Z]{1,3})(\d+)", coordinate.replace("$", ""))
        if match:
            column = 0
            for char in match.group(1):
                column = column * 26 + ord(char) - 64
            row_columns.setdefault((sheet, int(match.group(2))), []).append(column)
    interval_rows: dict[tuple[str, int, int], list[int]] = {}
    for (sheet, row), columns in row_columns.items():
        ordered = sorted(set(columns))
        start = previous = ordered[0]
        for column in ordered[1:] + [None]:
            if column is not None and column == previous + 1:
                previous = column
                continue
            interval_rows.setdefault((sheet, start, previous), []).append(row)
            if column is not None:
                start = previous = column
    formula_ranges: list[dict[str, str]] = []
    for (sheet, start_col, end_col), rows in sorted(interval_rows.items()):
        ordered_rows = sorted(set(rows))
        start_row = previous_row = ordered_rows[0]
        for row in ordered_rows[1:] + [None]:
            if row is not None and row == previous_row + 1:
                previous_row = row
                continue
            start_ref = f"{get_column_letter(start_col)}{start_row}"
            end_ref = f"{get_column_letter(end_col)}{previous_row}"
            formula_ranges.append(
                {
                    "sheet": sheet,
                    "range": start_ref if start_ref == end_ref else f"{start_ref}:{end_ref}",
                    "label": "确定性生成器冻结公式",
                }
            )
            if row is not None:
                start_row = previous_row = row
    payload = {
        "workflow": workflow,
        "workbook": workbook.name,
        "workbook_sha256": sha256(workbook),
        "required_sheets": required_sheets,
        "required_formulas": sorted(set(required_formulas)),
        "required_formula_map": formula_map,
        "formula_count_required": len(set(required_formulas)),
        "formula_ranges": formula_ranges,
        "key_outputs": parity,
        "check_status": check_status,
        "parity": parity,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def audit_contract(workbook_path: Path, contract_path: Path) -> dict[str, Any]:
    contract = json.loads(contract_path.read_text(encoding="utf-8"))
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    errors: list[str] = []
    missing_sheets = [name for name in contract["required_sheets"] if name not in wb.sheetnames]
    if missing_sheets:
        errors.append("缺少工作表：" + "、".join(missing_sheets))
    checked = 0
    for ref in contract["required_formulas"]:
        if "!" not in ref:
            errors.append(f"无效公式引用：{ref}")
            continue
        sheet, coordinate = ref.rsplit("!", 1)
        if sheet not in wb.sheetnames:
            continue
        checked += 1
        cell = wb[sheet][coordinate]
        if cell.data_type != "f":
            errors.append(f"派生单元格被硬编码：{ref}")
        elif cell.value != contract.get("required_formula_map", {}).get(ref):
            errors.append(f"公式与确定性生成合约不一致：{ref}")
        elif not isinstance(cell.value, str) or (
            "!" not in cell.value
            and not re.search(r"\$?[A-Z]{1,3}\$?\d+", cell.value)
        ):
            errors.append(f"公式未连接输入或计算链：{ref}")
    actual_hash = sha256(workbook_path)
    if actual_hash != contract.get("workbook_sha256"):
        errors.append("工作簿哈希与生成合约不一致")
    return {
        "workflow": contract.get("workflow"),
        "status": "PASS" if not errors else "FAIL",
        "artifact_sha256": actual_hash,
        "required_formula_count": len(contract["required_formulas"]),
        "checked_formula_count": checked,
        "errors": errors,
    }
