#!/usr/bin/env python3

from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


MODULE_PATH = Path(__file__).with_name("audit_formula_semantics.py")
SPEC = importlib.util.spec_from_file_location("audit_formula_semantics", MODULE_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
SPEC.loader.exec_module(MODULE)

MATERIALIZE_PATH = Path(__file__).with_name("materialize_formula_contract.py")
MATERIALIZE_SPEC = importlib.util.spec_from_file_location("materialize_formula_contract", MATERIALIZE_PATH)
MATERIALIZE = importlib.util.module_from_spec(MATERIALIZE_SPEC)
assert MATERIALIZE_SPEC and MATERIALIZE_SPEC.loader
MATERIALIZE_SPEC.loader.exec_module(MATERIALIZE)


class FormulaSemanticAuditTests(unittest.TestCase):
    def make_workbook(self, formula: str) -> Path:
        root = Path(tempfile.mkdtemp())
        path = root / "model.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "估值假设"
        labels = ["股权成本", "税后债务成本", "权益权重", "债务权重", "WACC"]
        for row, label in enumerate(labels, start=10):
            sheet.cell(row=row, column=2, value=label)
        sheet["C10"] = 0.10
        sheet["C11"] = 0.03
        sheet["C12"] = 0.80
        sheet["C13"] = 0.20
        sheet["C14"] = formula
        workbook.save(path)
        workbook.close()
        return path

    def contract(self) -> dict:
        return {
            "workflow": "dcf",
            "fields": {
                "cost_of_equity": {"kind": "input", "unit": "percent"},
                "after_tax_cost_of_debt": {"kind": "input", "unit": "percent"},
                "equity_weight": {"kind": "input", "unit": "percent"},
                "debt_weight": {"kind": "input", "unit": "percent"},
                "wacc": {
                    "kind": "output_formula",
                    "unit": "percent",
                    "required_dependencies": [
                        "equity_weight",
                        "cost_of_equity",
                        "debt_weight",
                        "after_tax_cost_of_debt",
                    ],
                    "allowed_dependencies": [
                        "equity_weight",
                        "cost_of_equity",
                        "debt_weight",
                        "after_tax_cost_of_debt",
                    ],
                },
            },
            "cell_map": {
                "cost_of_equity": "估值假设!C10",
                "after_tax_cost_of_debt": "估值假设!C11",
                "equity_weight": "估值假设!C12",
                "debt_weight": "估值假设!C13",
                "wacc": "估值假设!C14",
            },
        }

    def test_correct_formula_passes(self) -> None:
        result = MODULE.audit(self.make_workbook("=C12*C10+C13*C11"), self.contract())
        self.assertEqual(result["status"], "PASS")

    def test_text_label_reference_fails(self) -> None:
        result = MODULE.audit(self.make_workbook("=B12*B10+B13*B11"), self.contract())
        self.assertEqual(result["status"], "FAIL")
        self.assertTrue(any("文本单元格" in item for item in result["errors"]))

    def test_wrong_numeric_assumption_fails(self) -> None:
        contract = self.contract()
        contract["fields"]["other_rate"] = {"kind": "input", "unit": "percent"}
        contract["cell_map"]["other_rate"] = "估值假设!C9"
        path = self.make_workbook("=C12*C10+C13*C9")
        workbook = load_workbook(path)
        workbook["估值假设"]["C9"] = 0.05
        workbook.save(path)
        workbook.close()
        result = MODULE.audit(path, contract)
        self.assertEqual(result["status"], "FAIL")
        self.assertTrue(any("缺少必需依赖" in item for item in result["errors"]))
        self.assertTrue(any("未允许依赖" in item for item in result["errors"]))

    def test_self_reference_fails(self) -> None:
        result = MODULE.audit(self.make_workbook("=C14+C12*C10+C13*C11"), self.contract())
        self.assertEqual(result["status"], "FAIL")
        self.assertTrue(any("禁止依赖" in item or "循环" in item for item in result["errors"]))

    def test_materializer_compiles_field_ids_not_label_coordinates(self) -> None:
        source = self.make_workbook("=0")
        output = source.with_name("compiled.xlsx")
        contract = self.contract()
        contract["fields"]["wacc"]["formula_template"] = (
            "={equity_weight}*{cost_of_equity}"
            "+{debt_weight}*{after_tax_cost_of_debt}"
        )
        MATERIALIZE.materialize(source, output, contract)
        workbook = load_workbook(output, data_only=False)
        try:
            formula = workbook["估值假设"]["C14"].value
        finally:
            workbook.close()
        self.assertEqual(
            formula,
            "='估值假设'!$C$12*'估值假设'!$C$10"
            "+'估值假设'!$C$13*'估值假设'!$C$11",
        )


if __name__ == "__main__":
    unittest.main(verbosity=2)
