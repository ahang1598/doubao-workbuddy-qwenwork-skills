#!/usr/bin/env python3
"""Audit workbook formulas against stable business-field dependencies."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


CELL_RE = re.compile(r"^\$?([A-Z]{1,3})\$?(\d+)$", re.IGNORECASE)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def split_address(address: str, default_sheet: str | None = None) -> tuple[str, str]:
    text = str(address).strip()
    if "!" in text:
        sheet, coordinate = text.rsplit("!", 1)
        sheet = sheet.strip("'").replace("''", "'")
    elif default_sheet:
        sheet, coordinate = default_sheet, text
    else:
        raise ValueError(f"地址缺少工作表：{address}")
    coordinate = coordinate.replace("$", "").upper()
    if not CELL_RE.match(coordinate):
        raise ValueError(f"不是单一单元格地址：{address}")
    return sheet, coordinate


def canonical(address: str, default_sheet: str | None = None) -> str:
    sheet, coordinate = split_address(address, default_sheet)
    return f"{sheet}!{coordinate}"


def range_addresses(token: str, default_sheet: str) -> list[str]:
    text = token.strip()
    if text.startswith("["):
        return []
    if "!" in text:
        sheet_text, range_text = text.rsplit("!", 1)
        sheet = sheet_text.strip("'").replace("''", "'")
    else:
        sheet, range_text = default_sheet, text
    range_text = range_text.replace("$", "").upper()
    if ":" not in range_text:
        if not CELL_RE.match(range_text):
            return []
        return [f"{sheet}!{range_text}"]
    start, end = range_text.split(":", 1)
    if not CELL_RE.match(start) or not CELL_RE.match(end):
        return []
    min_col, min_row, max_col, max_row = range_boundaries(f"{start}:{end}")
    return [
        f"{sheet}!{get_column_letter(col)}{row}"
        for row in range(min_row, max_row + 1)
        for col in range(min_col, max_col + 1)
    ]


def formula_references(formula: str, default_sheet: str) -> list[str]:
    references: list[str] = []
    for item in Tokenizer(formula).items:
        if item.type == "OPERAND" and item.subtype == "RANGE":
            references.extend(range_addresses(item.value, default_sheet))
    return list(dict.fromkeys(references))


def is_text_label(value: Any) -> bool:
    return isinstance(value, str) and not value.startswith("=") and value.strip() != ""


def find_cycles(graph: dict[str, set[str]]) -> list[list[str]]:
    cycles: list[list[str]] = []
    visiting: list[str] = []
    state: dict[str, int] = {}

    def visit(node: str) -> None:
        status = state.get(node, 0)
        if status == 1:
            index = visiting.index(node)
            cycle = visiting[index:] + [node]
            if cycle not in cycles:
                cycles.append(cycle)
            return
        if status == 2:
            return
        state[node] = 1
        visiting.append(node)
        for dependency in sorted(graph.get(node, set())):
            if dependency in graph:
                visit(dependency)
        visiting.pop()
        state[node] = 2

    for field_id in sorted(graph):
        visit(field_id)
    return cycles


def as_set(value: Any) -> set[str]:
    if value is None:
        return set()
    if isinstance(value, str):
        return {value}
    if isinstance(value, Iterable):
        return {str(item) for item in value}
    return set()


def audit(workbook_path: Path, contract: dict[str, Any]) -> dict[str, Any]:
    fields = contract.get("fields") or {}
    cell_map = contract.get("cell_map") or {}
    errors: list[str] = []
    warnings: list[str] = []
    lineage: dict[str, Any] = {}

    if not isinstance(fields, dict) or not fields:
        errors.append("formula-contract缺少fields")
    if not isinstance(cell_map, dict) or not cell_map:
        errors.append("formula-contract缺少cell_map")

    normalized_map: dict[str, str] = {}
    reverse_map: dict[str, str] = {}
    for field_id, address in cell_map.items():
        try:
            normalized = canonical(str(address))
        except ValueError as exc:
            errors.append(f"{field_id}映射无效：{exc}")
            continue
        if normalized in reverse_map:
            errors.append(f"单元格重复映射：{normalized}同时映射到{reverse_map[normalized]}和{field_id}")
        normalized_map[str(field_id)] = normalized
        reverse_map[normalized] = str(field_id)

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    graph: dict[str, set[str]] = {}
    try:
        for field_id, specification in fields.items():
            if field_id not in normalized_map:
                errors.append(f"{field_id}缺少cell_map映射")
                continue
            sheet_name, coordinate = split_address(normalized_map[field_id])
            if sheet_name not in workbook.sheetnames:
                errors.append(f"{field_id}引用不存在的工作表：{sheet_name}")
                continue
            cell = workbook[sheet_name][coordinate]
            kind = str((specification or {}).get("kind", ""))
            formula = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
            if kind in {"formula", "output_formula"} and formula is None:
                errors.append(f"{field_id}必须是公式，但{normalized_map[field_id]}不是公式")
                continue
            if formula is None:
                lineage[field_id] = {
                    "cell": normalized_map[field_id],
                    "kind": kind,
                    "formula": None,
                    "dependencies": [],
                }
                continue

            references = formula_references(formula, sheet_name)
            dependencies: set[str] = set()
            unmapped: list[str] = []
            text_references: list[str] = []
            for address in references:
                dependency = reverse_map.get(address)
                if dependency:
                    dependencies.add(dependency)
                else:
                    unmapped.append(address)
                ref_sheet, ref_coordinate = split_address(address)
                if ref_sheet in workbook.sheetnames and is_text_label(workbook[ref_sheet][ref_coordinate].value):
                    text_references.append(address)

            required = as_set(specification.get("required_dependencies"))
            allowed = as_set(specification.get("allowed_dependencies"))
            forbidden = as_set(specification.get("forbidden_dependencies")) | {field_id}
            missing = sorted(required - dependencies)
            unexpected = sorted(dependencies - allowed) if allowed else []
            prohibited = sorted(dependencies.intersection(forbidden))
            if missing:
                errors.append(f"{field_id}缺少必需依赖：{', '.join(missing)}")
            if unexpected:
                errors.append(f"{field_id}包含未允许依赖：{', '.join(unexpected)}")
            if prohibited:
                errors.append(f"{field_id}包含禁止依赖：{', '.join(prohibited)}")
            if unmapped and not specification.get("allow_unmapped_dependencies", False):
                errors.append(f"{field_id}包含未映射引用：{', '.join(sorted(unmapped))}")
            if text_references and not specification.get("allow_text_references", False):
                errors.append(f"{field_id}引用文本单元格：{', '.join(sorted(text_references))}")

            required_units = specification.get("required_dependency_units") or {}
            for dependency, expected_unit in required_units.items():
                actual_unit = (fields.get(dependency) or {}).get("unit")
                if actual_unit != expected_unit:
                    errors.append(
                        f"{field_id}依赖{dependency}单位错误：期望{expected_unit}，实际{actual_unit or '未声明'}"
                    )

            graph[str(field_id)] = dependencies
            lineage[str(field_id)] = {
                "cell": normalized_map[field_id],
                "kind": kind,
                "unit": specification.get("unit"),
                "formula": formula,
                "references": references,
                "dependencies": sorted(dependencies),
                "unmapped_references": sorted(unmapped),
                "text_references": sorted(text_references),
            }

        cycles = find_cycles(graph)
        for cycle in cycles:
            errors.append(f"公式依赖循环：{' -> '.join(cycle)}")
    finally:
        workbook.close()

    return {
        "workflow": contract.get("workflow"),
        "artifact_path": str(workbook_path.resolve()),
        "artifact_sha256": sha256_file(workbook_path),
        "status": "FAIL" if errors else "PASS",
        "errors": list(dict.fromkeys(errors)),
        "warnings": list(dict.fromkeys(warnings)),
        "field_count": len(fields),
        "mapped_field_count": len(normalized_map),
        "formula_lineage": lineage,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("formula_contract")
    parser.add_argument("output")
    args = parser.parse_args()
    workbook_path = Path(args.workbook).resolve()
    contract = json.loads(Path(args.formula_contract).read_text(encoding="utf-8"))
    result = audit(workbook_path, contract)
    Path(args.output).write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
