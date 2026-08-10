#!/usr/bin/env python3
"""Portable helpers for formula-workbook recalculation and audit."""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional, Union

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


ERROR_CODES = {
    "#REF!",
    "#DIV/0!",
    "#VALUE!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
    "#SPILL!",
    "#CALC!",
    "#FIELD!",
    "#BLOCKED!",
    "#CONNECT!",
    "#GETTING_DATA",
}


FORMULA_FUNCTION_RE = re.compile(r"(?<![A-Z0-9_.])([A-Z][A-Z0-9_.]*)\s*\(", re.IGNORECASE)
SHEET_REFERENCE_RE = re.compile(
    r"(?:'((?:[^']|'')+)'|([A-Za-z0-9_.\u4e00-\u9fff]+))!\$?[A-Z]{1,3}\$?\d+",
    re.IGNORECASE,
)


@dataclass
class Recalculation:
    status: str
    engine: str
    workbook_path: Optional[Path]
    message: Optional[str] = None
    stdout: str = ""
    stderr: str = ""
    temp_root: Optional[Path] = None


def sha256_file(path: Union[str, Path]) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def formula_text(value: Any) -> Optional[str]:
    if isinstance(value, str) and value.startswith("="):
        return value
    text = getattr(value, "text", None)
    if isinstance(text, str) and text.startswith("="):
        return text
    return None


def formula_snapshot(path: Union[str, Path]) -> dict[str, str]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    formulas: dict[str, str] = {}
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    formula = formula_text(cell.value)
                    if formula is not None:
                        formulas[f"{sheet.title}!{cell.coordinate}"] = formula
    finally:
        workbook.close()
    return formulas


def formula_hash(formulas: dict[str, str]) -> str:
    payload = json.dumps(formulas, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _strip_excel_literals(formula: str) -> tuple[str, str, bool, bool]:
    """Replace quoted strings and quoted sheet identifiers for cheap syntax linting."""
    output: list[str] = []
    reference_output: list[str] = []
    index = 0
    in_double = False
    in_single = False
    while index < len(formula):
        char = formula[index]
        if char == '"' and not in_single:
            if in_double and index + 1 < len(formula) and formula[index + 1] == '"':
                output.extend((" ", " "))
                reference_output.extend((" ", " "))
                index += 2
                continue
            in_double = not in_double
            output.append(" ")
            reference_output.append(" ")
        elif char == "'" and not in_double:
            if in_single and index + 1 < len(formula) and formula[index + 1] == "'":
                output.extend((" ", " "))
                reference_output.extend(("'", "'"))
                index += 2
                continue
            in_single = not in_single
            output.append(" ")
            reference_output.append(char)
        else:
            output.append(" " if in_double or in_single else char)
            reference_output.append(" " if in_double else char)
        index += 1
    return "".join(output), "".join(reference_output), not in_double, not in_single


def lint_formula(
    formula: str,
    *,
    sheet_names: Iterable[str],
    forbidden_functions: Iterable[str] = (),
    warn_functions: Iterable[str] = (),
) -> tuple[list[str], list[str]]:
    """Cheap pre-recalculation lint. LibreOffice remains the execution authority."""
    errors: list[str] = []
    warnings: list[str] = []
    if not isinstance(formula, str) or not formula.startswith("="):
        return ["公式必须以=开头"], warnings
    body, reference_body, double_quotes_closed, single_quotes_closed = _strip_excel_literals(formula[1:])
    if not double_quotes_closed:
        errors.append("双引号字符串未闭合")
    if not single_quotes_closed:
        errors.append("单引号工作表名未闭合")
    if not body.strip():
        errors.append("公式主体为空")
    depth = 0
    for char in body:
        if char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
            if depth < 0:
                errors.append("右括号多于左括号")
                break
    if depth > 0:
        errors.append("左括号未闭合")
    if re.search(r"(?:\+=|-=|\*=|/=)", body):
        errors.append("出现非Excel运算符（+=、-=、*=或/=）")
    if re.search(r"[+\-*/^&=<>:,]\s*$", body.strip()):
        errors.append("公式以运算符或分隔符结尾")

    known_sheets = set(sheet_names)
    for match in SHEET_REFERENCE_RE.finditer(reference_body):
        referenced = (match.group(1) or match.group(2) or "").replace("''", "'")
        if referenced and referenced not in known_sheets:
            errors.append(f"引用不存在的工作表：{referenced}")

    functions = {match.group(1).upper() for match in FORMULA_FUNCTION_RE.finditer(body)}
    for name in sorted(functions.intersection({item.upper() for item in forbidden_functions})):
        errors.append(f"使用禁用函数：{name}")
    for name in sorted(functions.intersection({item.upper() for item in warn_functions})):
        warnings.append(f"使用难审计函数：{name}")
    return list(dict.fromkeys(errors)), list(dict.fromkeys(warnings))


def inspect_workbook_structure(path: Union[str, Path]) -> dict[str, Any]:
    """Return a machine-readable structural profile without changing the workbook."""
    formula_book = load_workbook(path, data_only=False, read_only=False)
    cached_book = load_workbook(path, data_only=True, read_only=False)
    try:
        sheets: list[dict[str, Any]] = []
        all_cached_errors: list[dict[str, Any]] = []
        for sheet in formula_book.worksheets:
            cached_sheet = cached_book[sheet.title]
            formulas: list[str] = []
            comments = 0
            nonempty = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        nonempty += 1
                    if formula_text(cell.value) is not None:
                        formulas.append(cell.coordinate)
                    if cell.comment is not None:
                        comments += 1
                    cached_value = cached_sheet[cell.coordinate].value
                    if cached_sheet[cell.coordinate].data_type == "e" or (
                        isinstance(cached_value, str)
                        and (cached_value.strip() in ERROR_CODES or cached_value.strip().startswith("Err:"))
                    ):
                        all_cached_errors.append(
                            {"address": f"{sheet.title}!{cell.coordinate}", "value": cached_value}
                        )
            validations = []
            for validation in sheet.data_validations.dataValidation:
                validations.append(
                    {
                        "type": validation.type,
                        "sqref": str(validation.sqref),
                        "operator": validation.operator,
                        "formula1": validation.formula1,
                        "formula2": validation.formula2,
                    }
                )
            sheets.append(
                {
                    "name": sheet.title,
                    "state": sheet.sheet_state,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "nonempty_cell_count": nonempty,
                    "formula_count": len(formulas),
                    "formula_cells": formulas,
                    "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
                    "hidden_rows": [index for index, dim in sheet.row_dimensions.items() if dim.hidden],
                    "hidden_columns": [index for index, dim in sheet.column_dimensions.items() if dim.hidden],
                    "data_validations": validations,
                    "table_names": sorted(sheet.tables.keys()),
                    "chart_count": len(sheet._charts),
                    "comment_count": comments,
                    "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                    "auto_filter_ref": sheet.auto_filter.ref,
                    "print_area": str(sheet.print_area) if sheet.print_area else None,
                }
            )
        calc = getattr(formula_book, "calculation", None)
        defined_names = []
        try:
            defined_names = sorted(item.name for item in formula_book.defined_names.values())
        except Exception:
            defined_names = []
        return {
            "artifact_path": str(Path(path).resolve()),
            "artifact_sha256": sha256_file(path),
            "sheet_count": len(sheets),
            "sheets": sheets,
            "formula_count": sum(item["formula_count"] for item in sheets),
            "cached_error_count": len(all_cached_errors),
            "cached_errors": all_cached_errors,
            "external_relationship_count": len(getattr(formula_book, "_external_links", []) or []),
            "defined_names": defined_names,
            "calculation_mode": getattr(calc, "calcMode", None),
            "full_calculation_on_load": getattr(calc, "fullCalcOnLoad", None),
            "force_full_calculation": getattr(calc, "forceFullCalc", None),
        }
    finally:
        cached_book.close()
        formula_book.close()


def resolve_soffice(explicit: Optional[str] = None) -> Optional[str]:
    candidates = [
        explicit,
        __import__("os").environ.get("SOFFICE_PATH"),
        shutil.which("soffice"),
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/lib/libreoffice/program/soffice",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).is_file():
            return str(Path(candidate).resolve())
    return None


def recalculate_with_libreoffice(
    path: Union[str, Path], timeout: int = 60, soffice_path: Optional[str] = None
) -> Recalculation:
    soffice = resolve_soffice(soffice_path)
    if not soffice:
        return Recalculation("UNAVAILABLE", "none", None, "soffice not found")
    source = Path(path).resolve()
    temp_root = Path(tempfile.mkdtemp(prefix="finance-workbook-recalc-"))
    input_dir = temp_root / "input"
    output_dir = temp_root / "output"
    profile_dir = temp_root / "profile"
    input_dir.mkdir()
    output_dir.mkdir()
    profile_dir.mkdir()
    staged = input_dir / source.name
    shutil.copy2(source, staged)
    profile_uri = profile_dir.resolve().as_uri()
    command = [
        soffice,
        f"-env:UserInstallation={profile_uri}",
        "--headless",
        "--norestore",
        "--nodefault",
        "--nofirststartwizard",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(output_dir),
        str(staged),
    ]
    try:
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=timeout,
            check=False,
        )
    except subprocess.TimeoutExpired as exc:
        return Recalculation(
            "TIMEOUT",
            "libreoffice",
            None,
            f"LibreOffice recalculation exceeded {timeout}s",
            exc.stdout or "",
            exc.stderr or "",
            temp_root,
        )
    recalculated = output_dir / source.name
    if completed.returncode != 0 or not recalculated.is_file():
        return Recalculation(
            "FAILED",
            "libreoffice",
            None,
            f"LibreOffice exited with {completed.returncode}",
            completed.stdout,
            completed.stderr,
            temp_root,
        )
    return Recalculation(
        "SUCCESS",
        "libreoffice",
        recalculated,
        None,
        completed.stdout,
        completed.stderr,
        temp_root,
    )


def scan_cached_errors(path: Union[str, Path], formula_addresses: Iterable[str]) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    errors: list[dict[str, Any]] = []
    try:
        for address in formula_addresses:
            sheet_name, coordinate = address.rsplit("!", 1)
            if sheet_name not in workbook.sheetnames:
                errors.append({"address": address, "value": "#REF!", "kind": "missing_sheet"})
                continue
            cell = workbook[sheet_name][coordinate]
            value = cell.value
            is_error = cell.data_type == "e" or (
                isinstance(value, str)
                and (value.strip() in ERROR_CODES or value.strip().startswith("Err:"))
            )
            if is_error:
                errors.append({"address": address, "value": value, "kind": "formula_error"})
    finally:
        workbook.close()
    return errors


def direct_circular_addresses(formulas: dict[str, str]) -> list[str]:
    results: list[str] = []
    for full_address, formula in formulas.items():
        sheet_name, address = full_address.rsplit("!", 1)
        match = re.fullmatch(r"([A-Z]{1,3})(\d+)", address.replace("$", ""))
        if not match:
            continue
        stripped = re.sub(
            r"'[^']+'!\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?",
            "",
            formula,
        )
        stripped = re.sub(
            r"[A-Za-z0-9_\u4e00-\u9fff]+!\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?",
            "",
            stripped,
        )
        self_pattern = re.compile(
            rf"(^|[^A-Z0-9_])\$?{match.group(1)}\$?{match.group(2)}([^0-9]|$)"
        )
        if self_pattern.search(stripped):
            results.append(f"{sheet_name}!{address}")
    return results


def external_formula_addresses(formulas: dict[str, str]) -> list[str]:
    return [address for address, formula in formulas.items() if re.search(r"\[[^\]]+\]", formula)]


def range_cells(range_text: str) -> Iterable[tuple[int, int]]:
    min_col, min_row, max_col, max_row = range_boundaries(range_text)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            yield row, col


def cached_value(workbook_path: Union[str, Path], sheet: str, cell: str) -> Any:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    try:
        return workbook[sheet][cell].value
    finally:
        workbook.close()


def audit_contract(
    workbook_path: Union[str, Path],
    contract: dict[str, Any],
    *,
    recalculate: str = "auto",
    timeout: int = 60,
    soffice_path: Optional[str] = None,
) -> dict[str, Any]:
    original = Path(workbook_path).resolve()
    errors: list[str] = []
    warnings: list[str] = []
    metrics: dict[str, Any] = {}
    formulas_before = formula_snapshot(original)
    pre_recalculation_structure = inspect_workbook_structure(original)
    original_book = load_workbook(original, data_only=False, read_only=False)
    sheet_names = list(original_book.sheetnames)
    external_relationship_count = len(getattr(original_book, "_external_links", []) or [])
    original_book.close()

    policy = contract.get("formula_policy") or {}
    lint_errors: list[dict[str, Any]] = []
    lint_warnings: list[dict[str, Any]] = []
    for address, formula in formulas_before.items():
        formula_errors, formula_warnings = lint_formula(
            formula,
            sheet_names=sheet_names,
            forbidden_functions=policy.get("forbidden_functions", []),
            warn_functions=policy.get("warn_functions", ["INDIRECT", "OFFSET"]),
        )
        if formula_errors:
            lint_errors.append({"address": address, "formula": formula, "issues": formula_errors})
        if formula_warnings:
            lint_warnings.append({"address": address, "formula": formula, "issues": formula_warnings})
    for item in lint_errors:
        errors.append(f"公式事前校验失败：{item['address']}（{'；'.join(item['issues'])}）")
    for item in lint_warnings:
        warnings.append(f"公式审计警告：{item['address']}（{'；'.join(item['issues'])}）")

    if not contract.get("workflow"):
        errors.append("工作簿审计合约缺少workflow")
    for field in ("required_sheets", "formula_ranges", "key_outputs"):
        if not isinstance(contract.get(field), list):
            errors.append(f"工作簿审计合约缺少{field}")
    for sheet in contract.get("required_sheets", []):
        if sheet not in sheet_names:
            errors.append(f"缺少必需工作表：{sheet}")

    if not formulas_before:
        errors.append("工作簿未检测到公式")
    circulars = direct_circular_addresses(formulas_before)
    for address in circulars:
        errors.append(f"直接循环引用：{address}")
    external_formulas = external_formula_addresses(formulas_before)
    for address in external_formulas:
        errors.append(f"外部工作簿链接：{address}")
    if external_relationship_count:
        errors.append(f"工作簿包含{external_relationship_count}个外部链接关系")

    formula_book = load_workbook(original, data_only=False, read_only=False)
    try:
        for item in contract.get("formula_ranges", []):
            sheet_name, range_text = item.get("sheet"), item.get("range")
            if sheet_name not in formula_book.sheetnames or not range_text:
                continue
            missing = 0
            sheet = formula_book[sheet_name]
            for row, col in range_cells(range_text):
                if formula_text(sheet.cell(row=row, column=col).value) is None:
                    missing += 1
            if missing:
                label = item.get("label", "关键计算区")
                errors.append(f"公式合约失败：{sheet_name}!{range_text}存在{missing}个非公式单元格（{label}）")
    finally:
        formula_book.close()

    recalculation = (
        Recalculation("SKIPPED", "none", original, "recalculation disabled")
        if recalculate == "off"
        else recalculate_with_libreoffice(original, timeout=timeout, soffice_path=soffice_path)
    )
    audit_path = recalculation.workbook_path or original
    if recalculation.status == "UNAVAILABLE":
        message = "未找到LibreOffice，无法独立重算公式"
        if recalculate == "required":
            errors.append(message)
        else:
            warnings.append(message)
    elif recalculation.status in {"FAILED", "TIMEOUT"}:
        errors.append(recalculation.message or "LibreOffice重算失败")
    elif recalculation.status == "SKIPPED":
        warnings.append("公式重算被显式关闭")

    formulas_after = formula_snapshot(audit_path)
    missing_after = sorted(set(formulas_before) - set(formulas_after))
    added_after = sorted(set(formulas_after) - set(formulas_before))
    if recalculation.status == "SUCCESS" and (missing_after or added_after):
        errors.append(
            f"重算前后公式地址集合发生变化：缺失{len(missing_after)}，新增{len(added_after)}"
        )

    cached_errors = scan_cached_errors(audit_path, formulas_after)
    for item in cached_errors:
        errors.append(f"公式错误：{item['address']}={item['value']}")

    structure = inspect_workbook_structure(audit_path)
    formula_error_addresses = {item["address"] for item in cached_errors}
    nonformula_errors = [
        item for item in structure["cached_errors"] if item["address"] not in formula_error_addresses
    ]
    for item in nonformula_errors:
        errors.append(f"工作簿错误值：{item['address']}={item['value']}")

    value_checks_available = recalculation.status == "SUCCESS"
    if not value_checks_available:
        warnings.append("关键输出数值和检查页状态未验证，因为没有成功的独立重算结果")
    cached_book = load_workbook(audit_path, data_only=True, read_only=False)
    formula_book = load_workbook(original, data_only=False, read_only=False)
    try:
        for item in contract.get("key_outputs", []):
            sheet_name, cell_ref = item.get("sheet"), item.get("cell")
            if sheet_name not in cached_book.sheetnames or not cell_ref:
                continue
            value = cached_book[sheet_name][cell_ref].value
            formula = formula_text(formula_book[sheet_name][cell_ref].value)
            if item.get("require_formula", True) and formula is None:
                errors.append(f"关键输出不是公式：{sheet_name}!{cell_ref}")
            expected = item.get("expected")
            tolerance = item.get("tolerance", 1e-6)
            if value_checks_available:
                if isinstance(expected, (int, float)) and not isinstance(expected, bool):
                    if not isinstance(value, (int, float)) or isinstance(value, bool) or abs(value - expected) > tolerance:
                        errors.append(f"关键输出不一致：{sheet_name}!{cell_ref}={value}，预期={expected}")
                elif value != expected:
                    errors.append(f"关键输出不一致：{sheet_name}!{cell_ref}={value}，预期={expected}")

        check = contract.get("check_status")
        if isinstance(check, dict):
            sheet_name, cell_ref = check.get("sheet"), check.get("cell")
            if sheet_name not in cached_book.sheetnames:
                errors.append(f"缺少检查状态工作表：{sheet_name}")
            elif cell_ref:
                value = cached_book[sheet_name][cell_ref].value
                formula = formula_text(formula_book[sheet_name][cell_ref].value)
                metrics["check_status_value"] = value
                metrics["check_status_formula"] = formula
                if check.get("require_formula", True) and formula is None:
                    errors.append(f"检查总状态不是公式：{sheet_name}!{cell_ref}")
                pass_value = check.get("pass_value", "PASS")
                if value_checks_available and value != pass_value:
                    errors.append(f"工作簿检查总状态为{value or '空白'}")
    finally:
        cached_book.close()
        formula_book.close()

    metrics.update(
        {
            "formula_count": len(formulas_before),
            "formula_hash_before": formula_hash(formulas_before),
            "formula_hash_after": formula_hash(formulas_after),
            "formula_error_count": len(cached_errors),
            "nonformula_error_count": len(nonformula_errors),
            "formula_lint_error_count": len(lint_errors),
            "formula_lint_warning_count": len(lint_warnings),
            "direct_circular_count": len(circulars),
            "external_link_formula_count": len(external_formulas),
            "external_relationship_count": external_relationship_count,
            "recalculation_status": recalculation.status,
            "recalculation_engine": recalculation.engine,
            "required_sheet_count": len(contract.get("required_sheets", [])),
            "present_required_sheet_count": len(
                [name for name in contract.get("required_sheets", []) if name in sheet_names]
            ),
            "formula_contract_count": len(contract.get("formula_ranges", [])),
            "key_output_count": len(contract.get("key_outputs", [])),
        }
    )
    status = "FAIL" if errors else ("INCOMPLETE" if warnings else "PASS")
    output = {
        "status": status,
        "workflow": contract.get("workflow"),
        "artifact_type": "formula_workbook",
        "artifact_path": str(original),
        "artifact_sha256": sha256_file(original),
        "contract_path": contract.get("_contract_path"),
        "errors": errors,
        "warnings": warnings,
        "metrics": metrics,
        "formula_lint": {"errors": lint_errors, "warnings": lint_warnings},
        "workbook_structure_before_recalculation": pre_recalculation_structure,
        "workbook_structure_after_recalculation": structure,
    }
    if recalculation.temp_root:
        shutil.rmtree(recalculation.temp_root, ignore_errors=True)
    return output
