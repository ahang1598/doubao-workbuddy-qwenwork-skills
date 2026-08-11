#!/usr/bin/env python3
"""Build the single-user-deliverable, formula-driven LBO workbook."""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "common"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from deterministic_excel import mark_inputs, new_workbook, save_contract, style_sheet
from lbo_engine import run_case


def _style_text_sheet(ws: Any, widths: dict[int, float]) -> None:
    style_sheet(ws, widths)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _source_ids(value: Any) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return ""


def build(case: dict, workbook_path: Path, contract_path: Path) -> None:
    result = run_case(case)
    if result.get("validation_errors"):
        raise ValueError("; ".join(result["validation_errors"]))

    wb = new_workbook("杠杆收购（LBO）：单一Excel决策与审计模型")
    required: list[str] = []
    company = case["company"]
    years = case["years"]
    tranches = sorted(
        case["debt_tranches"], key=lambda item: item.get("cash_sweep_priority", 999)
    )
    exit_years = case["exit"]["years"]
    exit_multiples = case["exit"]["multiples"]
    base_exit_year = 5 if 5 in exit_years else exit_years[len(exit_years) // 2]
    base_exit_multiple = exit_multiples[len(exit_multiples) // 2]

    cover = wb["封面"]
    cover["A4"] = "本工作簿是唯一正式用户交付物；交易逻辑、回报、风险、来源和检查均在簿内。"
    cover["A6"] = "目标公司"
    cover["B6"] = f"{company['name']}（{company['ticker']}）"
    cover["A7"] = "估值/交易基准日"
    cover["B7"] = case["as_of_date"]
    cover["A8"] = "报告币种"
    cover["B8"] = company["currency"]
    cover["A9"] = "默认退出窗口"
    cover["B9"] = "第3年至第7年"
    cover["A10"] = "默认交付策略"
    cover["B10"] = "只交付Excel；独立Markdown仅在用户明确要求时额外生成"
    cover.column_dimensions["A"].width = 24
    cover.column_dimensions["B"].width = 60

    assumptions = wb.create_sheet("交易与融资假设")
    assumptions.append(["字段", "数值", "单位/说明", "来源/假设编号"])
    entry = case["entry"]
    assumption_rows = [
        ("股权购买价", entry.get("equity_purchase_price", 0), "金额", "交易假设"),
        ("需偿还债务", entry.get("debt_to_refinance", 0), "金额", "交易假设"),
        ("目标现金使用", entry.get("target_cash_used", 0), "金额", "交易假设"),
        ("交易费用", entry.get("transaction_fees", 0), "金额", "交易假设"),
        ("融资费用", entry.get("financing_fees", 0), "金额", "融资假设"),
        ("最低现金", entry.get("minimum_cash", 0), "金额", "流动性假设"),
        ("管理层滚存", entry.get("management_rollover", 0), "金额", "交易假设"),
        ("进入EBITDA", entry.get("entry_ebitda", 0), "金额", "经营假设"),
        ("退出费用率", case["exit"].get("exit_fee_rate", 0), "百分比", "退出假设"),
        ("退出债务类调整", case["exit"].get("exit_debt_like_adjustments", 0), "金额", "退出假设"),
        ("交易基准日", date.fromisoformat(case["as_of_date"]), "日期", "估值基准"),
        ("收购现金", entry.get("cash_acquired", 0), "金额", "交易假设"),
        ("少数股东权益", entry.get("minority_interest", 0), "金额", "估值调整"),
        ("优先股", entry.get("preferred_stock", 0), "金额", "估值调整"),
        ("其他债务类项目", entry.get("other_debt_like", 0), "金额", "估值调整"),
    ]
    for row in assumption_rows:
        assumptions.append(row)
    assumptions.append(
        [
            "进入企业价值",
            "=B2+B3-B13+B14+B15+B16",
            "金额",
            "股权购买价+净债务+类债务调整",
        ]
    )
    assumptions.append(
        [
            "进入倍数",
            '=IFERROR(B17/B9,"")',
            "倍数",
            "进入企业价值/进入EBITDA",
        ]
    )
    mark_inputs(assumptions, [f"B{i}" for i in range(2, 17)])
    required += ["交易与融资假设!B17", "交易与融资假设!B18"]
    style_sheet(assumptions, {1: 28, 2: 20, 3: 20, 4: 32})
    assumptions["B12"].number_format = "yyyy-mm-dd"

    su = wb.create_sheet("Sources & Uses")
    su.append(["用途", "金额", "资金来源", "金额"])
    su.append(["股权购买价", "='交易与融资假设'!B2", "新债务", "=SUM('债务假设'!B2:B200)"])
    su.append(["需偿还债务", "='交易与融资假设'!B3", "目标现金使用", "='交易与融资假设'!B4"])
    su.append(["优先股偿还", entry.get("preferred_stock_to_repay", 0), "管理层滚存", "='交易与融资假设'!B8"])
    su.append(["其他债务类偿还", entry.get("other_debt_like_to_repay", 0), "财务投资人股权", "=B9-SUM(D2:D4)"])
    su.append(["交易费用", "='交易与融资假设'!B5", "", ""])
    su.append(["融资费用", "='交易与融资假设'!B6", "", ""])
    su.append(["最低现金", "='交易与融资假设'!B7", "", ""])
    su.append(["总用途", "=SUM(B2:B8)", "总来源", "=SUM(D2:D5)"])
    su.append(["平衡检查", "=D9-B9", "", ""])
    required += [
        f"Sources & Uses!{ref}"
        for ref in ("B2", "D2", "B3", "D3", "D4", "D5", "B6", "B7", "B8", "B9", "D9", "B10")
    ]
    style_sheet(su, {1: 26, 2: 18, 3: 26, 4: 18})

    debt_inputs = wb.create_sheet("债务假设")
    debt_inputs.append(
        [
            "债务层级",
            "期初余额",
            "现金利率",
            "PIK利率",
            "强制摊还率",
            "偿还优先级",
            "是否循环额度",
            "承诺额度",
            "到期年",
            "来源/假设编号",
        ]
    )
    for tranche in tranches:
        debt_inputs.append(
            [
                tranche["name"],
                tranche.get("opening_balance", 0),
                tranche.get("cash_interest_rate", 0),
                tranche.get("pik_rate", 0),
                tranche.get("mandatory_amortization_rate", 0),
                tranche.get("cash_sweep_priority", 0),
                1 if tranche.get("is_revolver", False) else 0,
                tranche.get("commitment", 0),
                tranche.get("maturity_year"),
                "融资条款",
            ]
        )
    mark_inputs(
        debt_inputs,
        [
            f"{col}{row}"
            for row in range(2, debt_inputs.max_row + 1)
            for col in "BCDEFGHI"
        ],
    )
    style_sheet(
        debt_inputs,
        {1: 24, 2: 16, 3: 14, 4: 14, 5: 16, 6: 14, 7: 14, 8: 14, 9: 12, 10: 24},
    )

    ops = wb.create_sheet("经营预测")
    ops.append(
        [
            "年度",
            "EBITDA",
            "同比增长",
            "现金税",
            "资本开支",
            "营运资本增加",
            "其他现金成本",
            "后续股权投入",
            "计划财务投资人分红",
            "利息前可用于偿债现金",
        ]
    )
    for item in years:
        r = ops.max_row + 1
        growth = (
            f'=IFERROR(B{r}/\'交易与融资假设\'!B9-1,"")'
            if r == 2
            else f'=IFERROR(B{r}/B{r-1}-1,"")'
        )
        ops.append(
            [
                item["year"],
                item.get("ebitda", 0),
                growth,
                item.get("cash_taxes", 0),
                item.get("capex", 0),
                item.get("change_nwc", 0),
                item.get("other_cash_costs", 0),
                item.get("follow_on_equity", 0),
                item.get("sponsor_distribution", 0),
                f"=B{r}-SUM(D{r}:G{r})+H{r}/('Sources & Uses'!D5/('Sources & Uses'!D5+'交易与融资假设'!B8))",
            ]
        )
        required += [f"经营预测!C{r}", f"经营预测!J{r}"]
    mark_inputs(
        ops,
        [
            f"{c}{r}"
            for r in range(2, ops.max_row + 1)
            for c in "ABDEFGHI"
        ],
    )
    style_sheet(ops, {1: 12, 2: 16, 3: 14, 4: 15, 5: 15, 6: 17, 7: 17, 8: 17, 9: 22, 10: 22})

    schedule = wb.create_sheet("分层债务")
    schedule.append(
        [
            "年度",
            "债务层级",
            "期初余额",
            "现金利息",
            "PIK利息",
            "强制摊还",
            "现金扫款",
            "循环额度提款",
            "期末余额",
        ]
    )
    tranche_count = len(tranches)
    for year_index, _year in enumerate(years):
        for tranche_index, tranche in enumerate(tranches):
            r = schedule.max_row + 1
            debt_row = tranche_index + 2
            opening = (
                f"='债务假设'!B{debt_row}"
                if year_index == 0
                else f"='分层债务'!I{r-tranche_count}"
            )
            ops_row = year_index + 2
            prior_sweep = "0" if tranche_index == 0 else f"SUM(G{r-tranche_index}:G{r-1})"
            total_interest_rows = f"SUMIFS($D$2:$D$500,$A$2:$A$500,A{r})"
            total_mandatory_rows = f"SUMIFS($F$2:$F$500,$A$2:$A$500,A{r})"
            available_cash = f"'现金与利息'!B{year_index + 2}"
            draw_formula = (
                f"=MIN(MAX(0,{total_interest_rows}+{total_mandatory_rows}-{available_cash}),MAX(0,'债务假设'!H{debt_row}-C{r}-E{r}))"
                if tranche.get("is_revolver")
                else f"='债务假设'!H{debt_row}*0"
            )
            schedule.append(
                [
                    f"='经营预测'!A{ops_row}",
                    f"='债务假设'!A{debt_row}",
                    opening,
                    f"=C{r}*'债务假设'!C{debt_row}",
                    f"=C{r}*'债务假设'!D{debt_row}",
                    f'=MIN(C{r}+E{r},IF(AND(\'债务假设\'!I{debt_row}<>"",A{r}>=\'债务假设\'!I{debt_row}),C{r}+E{r},C{r}*\'债务假设\'!E{debt_row}))',
                    f"=MIN(MAX(0,{available_cash}-{total_interest_rows}-{total_mandatory_rows}-{prior_sweep}),MAX(0,C{r}+E{r}-F{r}))",
                    draw_formula,
                    f"=MAX(0,C{r}+E{r}+H{r}-F{r}-G{r})",
                ]
            )
            required += [f"分层债务!{c}{r}" for c in "ACDEFGHI"]
    style_sheet(schedule, {1: 12, 2: 24, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16, 8: 17, 9: 16})

    cash = wb.create_sheet("现金与利息")
    cash.append(
        [
            "年度",
            "利息前现金",
            "现金利息",
            "PIK利息",
            "强制摊还",
            "现金扫款",
            "循环额度提款",
            "期末债务",
            "最低现金",
            "期末现金",
            "期末净债务",
            "流动性检查",
            "已实现财务投资人分红",
            "分红缺口",
        ]
    )
    for year_index, _year in enumerate(years):
        r = cash.max_row + 1
        ops_row = year_index + 2
        first = 2 + year_index * tranche_count
        last = first + tranche_count - 1
        cash.append(
            [
                f"='经营预测'!A{ops_row}",
                (
                    f"='经营预测'!J{ops_row}"
                    if year_index == 0
                    else f"='经营预测'!J{ops_row}+MAX(0,J{r-1}-I{r-1})"
                ),
                f"=SUM('分层债务'!D{first}:D{last})",
                f"=SUM('分层债务'!E{first}:E{last})",
                f"=SUM('分层债务'!F{first}:F{last})",
                f"=SUM('分层债务'!G{first}:G{last})",
                f"=SUM('分层债务'!H{first}:H{last})",
                f"=SUM('分层债务'!I{first}:I{last})",
                "='交易与融资假设'!B7",
                f"=I{r}+MAX(0,B{r}-C{r}-E{r}-F{r}+G{r})-M{r}/('Sources & Uses'!D5/('Sources & Uses'!D5+'交易与融资假设'!B8))",
                f"=H{r}-J{r}",
                f'=IF(B{r}-C{r}+G{r}>=E{r},"PASS","FAIL")',
                f"=MIN('经营预测'!I{ops_row},MAX(0,B{r}-C{r}-E{r}-F{r}+G{r})*('Sources & Uses'!D5/('Sources & Uses'!D5+'交易与融资假设'!B8)))",
                f"='经营预测'!I{ops_row}-M{r}",
            ]
        )
        required += [f"现金与利息!{c}{r}" for c in "ABCDEFGHIJKLMN"]
    style_sheet(cash, {1: 12, 2: 17, 3: 15, 4: 15, 5: 15, 6: 15, 7: 17, 8: 16, 9: 14, 10: 16, 11: 16, 12: 16, 13: 22, 14: 16})

    exits = wb.create_sheet("退出回报")
    exits.append(
        [
            "退出年",
            "退出倍数",
            "退出EBITDA",
            "企业价值",
            "期末债务",
            "期末现金",
            "退出债务类调整",
            "退出费用",
            "股权价值",
            "财务投资人持股",
            "财务投资人退出所得",
            "累计投入",
            "利润",
            "MOIC",
            "简化IRR",
            "XIRR",
        ]
    )
    exit_row_map: dict[tuple[int, float], int] = {}
    sponsor_equity_ref = "'Sources & Uses'!D5"
    for exit_year in exit_years:
        ops_row = next(
            i + 2 for i, item in enumerate(years) if int(item["year"]) == int(exit_year)
        )
        cash_row = next(
            i + 2 for i, item in enumerate(years) if int(item["year"]) == int(exit_year)
        )
        for multiple in exit_multiples:
            r = exits.max_row + 1
            exit_row_map[(int(exit_year), float(multiple))] = r
            exits.append(
                [
                    exit_year,
                    multiple,
                    f"='经营预测'!B{ops_row}",
                    f"=B{r}*C{r}",
                    f"='现金与利息'!H{cash_row}",
                    f"='现金与利息'!J{cash_row}",
                    "='交易与融资假设'!B11",
                    f"=D{r}*'交易与融资假设'!B10",
                    f"=MAX(0,D{r}-E{r}+F{r}-G{r}-H{r})",
                    f"={sponsor_equity_ref}/({sponsor_equity_ref}+'交易与融资假设'!B8)",
                    f"=I{r}*J{r}",
                    f"={sponsor_equity_ref}+SUM('经营预测'!H$2:INDEX('经营预测'!H:H,A{r}+1))",
                    f"=K{r}-L{r}+SUM('现金与利息'!M$2:INDEX('现金与利息'!M:M,A{r}+1))",
                    f'=IFERROR((K{r}+SUM(\'现金与利息\'!M$2:INDEX(\'现金与利息\'!M:M,A{r}+1)))/L{r},"")',
                    f'=IFERROR(N{r}^(1/A{r})-1,"")',
                    "",
                ]
            )
            required += [f"退出回报!{c}{r}" for c in "CDEFGHIJKLMNO"]
    mark_inputs(
        exits,
        [f"A{r}" for r in range(2, exits.max_row + 1)]
        + [f"B{r}" for r in range(2, exits.max_row + 1)],
    )
    style_sheet(exits, {1: 12, 2: 14, 3: 16, 4: 16, 5: 16, 6: 16, 7: 18, 8: 15, 9: 16, 10: 17, 11: 20, 12: 16, 13: 16, 14: 13, 15: 14, 16: 14})
    for row in range(2, exits.max_row + 1):
        exits[f"B{row}"].number_format = '0.00x'
        exits[f"J{row}"].number_format = "0.0%"
        exits[f"N{row}"].number_format = '0.00x'
        exits[f"O{row}"].number_format = "0.0%"
        exits[f"P{row}"].number_format = "0.0%"

    investor_cf = wb.create_sheet("投资人现金流")
    investor_cf.append(["退出情景", "行类型", "投入日", *[f"第{i}年" for i in range(1, len(years) + 1)]])
    xirr_refs: dict[int, tuple[int, int]] = {}
    for (exit_year, multiple), exit_row in exit_row_map.items():
        date_row = investor_cf.max_row + 1
        cash_row = date_row + 1
        key = f"Y{exit_year}-{multiple:.2f}x"
        date_values = [key, "日期", "='交易与融资假设'!B12"]
        cash_values = [key, "现金流", "=-'Sources & Uses'!D5"]
        for year in range(1, len(years) + 1):
            date_values.append(
                f"=DATE(YEAR('交易与融资假设'!B12)+{year},MONTH('交易与融资假设'!B12),DAY('交易与融资假设'!B12))"
            )
            op_row = year + 1
            if year <= exit_year:
                suffix = f"+'退出回报'!K{exit_row}" if year == exit_year else ""
                cash_values.append(
                    f"='现金与利息'!M{op_row}-'经营预测'!H{op_row}{suffix}"
                )
            else:
                cash_values.append(f"='经营预测'!H{op_row}*0")
        investor_cf.append(date_values)
        investor_cf.append(cash_values)
        xirr_refs[exit_row] = (date_row, cash_row)
        required += [
            f"投资人现金流!{investor_cf.cell(date_row, c).coordinate}"
            for c in range(3, investor_cf.max_column + 1)
        ]
        required += [
            f"投资人现金流!{investor_cf.cell(cash_row, c).coordinate}"
            for c in range(3, investor_cf.max_column + 1)
        ]
    style_sheet(investor_cf, {1: 18, 2: 12, 3: 16})
    last_cf_col = investor_cf.cell(1, investor_cf.max_column).column_letter
    for exit_row, (date_row, cash_row) in xirr_refs.items():
        exits[f"P{exit_row}"] = (
            f'=IFERROR(XIRR(\'投资人现金流\'!C{cash_row}:{last_cf_col}{cash_row},'
            f'\'投资人现金流\'!C{date_row}:{last_cf_col}{date_row}),"")'
        )
        required.append(f"退出回报!P{exit_row}")

    attribution = wb.create_sheet("回报归因")
    attribution.append(
        [
            "退出年",
            "退出倍数",
            "起始股权购买价",
            "EBITDA增长贡献",
            "倍数变化贡献",
            "净债务偿还贡献",
            "债务类调整变化",
            "退出费用影响",
            "预期退出股权价值",
            "实际退出股权价值",
            "勾稽差异",
        ]
    )
    for (exit_year, multiple), exit_row in exit_row_map.items():
        cash_row = next(
            i + 2 for i, item in enumerate(years) if int(item["year"]) == exit_year
        )
        r = attribution.max_row + 1
        attribution.append(
            [
                f"='退出回报'!A{exit_row}",
                f"='退出回报'!B{exit_row}",
                "='交易与融资假设'!B2",
                f"=('退出回报'!C{exit_row}-'交易与融资假设'!B9)*'交易与融资假设'!B18",
                f"='退出回报'!C{exit_row}*('退出回报'!B{exit_row}-'交易与融资假设'!B18)",
                f"=('交易与融资假设'!B3-'交易与融资假设'!B13)-'现金与利息'!K{cash_row}",
                f"=('交易与融资假设'!B14+'交易与融资假设'!B15+'交易与融资假设'!B16)-'退出回报'!G{exit_row}",
                f"=-'退出回报'!H{exit_row}",
                f"=SUM(C{r}:H{r})",
                f"='退出回报'!I{exit_row}",
                f"=J{r}-I{r}",
            ]
        )
        required += [f"回报归因!{c}{r}" for c in "ABCDEFGHIJK"]
    style_sheet(attribution, {1: 12, 2: 14, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18, 8: 17, 9: 20, 10: 20, 11: 16})

    base_exit_row = exit_row_map[(int(base_exit_year), float(base_exit_multiple))]
    sensitivity = wb.create_sheet("情景与敏感性")
    sensitivity.append(
        [
            "收购价因子/退出倍数",
            *[base_exit_multiple * factor for factor in (0.8, 0.9, 1.0, 1.1, 1.2)],
        ]
    )
    for i, purchase_factor in enumerate((0.8, 0.9, 1.0, 1.1, 1.2), start=2):
        sensitivity.cell(i, 1, purchase_factor)
        for j in range(2, 7):
            multiple_ref = sensitivity.cell(1, j).coordinate
            formula = (
                f'=IFERROR((MAX(0,\'退出回报\'!C{base_exit_row}*{multiple_ref}'
                f'-\'退出回报\'!E{base_exit_row}+\'退出回报\'!F{base_exit_row}'
                f'-\'退出回报\'!G{base_exit_row}-\'退出回报\'!C{base_exit_row}*{multiple_ref}*\'交易与融资假设\'!B10)'
                f'*((\'Sources & Uses\'!D5+\'交易与融资假设\'!B2*($A{i}-1))/'
                f'((\'Sources & Uses\'!D5+\'交易与融资假设\'!B2*($A{i}-1))+\'交易与融资假设\'!B8)))'
                f'/((\'Sources & Uses\'!D5+\'交易与融资假设\'!B2*($A{i}-1))),"")'
            )
            sensitivity.cell(i, j, formula)
            required.append(f"情景与敏感性!{sensitivity.cell(i, j).coordinate}")
    sensitivity["A8"] = "输出"
    sensitivity["B8"] = f"第{base_exit_year}年MOIC；每个格子重新计算收购价、财务投资人股权和持股比例。"
    mark_inputs(
        sensitivity,
        [f"{c}1" for c in "BCDEF"] + [f"A{r}" for r in range(2, 7)],
    )
    style_sheet(sensitivity, {1: 24, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16})
    for row in range(2, 7):
        sensitivity[f"A{row}"].number_format = "0.0x"
        for column in "BCDEF":
            sensitivity[f"{column}{row}"].number_format = "0.00x"

    target_return = wb.create_sheet("目标回报反推")
    target_return.append(
        [
            "目标IRR",
            "退出年",
            "退出倍数",
            "退出所得",
            "允许财务投资人股权",
            "最高股权购买价",
        ]
    )
    for target_irr in case.get("target_irrs", [0.20, 0.25, 0.30]):
        r = target_return.max_row + 1
        target_return.append(
            [
                target_irr,
                base_exit_year,
                base_exit_multiple,
                f"='退出回报'!K{base_exit_row}",
                f"=D{r}/(1+A{r})^B{r}",
                f"=E{r}+SUM('Sources & Uses'!D2:D4)-SUM('Sources & Uses'!B3:B8)",
            ]
        )
        required += [f"目标回报反推!{c}{r}" for c in "DEF"]
    mark_inputs(
        target_return,
        [f"{c}{r}" for r in range(2, target_return.max_row + 1) for c in "ABC"],
    )
    style_sheet(target_return, {1: 16, 2: 14, 3: 14, 4: 18, 5: 20, 6: 20})
    for row in range(2, target_return.max_row + 1):
        target_return[f"A{row}"].number_format = "0.0%"
        target_return[f"C{row}"].number_format = '0.00x'

    ledger = wb.create_sheet("假设依据")
    ledger.append(["假设编号", "主题", "内容", "模型使用位置"])
    for index, (key, value) in enumerate(
        (case.get("assumption_ledger") or {}).items(), start=1
    ):
        ledger.append([f"ASM-LBO-{index:03d}", key, value, "交易、经营、债务或退出模型"])
    mark_inputs(ledger, [f"C{r}" for r in range(2, ledger.max_row + 1)])
    _style_text_sheet(ledger, {1: 18, 2: 28, 3: 80, 4: 34})

    history = wb.create_sheet("历史数据与口径")
    history.append(["项目", "数值", "期间/口径", "来源组"])
    field_sources = (case.get("provenance") or {}).get("field_sources") or {}
    history.append(["进入EBITDA", "='交易与融资假设'!B9", "进入估值口径", _source_ids(field_sources.get("operating_case"))])
    history.append(["进入债务", "='交易与融资假设'!B3", "交易基准日需偿还债务", _source_ids(field_sources.get("entry"))])
    history.append(["收购现金", "='交易与融资假设'!B13", "交易可使用现金", _source_ids(field_sources.get("entry"))])
    history.append(["进入企业价值", "='交易与融资假设'!B17", "股权价值到企业价值桥", _source_ids(field_sources.get("entry"))])
    required += [f"历史数据与口径!B{r}" for r in range(2, history.max_row + 1)]
    style_sheet(history, {1: 26, 2: 20, 3: 34, 4: 26})

    risks = wb.create_sheet("风险与失效条件")
    risks.append(["类型", "编号", "内容", "处理要求"])
    risk_items = case.get("invalidation_conditions") or [
        "EBITDA低于基准经营路径",
        "现金利率或再融资成本显著高于债务假设",
        "退出倍数低于敏感性下沿",
    ]
    for index, item in enumerate(risk_items, start=1):
        risks.append(["失效条件", f"INV-{index:03d}", item, "更新对应输入并完整重跑模型"])
    for index, item in enumerate(result.get("blocking_issues", []), start=1):
        risks.append(["阻断事项", f"BLK-{index:03d}", item, "修复前不得发布MOIC或IRR"])
    for index, item in enumerate(result.get("warnings", []), start=1):
        risks.append(["警告", f"WRN-{index:03d}", item, "评估影响并在结论页披露"])
    mark_inputs(risks, [f"C{r}" for r in range(2, risks.max_row + 1)])
    _style_text_sheet(risks, {1: 14, 2: 14, 3: 82, 4: 38})

    sources = wb.create_sheet("数据来源")
    sources.append(["来源ID", "标题", "日期", "URL/定位", "备注"])
    for item in (case.get("provenance") or {}).get("sources", []):
        sources.append(
            [
                item.get("source_id"),
                item.get("title"),
                item.get("date"),
                item.get("url") or item.get("location"),
                item.get("notes"),
            ]
        )
    mark_inputs(
        sources,
        [f"{c}{r}" for r in range(2, sources.max_row + 1) for c in "ABCDE"],
    )
    _style_text_sheet(sources, {1: 18, 2: 44, 3: 16, 4: 68, 5: 40})

    checks = wb.create_sheet("模型检查")
    checks.append(["检查项", "实际值", "期望值", "差异", "容差", "状态", "修复提示"])
    checks.append(
        [
            "Sources & Uses平衡",
            "='Sources & Uses'!B10",
            0,
            "=B2-C2",
            0.000001,
            '=IF(ABS(D2)<=E2,"PASS","FAIL")',
            "检查交易用途及资金来源",
        ]
    )
    checks.append(
        [
            "期末债务非负",
            "=MIN('分层债务'!I2:I500)",
            0,
            "=MIN(0,B3-C3)",
            0,
            '=IF(B3>=C3,"PASS","FAIL")',
            "检查摊还、现金扫款及循环额度提款",
        ]
    )
    checks.append(
        [
            "流动性检查",
            '=COUNTIF(\'现金与利息\'!L2:L100,"FAIL")',
            0,
            "=B4-C4",
            0,
            '=IF(B4=0,"PASS","FAIL")',
            "现金不足时需补充循环额度或股权",
        ]
    )
    checks.append(
        [
            "退出回报公式",
            f"=COUNT('退出回报'!N2:P{exits.max_row})",
            (exits.max_row - 1) * 3,
            "=B5-C5",
            0,
            '=IF(B5=C5,"PASS","FAIL")',
            "MOIC、IRR、XIRR均须由公式生成",
        ]
    )
    checks.append(
        [
            "回报归因勾稽",
            f"=MAX(ABS('回报归因'!K2:K{attribution.max_row}))",
            0,
            "=B6-C6",
            0.000001,
            '=IF(ABS(D6)<=E6,"PASS","FAIL")',
            "检查EBITDA、倍数、净债务和费用贡献",
        ]
    )
    checks.append(
        [
            "敏感性公式数量",
            "=COUNT('情景与敏感性'!B2:F6)",
            25,
            "=B7-C7",
            0,
            '=IF(B7=C7,"PASS","FAIL")',
            "25个敏感性格均须为公式",
        ]
    )
    checks.append(
        [
            "模型总体状态",
            '=COUNTIF(F2:F7,"FAIL")',
            0,
            "=B8-C8",
            0,
            '=IF(B8=0,"PASS","FAIL")',
            "任一检查失败即阻断回报结论",
        ]
    )
    required += [
        f"模型检查!{c}{r}"
        for r in range(2, checks.max_row + 1)
        for c in "BCDF"
        if isinstance(checks[f"{c}{r}"].value, str)
        and checks[f"{c}{r}"].value.startswith("=")
    ]
    style_sheet(checks, {1: 28, 2: 18, 3: 18, 4: 16, 5: 12, 6: 14, 7: 52})

    summary = wb.create_sheet("交易摘要", 1)
    summary.append(["项目", "数值/结论", "解释", "追溯位置"])
    summary.append(["模型状态", "='模型检查'!F8", "只有PASS才允许发布MOIC、IRR和最高收购价", "模型检查"])
    summary.append(["目标公司", company["name"], company["ticker"], "封面"])
    summary.append(["交易基准日", case["as_of_date"], company["currency"], "交易与融资假设"])
    summary.append(["股权购买价", "='交易与融资假设'!B2", "金额", "交易与融资假设"])
    summary.append(["进入企业价值", "='交易与融资假设'!B17", "金额", "交易与融资假设"])
    summary.append(["进入倍数", "='交易与融资假设'!B18", "企业价值/EBITDA", "交易与融资假设"])
    summary.append(["财务投资人股权", "='Sources & Uses'!D5", "金额", "Sources & Uses"])
    summary.append(["基准退出年", base_exit_year, "持有期", "退出回报"])
    summary.append(["基准退出倍数", base_exit_multiple, "倍数", "退出回报"])
    summary.append(["基准MOIC", f"='退出回报'!N{base_exit_row}", "投资资本倍数", "退出回报"])
    summary.append(["基准XIRR", f"='退出回报'!P{base_exit_row}", "按实际周年日期计算", "退出回报"])
    summary.append(["基准利润", f"='退出回报'!M{base_exit_row}", "退出所得减累计投入", "退出回报"])
    summary.append(["核心回报来源", "详见回报归因", "EBITDA增长、倍数变化、净债务偿还及费用", "回报归因"])
    for r in (2, 5, 6, 7, 8, 11, 12, 13):
        required.append(f"交易摘要!B{r}")
    _style_text_sheet(summary, {1: 28, 2: 48, 3: 52, 4: 24})
    summary["B7"].number_format = '0.00x'
    summary["B10"].number_format = '0.00x'
    summary["B11"].number_format = '0.00x'
    summary["B12"].number_format = "0.0%"
    summary["B13"].number_format = "#,##0.00"

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(workbook_path)
    base_result = next(
        item
        for item in result["exit_results"]
        if int(item["exit_year"]) == int(base_exit_year)
        and abs(float(item["exit_multiple"]) - float(base_exit_multiple)) < 1e-9
    )
    save_contract(
        contract_path,
        "lbo",
        workbook_path,
        required,
        wb.sheetnames,
        [
            {
                "sheet": "交易摘要",
                "cell": "B11",
                "expected": base_result["moic"],
                "tolerance": 1e-8,
                "require_formula": True,
            },
            {
                "sheet": "交易摘要",
                "cell": "B12",
                "expected": base_result["xirr"],
                "tolerance": 1e-8,
                "require_formula": True,
            },
            {
                "sheet": "交易摘要",
                "cell": "B13",
                "expected": base_result["profit"],
                "tolerance": 1e-6,
                "require_formula": True,
            },
            {
                "sheet": "退出回报",
                "cell": f"E{base_exit_row}",
                "expected": base_result["ending_debt"],
                "tolerance": 1e-6,
                "require_formula": True,
            },
        ],
        {
            "sheet": "模型检查",
            "cell": "F8",
            "pass_value": "PASS",
            "require_formula": True,
        },
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="生成单一正式交付的LBO公式工作簿")
    parser.add_argument("case", type=Path)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--contract", type=Path, required=True)
    args = parser.parse_args()
    build(
        json.loads(args.case.read_text(encoding="utf-8")),
        args.workbook,
        args.contract,
    )
    for path, label in ((args.workbook, "LBO工作簿"), (args.contract, "LBO工作簿合约")):
        if not path.is_file():
            raise RuntimeError(f"{label}未生成：{path}")
        if path.stat().st_size <= 0:
            raise RuntimeError(f"{label}大小为0：{path}")
    workbook = load_workbook(args.workbook, read_only=True, data_only=False)
    workbook.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
