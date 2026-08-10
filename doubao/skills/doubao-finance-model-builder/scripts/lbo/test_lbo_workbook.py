#!/usr/bin/env python3
"""Failure-injection tests for the deterministic LBO workbook."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from audit_lbo_workbook import audit_contract
from build_lbo_workbook import build
from validate_case import _sample_case


class LboWorkbookTests(unittest.TestCase):
    def test_bundled_assets_pass_frozen_contract(self) -> None:
        skill_root = Path(__file__).resolve().parents[2]
        asset_root = skill_root / "assets" / "lbo"
        result = audit_contract(
            asset_root / "lbo-model-template.xlsx",
            asset_root / "workbook-contract-example.json",
        )
        self.assertEqual(result["status"], "PASS", result["errors"])
        self.assertGreater(result["required_formula_count"], 100)

    def test_bundled_example_rebuilds_and_passes(self) -> None:
        skill_root = Path(__file__).resolve().parents[2]
        payload = json.loads(
            (skill_root / "assets" / "lbo" / "example-case.json").read_text(
                encoding="utf-8"
            )
        )
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            build(payload, root / "lbo.xlsx", root / "contract.json")
            result = audit_contract(root / "lbo.xlsx", root / "contract.json")
            self.assertEqual(result["status"], "PASS", result["errors"])

    def test_generated_workbook_passes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            build(_sample_case(), root / "lbo.xlsx", root / "contract.json")
            result = audit_contract(root / "lbo.xlsx", root / "contract.json")
            self.assertEqual(result["status"], "PASS", result["errors"])
            self.assertGreater(result["required_formula_count"], 100)
            wb = load_workbook(root / "lbo.xlsx", read_only=True)
            for sheet in (
                "封面",
                "交易摘要",
                "假设依据",
                "历史数据与口径",
                "风险与失效条件",
                "数据来源",
                "模型检查",
            ):
                self.assertIn(sheet, wb.sheetnames)

    def test_hardcoded_debt_cell_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            build(_sample_case(), root / "lbo.xlsx", root / "contract.json")
            wb = load_workbook(root / "lbo.xlsx")
            wb["分层债务"]["H2"] = 600
            wb.save(root / "lbo.xlsx")
            result = audit_contract(root / "lbo.xlsx", root / "contract.json")
            self.assertEqual(result["status"], "FAIL")
            self.assertTrue(any("分层债务!H2" in item for item in result["errors"]))

    def test_hardcoded_return_cell_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            build(_sample_case(), root / "lbo.xlsx", root / "contract.json")
            wb = load_workbook(root / "lbo.xlsx")
            wb["退出回报"]["K2"] = 2.0
            wb.save(root / "lbo.xlsx")
            result = audit_contract(root / "lbo.xlsx", root / "contract.json")
            self.assertEqual(result["status"], "FAIL")

    def test_hardcoded_summary_value_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            build(_sample_case(), root / "lbo.xlsx", root / "contract.json")
            wb = load_workbook(root / "lbo.xlsx")
            wb["交易摘要"]["B8"] = 2.0
            wb.save(root / "lbo.xlsx")
            result = audit_contract(root / "lbo.xlsx", root / "contract.json")
            self.assertEqual(result["status"], "FAIL")


if __name__ == "__main__":
    unittest.main(verbosity=2)
