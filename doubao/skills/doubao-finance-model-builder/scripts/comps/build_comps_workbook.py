#!/usr/bin/env python3
"""Build the single-user-deliverable, formula-driven comps workbook."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "common"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from calculate_comps import (
    ALL_METRICS,
    BOOK_METRICS,
    EV_METRICS,
    FCF_YIELD_METRICS,
    METRIC_LABELS_CN,
    PE_METRICS,
    PS_METRICS,
    VALUATION_METRICS,
    calculate,
)
from deterministic_excel import (
    HEADER_FILL,
    INPUT_FILL,
    DEFAULT_FONT,
    mark_inputs,
    new_workbook,
    save_contract,
    style_sheet,
)


METRIC_COLUMNS = {
    "ltm_ev_revenue": "Y",
    "ntm_ev_revenue": "Z",
    "ltm_ev_ebitda": "AA",
    "ntm_ev_ebitda": "AB",
    "ltm_pe": "AC",
    "ntm_pe": "AD",
    "ltm_ps": "AE",
    "ntm_ps": "AF",
    "ltm_fcf_yield": "AG",
    "ntm_fcf_yield": "AH",
    "price_to_book": "AI",
    "ltm_roe": "AJ",
}
VALUATION_METRIC_ORDER = [metric for metric in ALL_METRICS if metric in VALUATION_METRICS]


def _source_ids(value: Any) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return ""


def _first_primary(payload: dict[str, Any]) -> str:
    profile = payload.get("valuation_profile") or {}
    for metric in profile.get("primary_metrics", []):
        if metric in VALUATION_METRICS:
            return metric
    return "price_to_book"


def _style_text_sheet(ws: Any, widths: dict[int, float]) -> None:
    style_sheet(ws, widths)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _implied_formula(
    metric: str,
    anchor_ref: str,
    fundamental_ref: str,
    shares_ref: str,
    net_debt_ref: str,
) -> str:
    if metric in EV_METRICS:
        return f'=IFERROR(({anchor_ref}*{fundamental_ref}-{net_debt_ref})/{shares_ref},"")'
    if metric in PE_METRICS or metric in PS_METRICS:
        return f'=IFERROR({anchor_ref}*{fundamental_ref}/{shares_ref},"")'
    if metric in FCF_YIELD_METRICS:
        return f'=IFERROR({fundamental_ref}/{anchor_ref}/{shares_ref},"")'
    if metric in BOOK_METRICS:
        return f'=IFERROR({anchor_ref}*{fundamental_ref},"")'
    raise ValueError(f"unsupported valuation metric: {metric}")


def build(payload: dict, workbook_path: Path, contract_path: Path) -> None:
    calculated = calculate(payload)
    wb = new_workbook("可比公司分析：单一Excel决策与审计模型")
    required: list[str] = []
    primary_metric = _first_primary(payload)
    source_companies = {item["ticker"]: item for item in payload["companies"]}
    target_calc = next(
        item for item in calculated["companies"] if item["classification"] == "Target"
    )

    cover = wb["封面"]
    cover["A4"] = "本工作簿是唯一正式用户交付物；定量结论、定性说明、来源和检查均在簿内。"
    cover["A6"] = "估值基准日"
    cover["B6"] = calculated["meta"]["valuation_date"]
    cover["A7"] = "目标公司"
    cover["B7"] = f"{target_calc['name']}（{target_calc['ticker']}）"
    cover["A8"] = "主估值指标"
    cover["B8"] = METRIC_LABELS_CN.get(primary_metric, primary_metric)
    cover["A9"] = "数据等级"
    cover["B9"] = calculated["meta"].get("data_tier")
    cover["A10"] = "默认交付策略"
    cover["B10"] = "只交付Excel；独立Markdown仅在用户明确要求时额外生成"
    cover.column_dimensions["A"].width = 24
    cover.column_dimensions["B"].width = 58

    # Raw inputs and formula-derived multiples remain the calculation spine.
    raw = wb.create_sheet("原始数据与计算")
    headers = [
        "公司",
        "代码",
        "分类",
        "股价",
        "稀释股数",
        "债务",
        "现金",
        "优先股",
        "少数股东权益",
        "债务类调整",
        "非经营投资",
        "LTM收入",
        "NTM收入",
        "LTM EBITDA",
        "NTM EBITDA",
        "LTM净利润",
        "NTM净利润",
        "LTM自由现金流",
        "NTM自由现金流",
        "普通股权益",
        "账面价值股数",
        "平均普通股权益",
        "市值",
        "企业价值",
        "LTM EV/收入",
        "NTM EV/收入",
        "LTM EV/EBITDA",
        "NTM EV/EBITDA",
        "LTM P/E",
        "NTM P/E",
        "LTM P/S",
        "NTM P/S",
        "LTM FCF收益率",
        "NTM FCF收益率",
        "P/B",
        "LTM ROE",
    ]
    raw.append(headers)
    for company in calculated["companies"]:
        source = source_companies[company["ticker"]]
        financials = company["financials"]
        r = raw.max_row + 1
        raw.append(
            [
                company["name"],
                company["ticker"],
                company["classification"],
                company["price"],
                company["diluted_shares"],
                source.get("debt", 0),
                source.get("cash", 0),
                source.get("preferred_equity", 0),
                source.get("noncontrolling_interest", 0),
                source.get("debt_like_adjustments", 0),
                source.get("non_operating_investments", 0),
                financials.get("ltm_revenue"),
                financials.get("ntm_revenue"),
                financials.get("ltm_ebitda"),
                financials.get("ntm_ebitda"),
                financials.get("ltm_net_income"),
                financials.get("ntm_net_income"),
                financials.get("ltm_fcf"),
                financials.get("ntm_fcf"),
                source.get("common_equity"),
                source.get("book_value_shares"),
                company.get("average_common_equity"),
                f"=D{r}*E{r}",
                f"=W{r}+F{r}-G{r}+H{r}+I{r}+J{r}-K{r}",
                f'=IFERROR(X{r}/L{r},"")',
                f'=IFERROR(X{r}/M{r},"")',
                f'=IFERROR(X{r}/N{r},"")',
                f'=IFERROR(X{r}/O{r},"")',
                f'=IFERROR(W{r}/P{r},"")',
                f'=IFERROR(W{r}/Q{r},"")',
                f'=IFERROR(W{r}/L{r},"")',
                f'=IFERROR(W{r}/M{r},"")',
                f'=IFERROR(R{r}/W{r},"")',
                f'=IFERROR(S{r}/W{r},"")',
                f'=IFERROR(D{r}/(T{r}/U{r}),"")',
                f'=IFERROR(P{r}/V{r},"")',
            ]
        )
        required += [
            f"原始数据与计算!{c}{r}"
            for c in [
                "W",
                "X",
                "Y",
                "Z",
                "AA",
                "AB",
                "AC",
                "AD",
                "AE",
                "AF",
                "AG",
                "AH",
                "AI",
                "AJ",
            ]
        ]
    mark_inputs(
        raw,
        [
            f"{c}{r}"
            for r in range(2, raw.max_row + 1)
            for c in [
                "A",
                "B",
                "C",
                "D",
                "E",
                "F",
                "G",
                "H",
                "I",
                "J",
                "K",
                "L",
                "M",
                "N",
                "O",
                "P",
                "Q",
                "R",
                "S",
                "T",
                "U",
                "V",
            ]
        ],
    )
    style_sheet(raw, {1: 22, 2: 14, 3: 12, 4: 14, 5: 14})

    company_rows = {raw[f"B{r}"].value: r for r in range(2, raw.max_row + 1)}
    target_row = company_rows[target_calc["ticker"]]

    target = wb.create_sheet("目标公司")
    target.append(["项目", "数值", "单位/口径", "来源定位"])
    target_rows = [
        ("公司", f"='原始数据与计算'!A{target_row}", "文本", "原始数据与计算"),
        ("代码", f"='原始数据与计算'!B{target_row}", "文本", "原始数据与计算"),
        ("股价", f"='原始数据与计算'!D{target_row}", "每股", "原始数据与计算"),
        ("完全稀释股数", f"='原始数据与计算'!E{target_row}", "股数", "标准化调整"),
        ("市值", f"='原始数据与计算'!W{target_row}", "金额", "市值与EV桥"),
        ("企业价值", f"='原始数据与计算'!X{target_row}", "金额", "市值与EV桥"),
        ("主估值指标", METRIC_LABELS_CN.get(primary_metric, primary_metric), "方法", "假设与方法"),
        ("当前主指标", f"='原始数据与计算'!{METRIC_COLUMNS[primary_metric]}{target_row}", "倍数/收益率", "交易倍数"),
    ]
    for row in target_rows:
        target.append(row)
        if isinstance(row[1], str) and row[1].startswith("="):
            required.append(f"目标公司!B{target.max_row}")
    style_sheet(target, {1: 26, 2: 24, 3: 20, 4: 24})

    screening = wb.create_sheet("同行筛选")
    screening.append(
        [
            "公司",
            "代码",
            "分类",
            "同行角色",
            "业务重合",
            "商业模式",
            "收入结构",
            "市值带",
            "加权得分",
            "数据质量",
            "纳入理由",
            "分类理由",
            "指标理由",
        ]
    )
    for company in calculated["companies"]:
        if company["classification"] == "Target":
            continue
        assessment = company.get("peer_assessment") or {}
        scores = assessment.get("scores") or {}
        screening.append(
            [
                company["name"],
                company["ticker"],
                company["classification"],
                company.get("peer_role"),
                scores.get("business_overlap"),
                scores.get("business_model"),
                scores.get("revenue_structure"),
                scores.get("market_cap_band"),
                assessment.get("weighted_score"),
                assessment.get("data_quality"),
                company.get("selection_rationale"),
                company.get("classification_rationale"),
                company.get("metric_rationale"),
            ]
        )
    mark_inputs(
        screening,
        [
            f"{c}{r}"
            for r in range(2, screening.max_row + 1)
            for c in "CDEFGHIJKLM"
        ],
    )
    _style_text_sheet(
        screening,
        {1: 20, 2: 14, 3: 12, 4: 22, 5: 12, 6: 12, 7: 12, 8: 12, 9: 14, 10: 14, 11: 34, 12: 34, 13: 34},
    )

    basis = wb.create_sheet("历史数据与口径")
    basis.append(
        [
            "公司",
            "代码",
            "股价日期",
            "股数日期",
            "资产负债表日",
            "预测快照日",
            "价格口径",
            "股本来源",
            "基本面来源",
            "公司行动来源",
        ]
    )
    for company in payload["companies"]:
        mapping = company.get("field_sources") or {}
        basis.append(
            [
                company.get("name"),
                company.get("ticker"),
                company.get("price_date"),
                company.get("share_count_date"),
                company.get("balance_sheet_date"),
                company.get("estimate_date"),
                company.get("price_basis"),
                _source_ids(mapping.get("diluted_shares")),
                _source_ids(mapping.get("primary_fundamentals")),
                _source_ids(mapping.get("corporate_actions")),
            ]
        )
    mark_inputs(
        basis,
        [
            f"{c}{r}"
            for r in range(2, basis.max_row + 1)
            for c in "CDEFGHIJ"
        ],
    )
    _style_text_sheet(basis, {1: 20, 2: 14, 3: 14, 4: 14, 5: 16, 6: 16, 7: 20, 8: 20, 9: 20, 10: 20})

    adjustments = wb.create_sheet("标准化调整")
    adjustments.append(
        [
            "公司",
            "报告股数",
            "期权增量",
            "限制性股票",
            "可转债增量",
            "其他稀释",
            "已完成发行",
            "已完成回购",
            "完全稀释股数",
            "债务类净调整",
        ]
    )
    for company in payload["companies"]:
        ticker = company["ticker"]
        rr = company_rows[ticker]
        r = adjustments.max_row + 1
        bridge = company.get("share_count_bridge") or {}
        adjustments.append(
            [
                company.get("name"),
                bridge.get("reported_basic_shares", company.get("diluted_shares")),
                bridge.get("incremental_options", 0),
                bridge.get("unvested_rsus", 0) + bridge.get("performance_shares", 0),
                bridge.get("convertible_incremental_shares", 0),
                bridge.get("other_dilution", 0),
                bridge.get("settled_issuance_shares", 0),
                bridge.get("settled_buyback_shares", 0),
                f"=B{r}+SUM(C{r}:G{r})-H{r}",
                f"='原始数据与计算'!X{rr}-'原始数据与计算'!W{rr}",
            ]
        )
        required += [f"标准化调整!I{r}", f"标准化调整!J{r}"]
    mark_inputs(
        adjustments,
        [
            f"{c}{r}"
            for r in range(2, adjustments.max_row + 1)
            for c in "BCDEFGH"
        ],
    )
    style_sheet(adjustments, {1: 20, 2: 15, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14, 9: 16, 10: 18})

    bridge = wb.create_sheet("市值与EV桥")
    bridge.append(
        [
            "公司",
            "分类",
            "股价",
            "稀释股数",
            "市值",
            "债务",
            "现金",
            "优先股",
            "少数股东",
            "债务类调整",
            "非经营投资",
            "净债务及类债务调整",
            "企业价值",
        ]
    )
    for company in calculated["companies"]:
        rr = company_rows[company["ticker"]]
        r = bridge.max_row + 1
        bridge.append(
            [
                f"='原始数据与计算'!A{rr}",
                f"='原始数据与计算'!C{rr}",
                f"='原始数据与计算'!D{rr}",
                f"='原始数据与计算'!E{rr}",
                f"=C{r}*D{r}",
                f"='原始数据与计算'!F{rr}",
                f"='原始数据与计算'!G{rr}",
                f"='原始数据与计算'!H{rr}",
                f"='原始数据与计算'!I{rr}",
                f"='原始数据与计算'!J{rr}",
                f"='原始数据与计算'!K{rr}",
                f"=F{r}-G{r}+H{r}+I{r}+J{r}-K{r}",
                f"=E{r}+L{r}",
            ]
        )
        required += [f"市值与EV桥!{c}{r}" for c in "ABCDEFGHIJKLM"]
    style_sheet(bridge, {1: 20, 2: 12, 3: 14, 4: 15, 5: 16, 6: 14, 7: 14, 8: 14, 9: 14, 10: 16, 11: 16, 12: 20, 13: 16})

    multiples = wb.create_sheet("交易倍数")
    multiples.append(["公司", "分类", *ALL_METRICS])
    for company in calculated["companies"]:
        rr = company_rows[company["ticker"]]
        r = multiples.max_row + 1
        row = [f"='原始数据与计算'!A{rr}", f"='原始数据与计算'!C{rr}"]
        row.extend(
            f"='原始数据与计算'!{METRIC_COLUMNS[metric]}{rr}"
            for metric in ALL_METRICS
        )
        multiples.append(row)
        required += [
            f"交易倍数!{multiples.cell(r, c).coordinate}"
            for c in range(1, multiples.max_column + 1)
        ]
    style_sheet(multiples, {1: 20, 2: 12})

    core = wb.create_sheet("核心样本")
    core.append(["公司", *ALL_METRICS])
    for company in calculated["companies"]:
        assessment = company.get("peer_assessment") or {}
        if (
            company["classification"] != "Core"
            or assessment.get("eligible_for_core_statistics") is not True
        ):
            continue
        rr = company_rows[company["ticker"]]
        r = core.max_row + 1
        core.append(
            [f"='原始数据与计算'!A{rr}"]
            + [
                f"='原始数据与计算'!{METRIC_COLUMNS[metric]}{rr}"
                for metric in ALL_METRICS
            ]
        )
        required += [
            f"核心样本!{core.cell(r, c).coordinate}"
            for c in range(1, core.max_column + 1)
        ]
    style_sheet(core)

    stats = wb.create_sheet("同行统计")
    stats.append(["指标", "样本数", "P25", "中位数", "P75", "是否稳健"])
    last_core = max(2, core.max_row)
    stats_rows: dict[str, int] = {}
    for metric_index, metric in enumerate(ALL_METRICS, start=2):
        r = stats.max_row + 1
        stats_rows[metric] = r
        letter = core.cell(1, metric_index).column_letter
        rng = f"'核心样本'!{letter}2:{letter}{last_core}"
        stats.append(
            [
                metric,
                f"=COUNT({rng})",
                f'=IFERROR(PERCENTILE({rng},0.25),"")',
                f'=IFERROR(MEDIAN({rng}),"")',
                f'=IFERROR(PERCENTILE({rng},0.75),"")',
                f'=IF(B{r}>=3,"是","否")',
            ]
        )
        required += [f"同行统计!{c}{r}" for c in "BCDEF"]
    style_sheet(stats, {1: 24, 2: 12, 3: 14, 4: 14, 5: 14, 6: 14})
    for metric, row in stats_rows.items():
        number_format = "0.0%" if metric in FCF_YIELD_METRICS or metric == "ltm_roe" else '0.00x'
        for column in "CDE":
            stats[f"{column}{row}"].number_format = number_format

    implied = wb.create_sheet("隐含估值")
    implied.append(
        [
            "指标",
            "P25锚",
            "中位数锚",
            "P75锚",
            "目标基本面",
            "P25隐含每股价值",
            "中位数隐含每股价值",
            "P75隐含每股价值",
            "当前股价",
            "中位数空间",
        ]
    )
    implied_rows: dict[str, int] = {}
    for metric in VALUATION_METRIC_ORDER:
        r = implied.max_row + 1
        implied_rows[metric] = r
        sr = stats_rows[metric]
        if metric in EV_METRICS:
            field_col = {"ltm_revenue": "L", "ntm_revenue": "M", "ltm_ebitda": "N", "ntm_ebitda": "O"}[EV_METRICS[metric]]
            fundamental = f"='原始数据与计算'!{field_col}{target_row}"
        elif metric in PE_METRICS:
            fundamental = f"='原始数据与计算'!{'P' if metric == 'ltm_pe' else 'Q'}{target_row}"
        elif metric in PS_METRICS:
            fundamental = f"='原始数据与计算'!{'L' if metric == 'ltm_ps' else 'M'}{target_row}"
        elif metric in FCF_YIELD_METRICS:
            fundamental = f"='原始数据与计算'!{'R' if metric == 'ltm_fcf_yield' else 'S'}{target_row}"
        else:
            fundamental = f"='原始数据与计算'!D{target_row}/'原始数据与计算'!AI{target_row}"
        shares_ref = f"'原始数据与计算'!E{target_row}"
        net_debt_ref = (
            f"('原始数据与计算'!X{target_row}-'原始数据与计算'!W{target_row})"
        )
        implied.append(
            [
                metric,
                f"='同行统计'!C{sr}",
                f"='同行统计'!D{sr}",
                f"='同行统计'!E{sr}",
                fundamental,
                _implied_formula(metric, f"B{r}", f"E{r}", shares_ref, net_debt_ref),
                _implied_formula(metric, f"C{r}", f"E{r}", shares_ref, net_debt_ref),
                _implied_formula(metric, f"D{r}", f"E{r}", shares_ref, net_debt_ref),
                f"='原始数据与计算'!D{target_row}",
                f'=IFERROR(G{r}/I{r}-1,"")',
            ]
        )
        required += [f"隐含估值!{c}{r}" for c in "BCDEFGHIJ"]
    style_sheet(implied, {1: 24, 2: 14, 3: 14, 4: 14, 5: 18, 6: 20, 7: 20, 8: 20, 9: 14, 10: 16})
    for metric, row in implied_rows.items():
        anchor_format = "0.0%" if metric in FCF_YIELD_METRICS else '0.00x'
        for column in "BCD":
            implied[f"{column}{row}"].number_format = anchor_format
        for column in "FGHI":
            implied[f"{column}{row}"].number_format = "0.00"
        implied[f"J{row}"].number_format = "0.0%"

    primary_implied_row = implied_rows[primary_metric]
    sensitivity = wb.create_sheet("情景与敏感性")
    sensitivity.append(["基本面因子/估值锚因子", 0.8, 0.9, 1.0, 1.1, 1.2])
    for i, fundamental_factor in enumerate((0.8, 0.9, 1.0, 1.1, 1.2), start=2):
        sensitivity.cell(i, 1, fundamental_factor)
        for j in range(2, 7):
            anchor_ref = f"'隐含估值'!C{primary_implied_row}*{sensitivity.cell(1, j).coordinate}"
            fundamental_ref = f"'隐含估值'!E{primary_implied_row}*$A{i}"
            formula = _implied_formula(
                primary_metric,
                anchor_ref,
                fundamental_ref,
                f"'原始数据与计算'!E{target_row}",
                f"('原始数据与计算'!X{target_row}-'原始数据与计算'!W{target_row})",
            )
            sensitivity.cell(i, j, formula)
            required.append(f"情景与敏感性!{sensitivity.cell(i, j).coordinate}")
    sensitivity["A8"] = "说明"
    sensitivity["B8"] = "每个格子均重新应用目标基本面与估值锚，不是对最终价格的静态粘贴。"
    mark_inputs(
        sensitivity,
        [f"{c}1" for c in "BCDEF"] + [f"A{r}" for r in range(2, 7)],
    )
    style_sheet(sensitivity, {1: 24, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16})
    for row in range(2, 7):
        sensitivity[f"A{row}"].number_format = "0.0x"
        for column in "BCDEF":
            sensitivity[f"{column}{row}"].number_format = "0.00"

    market_implied = wb.create_sheet("市场隐含预期")
    market_implied.append(["项目", "数值", "公式含义"])
    market_implied.append(
        [
            "当前主指标",
            f"='原始数据与计算'!{METRIC_COLUMNS[primary_metric]}{target_row}",
            "目标公司当前市场倍数或收益率",
        ]
    )
    market_implied.append(
        [
            "核心同行中位数",
            f"='同行统计'!D{stats_rows[primary_metric]}",
            "核心样本中位数",
        ]
    )
    market_implied.append(
        [
            "当前相对中位数溢折价",
            '=IFERROR(B2/B3-1,"")',
            "正数为溢价，负数为折价；收益率指标需反向解释",
        ]
    )
    required += ["市场隐含预期!B2", "市场隐含预期!B3", "市场隐含预期!B4"]
    style_sheet(market_implied, {1: 30, 2: 20, 3: 52})

    assumptions = wb.create_sheet("假设与方法")
    assumptions.append(["假设编号", "主题", "内容", "模型使用位置"])
    profile = payload.get("valuation_profile") or {}
    assumptions.append(["ASM-COMP-001", "主估值指标", METRIC_LABELS_CN.get(primary_metric, primary_metric), "隐含估值、情景与敏感性"])
    assumptions.append(["ASM-COMP-002", "次要指标", ", ".join(profile.get("secondary_metrics", [])), "交叉检查"])
    assumptions.append(["ASM-COMP-003", "弃用指标", ", ".join(profile.get("rejected_metrics", [])), "方法限制"])
    assumptions.append(["ASM-COMP-004", "溢折价原则", (payload.get("analysis_summary") or {}).get("premium_discount_rationale", ""), "结论摘要"])
    mark_inputs(assumptions, [f"C{r}" for r in range(2, assumptions.max_row + 1)])
    _style_text_sheet(assumptions, {1: 18, 2: 24, 3: 72, 4: 28})

    quality = wb.create_sheet("数据质量与失效条件")
    quality.append(["类型", "编号", "内容", "处理要求"])
    invalidation = (payload.get("analysis_summary") or {}).get("invalidation_conditions", [])
    for i, item in enumerate(invalidation, start=1):
        quality.append(["失效条件", f"INV-{i:03d}", item, "到达观察时点后更新模型并重新发布"])
    for i, item in enumerate(calculated.get("blocking_issues", []), start=1):
        quality.append(["阻断事项", f"BLK-{i:03d}", item, "修复前不得输出正式估值结论"])
    for i, item in enumerate(calculated.get("warnings", []), start=1):
        quality.append(["警告", f"WRN-{i:03d}", item, "评估影响并在摘要中披露"])
    mark_inputs(quality, [f"C{r}" for r in range(2, quality.max_row + 1)])
    _style_text_sheet(quality, {1: 14, 2: 14, 3: 80, 4: 36})

    sources = wb.create_sheet("数据来源")
    sources.append(["来源ID", "标题", "发布日期", "URL/定位", "备注"])
    for item in payload.get("source_ledger", []):
        sources.append(
            [
                item.get("source_id"),
                item.get("title"),
                item.get("publication_date"),
                item.get("url") or item.get("location"),
                item.get("notes"),
            ]
        )
    mark_inputs(
        sources,
        [f"{c}{r}" for r in range(2, sources.max_row + 1) for c in "ABCDE"],
    )
    _style_text_sheet(sources, {1: 18, 2: 42, 3: 16, 4: 68, 5: 38})

    narrative = wb.create_sheet("结论依据")
    narrative.append(["主题", "分析结论", "数值来源"])
    summary = payload.get("analysis_summary") or {}
    narrative.append(["总体结论", summary.get("conclusion", ""), "结论摘要及隐含估值"])
    narrative.append(["同行比较", summary.get("peer_comparison", ""), "同行筛选、交易倍数"])
    narrative.append(["溢折价逻辑", summary.get("premium_discount_rationale", ""), "假设与方法、市场隐含预期"])
    mark_inputs(narrative, [f"B{r}" for r in range(2, narrative.max_row + 1)])
    _style_text_sheet(narrative, {1: 20, 2: 90, 3: 34})

    checks = wb.create_sheet("模型检查")
    checks.append(["检查项", "实际值", "期望值", "差异", "容差", "状态", "修复提示"])
    checks.append(
        [
            "目标公司唯一",
            f'=COUNTIF(\'原始数据与计算\'!C2:C{raw.max_row},"Target")',
            1,
            "=B2-C2",
            0,
            '=IF(ABS(D2)<=E2,"PASS","FAIL")',
            "目标公司分类必须且只能出现一次",
        ]
    )
    checks.append(
        [
            "主指标有效核心样本",
            f"='同行统计'!B{stats_rows[primary_metric]}",
            3,
            "=B3-C3",
            0,
            '=IF(B3>=C3,"PASS","FAIL")',
            "至少需要三家有效核心可比公司",
        ]
    )
    checks.append(
        [
            "目标市值重算",
            f"='原始数据与计算'!W{target_row}",
            f"='原始数据与计算'!D{target_row}*'原始数据与计算'!E{target_row}",
            "=B4-C4",
            0.000001,
            '=IF(ABS(D4)<=E4,"PASS","FAIL")',
            "检查股价与完全稀释股数",
        ]
    )
    checks.append(
        [
            "目标EV重算",
            f"='原始数据与计算'!X{target_row}",
            f"='原始数据与计算'!W{target_row}+'原始数据与计算'!F{target_row}-'原始数据与计算'!G{target_row}+'原始数据与计算'!H{target_row}+'原始数据与计算'!I{target_row}+'原始数据与计算'!J{target_row}-'原始数据与计算'!K{target_row}",
            "=B5-C5",
            0.000001,
            '=IF(ABS(D5)<=E5,"PASS","FAIL")',
            "检查债务、现金、租赁及非经营投资口径",
        ]
    )
    checks.append(
        [
            "敏感性公式数量",
            "=COUNT('情景与敏感性'!B2:F6)",
            25,
            "=B6-C6",
            0,
            '=IF(ABS(D6)<=E6,"PASS","FAIL")',
            "25个敏感性格均须为公式",
        ]
    )
    checks.append(
        [
            "模型总体状态",
            '=COUNTIF(F2:F6,"FAIL")',
            0,
            "=B7-C7",
            0,
            '=IF(B7=0,"PASS","FAIL")',
            "任一检查失败即阻断结论",
        ]
    )
    required += [
        f"模型检查!{c}{r}"
        for r in range(2, checks.max_row + 1)
        for c in "BCDF"
        if isinstance(checks[f"{c}{r}"].value, str)
        and checks[f"{c}{r}"].value.startswith("=")
    ]
    style_sheet(checks, {1: 28, 2: 18, 3: 18, 4: 16, 5: 12, 6: 14, 7: 48})

    decision = wb.create_sheet("结论摘要", 1)
    decision.append(["项目", "数值/结论", "解释", "追溯位置"])
    decision.append(["模型状态", "='模型检查'!F7", "只有PASS才允许发布正式估值区间", "模型检查"])
    decision.append(["估值基准日", calculated["meta"]["valuation_date"], "统一价格、股本和汇率日期", "历史数据与口径"])
    decision.append(["目标公司", target_calc["name"], target_calc["ticker"], "目标公司"])
    decision.append(["主估值指标", METRIC_LABELS_CN.get(primary_metric, primary_metric), "由行业与阶段路由确定", "假设与方法"])
    decision.append(["当前股价", f"='隐含估值'!I{primary_implied_row}", "未复权估值日收盘价", "隐含估值"])
    decision.append(["P25隐含价值", f"=MIN('隐含估值'!F{primary_implied_row}:'隐含估值'!H{primary_implied_row})", "考虑收益率指标方向", "隐含估值"])
    decision.append(["中位数隐含价值", f"='隐含估值'!G{primary_implied_row}", "核心样本中位数锚", "隐含估值"])
    decision.append(["P75隐含价值", f"=MAX('隐含估值'!F{primary_implied_row}:'隐含估值'!H{primary_implied_row})", "考虑收益率指标方向", "隐含估值"])
    decision.append(["中位数上涨/下跌空间", "=IFERROR(B8/B6-1,\"\")", "相对估值基准日股价", "隐含估值"])
    decision.append(["结论", summary.get("conclusion", ""), "定性结论不得替代公式结果", "结论依据"])
    decision.append(["同行比较", summary.get("peer_comparison", ""), "解释增长、盈利、现金转化与风险", "结论依据"])
    decision.append(["溢折价逻辑", summary.get("premium_discount_rationale", ""), "禁止无证据百分比加减", "结论依据"])
    for r in (2, 6, 7, 8, 9, 10):
        if isinstance(decision[f"B{r}"].value, str) and decision[f"B{r}"].value.startswith("="):
            required.append(f"结论摘要!B{r}")
    mark_inputs(decision, ["B11", "B12", "B13"])
    _style_text_sheet(decision, {1: 28, 2: 62, 3: 52, 4: 24})
    for row in (6, 7, 8, 9):
        decision[f"B{row}"].number_format = "0.00"
    decision["B10"].number_format = "0.0%"
    decision.sheet_view.showGridLines = False

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(workbook_path)
    primary_values = calculated["implied_share_values"][primary_metric]
    save_contract(
        contract_path,
        "comps",
        workbook_path,
        required,
        wb.sheetnames,
        [
            {
                "sheet": "结论摘要",
                "cell": "B6",
                "expected": target_calc["price"],
                "tolerance": 1e-8,
                "require_formula": True,
            },
            {
                "sheet": "结论摘要",
                "cell": "B7",
                "expected": primary_values["low"],
                "tolerance": 1e-8,
                "require_formula": True,
            },
            {
                "sheet": "结论摘要",
                "cell": "B8",
                "expected": primary_values["median"],
                "tolerance": 1e-8,
                "require_formula": True,
            },
            {
                "sheet": "结论摘要",
                "cell": "B9",
                "expected": primary_values["high"],
                "tolerance": 1e-8,
                "require_formula": True,
            },
            {
                "sheet": "结论摘要",
                "cell": "B10",
                "expected": primary_values["median"] / target_calc["price"] - 1,
                "tolerance": 1e-8,
                "require_formula": True,
            },
        ],
        {
            "sheet": "模型检查",
            "cell": "F7",
            "pass_value": "PASS",
            "require_formula": True,
        },
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="生成单一正式交付的可比公司公式工作簿")
    parser.add_argument("input", type=Path)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--contract", type=Path, required=True)
    args = parser.parse_args()
    build(
        json.loads(args.input.read_text(encoding="utf-8")),
        args.workbook,
        args.contract,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
