#!/usr/bin/env python3
"""Failure-injection tests for the deterministic comps workbook."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from audit_comps_workbook import audit_contract
from build_comps_workbook import build
from test_comps_quality import sample_payload


class CompsWorkbookTests(unittest.TestCase):
    def test_bundled_assets_pass_frozen_contract(self) -> None:
        skill_root = Path(__file__).resolve().parents[2]
        asset_root = skill_root / "assets" / "comps"
        result = audit_contract(
            asset_root / "comps-model-template.xlsx",
            asset_root / "workbook-contract-example.json",
        )
        self.assertEqual(result["status"], "PASS", result["errors"])
        self.assertGreater(result["required_formula_count"], 100)

    def test_bundled_example_rebuilds_and_passes(self) -> None:
        skill_root = Path(__file__).resolve().parents[2]
        payload = json.loads(
            (skill_root / "assets" / "comps" / "example-input.json").read_text(
                encoding="utf-8"
            )
        )
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            build(payload, root / "comps.xlsx", root / "contract.json")
            result = audit_contract(root / "comps.xlsx", root / "contract.json")
            self.assertEqual(result["status"], "PASS", result["errors"])

    def test_generated_workbook_passes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            build(sample_payload(), root / "comps.xlsx", root / "contract.json")
            result = audit_contract(root / "comps.xlsx", root / "contract.json")
            self.assertEqual(result["status"], "PASS", result["errors"])
            self.assertGreater(result["required_formula_count"], 100)
            wb = load_workbook(root / "comps.xlsx", read_only=True)
            for sheet in (
                "封面",
                "结论摘要",
                "假设与方法",
                "数据质量与失效条件",
                "数据来源",
                "结论依据",
                "模型检查",
            ):
                self.assertIn(sheet, wb.sheetnames)

    def test_hardcoded_multiple_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            build(sample_payload(), root / "comps.xlsx", root / "contract.json")
            wb = load_workbook(root / "comps.xlsx")
            wb["原始数据与计算"]["AI2"] = 1.0
            wb.save(root / "comps.xlsx")
            result = audit_contract(root / "comps.xlsx", root / "contract.json")
            self.assertEqual(result["status"], "FAIL")
            self.assertTrue(any("原始数据与计算!AI2" in item for item in result["errors"]))

    def test_hardcoded_sensitivity_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            build(sample_payload(), root / "comps.xlsx", root / "contract.json")
            wb = load_workbook(root / "comps.xlsx")
            wb["情景与敏感性"]["C3"] = 12.0
            wb.save(root / "comps.xlsx")
            result = audit_contract(root / "comps.xlsx", root / "contract.json")
            self.assertEqual(result["status"], "FAIL")

    def test_hardcoded_summary_value_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            build(sample_payload(), root / "comps.xlsx", root / "contract.json")
            wb = load_workbook(root / "comps.xlsx")
            wb["结论摘要"]["B5"] = 10.0
            wb.save(root / "comps.xlsx")
            result = audit_contract(root / "comps.xlsx", root / "contract.json")
            self.assertEqual(result["status"], "FAIL")


if __name__ == "__main__":
    unittest.main(verbosity=2)
