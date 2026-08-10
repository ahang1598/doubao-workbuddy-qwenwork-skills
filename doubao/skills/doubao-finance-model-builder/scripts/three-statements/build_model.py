#!/usr/bin/env python3
"""Generate a formula-driven Chinese three-statement model with OpenPyXL."""

from __future__ import annotations

import argparse
import hashlib
import json
from copy import copy
from datetime import date
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


CASES = ("Base", "Bull", "Bear")
CASE_CN = {"Base": "基准", "Bull": "乐观", "Bear": "悲观"}
DRIVERS = (
    ("revenue_growth", "收入增速", "%"),
    ("gross_margin", "毛利率", "%"),
    ("sga_pct_revenue", "销售及管理费用率", "%"),
    ("rnd_pct_revenue", "研发费用率", "%"),
    ("other_opex_pct_revenue", "其他经营费用率", "%"),
    ("tax_rate", "所得税率", "%"),
    ("dso", "应收账款周转天数（DSO）", "天"),
    ("inventory_days", "存货周转天数", "天"),
    ("dpo", "应付账款周转天数（DPO）", "天"),
    ("other_current_assets_pct_revenue", "其他流动资产/收入", "%"),
    ("other_current_liabilities_pct_revenue", "其他流动负债/收入", "%"),
    ("capex_pct_revenue", "资本开支/收入", "%"),
    ("depreciation_rate_on_opening_ppe", "折旧摊销/期初固定资产净额", "%"),
    ("debt_interest_rate", "债务利率", "%"),
    ("cash_interest_rate", "现金收益率", "%"),
    ("dividend_payout", "股利支付率", "%"),
    ("debt_draw", "新增借款", "金额"),
    ("debt_repayment", "偿还债务", "金额"),
    ("share_issuance", "股权融资", "金额"),
    ("minimum_cash", "最低现金余额", "金额"),
)

NAVY = "17365D"
LIGHT_BLUE = "D9EAF7"
FORECAST = "E2F0D9"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"
BLUE = "0000FF"
GREEN = "008000"
RED = "9C0006"
PASS = "C6EFCE"
FAIL = "FFC7CE"
THIN = Side(style="thin", color="7F7F7F")
AMOUNT_FMT = '#,##0.0;[Red](#,##0.0);-'
PCT_FMT = '0.0%;[Red](0.0%);-'


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def emit_support_artifacts(
    workbook_path: Path,
    support_dir: Path,
    *,
    workflow: str,
    generator: str,
    input_path: Path | None,
    required_sheets: list[str],
    lineage: list[dict[str, Any]],
) -> None:
    """Create deterministic support metadata from the saved workbook."""
    support_dir.mkdir(parents=True, exist_ok=True)
    workbook_bytes = workbook_path.read_bytes()
    workbook_hash = hashlib.sha256(workbook_bytes).hexdigest()
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        formula_addresses: list[str] = []
        formula_ranges: list[dict[str, Any]] = []
        sheet_formula_counts: dict[str, int] = {}
        for ws in wb.worksheets:
            cells = [
                cell
                for row in ws.iter_rows()
                for cell in row
                if cell.data_type == "f"
                or (isinstance(cell.value, str) and cell.value.startswith("="))
            ]
            sheet_formula_counts[ws.title] = len(cells)
            formula_addresses.extend(f"{ws.title}!{cell.coordinate}" for cell in cells)
            if cells:
                min_row = min(cell.row for cell in cells)
                max_row = max(cell.row for cell in cells)
                min_col = min(cell.column for cell in cells)
                max_col = max(cell.column for cell in cells)
                formula_ranges.append(
                    {
                        "sheet": ws.title,
                        "range": (
                            f"{get_column_letter(min_col)}{min_row}:"
                            f"{get_column_letter(max_col)}{max_row}"
                        ),
                        "formula_count": len(cells),
                    }
                )
        missing_sheets = [name for name in required_sheets if name not in wb.sheetnames]
        check_formula = wb["检查"]["B2"].value if "检查" in wb.sheetnames else None
        check_status = {
            "sheet": "检查",
            "cell": "B2",
            "pass_value": "PASS",
            "require_formula": True,
            "formula": check_formula,
        }
    finally:
        wb.close()

    formula_count = len(formula_addresses)
    contract = {
        "workflow": workflow,
        "workbook": workbook_path.name,
        "workbook_sha256": workbook_hash,
        "required_sheets": required_sheets,
        "formula_ranges": formula_ranges,
        "required_formulas": formula_addresses,
        "formula_count": formula_count,
        "check_status": check_status,
    }
    _write_json(support_dir / "workbook-contract.json", contract)
    _write_json(support_dir / "cell-lineage.json", lineage)

    required_sheets_status = "PASS" if not missing_sheets else "FAIL"
    formula_status = (
        "PASS"
        if formula_count > 0
        and isinstance(check_formula, str)
        and check_formula.startswith("=")
        else "FAIL"
    )
    build_status = (
        "PASS"
        if required_sheets_status == "PASS" and formula_status == "PASS"
        else "FAIL"
    )
    build_audit = {
        "workbook": workbook_path.name,
        "sha256": workbook_hash,
        "bytes": workbook_path.stat().st_size,
        "engine": "Python/OpenPyXL",
        "artifact_tool_used": False,
        "required_sheets_status": required_sheets_status,
        "missing_required_sheets": missing_sheets,
        "formula_contract_status": formula_status,
        "formula_count": formula_count,
        "formula_ranges": formula_ranges,
        "sheet_formula_counts": sheet_formula_counts,
        "check_status": check_status,
        "build_status": build_status,
        "notes": [
            "支持文件由最终保存的工作簿回读生成。",
            "OpenPyXL写入公式但不计算缓存值；交付前仍需Excel或LibreOffice隔离重算。",
        ],
    }
    _write_json(support_dir / "workbook-build-audit.json", build_audit)

    input_hash = None
    if input_path is not None and input_path.is_file():
        input_hash = hashlib.sha256(input_path.read_bytes()).hexdigest()
    run_record = {
        "task_id": workbook_path.stem,
        "workflows": [workflow],
        "generator": generator,
        "engine": "Python/OpenPyXL",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "input": {
            "path": str(input_path) if input_path is not None else None,
            "sha256": input_hash,
        },
        "output": {
            "path": str(workbook_path),
            "sha256": workbook_hash,
            "bytes": workbook_path.stat().st_size,
        },
        "stages": {
            "input_validated": "PASS",
            "formula_materialized": formula_status,
            "required_sheets_verified": required_sheets_status,
            "artifact_verified": build_status,
            "delivery_recalculated": "INCOMPLETE",
        },
        "model_status": (
            "INCOMPLETE" if build_status == "PASS" else "FAIL"
        ),
        "hard_failures": (
            []
            if build_status == "PASS"
            else ["工作簿构建期结构或公式检查失败"]
        ),
        "warnings": [
            "尚未执行隔离重算与交付审计；本记录不将检查页缓存值声明为PASS。"
        ],
    }
    _write_json(support_dir / "run-record.json", run_record)
    _write_json(
        support_dir / "artifact-manifest.json",
        {
            "hero": {
                "path": workbook_path.name,
                "type": "formula_workbook",
                "bytes": workbook_path.stat().st_size,
                "sha256": workbook_hash,
            },
            "support": [
                "workbook-contract.json",
                "cell-lineage.json",
                "workbook-build-audit.json",
                "run-record.json",
            ],
        },
    )


def finite(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def validate_input(data: dict[str, Any]) -> None:
    if data.get("model_type", "growth") != "growth":
        raise ValueError("build_model.py only accepts model_type=growth")
    for key in ("company", "market", "ticker", "accounting_standard", "currency", "units"):
        if not data.get(key):
            raise ValueError(f"{key} is required")
    years = data.get("forecast_years")
    if not isinstance(data.get("actual_year"), int) or not isinstance(years, list) or not years:
        raise ValueError("actual_year and forecast_years are required")
    if data.get("selected_scenario") not in CASES:
        raise ValueError("selected_scenario must be Base, Bull or Bear")
    n = len(years)
    for section in ("historical_income_statement", "opening_balance_sheet"):
        obj = data.get(section)
        if not isinstance(obj, dict) or any(not finite(v) for v in obj.values()):
            raise ValueError(f"{section} must contain finite numeric values")
    for case in CASES:
        scenario = data.get("scenarios", {}).get(case)
        if not isinstance(scenario, dict):
            raise ValueError(f"missing scenario {case}")
        for key, _, _ in DRIVERS:
            values = scenario.get(key)
            if not isinstance(values, list) or len(values) != n or any(not finite(v) for v in values):
                raise ValueError(f"{case}.{key} must contain {n} finite values")
    bs = data["opening_balance_sheet"]
    assets = sum(bs[k] for k in ("cash", "accounts_receivable", "inventory", "other_current_assets", "net_ppe", "other_noncurrent_assets"))
    le = sum(bs[k] for k in ("accounts_payable", "other_current_liabilities", "debt", "other_noncurrent_liabilities", "share_capital", "retained_earnings"))
    if abs(assets - le) > 0.01:
        raise ValueError(f"opening balance sheet does not balance: {assets} vs {le}")


def _title(ws, last_col: int, text: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(1, 1, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28


def _header(ws, row: int, last_col: int) -> None:
    for cell in ws[row][:last_col]:
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(name="Arial", bold=True)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = Border(bottom=THIN)


def _section(ws, row: int, last_col: int, label: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row, 1, label)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Arial", bold=True, color=WHITE)


def _formula(cell, expression: str, cross_sheet: bool = False) -> None:
    cell.value = expression
    cell.font = Font(name="Arial", size=10, color=GREEN if cross_sheet else "000000")


def _statement(ws, title: str, data: dict[str, Any], last_col: int, max_row: int) -> None:
    _title(ws, last_col, title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"
    ws.cell(3, 1, f"项目（{data['currency']} {data['units']}）")
    ws.cell(3, 2, date(data["actual_year"], 1, 1)).number_format = 'yyyy"A"'
    for i, year in enumerate(data["forecast_years"], start=3):
        ws.cell(3, i, date(year, 1, 1)).number_format = 'yyyy"E"'
        ws.cell(3, i).fill = PatternFill("solid", fgColor=FORECAST)
    _header(ws, 3, last_col)
    ws.column_dimensions["A"].width = 35
    for col in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=last_col):
        for cell in row:
            font = copy(cell.font)
            font.name, font.sz = "Arial", 10
            cell.font = font


def _labels(ws, items: list[tuple[int, str, int]]) -> None:
    for row, label, indent in items:
        ws.cell(row, 1, label).alignment = Alignment(indent=indent)


def _total_border(ws, rows: list[int], last_col: int) -> None:
    for row in rows:
        for col in range(1, last_col + 1):
            ws.cell(row, col).border = Border(top=THIN)


def build(
    data: dict[str, Any],
    output_path: Path,
    support_dir: Path | None = None,
    input_path: Path | None = None,
) -> None:
    validate_input(data)
    years = data["forecast_years"]
    n = len(years)
    last_col = 2 + n
    hist = data["historical_income_statement"]
    obs = data["opening_balance_sheet"]

    wb = Workbook()
    wb.remove(wb.active)
    names = ("封面", "假设", "经营驱动", "明细预测", "利润表", "资产负债表", "现金流量表", "检查", "来源")
    sheets = {name: wb.create_sheet(name) for name in names}
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    for ws in sheets.values():
        ws.sheet_view.showGridLines = False

    # 假设
    ass = sheets["假设"]
    case_starts = {"Base": 4, "Bull": 4 + n, "Bear": 4 + 2 * n}
    selected_start = 4 + 3 * n
    ass_last_col = selected_start + n - 1
    _title(ass, ass_last_col, "预测假设与情景选择")
    ass["A3"], ass["B3"] = "当前情景", CASE_CN[data["selected_scenario"]]
    ass["A4"], ass["B4"] = "颜色说明", "蓝色=可编辑输入；绿色=跨工作表公式"
    ass["B3"].font = Font(color=BLUE, bold=True)
    validation = DataValidation(type="list", formula1='"基准,乐观,悲观"')
    ass.add_data_validation(validation)
    validation.add(ass["B3"])
    ass.append([])
    ass.cell(6, 1, "预测驱动")
    ass.cell(6, 2, "单位")
    ass.cell(6, 3, "备注")
    for case, start in case_starts.items():
        ass.merge_cells(start_row=5, start_column=start, end_row=5, end_column=start + n - 1)
        ass.cell(5, start, CASE_CN[case])
        for col in range(start, start + n):
            ass.cell(5, col).fill = PatternFill("solid", fgColor=NAVY)
            ass.cell(5, col).font = Font(color=WHITE, bold=True)
        for i, year in enumerate(years):
            ass.cell(6, start + i, date(year, 1, 1)).number_format = 'yyyy"E"'
    ass.merge_cells(start_row=5, start_column=selected_start, end_row=5, end_column=selected_start + n - 1)
    ass.cell(5, selected_start, "当前情景")
    for col in range(selected_start, selected_start + n):
        ass.cell(5, col).fill = PatternFill("solid", fgColor="548235")
        ass.cell(5, col).font = Font(color=WHITE, bold=True)
        ass.cell(6, col, date(years[col - selected_start], 1, 1)).number_format = 'yyyy"E"'
    _header(ass, 6, ass_last_col)
    driver_rows: dict[str, int] = {}
    for offset, (key, label, unit) in enumerate(DRIVERS):
        row = 7 + offset
        driver_rows[key] = row
        ass.cell(row, 1, label)
        ass.cell(row, 2, unit)
        ass.cell(row, 3, "通用回退假设；有经营KPI时应替换" if key == "revenue_growth" else "")
        for case, start in case_starts.items():
            for i, value in enumerate(data["scenarios"][case][key]):
                c = ass.cell(row, start + i, value)
                c.font = Font(color=BLUE)
        for i in range(n):
            base = f"{get_column_letter(case_starts['Base'] + i)}{row}"
            bull = f"{get_column_letter(case_starts['Bull'] + i)}{row}"
            bear = f"{get_column_letter(case_starts['Bear'] + i)}{row}"
            _formula(ass.cell(row, selected_start + i), f'=IF($B$3="基准",{base},IF($B$3="乐观",{bull},{bear}))')
        fmt = PCT_FMT if unit == "%" else ("0.0" if unit == "天" else AMOUNT_FMT)
        for col in range(4, ass_last_col + 1):
            ass.cell(row, col).number_format = fmt
    ass.freeze_panes = "D7"
    for col, width in {1: 38, 2: 14, 3: 42}.items():
        ass.column_dimensions[get_column_letter(col)].width = width
    for col in range(4, ass_last_col + 1):
        ass.column_dimensions[get_column_letter(col)].width = 12

    def sel(key: str, i: int) -> str:
        return f"'假设'!{get_column_letter(selected_start + i)}{driver_rows[key]}"

    # 经营驱动
    drv = sheets["经营驱动"]
    _statement(drv, "经营驱动", data, last_col, 8)
    _labels(drv, [(5, "营业收入", 0), (6, "同比增速", 1), (8, "说明", 0)])
    drv["B5"] = hist["revenue"]
    drv["B5"].font = Font(color=BLUE)
    for i in range(n):
        col = get_column_letter(3 + i)
        prev = get_column_letter(2 + i)
        _formula(drv[f"{col}5"], f"={prev}5*(1+{sel('revenue_growth', i)})", True)
        _formula(drv[f"{col}6"], f"={sel('revenue_growth', i)}", True)
    for cell in drv[5][1:last_col]:
        cell.number_format = AMOUNT_FMT
    for cell in drv[6][1:last_col]:
        cell.number_format = PCT_FMT
    drv.merge_cells(start_row=8, start_column=2, end_row=8, end_column=last_col)
    drv["B8"] = "收入增速为通用回退方法；应优先替换为销量、价格、客户、产能、门店或订单驱动。"
    drv["B8"].fill = PatternFill("solid", fgColor=YELLOW)
    drv["B8"].alignment = Alignment(wrap_text=True)

    # 明细预测
    sch = sheets["明细预测"]
    _statement(sch, "明细预测", data, last_col, 38)
    for row, label in ((5, "营运资本"), (15, "固定资产净额滚动"), (22, "债务与利息"), (30, "权益滚动")):
        _section(sch, row, last_col, label)
    _labels(sch, [
        (6, "应收账款", 0), (7, "存货", 0), (8, "其他流动资产", 0), (9, "应付账款", 0), (10, "其他流动负债", 0),
        (11, "应收账款周转天数（DSO）", 1), (12, "存货周转天数", 1), (13, "应付账款周转天数（DPO）", 1),
        (16, "期初固定资产净额", 0), (17, "资本开支", 0), (18, "折旧与摊销", 0), (19, "期末固定资产净额", 0),
        (23, "期初债务", 0), (24, "新增借款", 0), (25, "偿还债务", 0), (26, "期末债务", 0), (27, "利息费用", 0),
        (31, "期初留存收益", 0), (32, "净利润", 0), (33, "股利", 0), (34, "期末留存收益", 0),
        (36, "期初股本及资本公积", 0), (37, "股权融资", 0), (38, "期末股本及资本公积", 0),
    ])
    for row, value in {6: obs["accounts_receivable"], 7: obs["inventory"], 8: obs["other_current_assets"], 9: obs["accounts_payable"], 10: obs["other_current_liabilities"], 19: obs["net_ppe"], 26: obs["debt"], 34: obs["retained_earnings"], 38: obs["share_capital"]}.items():
        sch.cell(row, 2, value).font = Font(color=BLUE)
    for i in range(n):
        col, prev = get_column_letter(3 + i), get_column_letter(2 + i)
        formulas = {
            6: f"='经营驱动'!{col}5*{sel('dso', i)}/365",
            7: f"=-'利润表'!{col}6*{sel('inventory_days', i)}/365",
            8: f"='经营驱动'!{col}5*{sel('other_current_assets_pct_revenue', i)}",
            9: f"=-'利润表'!{col}6*{sel('dpo', i)}/365",
            10: f"='经营驱动'!{col}5*{sel('other_current_liabilities_pct_revenue', i)}",
            11: f"={sel('dso', i)}", 12: f"={sel('inventory_days', i)}", 13: f"={sel('dpo', i)}",
            16: f"={prev}19", 17: f"='经营驱动'!{col}5*{sel('capex_pct_revenue', i)}",
            18: f"={col}16*{sel('depreciation_rate_on_opening_ppe', i)}", 19: f"={col}16+{col}17-{col}18",
            23: f"={prev}26", 24: f"={sel('debt_draw', i)}", 25: f"={sel('debt_repayment', i)}", 26: f"={col}23+{col}24-{col}25",
            27: f"={col}23*{sel('debt_interest_rate', i)}", 31: f"={prev}34", 32: f"='利润表'!{col}22",
            33: f"=MAX(0,{col}32*{sel('dividend_payout', i)})", 34: f"={col}31+{col}32-{col}33",
            36: f"={prev}38", 37: f"={sel('share_issuance', i)}", 38: f"={col}36+{col}37",
        }
        for row, formula in formulas.items():
            _formula(sch.cell(row, 3 + i), formula, "'" in formula)
    for row in range(6, 39):
        for col in range(2, last_col + 1):
            sch.cell(row, col).number_format = "0.0" if row in (11, 12, 13) else AMOUNT_FMT

    # 利润表
    inc = sheets["利润表"]
    _statement(inc, "利润表预测", data, last_col, 23)
    _labels(inc, [(5, "营业收入", 0), (6, "营业成本", 0), (7, "毛利润", 0), (8, "毛利率", 1), (10, "销售及管理费用", 0), (11, "研发费用", 0), (12, "其他经营费用", 0), (13, "EBITDA", 0), (14, "EBITDA利润率", 1), (15, "折旧与摊销", 0), (16, "营业利润（EBIT）", 0), (17, "利息收入", 0), (18, "利息费用", 0), (19, "其他非经营项目", 0), (20, "税前利润", 0), (21, "所得税费用", 0), (22, "净利润", 0), (23, "净利率", 1)])
    actual_values = {5: hist["revenue"], 6: hist["cogs"], 10: hist["sga"], 11: hist["rnd"], 12: hist["other_opex"], 15: hist["depreciation_amortization"], 17: hist["interest_income"], 18: hist["interest_expense"], 19: hist["other_nonoperating"], 21: hist["income_tax"]}
    for row, value in actual_values.items():
        inc.cell(row, 2, value)
    for cell, formula in {"B7": "=SUM(B5:B6)", "B8": "=B7/B5", "B13": "=SUM(B7,B10:B12)", "B14": "=B13/B5", "B16": "=SUM(B13,B15)", "B20": "=SUM(B16:B19)", "B22": "=SUM(B20:B21)", "B23": "=B22/B5"}.items():
        _formula(inc[cell], formula)
    for i in range(n):
        col, prev = get_column_letter(3 + i), get_column_letter(2 + i)
        formulas = {
            5: f"='经营驱动'!{col}5", 6: f"=-{col}5*(1-{sel('gross_margin', i)})", 7: f"=SUM({col}5:{col}6)",
            8: f"={col}7/{col}5", 10: f"=-{col}5*{sel('sga_pct_revenue', i)}", 11: f"=-{col}5*{sel('rnd_pct_revenue', i)}",
            12: f"=-{col}5*{sel('other_opex_pct_revenue', i)}", 13: f"=SUM({col}7,{col}10:{col}12)", 14: f"={col}13/{col}5",
            15: f"=-'明细预测'!{col}18", 16: f"=SUM({col}13,{col}15)", 17: f"='资产负债表'!{prev}5*{sel('cash_interest_rate', i)}",
            18: f"=-'明细预测'!{col}27", 19: "=0", 20: f"=SUM({col}16:{col}19)",
            21: f"=-MAX(0,{col}20*{sel('tax_rate', i)})", 22: f"=SUM({col}20:{col}21)", 23: f"={col}22/{col}5",
        }
        for row, formula in formulas.items():
            _formula(inc.cell(row, 3 + i), formula, "'" in formula)
    for row in range(5, 24):
        for col in range(2, last_col + 1):
            inc.cell(row, col).number_format = PCT_FMT if row in (8, 14, 23) else AMOUNT_FMT
    _total_border(inc, [7, 13, 16, 20, 22], last_col)

    # 现金流量表
    cfs = sheets["现金流量表"]
    _statement(cfs, "现金流量表预测", data, last_col, 25)
    _labels(cfs, [(5, "净利润", 0), (6, "折旧与摊销", 0), (7, "应收账款变动", 0), (8, "存货变动", 0), (9, "其他流动资产变动", 0), (10, "应付账款变动", 0), (11, "其他流动负债变动", 0), (12, "经营活动现金流", 0), (14, "资本开支", 0), (15, "投资活动现金流", 0), (17, "新增借款", 0), (18, "偿还债务", 0), (19, "股权融资", 0), (20, "支付股利", 0), (21, "融资活动现金流", 0), (23, "现金净变动", 0), (24, "期初现金", 0), (25, "期末现金", 0)])
    for i in range(n):
        col, prev = get_column_letter(3 + i), get_column_letter(2 + i)
        formulas = {
            5: f"='利润表'!{col}22", 6: f"='明细预测'!{col}18", 7: f"='明细预测'!{prev}6-'明细预测'!{col}6",
            8: f"='明细预测'!{prev}7-'明细预测'!{col}7", 9: f"='明细预测'!{prev}8-'明细预测'!{col}8",
            10: f"='明细预测'!{col}9-'明细预测'!{prev}9", 11: f"='明细预测'!{col}10-'明细预测'!{prev}10",
            12: f"=SUM({col}5:{col}11)", 14: f"=-'明细预测'!{col}17", 15: f"=SUM({col}14:{col}14)",
            17: f"='明细预测'!{col}24", 18: f"=-'明细预测'!{col}25", 19: f"='明细预测'!{col}37",
            20: f"=-'明细预测'!{col}33", 21: f"=SUM({col}17:{col}20)", 23: f"=SUM({col}12,{col}15,{col}21)",
            24: f"='资产负债表'!B5" if i == 0 else f"={prev}25", 25: f"=SUM({col}23:{col}24)",
        }
        for row, formula in formulas.items():
            _formula(cfs.cell(row, 3 + i), formula, "'" in formula)
    for row in range(5, 26):
        for col in range(2, last_col + 1):
            cfs.cell(row, col).number_format = AMOUNT_FMT
    _total_border(cfs, [12, 15, 21, 23, 25], last_col)

    # 资产负债表
    bs = sheets["资产负债表"]
    _statement(bs, "资产负债表预测", data, last_col, 28)
    _labels(bs, [(5, "货币资金及现金等价物", 0), (6, "应收账款", 0), (7, "存货", 0), (8, "其他流动资产", 0), (9, "流动资产合计", 0), (11, "固定资产净额", 0), (12, "其他非流动资产", 0), (13, "资产总计", 0), (15, "应付账款", 0), (16, "其他流动负债", 0), (17, "流动负债合计", 0), (19, "有息债务", 0), (20, "其他非流动负债", 0), (21, "负债合计", 0), (23, "股本及资本公积", 0), (24, "留存收益", 0), (25, "所有者权益合计", 0), (27, "负债和所有者权益合计", 0), (28, "资产负债平衡检查", 1)])
    for row, key in {5: "cash", 6: "accounts_receivable", 7: "inventory", 8: "other_current_assets", 11: "net_ppe", 12: "other_noncurrent_assets", 15: "accounts_payable", 16: "other_current_liabilities", 19: "debt", 20: "other_noncurrent_liabilities", 23: "share_capital", 24: "retained_earnings"}.items():
        bs.cell(row, 2, obs[key])
    for cell, formula in {"B9": "=SUM(B5:B8)", "B13": "=SUM(B9,B11:B12)", "B17": "=SUM(B15:B16)", "B21": "=SUM(B17,B19:B20)", "B25": "=SUM(B23:B24)", "B27": "=SUM(B21,B25)", "B28": "=B13-B27"}.items():
        _formula(bs[cell], formula)
    for i in range(n):
        col, prev = get_column_letter(3 + i), get_column_letter(2 + i)
        formulas = {5: f"='现金流量表'!{col}25", 6: f"='明细预测'!{col}6", 7: f"='明细预测'!{col}7", 8: f"='明细预测'!{col}8", 9: f"=SUM({col}5:{col}8)", 11: f"='明细预测'!{col}19", 12: f"={prev}12", 13: f"=SUM({col}9,{col}11:{col}12)", 15: f"='明细预测'!{col}9", 16: f"='明细预测'!{col}10", 17: f"=SUM({col}15:{col}16)", 19: f"='明细预测'!{col}26", 20: f"={prev}20", 21: f"=SUM({col}17,{col}19:{col}20)", 23: f"='明细预测'!{col}38", 24: f"='明细预测'!{col}34", 25: f"=SUM({col}23:{col}24)", 27: f"=SUM({col}21,{col}25)", 28: f"={col}13-{col}27"}
        for row, formula in formulas.items():
            _formula(bs.cell(row, 3 + i), formula, "'" in formula)
    for row in range(5, 29):
        for col in range(2, last_col + 1):
            bs.cell(row, col).number_format = AMOUNT_FMT
    _total_border(bs, [9, 13, 17, 21, 25, 27, 28], last_col)

    # 检查
    checks = sheets["检查"]
    _title(checks, 8, "模型检查")
    checks["A2"], checks["B2"] = "模型状态", None
    headers = ("检查项", "期间", "实际值", "预期值", "差异", "容差", "状态", "修复提示")
    for col, value in enumerate(headers, start=1):
        checks.cell(3, col, value)
    _header(checks, 3, 8)
    row = 4

    def add_check(name: str, period: int, actual: str, expected: str, tolerance: float, hint: str, warning: bool = False) -> None:
        nonlocal row
        checks.cell(row, 1, name)
        checks.cell(row, 2, period)
        _formula(checks.cell(row, 3), actual, True)
        _formula(checks.cell(row, 4), expected, "'" in expected)
        _formula(checks.cell(row, 5), f"=C{row}-D{row}")
        checks.cell(row, 6, tolerance)
        status = f'=IF(C{row}>=D{row},"正常","警告")' if warning else f'=IF(ABS(E{row})<=F{row},"正常","失败")'
        _formula(checks.cell(row, 7), status)
        checks.cell(row, 8, hint)
        row += 1

    for i, year in enumerate(years):
        col = get_column_letter(3 + i)
        add_check("资产负债表平衡", year, f"='资产负债表'!{col}13", f"='资产负债表'!{col}27", 0.01, "追踪资产与负债、权益差异")
        add_check("现金流期末现金与资产负债表一致", year, f"='现金流量表'!{col}25", f"='资产负债表'!{col}5", 0.01, "追踪现金滚动")
        add_check("固定资产滚动", year, f"='明细预测'!{col}19", f"='明细预测'!{col}16+'明细预测'!{col}17-'明细预测'!{col}18", 0.01, "检查资本开支及折旧")
        add_check("债务滚动", year, f"='明细预测'!{col}26", f"='明细预测'!{col}23+'明细预测'!{col}24-'明细预测'!{col}25", 0.01, "检查借款与偿还")
        add_check("留存收益滚动", year, f"='明细预测'!{col}34", f"='明细预测'!{col}31+'明细预测'!{col}32-'明细预测'!{col}33", 0.01, "检查净利润与股利")
        add_check("股本滚动", year, f"='明细预测'!{col}38", f"='明细预测'!{col}36+'明细预测'!{col}37", 0.01, "检查股权融资")
        add_check("最低现金余额", year, f"='资产负债表'!{col}5", f"={sel('minimum_cash', i)}", 0, "增加透明融资计划", True)
    _formula(checks["B2"], f'=IF(COUNTIF(G4:G{row-1},"失败")>0,"FAIL",IF(COUNTIF(G4:G{row-1},"警告")>0,"INCOMPLETE","PASS"))')
    checks["B2"].fill = PatternFill("solid", fgColor=YELLOW)
    checks["B2"].font = Font(color=BLUE, bold=True)
    checks.freeze_panes = "A4"
    for col, width in enumerate((34, 14, 16, 16, 16, 12, 14, 48), start=1):
        checks.column_dimensions[get_column_letter(col)].width = width
    checks.conditional_formatting.add(f"G4:G{row-1}", FormulaRule(formula=['G4="正常"'], fill=PatternFill("solid", fgColor=PASS)))
    checks.conditional_formatting.add(f"G4:G{row-1}", FormulaRule(formula=['G4="失败"'], fill=PatternFill("solid", fgColor=FAIL)))

    # 来源
    src = sheets["来源"]
    _title(src, 9, "数据来源与审计底稿")
    source_headers = ("项目", "数值", "单位", "期间/截止日", "来源类型", "来源名称", "链接", "访问日期", "备注")
    for col, value in enumerate(source_headers, start=1):
        src.cell(3, col, value)
    _header(src, 3, 9)
    for r, item in enumerate(data.get("sources", []), start=4):
        values = (item.get("item", ""), item.get("value", ""), item.get("units", ""), item.get("period", ""), item.get("source_type", ""), item.get("source_name", ""), item.get("url", ""), item.get("accessed", ""), item.get("notes", ""))
        for col, value in enumerate(values, start=1):
            src.cell(r, col, value).alignment = Alignment(wrap_text=True, vertical="top")
    for col, width in enumerate((28, 18, 14, 18, 18, 22, 46, 16, 42), start=1):
        src.column_dimensions[get_column_letter(col)].width = width
    src.freeze_panes = "A4"

    # 封面
    cover = sheets["封面"]
    _title(cover, 8, f"{data['company']}（{data['ticker']}）— 三表预测模型")
    cover_rows = (
        ("公司", data["company"]), ("模型日期", data.get("model_date", date.today().isoformat())),
        ("上市市场/代码", f"{data['market']} / {data['ticker']}"), ("会计准则", data["accounting_standard"]),
        ("财年截止日", data.get("fiscal_year_end", "未提供")), ("币种/单位", f"{data['currency']} / {data['units']}"),
        ("历史截止期", data["actual_year"]), ("预测期间", f"{years[0]}-{years[-1]}"), ("当前情景", None),
        ("模型状态", None), ("模型用途", "非金融企业公式驱动三表预测起始模型"),
        ("重要提示", "公司经营KPI可得时，应以量价、客户、产能、门店或订单驱动替换通用收入增速。"),
    )
    for r, (label, value) in enumerate(cover_rows, start=3):
        cover.cell(r, 1, label).font = Font(bold=True)
        cover.cell(r, 2, value)
    _formula(cover["B11"], "='假设'!B3", True)
    _formula(cover["B12"], "='检查'!B2", True)
    cover.column_dimensions["A"].width = 24
    cover.column_dimensions["B"].width = 74
    cover["B14"].alignment = Alignment(wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    if support_dir is not None:
        emit_support_artifacts(
            output_path,
            support_dir,
            workflow="three_statements_growth",
            generator="scripts/three-statements/build_model.py",
            input_path=input_path,
            required_sheets=list(names),
            lineage=[
                {
                    "field": "scenarios",
                    "sheet": "假设",
                    "cell_or_range": f"A3:{get_column_letter(ass_last_col)}{6 + len(DRIVERS)}",
                    "role": "input_and_selector",
                    "source_ids": [],
                    "scenario": "Base/Bull/Bear",
                    "notes": "情景输入及当前情景公式选择",
                },
                {
                    "field": "revenue_forecast",
                    "sheet": "经营驱动",
                    "cell_or_range": f"B5:{get_column_letter(last_col)}6",
                    "role": "formula_output",
                    "source_ids": [],
                    "scenario": "selected",
                    "notes": "收入及增速滚动",
                },
                {
                    "field": "income_statement",
                    "sheet": "利润表",
                    "cell_or_range": f"A5:{get_column_letter(last_col)}23",
                    "role": "formula_output",
                    "source_ids": [],
                    "scenario": "selected",
                    "notes": "利润表预测",
                },
                {
                    "field": "balance_sheet",
                    "sheet": "资产负债表",
                    "cell_or_range": f"A5:{get_column_letter(last_col)}28",
                    "role": "formula_output",
                    "source_ids": [],
                    "scenario": "selected",
                    "notes": "资产负债表预测及平衡检查",
                },
                {
                    "field": "cash_flow_statement",
                    "sheet": "现金流量表",
                    "cell_or_range": f"A5:{get_column_letter(last_col)}25",
                    "role": "formula_output",
                    "source_ids": [],
                    "scenario": "selected",
                    "notes": "现金流量表预测及期末现金滚动",
                },
                {
                    "field": "model_check_status",
                    "sheet": "检查",
                    "cell_or_range": "B2",
                    "role": "formula_check",
                    "source_ids": [],
                    "scenario": "selected",
                    "notes": "由明细检查状态汇总",
                },
            ],
        )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--support-dir", type=Path)
    args = parser.parse_args()
    data = json.loads(args.input.read_text(encoding="utf-8"))
    build(data, args.output, args.support_dir, args.input)
    print(json.dumps({"output": str(args.output), "support_dir": str(args.support_dir) if args.support_dir else None, "engine": "openpyxl", "sheets": ["封面", "假设", "经营驱动", "明细预测", "利润表", "资产负债表", "现金流量表", "检查", "来源"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
