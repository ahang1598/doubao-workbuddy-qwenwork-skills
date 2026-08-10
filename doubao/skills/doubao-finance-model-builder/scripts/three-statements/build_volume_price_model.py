#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate a product-level volume/price three-statement model with OpenPyXL."""

from __future__ import annotations

import argparse
import json
from copy import copy
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from build_model import emit_support_artifacts


CASES = ("Base", "Bull", "Bear")
CASE_CN = {"Base": "基准", "Bull": "乐观", "Bear": "谨慎"}
GENERAL = (
    ("sga_pct_revenue", "销售及管理费用率", "%"), ("other_operating_items", "其他经营项目", "金额"),
    ("tax_rate", "有效税率", "%"), ("dso", "应收账款周转天数", "天"),
    ("inventory_days", "存货周转天数", "天"), ("dpo", "应付账款周转天数", "天"),
    ("other_current_assets_pct_revenue", "其他流动资产/收入", "%"),
    ("other_current_liabilities_pct_revenue", "其他流动负债/收入", "%"),
    ("capex_amount", "资本开支", "金额"), ("depreciation_amount", "折旧摊销", "金额"),
    ("debt_end", "期末有息债务", "金额"), ("debt_interest_rate", "债务利率", "%"),
    ("cash_interest_rate", "现金收益率", "%"), ("parent_attribution_pct", "归母利润占比", "%"),
    ("parent_dividend_payout", "母公司派息率", "%"), ("nci_dividend_payout", "少数股东派息率", "%"),
    ("share_issuance", "股权融资", "金额"), ("other_nonoperating", "其他非经营项目", "金额"),
    ("fx_effect", "汇率变动对现金影响", "金额"), ("other_investing", "其他投资现金流", "金额"),
    ("other_financing", "其他融资现金流", "金额"), ("revenue_reconciliation", "预测收入口径调节", "金额"),
    ("cogs_reconciliation", "预测成本口径调节", "金额"),
)
PRODUCT_METRICS = (
    ("production_volume", "产量"), ("sales_conversion_rate", "产销率"), ("sales_volume", "直接销量（可留空）"),
    ("realized_price", "实现价格"), ("price_fx_to_model_currency", "价格币种兑模型币种"),
    ("unit_cost", "单位成本"), ("cost_fx_to_model_currency", "成本币种兑模型币种"),
)
QUANTITY_MULTIPLIER = {"吨": 1, "千吨": 1000, "万吨": 10000, "公斤": .001, "件": 1, "千件": 1000, "万件": 10000, "百万件": 1000000}
MONEY_MULTIPLIER = {"元/吨": 1, "人民币/吨": 1, "美元/吨": 1, "港元/吨": 1, "万元/吨": 10000, "万美元/吨": 10000, "万港元/吨": 10000, "元/件": 1, "美元/件": 1, "港元/件": 1, "万元/件": 10000, "万美元/件": 10000}
NAVY, LIGHT_BLUE, FORECAST = "17365D", "D9EAF7", "E2F0D9"
BLUE, GREEN, WHITE, YELLOW = "0000FF", "008000", "FFFFFF", "FFF2CC"
PASS, FAIL = "C6EFCE", "FFC7CE"
THIN = Side(style="thin", color="7F7F7F")
AMOUNT_FMT, PCT_FMT = '#,##0.0;[Red](#,##0.0);-', '0.0%;[Red](0.0%);-'


def finite(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _require_array(value: Any, n: int, label: str, allow_null: bool = False) -> None:
    if not isinstance(value, list) or len(value) != n:
        raise ValueError(f"{label} must contain {n} values")
    if any(not (allow_null and item is None) and not finite(item) for item in value):
        raise ValueError(f"{label} contains a non-numeric value")


def validate(data: dict[str, Any]) -> None:
    if data.get("model_type") != "volume_price":
        raise ValueError("model_type must be volume_price")
    for key in ("company", "ticker", "market", "accounting_standard", "currency", "units"):
        if not data.get(key):
            raise ValueError(f"{key} is required")
    years = data.get("forecast_years")
    if not isinstance(data.get("actual_year"), int) or not isinstance(years, list) or not years:
        raise ValueError("actual_year and forecast_years are required")
    if data.get("selected_scenario") not in CASES:
        raise ValueError("selected_scenario must be Base, Bull or Bear")
    if not finite(data.get("amount_divisor")) or data["amount_divisor"] <= 0:
        raise ValueError("amount_divisor must be positive")
    n = len(years)
    for section in ("historical_income_statement", "opening_balance_sheet", "historical_cash_flow_statement"):
        obj = data.get(section)
        if not isinstance(obj, dict) or any(not finite(v) for v in obj.values()):
            raise ValueError(f"{section} must contain finite numeric values")
    for case in CASES:
        scenario = data.get("scenarios", {}).get(case)
        if not isinstance(scenario, dict):
            raise ValueError(f"missing scenario {case}")
        for key, _, _ in GENERAL:
            _require_array(scenario.get(key), n, f"{case}.{key}")
    products = data.get("products")
    if not isinstance(products, list) or not products:
        raise ValueError("products is required")
    ids: set[str] = set()
    for product in products:
        pid = product.get("id")
        if not isinstance(pid, str) or not pid or pid in ids:
            raise ValueError("product ids must be present and unique")
        ids.add(pid)
        for key in ("name", "quantity_unit", "price_unit", "cost_unit"):
            if not product.get(key):
                raise ValueError(f"{pid}.{key} is required")
        q = product.get("quantity_multiplier", QUANTITY_MULTIPLIER.get(product["quantity_unit"]))
        p = product.get("price_multiplier", MONEY_MULTIPLIER.get(product["price_unit"]))
        c = product.get("cost_multiplier", MONEY_MULTIPLIER.get(product["cost_unit"]))
        if not all(finite(v) and v > 0 for v in (q, p, c)):
            raise ValueError(f"{pid}: unsupported units; provide explicit multipliers")
        product["quantity_multiplier"], product["price_multiplier"], product["cost_multiplier"] = q, p, c
        for key in ("production_volume", "sales_volume", "realized_price", "price_fx_to_model_currency", "unit_cost", "cost_fx_to_model_currency", "reported_revenue", "reported_cogs"):
            if not finite(product.get("actual", {}).get(key)):
                raise ValueError(f"{pid}.actual.{key} is required")
        for case in CASES:
            scenario = product.get("scenarios", {}).get(case)
            if not isinstance(scenario, dict):
                raise ValueError(f"{pid} missing scenario {case}")
            for key, _ in PRODUCT_METRICS:
                _require_array(scenario.get(key), n, f"{pid}.{case}.{key}", key == "sales_volume" and product.get("sales_mode", "production_times_conversion") != "direct_sales")
    bs = data["opening_balance_sheet"]
    bs.setdefault("other_equity", 0)
    assets = sum(bs[k] for k in ("cash", "accounts_receivable", "inventory", "other_current_assets", "net_ppe", "other_noncurrent_assets"))
    le = sum(bs[k] for k in ("accounts_payable", "other_current_liabilities", "debt", "other_noncurrent_liabilities", "share_capital", "retained_earnings", "other_equity", "noncontrolling_interests"))
    if abs(assets - le) > data.get("tolerance", .1):
        raise ValueError("opening balance sheet does not balance")


def title(ws, last_col: int, text: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(1, 1, text)
    for cell in ws[1][:last_col]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    ws.row_dimensions[1].height = 28


def header(ws, row: int, last_col: int) -> None:
    for cell in ws[row][:last_col]:
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(name="Arial", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=THIN)


def formula(cell, value: str, cross: bool = False) -> None:
    cell.value = value
    cell.font = Font(name="Arial", size=10, color=GREEN if cross else "000000")


def periods(ws, data: dict[str, Any], actual_col: int, forecast_col: int, last_col: int) -> None:
    ws.cell(3, 1, "语义键")
    ws.cell(3, 2, "项目")
    ws.cell(3, 3, "单位")
    ws.cell(3, actual_col, date(data["actual_year"], 1, 1)).number_format = 'yyyy"A"'
    for i, year in enumerate(data["forecast_years"]):
        c = ws.cell(3, forecast_col + i, date(year, 1, 1))
        c.number_format = 'yyyy"E"'
        c.fill = PatternFill("solid", fgColor=FORECAST)
    header(ws, 3, last_col)


def semantic_row(ws, mapping: dict[str, int], row: int, key: str, label: str, unit: str) -> int:
    if key in mapping:
        raise ValueError(f"duplicate semantic key {key}")
    mapping[key] = row
    ws.cell(row, 1, key)
    ws.cell(row, 2, label)
    ws.cell(row, 3, unit)
    return row


def style_model_sheet(ws, last_col: int, last_row: int) -> None:
    ws.freeze_panes = "D4"
    ws.sheet_view.showGridLines = False
    for col, width in ((1, 42), (2, 32), (3, 20)):
        ws.column_dimensions[get_column_letter(col)].width = width
    for col in range(4, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for row in ws.iter_rows(min_row=2, max_row=last_row, max_col=last_col):
        for cell in row:
            font = copy(cell.font)
            font.name, font.sz = "Arial", 10
            cell.font = font


def build(
    data: dict[str, Any],
    output: Path,
    support_dir: Path | None = None,
    input_path: Path | None = None,
) -> None:
    validate(data)
    n = len(data["forecast_years"])
    actual_col, forecast_col = 4, 5
    last_col = forecast_col + n - 1
    tolerance = data.get("tolerance", .1)
    max_multiple = data.get("max_revenue_multiple", 10)
    hist, obs, hcf = data["historical_income_statement"], data["opening_balance_sheet"], data["historical_cash_flow_statement"]

    wb = Workbook()
    wb.remove(wb.active)
    names = ("封面", "假设", "产品明细", "利润表", "资产负债表", "现金流量表", "检查", "来源")
    sheets = {name: wb.create_sheet(name) for name in names}
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    for ws in sheets.values():
        ws.sheet_view.showGridLines = False

    # Assumptions: one visible row per case/driver/product.
    ass = sheets["假设"]
    title(ass, last_col, "量价模型假设与情景")
    ass["A3"], ass["B3"] = "当前情景", CASE_CN[data["selected_scenario"]]
    ass["A4"], ass["B4"] = "金额输出除数", data["amount_divisor"]
    ass["B3"].font = ass["B4"].font = Font(color=BLUE, bold=True)
    dv = DataValidation(type="list", formula1='"基准,乐观,谨慎"')
    ass.add_data_validation(dv)
    dv.add(ass["B3"])
    for col, value in enumerate(("语义键", "假设项目", "单位", "情景", *data["forecast_years"]), start=1):
        ass.cell(6, col, value)
    header(ass, 6, last_col)
    arows: dict[str, int] = {}
    row = 7
    for case in CASES:
        for key, label, unit in GENERAL:
            semantic_row(ass, arows, row, f"general.{case}.{key}", label, unit)
            ass.cell(row, 4, CASE_CN[case])
            for i, value in enumerate(data["scenarios"][case][key]):
                ass.cell(row, forecast_col + i, value).font = Font(color=BLUE)
            row += 1
    for product in data["products"]:
        for case in CASES:
            for key, label in PRODUCT_METRICS:
                unit = "%" if key == "sales_conversion_rate" else (product["price_unit"] if key == "realized_price" else product["cost_unit"] if key == "unit_cost" else f"{data['currency']}/币种" if "fx" in key else product["quantity_unit"])
                semantic_row(ass, arows, row, f"product.{product['id']}.{case}.{key}", f"{product['name']}—{label}", unit)
                ass.cell(row, 4, CASE_CN[case])
                for i, value in enumerate(product["scenarios"][case][key]):
                    ass.cell(row, forecast_col + i, value).font = Font(color=BLUE)
                row += 1
    style_model_sheet(ass, last_col, row)
    ass.freeze_panes = "E7"

    def acell(key: str, i: int) -> str:
        return f"'假设'!{get_column_letter(forecast_col+i)}{arows[key]}"

    def selected_general(key: str, i: int) -> str:
        return f'=IF(\'假设\'!$B$3="基准",{acell(f"general.Base.{key}", i)},IF(\'假设\'!$B$3="乐观",{acell(f"general.Bull.{key}", i)},{acell(f"general.Bear.{key}", i)}))'

    def selected_product(pid: str, key: str, i: int) -> str:
        return f'=IF(\'假设\'!$B$3="基准",{acell(f"product.{pid}.Base.{key}", i)},IF(\'假设\'!$B$3="乐观",{acell(f"product.{pid}.Bull.{key}", i)},{acell(f"product.{pid}.Bear.{key}", i)}))'

    # Product detail and explicit unit bridge.
    products = sheets["产品明细"]
    title(products, last_col, "产品级量价与单位成本明细")
    periods(products, data, actual_col, forecast_col, last_col)
    prows: dict[str, int] = {}
    product_groups: dict[str, dict[str, int]] = {}
    row = 5
    for product in data["products"]:
        products.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
        products.cell(row, 1, f"{product['name']}（{product.get('category', '产品')}）")
        for c in products[row][:last_col]:
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.font = Font(color=WHITE, bold=True)
        row += 1
        rows: dict[str, int] = {}
        definitions = (
            ("production_volume", "产量", product["quantity_unit"]), ("sales_conversion_rate", "产销率", "%"),
            ("sales_volume", "销量", product["quantity_unit"]), ("quantity_multiplier", "数量换算系数", "基础数量/输入单位"),
            ("realized_price", "实现价格", product["price_unit"]), ("price_multiplier", "价格单位换算系数", "基础金额/报价单位"),
            ("price_fx_to_model_currency", "价格币种兑模型币种", f"{data['currency']}/价格币种"),
            ("calculated_revenue", "量价计算收入", data["units"]), ("reported_revenue", "已披露收入", data["units"]),
            ("revenue_difference", "历史收入差异", data["units"]), ("unit_cost", "单位成本", product["cost_unit"]),
            ("cost_multiplier", "成本单位换算系数", "基础金额/成本单位"),
            ("cost_fx_to_model_currency", "成本币种兑模型币种", f"{data['currency']}/成本币种"),
            ("calculated_cogs", "量本计算成本", data["units"]), ("reported_cogs", "已披露成本", data["units"]),
            ("cogs_difference", "历史成本差异", data["units"]), ("gross_profit", "毛利", data["units"]), ("gross_margin", "毛利率", "%"),
        )
        for key, label, unit in definitions:
            rows[key] = semantic_row(products, prows, row, f"product.{product['id']}.{key}", f"{product['name']}—{label}", unit)
            row += 1
        product_groups[product["id"]] = rows
        actual = product["actual"]
        values = {
            "production_volume": actual["production_volume"], "sales_conversion_rate": actual["sales_volume"] / actual["production_volume"] if actual["production_volume"] else 0,
            "sales_volume": actual["sales_volume"], "quantity_multiplier": product["quantity_multiplier"], "realized_price": actual["realized_price"],
            "price_multiplier": product["price_multiplier"], "price_fx_to_model_currency": actual["price_fx_to_model_currency"],
            "reported_revenue": actual["reported_revenue"], "unit_cost": actual["unit_cost"], "cost_multiplier": product["cost_multiplier"],
            "cost_fx_to_model_currency": actual["cost_fx_to_model_currency"], "reported_cogs": actual["reported_cogs"],
        }
        for key, value in values.items():
            products.cell(rows[key], actual_col, value)
        c = get_column_letter(actual_col)
        formula(products.cell(rows["calculated_revenue"], actual_col), f"={c}{rows['sales_volume']}*{c}{rows['quantity_multiplier']}*{c}{rows['realized_price']}*{c}{rows['price_multiplier']}*{c}{rows['price_fx_to_model_currency']}/'假设'!$B$4", True)
        formula(products.cell(rows["revenue_difference"], actual_col), f"={c}{rows['calculated_revenue']}-{c}{rows['reported_revenue']}")
        formula(products.cell(rows["calculated_cogs"], actual_col), f"={c}{rows['sales_volume']}*{c}{rows['quantity_multiplier']}*{c}{rows['unit_cost']}*{c}{rows['cost_multiplier']}*{c}{rows['cost_fx_to_model_currency']}/'假设'!$B$4", True)
        formula(products.cell(rows["cogs_difference"], actual_col), f"={c}{rows['calculated_cogs']}-{c}{rows['reported_cogs']}")
        formula(products.cell(rows["gross_profit"], actual_col), f"={c}{rows['calculated_revenue']}-{c}{rows['calculated_cogs']}")
        formula(products.cell(rows["gross_margin"], actual_col), f"=IFERROR({c}{rows['gross_profit']}/{c}{rows['calculated_revenue']},0)")
        for i in range(n):
            c = get_column_letter(forecast_col + i)
            formula(products.cell(rows["production_volume"], forecast_col + i), selected_product(product["id"], "production_volume", i), True)
            formula(products.cell(rows["sales_conversion_rate"], forecast_col + i), selected_product(product["id"], "sales_conversion_rate", i), True)
            if product.get("sales_mode", "production_times_conversion") == "direct_sales":
                formula(products.cell(rows["sales_volume"], forecast_col + i), selected_product(product["id"], "sales_volume", i), True)
            else:
                formula(products.cell(rows["sales_volume"], forecast_col + i), f"={c}{rows['production_volume']}*{c}{rows['sales_conversion_rate']}")
            products.cell(rows["quantity_multiplier"], forecast_col + i, product["quantity_multiplier"])
            formula(products.cell(rows["realized_price"], forecast_col + i), selected_product(product["id"], "realized_price", i), True)
            products.cell(rows["price_multiplier"], forecast_col + i, product["price_multiplier"])
            formula(products.cell(rows["price_fx_to_model_currency"], forecast_col + i), selected_product(product["id"], "price_fx_to_model_currency", i), True)
            formula(products.cell(rows["calculated_revenue"], forecast_col + i), f"={c}{rows['sales_volume']}*{c}{rows['quantity_multiplier']}*{c}{rows['realized_price']}*{c}{rows['price_multiplier']}*{c}{rows['price_fx_to_model_currency']}/'假设'!$B$4", True)
            formula(products.cell(rows["unit_cost"], forecast_col + i), selected_product(product["id"], "unit_cost", i), True)
            products.cell(rows["cost_multiplier"], forecast_col + i, product["cost_multiplier"])
            formula(products.cell(rows["cost_fx_to_model_currency"], forecast_col + i), selected_product(product["id"], "cost_fx_to_model_currency", i), True)
            formula(products.cell(rows["calculated_cogs"], forecast_col + i), f"={c}{rows['sales_volume']}*{c}{rows['quantity_multiplier']}*{c}{rows['unit_cost']}*{c}{rows['cost_multiplier']}*{c}{rows['cost_fx_to_model_currency']}/'假设'!$B$4", True)
            formula(products.cell(rows["gross_profit"], forecast_col + i), f"={c}{rows['calculated_revenue']}-{c}{rows['calculated_cogs']}")
            formula(products.cell(rows["gross_margin"], forecast_col + i), f"=IFERROR({c}{rows['gross_profit']}/{c}{rows['calculated_revenue']},0)")
        row += 1
    products.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    products.cell(row, 1, "合计与口径调节")
    for c in products[row][:last_col]:
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(color=WHITE, bold=True)
    row += 1
    totals: dict[str, int] = {}
    for key, label, unit in (("calculated_revenue", "产品量价收入合计", data["units"]), ("revenue_reconciliation", "收入口径调节", data["units"]), ("statement_revenue", "利润表收入", data["units"]), ("calculated_cogs", "产品量本成本合计", data["units"]), ("cogs_reconciliation", "成本口径调节", data["units"]), ("statement_cogs", "利润表销售成本", data["units"]), ("gross_profit", "综合毛利", data["units"]), ("gross_margin", "综合毛利率", "%")):
        totals[key] = semantic_row(products, prows, row, f"total.{key}", label, unit)
        row += 1
    revenue_rows = [g["calculated_revenue"] for g in product_groups.values()]
    cogs_rows = [g["calculated_cogs"] for g in product_groups.values()]
    for i in range(-1, n):
        col_num = actual_col if i < 0 else forecast_col + i
        c = get_column_letter(col_num)
        formula(products.cell(totals["calculated_revenue"], col_num), f"=SUM({','.join(c+str(r) for r in revenue_rows)})")
        formula(products.cell(totals["calculated_cogs"], col_num), f"=SUM({','.join(c+str(r) for r in cogs_rows)})")
        if i >= 0:
            formula(products.cell(totals["revenue_reconciliation"], col_num), selected_general("revenue_reconciliation", i), True)
            formula(products.cell(totals["cogs_reconciliation"], col_num), selected_general("cogs_reconciliation", i), True)
    products.cell(totals["revenue_reconciliation"], actual_col, hist["revenue"] - sum(p["actual"]["reported_revenue"] for p in data["products"]))
    products.cell(totals["cogs_reconciliation"], actual_col, -hist["cogs"] - sum(p["actual"]["reported_cogs"] for p in data["products"]))
    for i in range(-1, n):
        col_num = actual_col if i < 0 else forecast_col + i
        c = get_column_letter(col_num)
        formula(products.cell(totals["statement_revenue"], col_num), f"={c}{totals['calculated_revenue']}+{c}{totals['revenue_reconciliation']}")
        formula(products.cell(totals["statement_cogs"], col_num), f"={c}{totals['calculated_cogs']}+{c}{totals['cogs_reconciliation']}")
        formula(products.cell(totals["gross_profit"], col_num), f"={c}{totals['statement_revenue']}-{c}{totals['statement_cogs']}")
        formula(products.cell(totals["gross_margin"], col_num), f"=IFERROR({c}{totals['gross_profit']}/{c}{totals['statement_revenue']},0)")
    for r in range(5, row):
        for c in range(actual_col, last_col + 1):
            products.cell(r, c).number_format = PCT_FMT if str(products.cell(r, 1).value).endswith(("gross_margin", "sales_conversion_rate")) else AMOUNT_FMT
    style_model_sheet(products, last_col, row)

    def setup_statement(name: str, title_text: str):
        ws = sheets[name]
        title(ws, last_col, title_text)
        periods(ws, data, actual_col, forecast_col, last_col)
        return ws

    # Income statement.
    inc = setup_statement("利润表", "利润表预测")
    irows: dict[str, int] = {}
    row = 5
    for key, label, unit in (("revenue", "营业收入", data["units"]), ("cogs", "销售成本", data["units"]), ("gross_profit", "毛利", data["units"]), ("gross_margin", "毛利率", "%"), ("sga", "销售及管理费用", data["units"]), ("other_operating_items", "其他经营项目", data["units"]), ("ebitda", "EBITDA", data["units"]), ("depreciation_amortization", "折旧摊销", data["units"]), ("ebit", "营业利润（EBIT）", data["units"]), ("interest_income", "利息收入", data["units"]), ("interest_expense", "利息费用", data["units"]), ("other_nonoperating", "其他非经营项目", data["units"]), ("pretax_income", "税前利润", data["units"]), ("income_tax", "所得税费用", data["units"]), ("net_income", "净利润", data["units"]), ("net_income_parent", "归母净利润", data["units"]), ("net_income_nci", "少数股东损益", data["units"]), ("attribution_check", "净利润归属检查", data["units"]), ("parent_dividend", "母公司股利", data["units"]), ("nci_dividend", "少数股东股利", data["units"])):
        irows[f"is.{key}"] = semantic_row(inc, irows, row, f"is.{key}", label, unit)
        row += 1
    for key in ("revenue", "cogs", "sga", "other_operating_items", "depreciation_amortization", "interest_income", "interest_expense", "other_nonoperating", "income_tax", "net_income_parent", "net_income_nci"):
        inc.cell(irows[f"is.{key}"], actual_col, hist[key])
    c = get_column_letter(actual_col)
    for key, value in {
        "gross_profit": f"=SUM({c}{irows['is.revenue']}:{c}{irows['is.cogs']})", "gross_margin": f"={c}{irows['is.gross_profit']}/{c}{irows['is.revenue']}",
        "ebitda": f"=SUM({c}{irows['is.gross_profit']},{c}{irows['is.sga']}:{c}{irows['is.other_operating_items']})",
        "ebit": f"={c}{irows['is.ebitda']}+{c}{irows['is.depreciation_amortization']}",
        "pretax_income": f"=SUM({c}{irows['is.ebit']},{c}{irows['is.interest_income']}:{c}{irows['is.other_nonoperating']})",
        "net_income": f"={c}{irows['is.pretax_income']}+{c}{irows['is.income_tax']}",
        "attribution_check": f"={c}{irows['is.net_income']}-{c}{irows['is.net_income_parent']}-{c}{irows['is.net_income_nci']}",
    }.items():
        formula(inc.cell(irows[f"is.{key}"], actual_col), value)
    inc.cell(irows["is.parent_dividend"], actual_col, hcf["parent_dividends"])
    inc.cell(irows["is.nci_dividend"], actual_col, hcf["nci_dividends"])
    for i in range(n):
        col, prev = get_column_letter(forecast_col + i), get_column_letter(actual_col + i)
        fs = {
            "revenue": f"='产品明细'!{col}{totals['statement_revenue']}", "cogs": f"=-'产品明细'!{col}{totals['statement_cogs']}",
            "gross_profit": f"=SUM({col}{irows['is.revenue']}:{col}{irows['is.cogs']})", "gross_margin": f"={col}{irows['is.gross_profit']}/{col}{irows['is.revenue']}",
            "sga": f"=-{col}{irows['is.revenue']}*({selected_general('sga_pct_revenue', i)[1:]})",
            "other_operating_items": selected_general("other_operating_items", i), "ebitda": f"=SUM({col}{irows['is.gross_profit']},{col}{irows['is.sga']}:{col}{irows['is.other_operating_items']})",
            "depreciation_amortization": f"=-({selected_general('depreciation_amount', i)[1:]})", "ebit": f"={col}{irows['is.ebitda']}+{col}{irows['is.depreciation_amortization']}",
            "interest_income": f"='资产负债表'!{prev}5*({selected_general('cash_interest_rate', i)[1:]})",
            "interest_expense": f"=-'资产负债表'!{prev}14*({selected_general('debt_interest_rate', i)[1:]})",
            "other_nonoperating": selected_general("other_nonoperating", i), "pretax_income": f"=SUM({col}{irows['is.ebit']},{col}{irows['is.interest_income']}:{col}{irows['is.other_nonoperating']})",
            "income_tax": f"=-MAX(0,{col}{irows['is.pretax_income']}*({selected_general('tax_rate', i)[1:]}))", "net_income": f"={col}{irows['is.pretax_income']}+{col}{irows['is.income_tax']}",
            "net_income_parent": f"={col}{irows['is.net_income']}*({selected_general('parent_attribution_pct', i)[1:]})",
            "net_income_nci": f"={col}{irows['is.net_income']}-{col}{irows['is.net_income_parent']}",
            "attribution_check": f"={col}{irows['is.net_income']}-{col}{irows['is.net_income_parent']}-{col}{irows['is.net_income_nci']}",
            "parent_dividend": f"=MAX(0,{col}{irows['is.net_income_parent']}*({selected_general('parent_dividend_payout', i)[1:]}))",
            "nci_dividend": f"=MAX(0,{col}{irows['is.net_income_nci']}*({selected_general('nci_dividend_payout', i)[1:]}))",
        }
        for key, value in fs.items():
            formula(inc.cell(irows[f"is.{key}"], forecast_col + i), value, "'" in value)
    style_model_sheet(inc, last_col, row)

    # Balance sheet.
    bs = setup_statement("资产负债表", "资产负债表预测")
    brows: dict[str, int] = {}
    row = 5
    for key, label in (("cash", "货币资金"), ("accounts_receivable", "应收账款"), ("inventory", "存货"), ("other_current_assets", "其他流动资产"), ("current_assets", "流动资产合计"), ("net_ppe", "固定资产净额"), ("other_noncurrent_assets", "其他非流动资产"), ("total_assets", "资产总计"), ("accounts_payable", "应付账款"), ("other_current_liabilities", "其他流动负债"), ("current_liabilities", "流动负债合计"), ("debt", "有息债务"), ("other_noncurrent_liabilities", "其他非流动负债"), ("total_liabilities", "负债合计"), ("share_capital", "股本及资本公积"), ("retained_earnings", "留存收益"), ("other_equity", "其他权益"), ("noncontrolling_interests", "少数股东权益"), ("total_equity", "权益合计"), ("liabilities_equity", "负债及权益总计"), ("balance_check", "平衡检查")):
        brows[f"bs.{key}"] = semantic_row(bs, brows, row, f"bs.{key}", label, data["units"])
        row += 1
    for key in ("cash", "accounts_receivable", "inventory", "other_current_assets", "net_ppe", "other_noncurrent_assets", "accounts_payable", "other_current_liabilities", "debt", "other_noncurrent_liabilities", "share_capital", "retained_earnings", "other_equity", "noncontrolling_interests"):
        bs.cell(brows[f"bs.{key}"], actual_col, obs[key])
    c = get_column_letter(actual_col)
    for key, value in {"current_assets": f"=SUM({c}{brows['bs.cash']}:{c}{brows['bs.other_current_assets']})", "total_assets": f"=SUM({c}{brows['bs.current_assets']},{c}{brows['bs.net_ppe']}:{c}{brows['bs.other_noncurrent_assets']})", "current_liabilities": f"=SUM({c}{brows['bs.accounts_payable']}:{c}{brows['bs.other_current_liabilities']})", "total_liabilities": f"=SUM({c}{brows['bs.current_liabilities']},{c}{brows['bs.debt']}:{c}{brows['bs.other_noncurrent_liabilities']})", "total_equity": f"=SUM({c}{brows['bs.share_capital']}:{c}{brows['bs.noncontrolling_interests']})", "liabilities_equity": f"={c}{brows['bs.total_liabilities']}+{c}{brows['bs.total_equity']}", "balance_check": f"={c}{brows['bs.total_assets']}-{c}{brows['bs.liabilities_equity']}"}.items():
        formula(bs.cell(brows[f"bs.{key}"], actual_col), value)

    # Cash flow rows are created before linked formulas.
    cfs = setup_statement("现金流量表", "现金流量表预测")
    crows: dict[str, int] = {}
    row_cfs = 5
    for key, label in (("net_income", "净利润"), ("depreciation_amortization", "折旧摊销"), ("change_ar", "应收账款减少/(增加)"), ("change_inventory", "存货减少/(增加)"), ("change_other_current_assets", "其他流动资产减少/(增加)"), ("change_ap", "应付账款增加/(减少)"), ("change_other_current_liabilities", "其他流动负债增加/(减少)"), ("operating_cash_flow", "经营活动现金流"), ("capex", "资本开支"), ("other_investing", "其他投资现金流"), ("investing_cash_flow", "投资活动现金流"), ("debt_change", "债务净变化"), ("share_issuance", "股权融资"), ("dividends", "支付股利"), ("other_financing", "其他融资现金流"), ("financing_cash_flow", "融资活动现金流"), ("net_change_cash", "现金净变动"), ("beginning_cash", "期初现金"), ("fx_effect", "汇率影响"), ("ending_cash", "期末现金")):
        crows[f"cfs.{key}"] = semantic_row(cfs, crows, row_cfs, f"cfs.{key}", label, data["units"])
        row_cfs += 1
    historical_cfs = {"net_income": hist["net_income_parent"] + hist["net_income_nci"], "depreciation_amortization": -hist["depreciation_amortization"], "operating_cash_flow": hcf["operating_cash_flow"], "capex": hcf["capex"], "other_investing": hcf["other_investing"], "investing_cash_flow": hcf["investing_cash_flow"], "debt_change": hcf["debt_change"], "share_issuance": hcf["share_issuance"], "dividends": -(hcf["parent_dividends"] + hcf["nci_dividends"]), "other_financing": hcf["other_financing"], "financing_cash_flow": hcf["financing_cash_flow"], "net_change_cash": hcf["net_change_cash"], "beginning_cash": hcf["beginning_cash"], "fx_effect": hcf["fx_effect"], "ending_cash": hcf["ending_cash"]}
    for key, value in historical_cfs.items():
        cfs.cell(crows[f"cfs.{key}"], actual_col, value)

    # Forecast balance sheet and CFS chain.
    for i in range(n):
        col, prev = get_column_letter(forecast_col + i), get_column_letter(actual_col + i)
        mode = data.get("policies", {}).get("working_capital_mode", "days")
        formula(
            inc.cell(irows["is.interest_income"], forecast_col + i),
            f"='资产负债表'!{prev}{brows['bs.cash']}*({selected_general('cash_interest_rate', i)[1:]})",
            True,
        )
        formula(
            inc.cell(irows["is.interest_expense"], forecast_col + i),
            f"=-'资产负债表'!{prev}{brows['bs.debt']}*({selected_general('debt_interest_rate', i)[1:]})",
            True,
        )
        bfs = {"cash": f"='现金流量表'!{col}{crows['cfs.ending_cash']}"}
        if mode == "balance":
            for key in ("accounts_receivable", "inventory", "other_current_assets", "accounts_payable", "other_current_liabilities"):
                bfs[key] = f"={prev}{brows[f'bs.{key}']}"
        else:
            bfs.update({
                "accounts_receivable": f"='利润表'!{col}{irows['is.revenue']}*({selected_general('dso', i)[1:]})/365",
                "inventory": f"=-'利润表'!{col}{irows['is.cogs']}*({selected_general('inventory_days', i)[1:]})/365",
                "other_current_assets": f"='利润表'!{col}{irows['is.revenue']}*({selected_general('other_current_assets_pct_revenue', i)[1:]})",
                "accounts_payable": f"=-'利润表'!{col}{irows['is.cogs']}*({selected_general('dpo', i)[1:]})/365",
                "other_current_liabilities": f"='利润表'!{col}{irows['is.revenue']}*({selected_general('other_current_liabilities_pct_revenue', i)[1:]})",
            })
        bfs.update({
            "current_assets": f"=SUM({col}{brows['bs.cash']}:{col}{brows['bs.other_current_assets']})",
            "net_ppe": f"={prev}{brows['bs.net_ppe']}+({selected_general('capex_amount', i)[1:]})-({selected_general('depreciation_amount', i)[1:]})",
            "other_noncurrent_assets": f"={prev}{brows['bs.other_noncurrent_assets']}-({selected_general('other_investing', i)[1:]})",
            "total_assets": f"=SUM({col}{brows['bs.current_assets']},{col}{brows['bs.net_ppe']}:{col}{brows['bs.other_noncurrent_assets']})",
            "current_liabilities": f"=SUM({col}{brows['bs.accounts_payable']}:{col}{brows['bs.other_current_liabilities']})",
            "debt": selected_general("debt_end", i), "other_noncurrent_liabilities": f"={prev}{brows['bs.other_noncurrent_liabilities']}",
            "total_liabilities": f"=SUM({col}{brows['bs.current_liabilities']},{col}{brows['bs.debt']}:{col}{brows['bs.other_noncurrent_liabilities']})",
            "share_capital": f"={prev}{brows['bs.share_capital']}+({selected_general('share_issuance', i)[1:]})",
            "retained_earnings": f"={prev}{brows['bs.retained_earnings']}+'利润表'!{col}{irows['is.net_income_parent']}-'利润表'!{col}{irows['is.parent_dividend']}",
            "other_equity": f"={prev}{brows['bs.other_equity']}+({selected_general('fx_effect', i)[1:]})",
            "noncontrolling_interests": f"={prev}{brows['bs.noncontrolling_interests']}+'利润表'!{col}{irows['is.net_income_nci']}-'利润表'!{col}{irows['is.nci_dividend']}",
            "total_equity": f"=SUM({col}{brows['bs.share_capital']}:{col}{brows['bs.noncontrolling_interests']})",
            "liabilities_equity": f"={col}{brows['bs.total_liabilities']}+{col}{brows['bs.total_equity']}",
            "balance_check": f"={col}{brows['bs.total_assets']}-{col}{brows['bs.liabilities_equity']}",
        })
        for key, value in bfs.items():
            formula(bs.cell(brows[f"bs.{key}"], forecast_col + i), value, "'" in value)
        cfs_formulas = {
            "net_income": f"='利润表'!{col}{irows['is.net_income']}", "depreciation_amortization": f"=-'利润表'!{col}{irows['is.depreciation_amortization']}",
            "change_ar": f"='资产负债表'!{prev}{brows['bs.accounts_receivable']}-'资产负债表'!{col}{brows['bs.accounts_receivable']}",
            "change_inventory": f"='资产负债表'!{prev}{brows['bs.inventory']}-'资产负债表'!{col}{brows['bs.inventory']}",
            "change_other_current_assets": f"='资产负债表'!{prev}{brows['bs.other_current_assets']}-'资产负债表'!{col}{brows['bs.other_current_assets']}",
            "change_ap": f"='资产负债表'!{col}{brows['bs.accounts_payable']}-'资产负债表'!{prev}{brows['bs.accounts_payable']}",
            "change_other_current_liabilities": f"='资产负债表'!{col}{brows['bs.other_current_liabilities']}-'资产负债表'!{prev}{brows['bs.other_current_liabilities']}",
            "operating_cash_flow": f"=SUM({col}{crows['cfs.net_income']}:{col}{crows['cfs.change_other_current_liabilities']})",
            "capex": f"=-({selected_general('capex_amount', i)[1:]})", "other_investing": selected_general("other_investing", i),
            "investing_cash_flow": f"=SUM({col}{crows['cfs.capex']}:{col}{crows['cfs.other_investing']})",
            "debt_change": f"='资产负债表'!{col}{brows['bs.debt']}-'资产负债表'!{prev}{brows['bs.debt']}",
            "share_issuance": selected_general("share_issuance", i),
            "dividends": f"=-SUM('利润表'!{col}{irows['is.parent_dividend']},'利润表'!{col}{irows['is.nci_dividend']})",
            "other_financing": selected_general("other_financing", i),
            "financing_cash_flow": f"=SUM({col}{crows['cfs.debt_change']}:{col}{crows['cfs.other_financing']})",
            "net_change_cash": f"=SUM({col}{crows['cfs.operating_cash_flow']},{col}{crows['cfs.investing_cash_flow']},{col}{crows['cfs.financing_cash_flow']})",
            "beginning_cash": f"='资产负债表'!{prev}{brows['bs.cash']}", "fx_effect": selected_general("fx_effect", i),
            "ending_cash": f"=SUM({col}{crows['cfs.net_change_cash']},{col}{crows['cfs.beginning_cash']},{col}{crows['cfs.fx_effect']})",
        }
        for key, value in cfs_formulas.items():
            formula(cfs.cell(crows[f"cfs.{key}"], forecast_col + i), value, "'" in value)
    style_model_sheet(bs, last_col, row)
    style_model_sheet(cfs, last_col, row_cfs)
    for ws, last_row in ((inc, max(irows.values())), (bs, max(brows.values())), (cfs, max(crows.values()))):
        for r in range(5, last_row + 1):
            for c in range(actual_col, last_col + 1):
                ws.cell(r, c).number_format = PCT_FMT if ws.cell(r, 3).value == "%" else AMOUNT_FMT

    # Checks.
    checks = sheets["检查"]
    title(checks, 8, "模型检查与硬阻断")
    checks["A2"], checks["B2"] = "模型状态", None
    for col, value in enumerate(("检查项", "期间", "实际值", "预期值", "差异", "容差", "状态", "修复提示"), start=1):
        checks.cell(3, col, value)
    header(checks, 3, 8)
    check_row = 4

    def add_check(name: str, period: int, actual: str, expected: str, hint: str, status: str | None = None) -> None:
        nonlocal check_row
        checks.cell(check_row, 1, name)
        checks.cell(check_row, 2, period)
        formula(checks.cell(check_row, 3), actual, True)
        formula(checks.cell(check_row, 4), expected, True)
        formula(checks.cell(check_row, 5), f"=C{check_row}-D{check_row}")
        checks.cell(check_row, 6, tolerance)
        formula(checks.cell(check_row, 7), status or f'=IF(ABS(E{check_row})<=F{check_row},"通过","失败")')
        checks.cell(check_row, 8, hint)
        check_row += 1

    add_check("历史资产负债表平衡", data["actual_year"], f"='资产负债表'!D{brows['bs.total_assets']}", f"='资产负债表'!D{brows['bs.liabilities_equity']}", "返回历史科目映射")
    add_check("历史净利润归属", data["actual_year"], f"='利润表'!D{irows['is.net_income']}", f"=SUM('利润表'!D{irows['is.net_income_parent']},'利润表'!D{irows['is.net_income_nci']})", "检查归属口径")
    add_check("历史现金勾稽", data["actual_year"], f"='现金流量表'!D{crows['cfs.ending_cash']}", f"='资产负债表'!D{brows['bs.cash']}", "核对现金")
    add_check("历史产品收入勾稽", data["actual_year"], f"='产品明细'!D{totals['statement_revenue']}", f"='利润表'!D{irows['is.revenue']}", "检查收入口径调节")
    add_check("历史产品成本勾稽", data["actual_year"], f"='产品明细'!D{totals['statement_cogs']}", f"=-'利润表'!D{irows['is.cogs']}", "检查成本口径调节")
    for i, year in enumerate(data["forecast_years"]):
        col, prev = get_column_letter(forecast_col + i), get_column_letter(actual_col + i)
        add_check("资产负债表平衡", year, f"='资产负债表'!{col}{brows['bs.total_assets']}", f"='资产负债表'!{col}{brows['bs.liabilities_equity']}", "追踪现金、营运资本和权益")
        add_check("现金勾稽", year, f"='现金流量表'!{col}{crows['cfs.ending_cash']}", f"='资产负债表'!{col}{brows['bs.cash']}", "追踪期末现金")
        add_check("固定资产滚动", year, f"='资产负债表'!{col}{brows['bs.net_ppe']}", f"='资产负债表'!{prev}{brows['bs.net_ppe']}+({selected_general('capex_amount', i)[1:]})-({selected_general('depreciation_amount', i)[1:]})", "检查CapEx与折旧")
        add_check("净利润归属", year, f"='利润表'!{col}{irows['is.net_income']}", f"=SUM('利润表'!{col}{irows['is.net_income_parent']},'利润表'!{col}{irows['is.net_income_nci']})", "复核归母和少数股东")
        add_check("收入产品勾稽", year, f"='利润表'!{col}{irows['is.revenue']}", f"='产品明细'!{col}{totals['statement_revenue']}", "追踪产品量价")
        add_check("成本产品勾稽", year, f"=-'利润表'!{col}{irows['is.cogs']}", f"='产品明细'!{col}{totals['statement_cogs']}", "追踪单位成本")
        r = check_row
        add_check("分红现金符号", year, f"='现金流量表'!{col}{crows['cfs.dividends']}", f"=-SUM('利润表'!{col}{irows['is.parent_dividend']},'利润表'!{col}{irows['is.nci_dividend']})", "支付股利必须为负", f'=IF(AND(ABS(E{r})<=F{r},C{r}<=0),"通过","失败")')
        r = check_row
        add_check("收入数量级", year, f"='利润表'!{col}{irows['is.revenue']}/'利润表'!D{irows['is.revenue']}", f"={max_multiple}", "检查单位换算", f'=IF(AND(C{r}>0,C{r}<=D{r}),"通过","失败")')
    formula(checks["B2"], f'=IF(COUNTIF(G4:G{check_row-1},"失败")>0,"FAIL",IF(COUNTIF(G4:G{check_row-1},"警告")>0,"INCOMPLETE","PASS"))')
    checks["B2"].fill = PatternFill("solid", fgColor=YELLOW)
    checks["B2"].font = Font(color=BLUE, bold=True)
    checks.freeze_panes = "A4"
    for col, width in enumerate((34, 14, 16, 16, 16, 12, 14, 48), start=1):
        checks.column_dimensions[get_column_letter(col)].width = width
    checks.conditional_formatting.add(f"G4:G{check_row-1}", FormulaRule(formula=['G4="通过"'], fill=PatternFill("solid", fgColor=PASS)))
    checks.conditional_formatting.add(f"G4:G{check_row-1}", FormulaRule(formula=['G4="失败"'], fill=PatternFill("solid", fgColor=FAIL)))

    # Sources and cover.
    src = sheets["来源"]
    title(src, 11, "原始数据与来源记录")
    for col, value in enumerate(("来源ID", "项目", "数值", "单位", "期间/日期", "发布日期", "来源类型", "来源名称", "URL", "页码/位置", "备注"), start=1):
        src.cell(3, col, value)
    header(src, 3, 11)
    for r, item in enumerate(data.get("sources", []), start=4):
        values = (item.get("source_id", ""), item.get("item", ""), item.get("value", ""), item.get("units", ""), item.get("period", ""), item.get("published", ""), item.get("source_type", ""), item.get("source_name", ""), item.get("url", ""), item.get("location", ""), item.get("notes", ""))
        for col, value in enumerate(values, start=1):
            src.cell(r, col, value).alignment = Alignment(wrap_text=True, vertical="top")
    src.freeze_panes = "A4"
    for col in range(1, 12):
        src.column_dimensions[get_column_letter(col)].width = 48 if col == 9 else (28 if col in (10, 11) else 20)

    cover = sheets["封面"]
    title(cover, 2, f"{data['company']}（{data['ticker']}）— 三表联动量价模型")
    cover_data = (("公司", data["company"]), ("代码/市场", f"{data['ticker']} / {data['market']}"), ("会计准则", data["accounting_standard"]), ("币种/单位", f"{data['currency']} / {data['units']}"), ("模型类型", "volume_price"), ("基准年", f"{data['actual_year']}A"), ("预测区间", f"{data['forecast_years'][0]}E–{data['forecast_years'][-1]}E"), ("当前情景", None), ("金额输出除数", None), ("营运资本模式", data.get("policies", {}).get("working_capital_mode", "days")), ("模型状态", None), ("状态说明", "检查页任一失败即FAIL；不得手填PASS"))
    for r, (label, value) in enumerate(cover_data, start=3):
        cover.cell(r, 1, label).font = Font(bold=True)
        cover.cell(r, 2, value)
    formula(cover["B10"], "='假设'!B3", True)
    formula(cover["B11"], "='假设'!B4", True)
    formula(cover["B13"], "='检查'!B2", True)
    cover.column_dimensions["A"].width = 24
    cover.column_dimensions["B"].width = 72

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    if support_dir is not None:
        emit_support_artifacts(
            output,
            support_dir,
            workflow="three_statements_volume_price",
            generator="scripts/three-statements/build_volume_price_model.py",
            input_path=input_path,
            required_sheets=list(names),
            lineage=[
                {
                    "field": "general_scenarios",
                    "sheet": "假设",
                    "cell_or_range": f"A3:{get_column_letter(last_col)}{max(arows.values())}",
                    "role": "input_and_selector",
                    "source_ids": [],
                    "scenario": "Base/Bull/Bear",
                    "notes": "通用及产品级三情景输入",
                },
                {
                    "field": "product_volume_price_cost",
                    "sheet": "产品明细",
                    "cell_or_range": f"A5:{get_column_letter(last_col)}{max(prows.values())}",
                    "role": "formula_output",
                    "source_ids": [
                        sid
                        for product in data["products"]
                        for sid in product.get("source_ids", [])
                    ],
                    "scenario": "selected",
                    "notes": "数量、单位换算、价格、成本、汇率至收入和成本",
                },
                {
                    "field": "income_statement",
                    "sheet": "利润表",
                    "cell_or_range": f"A5:{get_column_letter(last_col)}{max(irows.values())}",
                    "role": "formula_output",
                    "source_ids": data.get("source_map", {}).get("historical_financials", []),
                    "scenario": "selected",
                    "notes": "利润表预测",
                },
                {
                    "field": "balance_sheet",
                    "sheet": "资产负债表",
                    "cell_or_range": f"A5:{get_column_letter(last_col)}{max(brows.values())}",
                    "role": "formula_output",
                    "source_ids": data.get("source_map", {}).get("historical_financials", []),
                    "scenario": "selected",
                    "notes": "营运资本、固定资产、债务和权益滚动",
                },
                {
                    "field": "cash_flow_statement",
                    "sheet": "现金流量表",
                    "cell_or_range": f"A5:{get_column_letter(last_col)}{max(crows.values())}",
                    "role": "formula_output",
                    "source_ids": data.get("source_map", {}).get("historical_financials", []),
                    "scenario": "selected",
                    "notes": "现金流量表及期末现金滚动",
                },
                {
                    "field": "model_check_status",
                    "sheet": "检查",
                    "cell_or_range": "B2",
                    "role": "formula_check",
                    "source_ids": [],
                    "scenario": "selected",
                    "notes": "由明细检查状态汇总",
                },
            ],
        )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--support-dir", type=Path)
    args = parser.parse_args()
    data = json.loads(args.input.read_text(encoding="utf-8"))
    build(data, args.output, args.support_dir, args.input)
    print(json.dumps({"output": str(args.output), "support_dir": str(args.support_dir) if args.support_dir else None, "engine": "openpyxl", "sheets": ["封面", "假设", "产品明细", "利润表", "资产负债表", "现金流量表", "检查", "来源"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
