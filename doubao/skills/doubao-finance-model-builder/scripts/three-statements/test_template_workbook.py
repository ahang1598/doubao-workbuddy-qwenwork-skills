#!/usr/bin/env python3
"""Regression checks for the bundled three-statement example workbook."""

from __future__ import annotations

import unittest
from pathlib import Path

from openpyxl import load_workbook


TEMPLATE = Path(__file__).resolve().parents[2] / "assets" / "three-statements" / "three-statements-model-template.xlsx"


class ThreeStatementTemplateTests(unittest.TestCase):
    def test_template_has_required_sheets_and_formula_chain(self) -> None:
        workbook = load_workbook(TEMPLATE, data_only=False, read_only=False)
        required = {
            "封面", "数据来源", "假设依据", "经营驱动", "利润表",
            "资产负债表", "现金流量表", "债务与利息", "模型检查",
        }
        self.assertTrue(required.issubset(set(workbook.sheetnames)))
        self.assertEqual(workbook["利润表"]["D4"].data_type, "f")
        self.assertEqual(workbook["资产负债表"]["D4"].data_type, "f")
        self.assertEqual(workbook["现金流量表"]["D17"].data_type, "f")
        self.assertEqual(workbook["模型检查"]["F9"].data_type, "f")

    def test_template_cached_checks_pass(self) -> None:
        workbook = load_workbook(TEMPLATE, data_only=True, read_only=False)
        self.assertEqual(workbook["模型检查"]["F9"].value, "PASS")
        self.assertEqual(
            [workbook["模型检查"][f"F{row}"].value for row in range(4, 9)],
            ["PASS", "PASS", "PASS", "PASS", "PASS"],
        )


if __name__ == "__main__":
    unittest.main(verbosity=2)
