#!/usr/bin/env python3
"""Portable direct audit for the three-statement formula workbook."""

from __future__ import annotations

import argparse
import json
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import sys

COMMON = Path(__file__).resolve().parents[1] / "common"
sys.path.insert(0, str(COMMON))

from portable_workbook_audit import (  # noqa: E402
    audit_contract,
    formula_snapshot,
    formula_text,
    recalculate_with_libreoffice,
)


REQUIRED_SHEETS = ["封面", "假设", "利润表", "资产负债表", "现金流量表", "检查", "来源"]


def finite(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("output")
    parser.add_argument("--recalculate", choices=["auto", "required", "off"], default="auto")
    parser.add_argument("--timeout", type=int, default=60)
    args = parser.parse_args()

    contract = {
        "workflow": "three_statements",
        "required_sheets": REQUIRED_SHEETS,
        "formula_ranges": [],
        "key_outputs": [],
        "check_status": {
            "sheet": "检查",
            "cell": "B2",
            "pass_value": "PASS",
            "require_formula": True,
        },
    }
    result = audit_contract(
        args.workbook,
        contract,
        recalculate=args.recalculate,
        timeout=args.timeout,
    )
    errors = result["errors"]
    warnings = result["warnings"]
    metrics = result["metrics"]
    formulas_all = formula_snapshot(args.workbook)
    formulas: dict[str, dict[str, str]] = {}
    for full_address, formula in formulas_all.items():
        sheet_name, coordinate = full_address.rsplit("!", 1)
        formulas.setdefault(sheet_name, {})[coordinate] = formula

    recalculation = (
        recalculate_with_libreoffice(args.workbook, timeout=args.timeout)
        if args.recalculate != "off"
        else None
    )
    cached_path = recalculation.workbook_path if recalculation and recalculation.workbook_path else Path(args.workbook)
    cached = load_workbook(cached_path, data_only=True, read_only=False)
    original = load_workbook(args.workbook, data_only=False, read_only=False)
    duplicate_keys = 0
    try:
        for sheet in original.worksheets:
            seen = set()
            for row in range(1, sheet.max_row + 1):
                key = sheet.cell(row=row, column=1).value
                if not isinstance(key, str) or not re.match(r"^(is|bs|cfs|product|total)\.", key):
                    continue
                if key in seen:
                    duplicate_keys += 1
                    errors.append(f"重复语义键：{sheet.title}!{key}")
                seen.add(key)

        if "检查" in original.sheetnames and "检查" in cached.sheetnames:
            source = original["检查"]
            values = cached["检查"]
            overall_formula = formula_text(source["B2"].value)
            if not overall_formula or "COUNTIF" not in overall_formula.upper():
                errors.append("检查页总状态不是引用明细状态的公式")
            failed_rows = 0
            non_formula_status_rows = 0
            for row in range(4, max(source.max_row, values.max_row) + 1):
                label = values.cell(row=row, column=1).value
                if not label:
                    continue
                actual = values.cell(row=row, column=3).value
                expected = values.cell(row=row, column=4).value
                tolerance = values.cell(row=row, column=6).value
                status = values.cell(row=row, column=7).value
                status_formula = formula_text(source.cell(row=row, column=7).value)
                normalized_status_formula = (status_formula or "").replace("$", "")
                equality_check = re.search(
                    rf"ABS\(E{row}\)<=F{row}",
                    normalized_status_formula,
                    flags=re.IGNORECASE,
                )
                numeric_failure = (
                    bool(equality_check)
                    and
                    finite(actual)
                    and finite(expected)
                    and finite(tolerance)
                    and abs(actual - expected) > tolerance
                )
                invalid_numeric = not (finite(actual) and finite(expected) and finite(tolerance))
                if status == "失败" or str(status or "").startswith("#") or numeric_failure or invalid_numeric:
                    failed_rows += 1
                    errors.append(f"检查失败：{label}")
                if status_formula is None:
                    non_formula_status_rows += 1
            if non_formula_status_rows:
                errors.append(f"检查页存在{non_formula_status_rows}个静态状态单元格")
            if values["B2"].value == "PASS" and failed_rows:
                errors.append("检查页出现假PASS")
            metrics["failed_check_rows"] = failed_rows

        if "产品明细" in original.sheetnames:
            sheet = original["产品明细"]
            row_map = {}
            for row in range(1, sheet.max_row + 1):
                key = sheet.cell(row=row, column=1).value
                if isinstance(key, str):
                    row_map[key] = row
            product_ids = {
                match.group(1)
                for key in row_map
                if (match := re.match(r"^product\.([^.]+)\.calculated_revenue$", key))
            }
            if not product_ids:
                errors.append("量价模式缺少产品语义行")
            required = [
                "sales_volume",
                "quantity_multiplier",
                "realized_price",
                "price_multiplier",
                "price_fx_to_model_currency",
                "calculated_revenue",
                "unit_cost",
                "cost_multiplier",
                "cost_fx_to_model_currency",
                "calculated_cogs",
            ]
            for product_id in sorted(product_ids):
                rows = {}
                for metric in required:
                    rows[metric] = row_map.get(f"product.{product_id}.{metric}")
                    if not rows[metric]:
                        errors.append(f"产品{product_id}缺少{metric}语义行")
                if any(not rows[metric] for metric in required):
                    continue
                for col_index in range(4, sheet.max_column + 1):
                    col = get_column_letter(col_index)
                    revenue_formula = formulas.get("产品明细", {}).get(
                        f"{col}{rows['calculated_revenue']}", ""
                    )
                    cost_formula = formulas.get("产品明细", {}).get(
                        f"{col}{rows['calculated_cogs']}", ""
                    )
                    revenue_tokens = [
                        f"{col}{rows['sales_volume']}",
                        f"{col}{rows['quantity_multiplier']}",
                        f"{col}{rows['realized_price']}",
                        f"{col}{rows['price_multiplier']}",
                        f"{col}{rows['price_fx_to_model_currency']}",
                        "'假设'!$B$4",
                    ]
                    cost_tokens = [
                        f"{col}{rows['sales_volume']}",
                        f"{col}{rows['quantity_multiplier']}",
                        f"{col}{rows['unit_cost']}",
                        f"{col}{rows['cost_multiplier']}",
                        f"{col}{rows['cost_fx_to_model_currency']}",
                        "'假设'!$B$4",
                    ]
                    if not all(token in revenue_formula for token in revenue_tokens):
                        errors.append(f"产品{product_id}的{col}列收入单位公式不完整")
                    if not all(token in cost_formula for token in cost_tokens):
                        errors.append(f"产品{product_id}的{col}列成本单位公式不完整")
            metrics["product_count"] = len(product_ids)

        if "现金流量表" in original.sheetnames and "现金流量表" in cached.sheetnames:
            source = original["现金流量表"]
            values = cached["现金流量表"]
            dividends_row = None
            for row in range(1, source.max_row + 1):
                if source.cell(row=row, column=1).value == "cfs.dividends":
                    dividends_row = row
                    break
                if source.cell(row=row, column=1).value == "支付股利" or source.cell(row=row, column=2).value == "支付股利":
                    dividends_row = row
                    break
            if dividends_row:
                for col in range(2, values.max_column + 1):
                    value = values.cell(row=dividends_row, column=col).value
                    if finite(value) and value > 1e-9:
                        errors.append(f"支付股利为正现金流：现金流量表第{dividends_row}行第{col}列")
            else:
                errors.append("现金流量表缺少可识别的支付股利行")
    finally:
        cached.close()
        original.close()
        if recalculation and recalculation.temp_root:
            shutil.rmtree(recalculation.temp_root, ignore_errors=True)

    metrics["duplicate_semantic_key_count"] = duplicate_keys
    result["status"] = "FAIL" if errors else ("INCOMPLETE" if warnings else "PASS")
    Path(args.output).write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    raise SystemExit(0 if result["status"] == "PASS" else 1)


if __name__ == "__main__":
    main()
